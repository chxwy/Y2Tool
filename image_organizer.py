
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinterdnd2 as tkdnd
import popup_utils
from PIL import Image, ImageTk
from smart_upscale_plan_dialog import show_smart_upscale_plan_dialog
import os
import sys
import shutil
import json
import threading
from pathlib import Path
import re
import warnings
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
import time
import pickle
import hashlib

# --- 调试输出过滤器 ---
DEBUG = False  # 切换为 True 可输出全部调试信息
# 需要过滤的 emoji（不包含❌，保留错误提示）
# 扩展过滤列表，加入 🧩 和 📊 以及常见调试前缀
_FILTER_EMOJIS = {'🔍','📏','✅','🎯','🏆','📦','🔧','ℹ️','🚫','🧩','📊'}
# 保留输出的关键字
_PROTECT_KEYWORDS = ('合并A列', '合并D列', 'A列和D列', '合并A列和D列', 'Exception', '错误', '失败')

import builtins as _builtins
_original_print = _builtins.print

def _filtered_print(*args, **kwargs):
    """过滤掉无用调试信息，保留关键日志/异常"""
    if not DEBUG and args and isinstance(args[0], str):
        first = str(args[0])
        # 如果包含需保护关键字，则直接输出
        if any(key in first for key in _PROTECT_KEYWORDS):
            return _original_print(*args, **kwargs)
        # 否则若包含过滤emoji，则跳过
        if any(em in first for em in _FILTER_EMOJIS):
            return
    _original_print(*args, **kwargs)

_builtins.print = _filtered_print

# 全局替换messagebox为popup_utils
messagebox.showinfo = popup_utils.showinfo
messagebox.showwarning = popup_utils.showwarning
messagebox.showerror = popup_utils.showerror
messagebox.askyesno = popup_utils.askyesno
messagebox.askyesnocancel = popup_utils.askyesnocancel
messagebox.askquestion = popup_utils.askquestion
from copy import copy

def _import_pandas():
    """延迟导入pandas"""
    global pd
    if 'pd' not in globals():
        import pandas as pd
    return pd

def _import_openpyxl():
    """延迟导入openpyxl相关模块"""
    global openpyxl, Font, PatternFill, Alignment, OpenpyxlImage, cm_to_EMU, get_column_letter
    if 'openpyxl' not in globals():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.utils.units import cm_to_EMU
        from openpyxl.utils import get_column_letter
    return openpyxl, Font, PatternFill, Alignment, OpenpyxlImage, cm_to_EMU

def _import_requests():
    """延迟导入requests相关模块"""
    global requests, base64, BytesIO, urllib
    if 'requests' not in globals():
        import requests
        import base64
        from io import BytesIO
        import urllib
    return requests, base64, BytesIO, urllib

# 禁用PIL的DecompressionBombWarning警告
warnings.filterwarnings("ignore", category=Image.DecompressionBombWarning)
# 增加PIL图片大小限制
Image.MAX_IMAGE_PIXELS = None

# 支持的图片格式（全局变量，用于多进程）
SUPPORTED_FORMATS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif'}

def get_app_directory():
    """获取应用程序所在目录"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的可执行文件
        return os.path.dirname(sys.executable)
    else:
        # 如果是Python脚本
        return os.path.dirname(os.path.abspath(__file__))

class SimpleConfigManager:
    """简单的配置管理器"""
    def __init__(self):
        self.config_file = os.path.join(get_app_directory(), "config.json")
        
    def load_config(self, section, default=None):
        """加载配置"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get(section, default if default is not None else {})
            return default if default is not None else {}
        except Exception as e:
            print(f"加载配置失败: {e}")
            return default if default is not None else {}
    
    def save_config(self, section, data, silent=False):
        """保存配置 - 使用原子性写入"""
        print(f"🔧 save_config 被调用: section={section}, data={data}, silent={silent}")
        try:
            config = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            print(f"📖 当前配置文件内容长度: {len(config)} 项")
            config[section] = data
            print(f"💾 准备保存到文件: {self.config_file}")
            
            # 使用临时文件实现原子性写入
            temp_file = self.config_file + '.tmp'
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            # 原子性重命名
            if os.path.exists(self.config_file):
                os.replace(temp_file, self.config_file)
            else:
                os.rename(temp_file, self.config_file)
                
            print(f"✅ 配置已成功保存到 {section}")
            if not silent:
                print(f"配置已保存到 {section}")
        except Exception as e:
            # 清理临时文件
            temp_file = self.config_file + '.tmp'
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except (OSError, IOError):
                    pass
            print(f"❌ 保存配置失败: {e}")
            import traceback
            traceback.print_exc()
            if not silent:
                print(f"保存配置失败: {e}")

class BigJPGUpscaler:
    """BigJPG图片放大API封装类"""
    
    def __init__(self, api_key, base_url="https://bigjpg.com/api/task/"):
        self.api_key = api_key
        self.base_url = base_url
        
        # 延迟导入requests模块
        requests, base64, BytesIO, urllib = _import_requests()
        
        self.session = requests.Session()
        # 设置全局请求头，包括下载时需要的浏览器头部
        self.session.headers.update({
            'X-API-KEY': api_key,
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })
        # 尝试从配置加载
        config_manager = SimpleConfigManager()
        self.imgbb_api_key = config_manager.load_config('imgbb_api_key', "5d170edec64cef15aefa2540d93724cc")
        self.imgbb_upload_url = "https://api.imgbb.com/1/upload"
        
        # 重试配置
        self.max_retries = 3
        self.retry_delay = 2  # 秒
        self.timeout = 30  # 请求超时时间
    
    def _make_request(self, method, url, **kwargs):
        """带重试机制的请求方法，使用指数退避和抖动"""
        import random
        kwargs.setdefault('timeout', self.timeout)
        
        for attempt in range(self.max_retries):
            try:
                if method.upper() == 'POST':
                    response = self.session.post(url, **kwargs)
                else:
                    response = self.session.get(url, **kwargs)
                
                # 检查HTTP状态码
                if response.status_code == 429:  # 请求过于频繁
                    if attempt < self.max_retries - 1:
                        # 指数退避 + 抖动
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.5)  # 10%-50%的抖动
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    else:
                        return {'error': 'API请求过于频繁，请稍后再试'}
                
                if response.status_code == 401:
                    return {'error': 'API密钥无效，请检查密钥是否正确'}
                
                if response.status_code == 402:
                    return {'error': '账户余额不足，请充值后再试'}
                
                if response.status_code >= 500:  # 服务器错误
                    if attempt < self.max_retries - 1:
                        # 指数退避 + 抖动
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.5)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    else:
                        return {'error': f'服务器错误 ({response.status_code})，请稍后再试'}
                
                # 解析响应
                response_data = response.json()
                
                return response_data
                
            except requests.exceptions.Timeout:
                if attempt < self.max_retries - 1:
                    # 指数退避 + 抖动
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                else:
                    return {'error': '请求超时，请检查网络连接'}
            
            except requests.exceptions.ConnectionError:
                if attempt < self.max_retries - 1:
                    # 指数退避 + 抖动
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                else:
                    return {'error': '网络连接失败，请检查网络设置'}
            
            except requests.exceptions.RequestException as e:
                if attempt < self.max_retries - 1:
                    # 指数退避 + 抖动
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                else:
                    return {'error': f'网络请求失败: {str(e)}'}
            
            except (json.JSONDecodeError, KeyError, ValueError) as e:
                return {'error': f'处理响应数据时出错: {str(e)}'}
            
            except Exception as e:
                return {'error': f'未知错误: {str(e)}'}
        
        return {'error': '请求失败，已达到最大重试次数'}
    
    def upload_to_imgbb(self, image_path, max_retries=3):
        """
        上传图片到ImgBB图床
        
        Args:
            image_path: 图片文件路径
            max_retries: 最大重试次数
        
        Returns:
            dict: 包含上传结果的字典，成功时包含'url'字段，失败时包含'error'字段
        """
        # 检查文件是否存在
        if not os.path.exists(image_path):
            return {'error': f'文件不存在: {image_path}'}
        
        # 检查文件大小（ImgBB限制32MB）
        file_size = os.path.getsize(image_path)
        if file_size > 32 * 1024 * 1024:  # 32MB
            return {'error': f'文件过大: {file_size / (1024*1024):.1f}MB，超过32MB限制'}
        
        for attempt in range(max_retries):
            try:
                # 准备上传数据
                with open(image_path, 'rb') as f:
                    files = {'image': f}
                    data = {'key': self.imgbb_api_key}
                    
                    # 发送上传请求 - 使用独立的requests而不是session，避免BigJPG的headers干扰
                    response = requests.post(
                        self.imgbb_upload_url,
                        files=files,
                        data=data,
                        timeout=30
                    )
                
                if response.status_code == 200:
                    result = response.json()
                    if result.get('success') and 'data' in result:
                        # ImgBB返回的URL在data.url字段中
                        image_url = result['data']['url']
                        return {'url': image_url, 'result': result}
                    else:
                        error_msg = result.get('error', {}).get('message', '上传失败')
                        if attempt < max_retries - 1:
                            print(f"ImgBB上传失败（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                            # 指数退避 + 抖动
                            import random
                            base_delay = 2 ** attempt
                            jitter = random.uniform(0.1, 0.5)
                            delay = base_delay * (1 + jitter)
                            time.sleep(delay)
                            continue
                        return {'error': f'ImgBB上传失败: {error_msg}'}
                else:
                    error_msg = f'HTTP {response.status_code}: {response.text}'
                    if attempt < max_retries - 1:
                        print(f"ImgBB上传失败（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                        # 指数退避 + 抖动
                        import random
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.5)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    return {'error': error_msg}
                    
            except requests.exceptions.Timeout:
                error_msg = '请求超时'
                if attempt < max_retries - 1:
                    print(f"ImgBB上传超时（尝试 {attempt + 1}/{max_retries}），正在重试...")
                    # 指数退避 + 抖动
                    import random
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                return {'error': error_msg}
            except requests.exceptions.ConnectionError:
                error_msg = '网络连接失败'
                if attempt < max_retries - 1:
                    print(f"ImgBB上传连接失败（尝试 {attempt + 1}/{max_retries}），正在重试...")
                    # 指数退避 + 抖动
                    import random
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                return {'error': error_msg}
            except requests.exceptions.RequestException as e:
                error_msg = f'网络请求失败: {str(e)}'
                if attempt < max_retries - 1:
                    print(f"ImgBB上传网络错误（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                    # 指数退避 + 抖动
                    import random
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                return {'error': error_msg}
            except (json.JSONDecodeError, KeyError) as e:
                error_msg = f'ImgBB响应数据解析失败: {str(e)}'
                return {'error': error_msg}
            except (IOError, OSError) as e:
                error_msg = f'文件读取失败: {str(e)}'
                print(f"[ImgBB错误] 文件操作失败: {error_msg}")
                return {'error': error_msg}
            except Exception as e:
                error_msg = f'ImgBB上传未知错误: {str(e)}'
                print(f"[ImgBB错误] 未知异常: {error_msg} (尝试 {attempt + 1}/{max_retries})")
                if attempt < max_retries - 1:
                    print(f"ImgBB上传出错（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                    # 指数退避 + 抖动
                    import random
                    base_delay = 2 ** attempt
                    jitter = random.uniform(0.1, 0.3)
                    delay = base_delay * (1 + jitter)
                    time.sleep(delay)
                    continue
                return {'error': error_msg}
        
        final_error = f'上传失败，已重试{max_retries}次'
        print(f"[ImgBB错误] {final_error}")
        return {'error': final_error}

    def upload_image(self, image_path, style='art', noise='3', x2='1', max_retries=3):
        """
        上传图片进行放大处理
        
        Args:
            image_path: 图片文件路径
            style: 'art' (卡通插画) 或 'photo' (照片)
            noise: '-1'(无), '0'(低), '1'(中), '2'(高), '3'(最高)
            x2: '1'(2x), '2'(4x), '3'(8x), '4'(16x)
            max_retries: BigJPG API最大重试次数
        
        Returns:
            dict: API响应结果，包含任务ID等信息
        """
        try:
            # 检查文件是否存在
            if not os.path.exists(image_path):
                return {'error': f'文件不存在: {image_path}'}
            
            # 第一步：上传图片到图床获取公开URL
            print(f"正在上传图片到图床: {os.path.basename(image_path)}")
            upload_result = self.upload_to_imgbb(image_path)
            
            if 'error' in upload_result:
                return {'error': f'ImgBB 图床上传失败: {upload_result["error"]}'}
            
            image_url = upload_result['url']
            print(f"图床上传成功，URL: {image_url}")
            
            # 第二步：使用图床URL调用BigJPG API（带重试机制）
            print("正在调用BigJPG API进行图片处理...")
            
            for attempt in range(max_retries):
                try:
                    # 准备请求数据 - 使用官方文档的JSON格式
                    data = {
                        'style': style,
                        'noise': noise,
                        'x2': x2,
                        'input': image_url,  # 使用图床URL而不是base64数据
                        'file_name': os.path.basename(image_path)  # 添加文件名参数
                    }
                    
                    # 发送请求
                    response = self.session.post(
                        self.base_url,
                        headers={
                            'X-API-KEY': self.api_key,
                            'Content-Type': 'application/json'
                        },
                        data=json.dumps(data),
                        timeout=30
                    )
                    
                    print(f"BigJPG API HTTP状态码: {response.status_code}")
                    print(f"BigJPG API响应内容: {response.text}")
                    
                    if response.status_code == 200:
                        try:
                            result = response.json()
                            print("BigJPG API调用成功")
                            print(f"BigJPG API响应解析: {result}")
                            
                            # 检查响应中是否包含错误状态
                            if 'status' in result and result['status'] == 'param_error':
                                return {
                                    'error': f'BigJPG API参数错误: {result}',
                                    'api_response': result
                                }
                            
                            return result
                        except json.JSONDecodeError as e:
                            return {
                                'error': f'BigJPG API响应解析失败: {e}',
                                'raw_response': response.text
                            }
                    else:
                        error_msg = f'BigJPG API调用失败 - HTTP {response.status_code}: {response.text}'
                        if attempt < max_retries - 1:
                            print(f"BigJPG API调用失败（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                            # 指数退避 + 抖动
                            import random
                            base_delay = 2 ** attempt
                            jitter = random.uniform(0.1, 0.3)
                            delay = base_delay * (1 + jitter)
                            time.sleep(delay)
                            continue
                        return {
                            'error': error_msg,
                            'status_code': response.status_code
                        }
                        
                except requests.exceptions.Timeout:
                    error_msg = 'BigJPG API请求超时'
                    if attempt < max_retries - 1:
                        print(f"BigJPG API超时（尝试 {attempt + 1}/{max_retries}），正在重试...")
                        # 指数退避 + 抖动
                        import random
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.3)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    return {'error': error_msg}
                except requests.exceptions.ConnectionError:
                    error_msg = 'BigJPG API连接失败'
                    if attempt < max_retries - 1:
                        print(f"BigJPG API连接失败（尝试 {attempt + 1}/{max_retries}），正在重试...")
                        # 指数退避 + 抖动
                        import random
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.3)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    return {'error': error_msg}
                except requests.exceptions.RequestException as e:
                    error_msg = f'BigJPG API网络请求失败: {str(e)}'
                    if attempt < max_retries - 1:
                        print(f"BigJPG API网络错误（尝试 {attempt + 1}/{max_retries}）: {error_msg}，正在重试...")
                        # 指数退避 + 抖动
                        import random
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.3)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    return {'error': error_msg}
            
            return {'error': f'BigJPG API调用失败，已重试{max_retries}次'}
            
        except FileNotFoundError:
            return {'error': f'文件不存在: {image_path}'}
        except PermissionError:
            return {'error': f'无权限访问文件: {image_path}'}
        except Exception as e:
            return {'error': f'上传图片时出错: {str(e)}'}
    

    
    def check_task_status(self, task_ids):
        """
        查询任务状态
        
        Args:
            task_ids: 任务ID列表或单个任务ID
        
        Returns:
            dict: 任务状态信息
        """
        try:
            if isinstance(task_ids, list):
                ids_str = ','.join(task_ids)
            else:
                ids_str = str(task_ids)
            
            url = f"{self.base_url}{ids_str}"
            return self._make_request('GET', url)
            
        except Exception as e:
            return {'error': f'查询任务状态时出错: {str(e)}'}
    
    # 兼容旧代码接口
    def check_status(self, task_ids):
        """
        兼容旧版本调用，实际调用 check_task_status
        Args:
            task_ids: 任务ID列表或单个任务ID
        Returns:
            dict: 任务状态信息
        """
        return self.check_task_status(task_ids)

    def retry_task(self, task_ids):
        """
        重试任务
        
        Args:
            task_ids: 任务ID列表或单个任务ID
        
        Returns:
            dict: 重试结果
        """
        try:
            if isinstance(task_ids, list):
                ids_str = ','.join(task_ids)
            else:
                ids_str = str(task_ids)
            
            url = f"{self.base_url}{ids_str}"
            return self._make_request('POST', url)
            
        except Exception as e:
            return {'error': f'重试任务时出错: {str(e)}'}
    
    def download_result(self, download_url, save_path, progress_callback=None):
        """
        下载处理完成的图片
        
        Args:
            download_url: 下载链接
            save_path: 保存路径
            progress_callback: 进度回调函数，接收(current, total)参数
        
        Returns:
            bool: 下载是否成功
        """
        try:
            # 创建保存目录
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            # 使用带重试机制下载文件
            for attempt in range(self.max_retries):
                try:
                    # 为下载请求添加特定的请求头，覆盖session的默认头部
                    download_headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                        'Referer': 'https://bigjpg.com/',
                        'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
                        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Connection': 'keep-alive',
                        'Upgrade-Insecure-Requests': '1',
                        'Sec-Fetch-Dest': 'image',
                        'Sec-Fetch-Mode': 'no-cors',
                        'Sec-Fetch-Site': 'cross-site'
                    }
                    
                    # 添加随机延迟，避免被识别为机器人
                    import random
                    time.sleep(random.uniform(0.5, 1.6))
                    
                    response = self.session.get(download_url, stream=True, timeout=60, headers=download_headers)
                    response.raise_for_status()
                    
                    # 获取文件总大小
                    total_size = int(response.headers.get('content-length', 0))
                    downloaded_size = 0
                    
                    with open(save_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                                downloaded_size += len(chunk)
                                
                                # 调用进度回调
                                if progress_callback and total_size > 0:
                                    progress_callback(downloaded_size, total_size)
                    
                    return True
                    
                except requests.exceptions.RequestException as e:
                    if attempt < self.max_retries - 1:
                        # 指数退避 + 抖动
                        import random
                        base_delay = 2 ** attempt
                        jitter = random.uniform(0.1, 0.3)
                        delay = base_delay * (1 + jitter)
                        time.sleep(delay)
                        continue
                    else:
                        raise requests.exceptions.RequestException(f'下载失败: {str(e)}')
            
        except requests.exceptions.RequestException as e:
            print(f"网络请求失败: {e}")
            print(f"下载URL: {download_url}")
            print(f"保存路径: {save_path}")
            return False
        except (IOError, OSError) as e:
            print(f"文件操作失败: {e}")
            print(f"保存路径: {save_path}")
            print(f"目录是否存在: {os.path.exists(os.path.dirname(save_path))}")
            print(f"目录权限检查: {os.access(os.path.dirname(save_path), os.W_OK) if os.path.exists(os.path.dirname(save_path)) else 'N/A'}")
            return False
        except Exception as e:
            print(f"下载文件时出现未知错误: {e}")
            print(f"错误类型: {type(e).__name__}")
            print(f"下载URL: {download_url}")
            print(f"保存路径: {save_path}")
            return False
    
    def download_and_save_image(self, download_url, original_image_path):
        """
        下载处理完成的图片并保存到指定位置
        
        Args:
            download_url: 下载链接
            original_image_path: 原始图片路径，用于生成保存路径
        
        Returns:
            bool: 下载是否成功
        """
        try:
            # 生成保存路径（在原文件名后添加_upscaled）
            dir_path = os.path.dirname(original_image_path)
            filename = os.path.basename(original_image_path)
            name, ext = os.path.splitext(filename)
            save_path = os.path.join(dir_path, f"{name}_upscaled{ext}")
            
            # 使用现有的下载方法
            return self.download_result(download_url, save_path)
            
        except Exception as e:
            print(f"保存图片时出错: {e}")
            return False



def parallel_search_files(file_list, search_terms):
    """优化的并行搜索文件工作函数"""
    found_files = []
    
    # 预编译搜索词为小写，避免重复转换
    lower_search_terms = [term.lower() for term in search_terms]
    
    for file_path in file_list:
        filename_without_ext = os.path.splitext(os.path.basename(file_path))[0].lower()
        
        # 使用更高效的字符串匹配
        for term in lower_search_terms:
            if filename_without_ext.startswith(term):
                found_files.append(file_path)
                break  # 找到匹配后立即跳出，避免重复检查
    
    return found_files

class ImageOrganizerApp:
    BASIC_SIZE_TOLERANCE_CM = 1  # 基础尺寸匹配容差(厘米)

    def __init__(self, root):
        self.root = root
        self.root.title("Y2订单辅助工具1.9")
        self.root.geometry("1320x800")
        self.root.minsize(1020, 600)
        
        # 设置现代化窗口背景色
        self.root.configure(bg='#FFFFFF')
        
        # 配置文件路径
        app_dir = get_app_directory()
        self.config_file = os.path.join(app_dir, "config.json")
        self.icon_path = "logo.ico"  # 图标文件路径
        
        # 应用状态
        self.source_folder = ""
        self.search_results = []
        self.selected_images = set()
        self.last_folder_name = ""  # 上次使用的文件夹名称
        self.last_date_check = None  # 上次日期检查的日期
        
        # 支持的图片格式
        self.supported_formats = SUPPORTED_FORMATS
        
        # 网格布局参数（响应式）- 提前初始化
        self.grid_columns = 6  # 默认每行显示6个图片
        self.min_item_width = 160  # 每个图片项的最小宽度
        self.current_row = 0
        self.current_col = 0
        
        # 性能优化相关
        self.file_cache = {}  # 文件缓存
        self.file_cache_max_size = 100  # 文件缓存最大条目数
        self.cache_file = "file_cache.pkl"  # 缓存文件
        self.last_scan_time = 0  # 上次扫描时间
        self.cpu_count = max(1, mp.cpu_count() - 1)  # 使用CPU核心数-1
        
        # 搜索结果缓存（1.6版本的优化）
        self.search_cache = {}  # 搜索结果缓存
        self.search_cache_max_size = 30  # 最大缓存条目数（增加缓存容量）
        
        # UI优化相关线程池（1.6版本的优化）
        from concurrent.futures import ThreadPoolExecutor
        self.thumbnail_executor = ThreadPoolExecutor(max_workers=min(8, self.cpu_count * 2), 
                                                   thread_name_prefix="thumbnail")
        self.info_executor = ThreadPoolExecutor(max_workers=min(4, self.cpu_count), 
                                              thread_name_prefix="fileinfo")
        self.search_executor = ThreadPoolExecutor(max_workers=min(6, self.cpu_count * 2), 
                                                thread_name_prefix="search")
        
        # Excel处理相关
        self.excel_data = None  # 存储Excel数据
        self.size_mapping = {}  # 尺寸映射规则，只从config.json加载用户预设内容
        self.sku_column = "SKU"  # SKU列名
        self.size_column = "尺寸"  # 尺寸列名
        
        # 记忆功能相关
        self.select_keyword_history = []  # 勾选包含的历史记录
        self.hide_keyword_history = []    # 隐藏不包含的历史记录
        self.max_history_count = 10       # 最大历史记录数量
        
        # 导出路径配置（默认为桌面）
        self.excel_export_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.image_export_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.upscale_export_path = ""  # 高清图片保存路径，空字符串表示使用源文件位置
        
        # 加工方式配置相关 - 初始化为空列表
        self.current_processing_entries = []
        
        # BigJPG图片放大相关配置
        self.bigjpg_api_key = ""  # API密钥，初始为空，从配置文件加载
        self.bigjpg_base_url = "https://bigjpg.com/api/task/"
        self.imgbb_api_key = "5d170edec64cef15aefa2540d93724cc"  # ImgBB图床API密钥
        self.upscale_tasks = {}  # 存储放大任务信息 {task_id: {file_path, original_name, status}}
        
        # 全局面单名称弹窗管理
        self.waybill_dialog = None  # 全局面单名称弹窗实例
        self.upscale_progress_window = None  # 进度窗口引用
        
        # 高清处理配置记忆功能
        self.upscale_config = {
            'style': 'art',    # 图片类型：'art'(卡通/插画) 或 'photo'(照片)
            'noise': '-1',      # 降噪程度：'-1'(无), '0'(低), '1'(中), '2'(高), '3'(最高)
            'x4': '2',         # 放大倍数：'1'(2x), '2'(4x), '3'(8x), '4'(16x)
            'rename_rule': '原文件名（高清）',  # 重命名规则，默认为"原文件名（高清）"
            # 移除超时配置 - 用户手动控制处理过程
        }
        
        # 智能高清处理配置
        self.smart_upscale_config = {
            'target_width': 8000,      # 目标宽度
            'target_height': 8000,     # 目标高度
            'skip_qualified': True,    # 跳过已达标图片
            'enabled': True            # 启用智能模式
        }
        
        # 延迟初始化标志 - 优化启动速度
        self._ui_initialized = False
        self._cache_loaded = False
        
        # 队列处理相关属性 - 解决多表格对话框重叠问题
        self._file_processing_queue = []  # 文件处理队列
        self._is_processing_queue = False  # 是否正在处理队列
        self._current_dialog_active = False  # 当前是否有活跃对话框
        
        # 配置管理器 - 简单的配置管理
        self.config_manager = SimpleConfigManager()
        
        # 加载配置（轻量级操作）
        self.load_config()
        
        # 创建界面（延迟加载重型组件）
        self.create_widgets()
        
        # 启用拖拽
        self.setup_drag_drop()
        
        # 延迟加载重型操作
        self.root.after(100, self._delayed_initialization)
        
    def _delayed_initialization(self):
        """延迟初始化重型操作，提升启动速度"""
        try:
            # 加载文件缓存（重型操作）
            if not self._cache_loaded:
                self.load_file_cache()
                self._cache_loaded = True
            
            # 自动加载加工方式配置（在界面创建完成后）
            self.auto_load_processing_config()
            
            # 标记UI初始化完成
            self._ui_initialized = True
            
            # 延迟检查更新（启动3秒后，静默模式）
            self.root.after(3000, self._auto_check_update)
            
        except Exception as e:
            print(f"延迟初始化时出错: {e}")
    
    def _auto_check_update(self):
        """自动检查更新（静默模式）"""
        try:
            from update_module import check_for_updates
            # 静默检查，只在有更新时提示
            check_for_updates(self.root, silent=True)
        except Exception as e:
            # 静默失败，不打扰用户
            print(f"自动检查更新失败: {e}")
        
    def setup_styles(self):
        """设置专业化UI样式"""
        import platform
        
        # 配置ttk样式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 根据操作系统选择合适的字体
        system = platform.system()
        if system == "Darwin":  # macOS
            default_font = ('SF Pro Display', 'Helvetica Neue', 'Arial')
            title_font = ('SF Pro Display', 18, 'bold')
            section_font = ('SF Pro Display', 12, 'bold')
            content_font = ('SF Pro Display', 14, 'bold')
            info_font = ('SF Pro Display', 10)
            path_font = ('SF Pro Display', 9)
            status_font = ('SF Pro Display', 10, 'bold')
            button_font = ('SF Pro Display', 11, 'bold')
            button_small_font = ('SF Pro Display', 10, 'bold')
        elif system == "Linux":
            default_font = ('Ubuntu', 'DejaVu Sans', 'Arial')
            title_font = ('Ubuntu', 18, 'bold')
            section_font = ('Ubuntu', 12, 'bold')
            content_font = ('Ubuntu', 14, 'bold')
            info_font = ('Ubuntu', 10)
            path_font = ('Ubuntu', 9)
            status_font = ('Ubuntu', 10, 'bold')
            button_font = ('Ubuntu', 11, 'bold')
            button_small_font = ('Ubuntu', 10, 'bold')
        else:  # Windows
            default_font = ('Microsoft YaHei UI', 'Segoe UI', 'Arial')
            title_font = ('Microsoft YaHei UI', 18, 'bold')
            section_font = ('Microsoft YaHei UI', 12, 'bold')
            content_font = ('Microsoft YaHei UI', 14, 'bold')
            info_font = ('Microsoft YaHei UI', 10)
            path_font = ('Microsoft YaHei UI', 9)
            status_font = ('Microsoft YaHei UI', 10, 'bold')
            button_font = ('Microsoft YaHei UI', 11, 'bold')
            button_small_font = ('Microsoft YaHei UI', 10, 'bold')
        
        # 专业主题配色
        style.configure('Main.TFrame', background='#FFFFFF')
        style.configure('Sidebar.TFrame', background='#F8F9FA', relief='solid', borderwidth=1)
        style.configure('Content.TFrame', background='#FFFFFF')
        style.configure('Card.TFrame', background='#FFFFFF', relief='solid', borderwidth=1)
        
        # 标题样式
        style.configure('AppTitle.TLabel', background='#F8F9FA', foreground='#212529', 
                       font=title_font)
        style.configure('SectionTitle.TLabel', background='#F8F9FA', foreground='#495057', 
                       font=section_font)
        style.configure('ContentTitle.TLabel', background='#FFFFFF', foreground='#212529', 
                       font=content_font)
        
        # 文本样式
        style.configure('Info.TLabel', background='#F8F9FA', foreground='#6C757D', 
                       font=info_font)
        style.configure('Path.TLabel', background='#F8F9FA', foreground='#0D6EFD', 
                       font=path_font)
        style.configure('Status.TLabel', background='#FFFFFF', foreground='#198754', 
                       font=status_font)
        
        # 输入框样式
        style.configure('Modern.TEntry', fieldbackground='#FFFFFF', 
                       borderwidth=1, relief='solid', padding=8)
        
        # 按钮样式 - 参考您应用的按钮设计
        style.configure('Primary.TButton', background='#0D6EFD', foreground='white',
                       font=button_font, padding=(20, 12), relief='flat')
        style.map('Primary.TButton',
                 background=[('active', '#0B5ED7'), ('pressed', '#0A58CA')])
        
        style.configure('Secondary.TButton', background='#6C757D', foreground='white',
                       font=button_small_font, padding=(15, 8), relief='flat')
        style.map('Secondary.TButton',
                 background=[('active', '#5C636A'), ('pressed', '#565E64')])
        
        style.configure('Success.TButton', background='#198754', foreground='white',
                       font=button_font, padding=(20, 12), relief='flat')
        style.map('Success.TButton',
                 background=[('active', '#157347'), ('pressed', '#146C43')])
        
        # 强调色按钮样式
        style.configure('Accent.TButton', background='#0D6EFD', foreground='white',
                       font=button_font, padding=(15, 8), relief='flat')
        style.map('Accent.TButton',
                 background=[('active', '#0B5ED7'), ('pressed', '#0A58CA')])
        
        # 滚动条样式
        style.configure('Modern.Vertical.TScrollbar', background='#DEE2E6', 
                        troughcolor='#F8F9FA', borderwidth=0, arrowcolor='#6C757D')
        style.configure('Modern.Horizontal.TScrollbar', background='#DEE2E6', 
                        troughcolor='#F8F9FA', borderwidth=0, arrowcolor='#6C757D')
        
    def create_widgets(self):
        """创建专业化左右分栏界面组件"""
        # 设置专业化主题样式
        self.setup_styles()
        
        # 主框架
        main_frame = ttk.Frame(self.root, padding="0", style="Main.TFrame")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # 左侧边栏 - 控制面板
        sidebar_frame = ttk.Frame(main_frame, style="Sidebar.TFrame", padding="20")
        sidebar_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 1))
        sidebar_frame.columnconfigure(0, weight=1)
        
        # 应用标题区域
        title_frame = ttk.Frame(sidebar_frame, style="Sidebar.TFrame")
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 30))
        title_frame.columnconfigure(0, weight=1)
        
        # 应用标题
        title_label = ttk.Label(title_frame, text="📁 Y2订单辅助工具", 
                               style="AppTitle.TLabel")
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        # 设置按钮（放在标题右侧）
        settings_btn = ttk.Button(title_frame, text="⚙", 
                                 style="Secondary.TButton",
                                 width=3,
                                 command=self.open_settings_page)
        settings_btn.grid(row=0, column=1, sticky=tk.E, padx=(10, 0))
        
        # 源文件夹区域
        folder_section = ttk.Label(sidebar_frame, text="📂 图库路径设置", 
                                  style="SectionTitle.TLabel")
        folder_section.grid(row=1, column=0, sticky=tk.W, pady=(0, 10))
        
        self.folder_var = tk.StringVar(value=self.source_folder or "拖拽文件夹到此处，或点击选择")
        self.folder_label = ttk.Label(sidebar_frame, textvariable=self.folder_var, 
                                     style="Path.TLabel", padding="10",
                                     background='#FFFFFF', relief='solid', borderwidth=1)
        self.folder_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 选择文件夹按钮
        self.refresh_btn = ttk.Button(sidebar_frame, text="📁 选择文件夹", 
                                     style="Secondary.TButton", command=self.refresh_folder)
        self.refresh_btn.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 30))
        
        # 搜索区域
        search_section = ttk.Label(sidebar_frame, text="🔍 搜索图片", 
                                  style="SectionTitle.TLabel")
        search_section.grid(row=4, column=0, sticky=tk.W, pady=(0, 10))
        
        # 搜索输入框
        self.search_var = tk.StringVar()
        self.search_entry = tk.Text(sidebar_frame, height=4, width=30,
                                   font=('Microsoft YaHei UI', 10),
                                   bg='#FFFFFF', fg='#212529',
                                   relief='solid', bd=1, padx=8, pady=8)
        self.search_entry.grid(row=5, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        self.search_entry.bind('<KeyRelease>', self.on_text_change)
        
        # 设置占位符文本
        placeholder_text = "输入图片编号，每行一个\n例如：\nCL6453-1\nCL6812-4"
        self.search_entry.insert('1.0', placeholder_text)
        self.search_entry.bind('<FocusIn>', self.on_text_focus_in)
        self.search_entry.bind('<FocusOut>', self.on_text_focus_out)
        self.search_entry.config(fg='#6C757D')
        
        # 去除序号尾缀按钮
        remove_suffix_btn = ttk.Button(sidebar_frame, text="🔧 去除序号尾缀", 
                                      style="Secondary.TButton", 
                                      command=self.remove_suffix)
        remove_suffix_btn.grid(row=6, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 搜索按钮
        search_btn = ttk.Button(sidebar_frame, text="🔍 开始搜索", 
                               style="Primary.TButton", 
                               command=self.start_search)
        search_btn.grid(row=7, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 进度条已移除，避免挤压下面的功能模块
        
        # 操作区域
        action_section = ttk.Label(sidebar_frame, text="📤 整理操作", 
                                  style="SectionTitle.TLabel")
        action_section.grid(row=9, column=0, sticky=tk.W, pady=(0, 10))
        
        ttk.Label(sidebar_frame, text="新文件夹名称：", style="Info.TLabel").grid(row=10, column=0, sticky=tk.W, pady=(0, 5))
        self.folder_name_var = tk.StringVar()
        self.folder_name_entry = ttk.Entry(sidebar_frame, textvariable=self.folder_name_var,
                                          style="Modern.TEntry", font=('Microsoft YaHei UI', 11))
        self.folder_name_entry.grid(row=11, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # 设置上次使用的文件夹名称
        if hasattr(self, 'last_folder_name') and self.last_folder_name:
            # 检测并更新文件夹名称中的日期
            updated_name = self.update_folder_name_date(self.last_folder_name)
            # 应用智能名称处理 - 检查今日重复并自动添加序号
            updated_name = self.get_smart_name(updated_name, 'image_packages')
            self.folder_name_var.set(updated_name)
        
        # 高清处理按钮
        self.upscale_btn = ttk.Button(sidebar_frame, text="✨ 高清处理", 
                                     style="Info.TButton", 
                                     command=self.start_upscale_process,
                                     state="disabled")
        self.upscale_btn.grid(row=12, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.organize_btn = ttk.Button(sidebar_frame, text="🚀 图片打包", 
                                      style="Success.TButton", command=self.organize_images)
        self.organize_btn.grid(row=13, column=0, sticky=(tk.W, tk.E))
        
        # 添加说明文字和设置按钮的容器
        note_frame = ttk.Frame(sidebar_frame, style="Content.TFrame")
        note_frame.grid(row=14, column=0, sticky=(tk.W, tk.E), pady=(8, 0))
        note_frame.columnconfigure(0, weight=1)
        
        # 添加说明文字
        note_label = ttk.Label(note_frame, text="注：图片仅复制并打包至指定位置，不会删除原图片", 
                              style="Info.TLabel", font=('Microsoft YaHei UI', 9))
        note_label.grid(row=0, column=0, sticky=tk.W)
        
        # 右侧内容区域 - 搜索结果
        content_frame = ttk.Frame(main_frame, style="Content.TFrame", padding="20")
        content_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(2, weight=1)  # 滚动区域占据剩余空间
        
        # 内容标题
        content_title = ttk.Label(content_frame, text="📋 搜索结果", 
                                 style="ContentTitle.TLabel")
        content_title.grid(row=0, column=0, sticky=tk.W, pady=(0, 20))
        
        # 固定的搜索结果统计框架（不滚动）
        self.stats_container = ttk.Frame(content_frame, style="Content.TFrame")
        self.stats_container.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        self.stats_container.columnconfigure(0, weight=1)
        
        # 创建滚动区域
        self.canvas = tk.Canvas(content_frame, bg='#FFFFFF', highlightthickness=0, 
                               relief='flat', bd=0)
        
        # 垂直滚动条
        self.v_scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=self.canvas.yview,
                                   style="Modern.Vertical.TScrollbar")
        
        # 横向滚动条
        self.h_scrollbar = ttk.Scrollbar(content_frame, orient="horizontal", command=self.canvas.xview,
                                   style="Modern.Horizontal.TScrollbar")
        
        self.scrollable_frame = ttk.Frame(self.canvas, style="Content.TFrame")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.on_scrollable_frame_configure(e)
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # 配置双向滚动
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)
        
        # 布局Canvas
        self.canvas.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 存储滚动条的显示状态
        self.v_scrollbar_visible = False
        self.h_scrollbar_visible = False
        
        # 绑定Canvas大小变化事件
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        
        # 绑定鼠标滚轮 - 优化版本
        self.bind_mousewheel(self.canvas)
        self.bind_mousewheel(self.root)
        self.bind_mousewheel(self.scrollable_frame)
        
        # 确保焦点能够接收滚轮事件
        self.canvas.focus_set()
        
        # 绑定键盘滚动
        self.root.bind("<Up>", lambda e: self.canvas.yview_scroll(-2, "units"))
        self.root.bind("<Down>", lambda e: self.canvas.yview_scroll(2, "units"))
        self.root.bind("<Left>", lambda e: self.canvas.xview_scroll(-2, "units"))   # 左箭头横向滚动
        self.root.bind("<Right>", lambda e: self.canvas.xview_scroll(2, "units"))   # 右箭头横向滚动
        self.root.bind("<Prior>", lambda e: self.canvas.yview_scroll(-10, "units"))  # Page Up
        self.root.bind("<Next>", lambda e: self.canvas.yview_scroll(10, "units"))   # Page Down
        self.root.bind("<Home>", lambda e: self.canvas.yview_moveto(0))              # Home
        self.root.bind("<End>", lambda e: self.canvas.yview_moveto(1))               # End
        
        # 绑定窗口大小变化事件
        self.root.bind('<Configure>', self.on_window_resize)
        
        # 绑定窗口关闭事件，保存配置
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 初始化网格列配置
        self.configure_grid_columns()
        
    def setup_drag_drop(self):
        """设置拖拽功能"""
        self.root.drop_target_register(tkdnd.DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.on_drop)
        
    def activate_window(self):
        """激活窗口到最前端"""
        try:
            # 将窗口置顶
            self.root.lift()
            # 获得焦点
            self.root.focus_force()
            # 确保窗口可见
            self.root.deiconify()
            # 在Windows系统上，额外设置窗口属性确保激活
            if hasattr(self.root, 'wm_attributes'):
                self.root.wm_attributes('-topmost', True)
                self.root.after(100, lambda: self.root.wm_attributes('-topmost', False))
        except Exception as e:
            print(f"激活窗口时出错: {e}")
        
    def on_drop(self, event):
        """处理拖拽事件"""
        files = self.root.tk.splitlist(event.data)
        if files:
            # 激活窗口到最前端
            self.activate_window()
            
            # 分离文件夹和Excel文件
            folders = [f for f in files if os.path.isdir(f)]
            excel_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls')) and os.path.isfile(f)]
            
            # 处理文件夹（只处理第一个）
            if folders:
                self.set_source_folder(folders[0])
            
            # 处理Excel文件
            if excel_files:
                if len(excel_files) == 1:
                    # 单个Excel文件，使用原有逻辑
                    self.process_excel_file(excel_files[0])
                else:
                    # 多个Excel文件，依次处理每个文件
                    self.process_multiple_excel_sequential(excel_files)
            elif not folders:
                messagebox.showwarning("警告", "请拖入文件夹或Excel文件！")
    
    def process_excel_file(self, excel_path):
        """处理Excel文件"""
        try:
            # 显示进度
            self.show_progress()
            self.update_progress(20, "正在读取Excel文件...")
            
            # 读取Excel文件
            pd = _import_pandas()
            df = pd.read_excel(excel_path)
            self.excel_data = df
            
            self.update_progress(50, "正在解析数据...")
            
            # 检查必要的列是否存在
            if self.sku_column not in df.columns:
                # 尝试自动识别SKU列
                possible_sku_cols = ['SKU', 'sku', 'Sku', '产品编号', '编号', '货号']
                found_sku_col = None
                for col in possible_sku_cols:
                    if col in df.columns:
                        found_sku_col = col
                        break
                
                if found_sku_col:
                    self.sku_column = found_sku_col
                else:
                    self.hide_progress()
                    messagebox.showerror("错误", f"未找到SKU列，请确保Excel中包含以下列名之一：{', '.join(possible_sku_cols)}")
                    return
            
            # 检查尺寸列
            if self.size_column not in df.columns:
                # 尝试自动识别尺寸列
                possible_size_cols = ['尺寸', '规格', 'size', 'Size', '加工尺寸']
                found_size_col = None
                for col in possible_size_cols:
                    if col in df.columns:
                        found_size_col = col
                        break
                
                if found_size_col:
                    self.size_column = found_size_col
            
            self.update_progress(80, "正在处理SKU数据...")
            
            # 获取所有SKU并自动填充到搜索框
            sku_list = df[self.sku_column].dropna().astype(str).tolist()
            if sku_list:
                # 立即进行SKU前缀分析
                analysis_result, processing_config = self.analyze_sku_prefixes(df)
                self.current_sku_analysis = (analysis_result, processing_config)
                
                # 处理SKU数据，去除序号尾缀等
                processed_sku_list = []
                for sku in sku_list:
                    sku_str = str(sku).strip()
                    # 检查是否包含换行符或空格分隔的多个SKU
                    if '\n' in sku_str or ' ' in sku_str:
                        # 先按换行符分割，再按空格分割
                        lines = sku_str.split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                if ' ' in line:
                                    # 如果行内还有空格，继续分割
                                    skus_in_line = line.split()
                                    for individual_sku in skus_in_line:
                                        individual_sku = individual_sku.strip()
                                        if individual_sku:
                                            # 去除数字尾缀
                                            processed_sku = re.sub(r'-\d+$', '', individual_sku)
                                            processed_sku_list.append(processed_sku)
                                else:
                                    # 去除数字尾缀
                                    processed_sku = re.sub(r'-\d+$', '', line)
                                    processed_sku_list.append(processed_sku)
                    else:
                        # 单个SKU的情况
                        processed_sku = re.sub(r'-\d+$', '', sku_str)
                        processed_sku_list.append(processed_sku)
                
                # 去重处理，保持顺序
                unique_skus = []
                seen = set()
                for sku in processed_sku_list:
                    if sku not in seen:
                        unique_skus.append(sku)
                        seen.add(sku)
                
                # 在多表格处理模式下，将SKU添加到汇总列表中
                if hasattr(self, '_is_multi_table_processing') and self._is_multi_table_processing:
                    # 初始化SKU收集列表（如果不存在）
                    if not hasattr(self, '_processed_files_skus'):
                        self._processed_files_skus = []
                    
                    # 添加到总的SKU列表中
                    self._processed_files_skus.extend(unique_skus)
                    print(f"📋 多表格处理模式：已收集文件 {os.path.basename(excel_path)} 的 {len(unique_skus)} 个SKU，等待最终汇总")
                else:
                    # 单表格模式：直接填入搜索框
                    sku_text = '\n'.join(unique_skus)
                    
                    # 清空搜索框并填入SKU数据
                    self.search_entry.delete('1.0', tk.END)
                    self.search_entry.insert('1.0', sku_text)
                    self.search_entry.config(fg='#212529')
                    
                    # 单表格处理完成，SKU已填入搜索框，用户可手动点击搜索按钮
                    print("✅ 单表格处理完成，SKU已填入搜索框，请手动点击搜索按钮开始搜索图片")
                
                # 检测是否有未匹配的SKU前缀
                unmatched_skus = self.check_unmatched_skus(sku_list)
                
                if unmatched_skus:
                    # 在多表格处理模式下，检查是否有活跃对话框
                    if hasattr(self, '_is_multi_table_processing') and self._is_multi_table_processing:
                        if hasattr(self, '_current_dialog_active') and self._current_dialog_active:
                            # 如果有活跃对话框，暂停处理，等待对话框关闭
                            return
                        # 标记对话框为活跃状态
                        self._current_dialog_active = True
                    
                    # 显示手动填写弹窗
                    self.hide_progress()
                    self.show_manual_processing_dialog(unmatched_skus, sku_list, df, excel_path)
                    return
                
                self.update_progress(100, "Excel处理完成！")
                
                # 如果有尺寸列，询问用户是否整理尺寸并另存桌面
                if self.size_column in df.columns:
                    result = self.show_size_processing_dialog(len(sku_list))
                    if result == "yes":
                        # 显示表格导出弹窗，传递SKU列表
                        custom_name = self.show_excel_export_dialog(sku_list)
                        if custom_name:  # 用户确认导出
                            # 记录导出设置，供后续文件使用
                            self.last_export_settings = {
                                'export': True,
                                'open_after_export': False,
                                'size_processing': 'yes'
                            }
                            self.process_and_export_excel(df, excel_path, custom_name)
                        else:
                            # 用户取消导出，记录设置
                            self.last_export_settings = {
                                'export': False,
                                'size_processing': 'yes'
                            }
                        # 用户选择处理尺寸后自动启动搜索（仅在非批量处理时且启用自动搜索时）
                        if not getattr(self, '_is_multi_table_processing', False) and getattr(self, 'auto_search_enabled', False):
                            self.root.after(500, self.start_search)
                    elif result == "process_and_open":
                        # 显示表格导出弹窗，传递SKU列表
                        custom_name = self.show_excel_export_dialog(sku_list)
                        if custom_name:  # 用户确认导出
                            # 记录导出设置，供后续文件使用
                            self.last_export_settings = {
                                'export': True,
                                'open_after_export': True,
                                'size_processing': 'process_and_open'
                            }
                            self.process_and_export_excel(df, excel_path, custom_name, True)
                        else:
                            # 用户取消导出，记录设置
                            self.last_export_settings = {
                                'export': False,
                                'size_processing': 'process_and_open'
                            }
                        # 用户选择处理并打开后自动启动搜索（仅在非批量处理时且启用自动搜索时）
                        if not getattr(self, '_is_multi_table_processing', False) and getattr(self, 'auto_search_enabled', False):
                            self.root.after(500, self.start_search)
                    else:
                        # 用户选择不处理尺寸，记录设置
                        self.last_export_settings = {
                            'export': False,
                            'size_processing': 'no'
                        }
                        # 用户选择不处理尺寸后自动启动搜索（仅在非批量处理时且启用自动搜索时）
                        if not getattr(self, '_is_multi_table_processing', False) and getattr(self, 'auto_search_enabled', False):
                            self.root.after(500, self.start_search)
                else:
                    messagebox.showinfo("成功", f"已成功处理Excel文件！\n找到 {len(sku_list)} 个SKU\n已自动填入搜索框")
            else:
                self.hide_progress()
                messagebox.showwarning("警告", "Excel文件中未找到有效的SKU数据")
                
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("错误", f"处理Excel文件时出错：{str(e)}")
        finally:
            self.hide_progress()
    
    def process_multiple_excel_merge_skus(self, excel_files):
        """处理多个Excel文件并合并所有SKU到搜索框（静默处理）"""
        try:
            self.show_progress()
            self.update_progress(10, "正在读取多个Excel文件...")
            
            pd = _import_pandas()
            all_skus = []
            processed_files = []
            
            for i, excel_path in enumerate(excel_files):
                try:
                    progress = 10 + (i * 50 // len(excel_files))
                    filename = os.path.basename(excel_path)
                    self.update_progress(progress, f"正在读取: {filename}")
                    
                    # 读取Excel文件
                    df = pd.read_excel(excel_path)
                    
                    # 检查SKU列
                    sku_column = self.sku_column
                    if sku_column not in df.columns:
                        # 尝试自动识别SKU列
                        possible_sku_cols = ['SKU', 'sku', 'Sku', '产品编号', '编号', '货号']
                        found_sku_col = None
                        for col in possible_sku_cols:
                            if col in df.columns:
                                found_sku_col = col
                                break
                        
                        if found_sku_col:
                            sku_column = found_sku_col
                        else:
                            # 静默跳过无SKU列的文件
                            continue
                    
                    # 提取SKU并处理包含多个SKU的单元格
                    raw_skus = df[sku_column].dropna().astype(str).tolist()
                    file_skus = []
                    for sku_cell in raw_skus:
                        # 处理包含换行符或空格分隔的多个SKU的单元格
                        if '\n' in sku_cell or ' ' in sku_cell:
                            # 先按换行符分割，再按空格分割
                            lines = sku_cell.split('\n')
                            for line in lines:
                                line = line.strip()
                                if line:
                                    if ' ' in line:
                                        # 如果行内还有空格，继续分割
                                        skus_in_line = line.split()
                                        for individual_sku in skus_in_line:
                                            individual_sku = individual_sku.strip()
                                            if individual_sku:
                                                # 去除数字尾缀
                                                processed_sku = re.sub(r'-\d+$', '', individual_sku)
                                                file_skus.append(processed_sku)
                                    else:
                                        # 去除数字尾缀
                                        processed_sku = re.sub(r'-\d+$', '', line)
                                        file_skus.append(processed_sku)
                        else:
                            # 单个SKU，去除数字尾缀
                            processed_sku = re.sub(r'-\d+$', '', sku_cell)
                            file_skus.append(processed_sku)
                    
                    all_skus.extend(file_skus)
                    processed_files.append(filename)
                    
                except Exception as e:
                    # 静默跳过出错的文件
                    continue
            
            if all_skus:
                self.update_progress(70, "正在汇总SKU数据...")
                
                # 去重并保持顺序
                unique_skus = []
                seen = set()
                for sku in all_skus:
                    if sku not in seen:
                        unique_skus.append(sku)
                        seen.add(sku)
                
                # 汇总填入搜索框
                self.update_progress(80, "正在填入搜索框...")
                sku_text = '\n'.join(unique_skus)
                
                # 清空搜索框并填入所有SKU数据
                self.search_entry.delete('1.0', tk.END)
                self.search_entry.insert('1.0', sku_text)
                self.search_entry.config(fg='#212529')
                
                # 自动启用去除尾缀词功能
                self.update_progress(85, "正在启用去除尾缀词功能...")
                self.remove_suffix()
                
                self.update_progress(90, "多表格SKU汇总完成！")
                
                # 直接开始依次处理表格
                self.process_multiple_excel_sequential_with_naming(excel_files)
                
        except Exception as e:
            # 静默处理错误，不显示错误提示
            pass
        finally:
            self.hide_progress()
    
    def process_multiple_excel_sequential(self, excel_files):
        """依次处理多个Excel文件，使用队列机制避免对话框重叠"""
        try:
            print(f"🔄 开始队列化处理 {len(excel_files)} 个Excel文件...")
            
            # 重置队列状态
            self._file_processing_queue = excel_files.copy()
            self._is_processing_queue = True
            self._current_dialog_active = False
            
            # 启动队列处理
            self.process_file_queue_sequential()
            
        except Exception as e:
            print(f"❌ 队列化处理Excel文件时出错: {e}")
            import traceback
            traceback.print_exc()
            # 重置队列状态
            self._file_processing_queue = []
            self._is_processing_queue = False
            self._current_dialog_active = False

    def process_file_queue_sequential(self):
        """按顺序处理队列中的文件，确保对话框不重叠"""
        try:
            # 检查队列是否为空或已停止处理
            if not self._is_processing_queue or not self._file_processing_queue:
                print("✅ 队列处理完成或已停止")
                self._reset_queue_state()
                return
            
            # 检查是否有活跃的对话框
            if self._current_dialog_active:
                print("⏳ 有活跃对话框，等待处理完成...")
                return
            
            # 获取队列中的下一个文件
            current_file = self._file_processing_queue.pop(0)
            print(f"📁 开始处理队列文件: {os.path.basename(current_file)}")
            
            # 处理单个文件
            self.process_single_file_in_queue(current_file)
            
        except Exception as e:
            print(f"❌ 队列处理时出错: {e}")
            import traceback
            traceback.print_exc()
            self._reset_queue_state()

    def process_single_file_in_queue(self, excel_path):
        """处理队列中的单个文件，管理对话框状态并收集SKU"""
        try:
            print(f"🔄 队列处理文件: {os.path.basename(excel_path)}")
            
            # 初始化SKU收集列表（如果不存在）
            if not hasattr(self, '_processed_files_skus'):
                self._processed_files_skus = []
            
            # 设置多表格处理标记，防止单个文件处理时清空搜索框和重复收集SKU
            self._is_multi_table_processing = True
            
            # 调用原有的文件处理逻辑
            # 注意：这里会触发show_manual_processing_dialog（如果有未匹配的SKU）
            # SKU收集将在process_excel_file中进行，但不会填入搜索框
            self.process_excel_file(excel_path)
            
            # 如果没有触发对话框，直接继续处理下一个文件
            if not self._current_dialog_active:
                self.continue_queue_processing()
            
        except Exception as e:
            print(f"❌ 队列处理单个文件时出错: {e}")
            import traceback
            traceback.print_exc()
            # 出错时继续处理下一个文件
            self.continue_queue_processing()

    def continue_queue_processing(self):
        """继续处理队列中的下一个文件"""
        try:
            print("🔄 继续队列处理...")
            
            # 重置对话框状态
            self._current_dialog_active = False
            
            # 延迟一小段时间后继续处理，确保当前对话框完全关闭
            self.root.after(100, self.process_file_queue_sequential)
            
        except Exception as e:
            print(f"❌ 继续队列处理时出错: {e}")
            import traceback
            traceback.print_exc()
            self._reset_queue_state()

    def _reset_queue_state(self):
        """重置队列状态并汇总所有处理过的文件的SKU"""
        try:
            print("🔄 重置队列状态...")
            
            # 在队列处理完成时，汇总所有SKU到搜索框
            if hasattr(self, '_processed_files_skus') and self._processed_files_skus:
                print("📋 开始汇总所有处理过的文件的SKU...")
                
                # 去重并保持顺序
                unique_skus = []
                seen = set()
                for sku in self._processed_files_skus:
                    if sku not in seen:
                        unique_skus.append(sku)
                        seen.add(sku)
                
                # 填入搜索框
                sku_text = '\n'.join(unique_skus)
                self.search_entry.delete('1.0', tk.END)
                self.search_entry.insert('1.0', sku_text)
                self.search_entry.config(fg='#212529')
                
                print(f"📋 已将所有 {len(unique_skus)} 个SKU填入搜索框")
                
                # 自动启用去除序号尾缀功能
                print("🔧 自动启用去除序号尾缀功能...")
                self.remove_suffix()
                
                # 多表格处理完成，SKU已填入搜索框，根据配置决定是否自动启动搜索
                if getattr(self, 'auto_search_enabled', False):
                    print("✅ 多表格处理完成，SKU已填入搜索框，自动启动搜索")
                    self.root.after(500, self.start_search)
                else:
                    print("✅ 多表格处理完成，SKU已填入搜索框，请手动点击搜索按钮开始搜索图片")
                
                # 显示多表格面单名称弹窗（如果有收集到的面单名称）
                if hasattr(self, '_waybill_names_collection') and self._waybill_names_collection:
                    try:
                        from waybill_name_dialog import WaybillNameDialog
                        
                        # 使用全局弹窗实例
                        if self.waybill_dialog is None:
                            self.waybill_dialog = WaybillNameDialog(self.root, self)  # 直接传入organizer_instance
                        
                        self.waybill_dialog.show_multiple_waybills(self._waybill_names_collection)
                        print(f"📋 显示 {len(self._waybill_names_collection)} 个面单名称")
                        
                        # 清空面单名称收集
                        self._waybill_names_collection = []
                    except Exception as e:
                        print(f"显示多表格面单名称弹窗时出错: {e}")
                
                # 清空临时SKU存储
                self._processed_files_skus = []
            
            self._file_processing_queue = []
            self._is_processing_queue = False
            self._current_dialog_active = False
            
            # 清除多表格处理标记
            if hasattr(self, '_is_multi_table_processing'):
                self._is_multi_table_processing = False
            
            print("✅ 队列状态已重置")
            
        except Exception as e:
            print(f"❌ 重置队列状态时出错: {e}")
            import traceback
            traceback.print_exc()

    def process_multiple_excel_sequential_with_naming(self, excel_files):
        """依次处理多个Excel文件，先询问导出选项，再询问表格命名"""
        if not excel_files:
            return
        
        # 设置多表格处理标记
        self._is_multi_table_processing = True
        
        # 先显示尺寸处理选择对话框（与单表格流程一致）
        # 计算总SKU数量用于显示
        total_sku_count = 0
        for excel_path in excel_files:
            try:
                df = pd.read_excel(excel_path)
                # 识别SKU列
                sku_column = None
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in ['sku', 'spu', '商品编码', '货号', '款号']):
                        sku_column = col
                        break
                if sku_column:
                    total_sku_count += len(df[sku_column].dropna().unique())
            except Exception as e:
                print(f"读取Excel文件时出错: {e}")
                continue
        
        # 显示尺寸处理选择对话框
        size_choice = self.show_size_processing_dialog(total_sku_count)
        if size_choice is None or size_choice == "no":
            # 用户取消或选择否，仍然自动启动搜索
            self.root.after(500, self.start_search)
            return  # 用户取消或选择否
        
        # 根据用户选择设置导出设置
        export_settings = {
            'export': True,
            'open_after_export': size_choice == "process_and_open",
            'size_processing': 'yes'
        }
        
        # 显示表格命名对话框
        table_names = self.show_table_naming_dialog(excel_files)
        if not table_names:  # 用户取消了命名
            return
        
        # 创建统一的进度条窗口
        progress_window = self.create_batch_progress_window(len(excel_files))
        
        try:
            # 依次处理所有文件，使用统一的导出设置
            for i, excel_path in enumerate(excel_files):
                table_name = table_names[i]
                self.current_table_name = table_name
                
                # 更新进度条
                filename = os.path.basename(excel_path)
                progress_text = f"正在处理第 {i+1}/{len(excel_files)} 个文件：{filename}"
                self.update_batch_progress(progress_window, i+1, len(excel_files), progress_text)
                
                # 使用统一的导出设置处理文件
                self.process_excel_file_with_settings(excel_path, export_settings, table_name)
            
            # 所有文件处理完成
            self.update_batch_progress(progress_window, len(excel_files), len(excel_files), "所有文件处理完成！")
            
            # 延迟关闭进度窗口，让用户看到完成信息
            self.root.after(2000, lambda: self.close_batch_progress(progress_window))
            
            # 批量处理完成，SKU已填入搜索框，根据配置决定是否自动启动搜索
            if getattr(self, 'auto_search_enabled', False):
                print("✅ 批量处理完成，SKU已填入搜索框，自动启动搜索")
                self.root.after(500, self.start_search)
            else:
                print("✅ 批量处理完成，SKU已填入搜索框，自动搜索已禁用")
            
        except Exception as e:
            self.close_batch_progress(progress_window)
            messagebox.showerror("错误", f"批量处理时出错: {str(e)}")
        
        # 清理临时设置
        if hasattr(self, 'current_table_name'):
            delattr(self, 'current_table_name')
        
        # 清除多表格处理标记
        self._is_multi_table_processing = False
    
    def create_batch_progress_window(self, total_files):
        """创建批量处理进度窗口"""
        progress_window = tk.Toplevel(self.root)
        progress_window.title("批量处理进度")
        progress_window.geometry("500x150")
        progress_window.resizable(False, False)
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # 立即隐藏窗口，避免在左上角显示
        progress_window.withdraw()
        
        # 居中显示 - 先更新布局但窗口仍然隐藏
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
        y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        # 设置好位置后再显示窗口，避免移动效果
        progress_window.deiconify()
        
        # 主框架
        main_frame = ttk.Frame(progress_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="正在批量处理Excel文件...", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(pady=(0, 15))
        
        # 进度条
        progress_bar = ttk.Progressbar(main_frame, mode='determinate', length=400)
        progress_bar.pack(pady=(0, 10))
        progress_bar['maximum'] = total_files
        
        # 状态文本
        status_label = ttk.Label(main_frame, text="准备开始处理...", 
                                font=('Microsoft YaHei UI', 10))
        status_label.pack()
        
        # 存储组件引用
        progress_window.progress_bar = progress_bar
        progress_window.status_label = status_label
        
        return progress_window
    
    def update_batch_progress(self, progress_window, current, total, text):
        """更新批量处理进度"""
        if progress_window and progress_window.winfo_exists():
            progress_window.progress_bar['value'] = current
            progress_window.status_label.config(text=text)
            progress_window.update()
    
    def close_batch_progress(self, progress_window):
        """关闭批量处理进度窗口"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
            # 添加延迟确保窗口完全关闭
            self.root.after(100, lambda: None)
    
    def show_table_naming_dialog(self, excel_files):
        """显示表格命名对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("表格命名")
        dialog.geometry("500x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.attributes('-topmost', True)
        dialog.focus_force()
        
        # 立即隐藏窗口，避免在左上角显示
        dialog.withdraw()
        
        # 居中显示 - 先更新布局但窗口仍然隐藏
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 设置好位置后再显示窗口，避免移动效果
        dialog.deiconify()
        
        result = {'names': None}
        
        # 主框架
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="为每个表格命名", font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 存储输入框的列表
        name_entries = []
        
        # 创建滚动区域
        canvas = tk.Canvas(main_frame, height=200)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 为每个Excel文件创建命名输入框
        for i, excel_path in enumerate(excel_files):
            file_frame = ttk.Frame(scrollable_frame)
            file_frame.pack(fill=tk.X, pady=5, padx=10)
            
            # 文件名标签
            filename = os.path.basename(excel_path)
            file_label = ttk.Label(file_frame, text=f"文件 {i+1}: {filename}")
            file_label.pack(anchor=tk.W)
            
            # 生成默认名称 - 优先使用记忆的导出名称
            if hasattr(self, 'last_excel_export_name') and self.last_excel_export_name:
                default_name = self.last_excel_export_name
            else:
                # 如果没有记忆的名称，使用原有逻辑
                base_name = os.path.splitext(filename)[0]
                default_name = f"{base_name}-表格{i+1}"
            
            # 名称输入框
            name_entry = ttk.Entry(file_frame, width=50)
            name_entry.insert(0, default_name)
            name_entry.pack(fill=tk.X, pady=(5, 0))
            name_entries.append(name_entry)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        def on_cancel():
            dialog.destroy()
        
        # 按钮
        ttk.Button(button_frame, text="确认", command=on_confirm).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="取消", command=on_cancel).pack(side=tk.RIGHT)
        
        # 等待对话框关闭
        dialog.wait_window()
        
        return result['names']
    
    def process_excel_file_with_settings(self, excel_path, export_settings, custom_name=None):
        """使用指定的导出设置处理Excel文件，进行完整的处理流程"""
        try:
            # 显示进度条
            self.show_progress()
            self.update_progress(10, f"正在处理: {os.path.basename(excel_path)}")
            
            # 读取Excel文件
            df = pd.read_excel(excel_path)
            
            # 识别SKU列
            self.update_progress(20, "正在识别SKU列...")
            sku_column = None
            for col in df.columns:
                if any(keyword in str(col).lower() for keyword in ['sku', 'spu', '商品编码', '货号', '款号']):
                    sku_column = col
                    break
            
            if sku_column is None:
                self.hide_progress()
                messagebox.showwarning("警告", f"在文件 {os.path.basename(excel_path)} 中未找到SKU列")
                return
            
            # 提取SKU
            self.update_progress(30, "正在提取SKU...")
            sku_list = []
            for index, row in df.iterrows():
                sku_value = str(row[sku_column]).strip()
                if sku_value and sku_value.lower() != 'nan':
                    sku_list.append(sku_value)
            
            if not sku_list:
                self.hide_progress()
                messagebox.showwarning("警告", f"在文件 {os.path.basename(excel_path)} 中未找到有效的SKU数据")
                return
            
            # 进行SKU前缀分析
            self.update_progress(35, "正在分析SKU前缀...")
            analysis_result, processing_config = self.analyze_sku_prefixes(sku_list)
            
            # 存储分析结果供后续使用
            self.current_sku_analysis = analysis_result
            print(f"🔍 SKU分析结果已存储:")
            print(f"  analysis_result: {analysis_result}")
            print(f"  processing_config: {processing_config}")
            print(f"  self.current_sku_analysis: {self.current_sku_analysis}")
            
            # 不修改搜索框内容，保持显示所有图片
            # 用户希望显示所有图片，而不是逐个搜索
            
            # 检查是否需要处理尺寸
            self.update_progress(50, "正在检查尺寸数据...")
            
            # 检查未匹配的SKU
            unmatched_skus = self.check_unmatched_skus(sku_list)
            
            if unmatched_skus:
                # 在多表格处理模式下，检查是否有活跃对话框
                if hasattr(self, '_is_multi_table_processing') and self._is_multi_table_processing:
                    if hasattr(self, '_current_dialog_active') and self._current_dialog_active:
                        # 如果有活跃对话框，暂停处理，等待对话框关闭
                        return
                    # 标记对话框为活跃状态
                    self._current_dialog_active = True
                
                self.hide_progress()
                # 显示手动处理对话框
                self.show_manual_processing_dialog(unmatched_skus, sku_list, df, excel_path)
            else:
                # 所有SKU都已匹配，继续处理
                self.update_progress(70, "正在处理尺寸数据...")
                self.continue_excel_processing_with_settings(sku_list, df, excel_path, export_settings, custom_name)
                
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("错误", f"处理Excel文件时出错: {str(e)}")
            print(f"处理Excel文件时出错: {str(e)}")
    
    def continue_excel_processing_with_settings(self, sku_list, df, excel_path, export_settings, custom_name=None):
        """使用指定设置继续Excel处理流程"""
        try:
            self.update_progress(80, "正在处理尺寸映射...")
            
            # 处理尺寸映射
            df_processed = self.process_size_mapping(df)
            
            # 根据导出设置决定是否导出
            if export_settings.get('export', False):
                self.update_progress(90, "正在导出Excel文件...")
                
                # 使用自定义名称或生成默认名称
                if custom_name:
                    export_name = custom_name
                    # 应用名称匹配功能
                    if hasattr(self, 'current_sku_analysis') and self.current_sku_analysis:
                        name_matching_config = self.get_name_matching_config()
                        print(f"🔍 [调试] 准备对自定义名称应用名称匹配:")
                        print(f"  原始名称: {export_name}")
                        print(f"  SKU分析数据: {self.current_sku_analysis}")
                        print(f"  名称匹配配置: {name_matching_config}")
                        original_name = export_name
                        export_name, reason = self.apply_name_matching(export_name, self.current_sku_analysis, name_matching_config)
                        print(f"  匹配后名称: {export_name}")
                        print(f"  匹配原因: {reason}")
                        if export_name != original_name:
                            self.show_auto_dismiss_message(f"文件名已应用名称匹配: {original_name} → {export_name}")
                            print(f"🎯 名称匹配应用成功: {reason}")
                        else:
                            print(f"⚠️ 名称未发生变化: {reason}")
                    else:
                        print(f"❌ [调试] 无法应用名称匹配 - SKU分析数据不存在")
                    print(f"📝 使用自定义文件名: {export_name}")
                else:
                    # 生成默认文件名
                    base_name = os.path.splitext(os.path.basename(excel_path))[0]
                    export_name = f"{base_name}_已整理尺寸"
                    # 应用名称匹配功能
                    if hasattr(self, 'current_sku_analysis') and self.current_sku_analysis:
                        name_matching_config = self.get_name_matching_config()
                        print(f"🔍 [调试] 准备对默认名称应用名称匹配:")
                        print(f"  原始名称: {export_name}")
                        print(f"  SKU分析数据: {self.current_sku_analysis}")
                        print(f"  名称匹配配置: {name_matching_config}")
                        original_name = export_name
                        export_name, reason = self.apply_name_matching(export_name, self.current_sku_analysis, name_matching_config)
                        print(f"  匹配后名称: {export_name}")
                        print(f"  匹配原因: {reason}")
                        if export_name != original_name:
                            self.show_auto_dismiss_message(f"文件名已应用名称匹配: {original_name} → {export_name}")
                            print(f"🎯 名称匹配应用成功: {reason}")
                        else:
                            print(f"⚠️ 名称未发生变化: {reason}")
                    else:
                        print(f"❌ [调试] 无法应用名称匹配 - SKU分析数据不存在")
                    print(f"📝 使用默认文件名: {export_name}")
                
                # 处理并导出Excel
                should_open = export_settings.get('open_after_export', False)
                self.process_and_export_excel(df_processed, excel_path, export_name, should_open)
                
                self.update_progress(100, f"文件 {export_name} 处理完成！")
            else:
                self.update_progress(100, "SKU提取完成！")
            
            # 延迟隐藏进度条，让用户看到完成信息
            self.root.after(1000, self.hide_progress)
            
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("错误", f"处理Excel文件时出错: {str(e)}")
            print(f"处理Excel文件时出错: {str(e)}")

    def show_excel_export_dialog(self, sku_list=None):
        """显示模块化表格导出确认弹窗 (1.8 动态命名系统)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("📊 导出名称模块化设置")
        dialog.geometry("550x450")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 立即隐藏并居中
        dialog.withdraw()
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        dialog.deiconify()
        
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="🧩 导出文件名", font=('Microsoft YaHei UI', 12, 'bold')).pack(pady=(0, 20))

        # --- 智能感知逻辑 ---
        # 1. 抓取缩写 (从当前文件夹名抓取，抓不到用预设)
        current_folder = self.folder_name_var.get().strip()
        import re
        abbrev_match = re.search(r'^([A-Z]{2,4})', current_folder)
        detected_prefix = abbrev_match.group(1) if abbrev_match else self.naming_center.get('business_abbreviation', 'CHX')
        if "急采" not in detected_prefix and "急采" in current_folder:
            detected_prefix = "急采" + detected_prefix

        # 2. 日期 (今天)
        from datetime import datetime
        current_date = datetime.now().strftime("%m-%d")

        # 3. 物流商 (记忆上次，没记忆用默认第一个)
        logistics_list = self.naming_center.get('logistics_providers', ["Y2尊祐", "Y1尚为"])
        last_logistics = self.naming_center.get('last_logistics_provider', logistics_list[0])

        # 4. 产品 (SKU分析结果)
        detected_product = "画"
        # 获取所有可选产品类型用于下拉框
        product_options = ["画", "窗帘", "床上三件套", "床笠"] # 基础预设
        
        # 从配置中动态加载更多产品类型
        matching_config = self.get_name_matching_config()
        if matching_config:
            for p_name in matching_config.values():
                if p_name and p_name not in product_options:
                    product_options.append(p_name)

        if hasattr(self, 'current_sku_analysis') and self.current_sku_analysis:
            # 尝试通过 apply_name_matching 逻辑获取产品名
            name_matching_config = matching_config
            analysis_result = self.current_sku_analysis[0] if isinstance(self.current_sku_analysis, tuple) else self.current_sku_analysis
            _, reason = self.apply_name_matching("temp", analysis_result, name_matching_config)
            
            # 从 reason 中提取产品名 (支持 "替换为 '...'" 和 "添加产品类型 '...'")
            prod_match = re.search(r"(?:替换为|添加产品类型) '([^']+)'", reason)
            if prod_match:
                detected_product = prod_match.group(1)
                if detected_product not in product_options:
                    product_options.insert(0, detected_product)

        # 5. 后缀 (记忆上次)
        suffix_list = self.naming_center.get('custom_suffixes', ["艺术家", "画家", "设计师"])
        last_suffix = self.naming_center.get('last_custom_suffix', suffix_list[0])

        # --- UI 组件 ---
        grid_frame = ttk.Frame(main_frame)
        grid_frame.pack(fill=tk.X, pady=10)
        grid_frame.columnconfigure(1, weight=1)

        # 模块变量
        prefix_var = tk.StringVar(value=detected_prefix)
        date_var = tk.StringVar(value=current_date)
        provider_var = tk.StringVar(value=last_logistics)
        product_var = tk.StringVar(value=detected_product)
        suffix_var = tk.StringVar(value=last_suffix)
        preview_var = tk.StringVar()

        def update_preview(*args):
            template = self.naming_center.get('naming_template', '{prefix}-{date}-{provider}-{product}-{suffix}')
            # 组装
            result = template.format(
                prefix=prefix_var.get().strip(),
                date=date_var.get().strip(),
                provider=provider_var.get().strip(),
                product=product_var.get().strip(),
                suffix=suffix_var.get().strip()
            )
            # 应用序号检查
            smart_result = self.get_smart_name(result, 'excel_exports')
            preview_var.set(smart_result)

        # 绑定更新
        for v in [prefix_var, date_var, provider_var, product_var, suffix_var]:
            v.trace_add("write", update_preview)

        # 1. 业务前缀列表 (从配置获取，并确保 detected_prefix 在首位)
        abbrev_options = self.naming_center.get('business_abbreviations', ["CHX"])
        if detected_prefix not in abbrev_options:
            abbrev_options = [detected_prefix] + abbrev_options
        else:
            # 移动 detected_prefix 到首位
            abbrev_options = [detected_prefix] + [a for a in abbrev_options if a != detected_prefix]

        # 渲染 Grid
        rows = [
            ("名称缩写：", prefix_var, abbrev_options),
            ("日期：", date_var, None),
            ("物流商：", provider_var, logistics_list),
            ("产品名称：", product_var, product_options),
            ("所属部门：", suffix_var, suffix_list)
        ]

        for i, (label, var, vals) in enumerate(rows):
            ttk.Label(grid_frame, text=label).grid(row=i, column=0, sticky=tk.W, pady=5)
            if vals:
                cb = ttk.Combobox(grid_frame, textvariable=var, values=vals)
                cb.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=(5, 0))
            else:
                ent = ttk.Entry(grid_frame, textvariable=var)
                ent.grid(row=i, column=1, sticky=tk.EW, pady=5, padx=(5, 0))

        # 预览区域
        preview_frame = ttk.LabelFrame(main_frame, text="👀 文件名预览 (可手动修改序号)", padding="10")
        preview_frame.pack(fill=tk.X, pady=20)
        
        # 使用 Entry 代替 Label，允许手动修改
        preview_entry = ttk.Entry(preview_frame, textvariable=preview_var, font=('Consolas', 10, 'bold'))
        preview_entry.pack(fill=tk.X, padx=5, pady=5)
        
        # 设置 Entry 样式使其看起来不像普通的输入框，但可编辑
        style = ttk.Style()
        style.configure('Preview.TEntry', foreground='#2c3e50')
        preview_entry.configure(style='Preview.TEntry')

        # 按钮区域
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)

        result = {'name': None}

        def on_confirm():
            final_name = preview_var.get()
            # 记录最后使用的选项
            self.naming_center['last_logistics_provider'] = provider_var.get().strip()
            self.naming_center['last_custom_suffix'] = suffix_var.get().strip()
            
            # 自动学习新选项 (服务商、后缀、名称缩写)
            if self.naming_center.get('auto_save_new_items', True):
                new_p = provider_var.get().strip()
                if new_p and new_p not in self.naming_center['logistics_providers']:
                    self.naming_center['logistics_providers'].append(new_p)
                
                new_s = suffix_var.get().strip()
                if new_s and new_s not in self.naming_center['custom_suffixes']:
                    self.naming_center['custom_suffixes'].append(new_s)

                new_a = prefix_var.get().strip().upper()
                if new_a:
                    # 移除可能的“急采”前缀进行判断
                    base_a = new_a.replace("急采", "")
                    if 'business_abbreviations' not in self.naming_center:
                        self.naming_center['business_abbreviations'] = []
                    if base_a and base_a not in self.naming_center['business_abbreviations']:
                        self.naming_center['business_abbreviations'].append(base_a)
            
            self.save_config()
            result['name'] = final_name
            dialog.destroy()

        ttk.Button(btn_frame, text="确认导出", command=on_confirm, style='Accent.TButton').pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(btn_frame, text="取消", command=lambda: dialog.destroy()).pack(side=tk.RIGHT)

        # 初始化预览
        update_preview()
        
        dialog.wait_window()
        return result['name']

    def process_and_export_excel(self, df, original_path, custom_name=None, open_after_export=False):
        """处理Excel数据并导出到用户配置的路径，保留图片对象"""
        try:
            # 验证并获取有效的导出路径
            export_dir = self.validate_and_reset_export_path('excel')
            
            # 生成新文件名
            if custom_name:
                # 使用自定义名称
                base_filename = f"{custom_name}.xlsx"
                export_name_for_waybill = custom_name  # 保存用于面单名称转换的导出名称
            else:
                # 使用原始文件名
                original_name = os.path.splitext(os.path.basename(original_path))[0]
                base_filename = f"{original_name}_已整理尺寸.xlsx"
                export_name_for_waybill = original_name  # 保存用于面单名称转换的导出名称
            
            # 检查文件名重复并生成唯一文件名
            export_path = os.path.join(export_dir, base_filename)
            counter = 2
            while os.path.exists(export_path):
                # 分离文件名和扩展名
                name_without_ext = os.path.splitext(base_filename)[0]
                ext = os.path.splitext(base_filename)[1]
                # 生成带后缀的文件名
                unique_filename = f"{name_without_ext}-{counter}{ext}"
                export_path = os.path.join(export_dir, unique_filename)
                counter += 1
            
            # 使用openpyxl直接处理Excel文件，保留图片对象
            from openpyxl import load_workbook
            
            # 加载原始工作簿
            workbook = load_workbook(original_path)
            worksheet = workbook.active
            
            # 处理浮动图片：将所有图片下移一行并设置H列图片固定尺寸
            if hasattr(worksheet, '_images') and worksheet._images:
                for image in worksheet._images:
                    # 获取图片的锚点信息
                    if hasattr(image, 'anchor') and hasattr(image.anchor, '_from'):
                        # 将图片的起始行下移一行
                        image.anchor._from.row += 1
                        if hasattr(image.anchor, 'to') and image.anchor.to:
                            image.anchor.to.row += 1
                        
                        # 检查图片是否在H列（第8列）
                        if image.anchor._from.col == 7:  # H列是第8列，索引为7
                            # 延迟导入openpyxl相关模块
                            openpyxl, Font, PatternFill, Alignment, OpenpyxlImage, cm_to_EMU = _import_openpyxl()
                            # 直接设置H列图片固定尺寸：宽度2.79厘米，高度4.69厘米
                            # openpyxl会自动处理尺寸设置，无需手动解锁纵横比
                            image.width = cm_to_EMU(2.79)
                            image.height = cm_to_EMU(4.69)
            
            # 在第一行插入新行
            worksheet.insert_rows(1)
            
            # 根据最终的文件名生成表格标题
            # 从export_path中提取最终的文件名（不含扩展名）作为表格标题
            final_filename = os.path.splitext(os.path.basename(export_path))[0]
            table_title = final_filename
            
            # 更新面单名称转换用的导出名称为最终文件名（包含序号）
            export_name_for_waybill = final_filename
            
            # 延迟导入openpyxl相关模块
            openpyxl, Font, PatternFill, Alignment, OpenpyxlImage, cm_to_EMU = _import_openpyxl()
            
            # 在新插入的第一行填入表格标题并合并A-D列
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = table_title
            # 设置标题字体为宋体、20号、居中对齐
            title_cell.font = Font(name='宋体', size=20)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            # 合并A-D列（第1列到第5列）
            worksheet.merge_cells('A1:D1')
            # 设置第一行行高为77
            worksheet.row_dimensions[1].height = 77
            
            # 设置各列宽度
            worksheet.column_dimensions['A'].width = 8.48
            worksheet.column_dimensions['B'].width = 20.81
            worksheet.column_dimensions['C'].width = 24.14
            worksheet.column_dimensions['D'].width = 45.36
            worksheet.column_dimensions['E'].width = 24.14 
            worksheet.column_dimensions['F'].width = 24.14
            worksheet.column_dimensions['G'].width = 24.14
            worksheet.column_dimensions['H'].width = 24.14
            
            # 设置第二行行高为44
            worksheet.row_dimensions[2].height = 44
            
            # 找到各列的索引（现在标题行变成了第二行）
            size_column_index = None
            processing_column_index = None
            material_column_index = None
            sku_column_index = None
            sequence_column_index = None
            quantity_column_index = None  # F列（产品数量列）
            order_no_column_index = None  # 订单号列
            
            # 查找列索引（现在第二行是标题行）
            for col_idx, cell in enumerate(worksheet[2], 1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    # 设置第二行标题字体为宋体、20号
                    cell.font = Font(name='宋体', size=20)
                    
                    if self.size_column and cell_value == self.size_column:
                        size_column_index = col_idx
                    elif cell_value in ['加工方式', '加工方法']:
                        processing_column_index = col_idx
                    elif cell_value in ['材质', '材料']:
                        material_column_index = col_idx
                    elif cell_value in ['SKU', 'sku']:
                        sku_column_index = col_idx
                    elif cell_value in ['序号', '编号', '序列']:
                        sequence_column_index = col_idx
                    elif cell_value in ['数量', '产品数量', '件数'] or col_idx == 6:  # F列是第6列
                        quantity_column_index = col_idx
                    elif cell_value in ['订单号', '订单编号', 'Order', 'OrderNo', 'Order_No']:
                        order_no_column_index = col_idx
            
            # 处理数据行（现在从第三行开始，因为插入了新行）
            sequence_number = 1
            total_quantity = 0  # 用于计算F列总数量
            processed_rows = set()  # 记录已处理的行，避免重复处理
            
            # 先收集所有需要拆分的多SKU行
            multi_sku_rows = []
            for row_idx in range(3, worksheet.max_row + 1):
                # 检查是否是空行
                is_empty_row = True
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        is_empty_row = False
                        break
                
                if not is_empty_row:
                    # 新的B列关键词检测逻辑
                    b_cell = worksheet.cell(row=row_idx, column=size_column_index) if size_column_index else None
                    b_content = str(b_cell.value).strip() if b_cell and b_cell.value else ""
                    
                    if b_content:
                        print(f"检查第{row_idx}行B列内容: {repr(b_content)}")
                        
                        # 统一的多行判断逻辑
                        is_multi_line = False
                        skuid_matches = []
                        size_list = []
                        sku_blocks = []
                        
                        # 优先检测SKUID格式（新格式）
                        if 'SKUID:' in b_content:
                            # 统计关键词出现次数
                            skuid_count = len(re.findall(r'SKUID:', b_content, re.IGNORECASE))
                            variants_count = len(re.findall(r'Variants:', b_content, re.IGNORECASE))
                            warehouse_count = len(re.findall(r'Warehouse:', b_content, re.IGNORECASE))
                            
                            keyword_counts = {
                                'SKUID': skuid_count,
                                'Variants': variants_count,
                                'Warehouse': warehouse_count
                            }
                            print(f"关键词统计: {keyword_counts}")
                            
                            # 提取SKUID列表
                            skuid_matches = re.findall(r'SKUID:(\d+)', b_content)
                            print(f"找到的SKUID: {skuid_matches}")
                            
                            # 判断是否为多行：SKUID数量>1 或 任一关键词出现次数>1
                            is_multi_line = len(skuid_matches) > 1 or any(count > 1 for count in keyword_counts.values())
                            
                            if is_multi_line and len(skuid_matches) > 1:
                                print(f"检测到多SKU行(SKUID格式): 第{row_idx}行，包含{len(skuid_matches)}个SKU")
                                # 从B列按SKUID分割生成sku_blocks
                                sku_blocks = re.split(r'(?=SKUID:)', b_content)
                                sku_blocks = [block.strip() for block in sku_blocks if block.strip()]
                                
                                # 🔧 修复多行SKU尺寸处理：直接从每个SKUID块中提取原始Variants内容
                                print("🔧 从每个SKUID块中提取原始Variants内容，保持原始格式")
                                
                                # 从每个sku_block中提取原始Variants内容
                                original_variants_list = []
                                for i, block in enumerate(sku_blocks):
                                    # 从block中提取Variants内容
                                    variants_match = re.search(r'Variants:\s*([^\n\r]+)', block)
                                    if variants_match:
                                        variants_content = variants_match.group(1).strip()
                                        original_variants_list.append(variants_content)
                                        print(f"✅ 从SKUID块{i+1}提取原始Variants: '{variants_content}'")
                                    else:
                                        # 如果没有找到Variants，尝试提取整个block的非SKUID部分
                                        non_skuid_content = re.sub(r'SKUID:\d+\s*', '', block).strip()
                                        if non_skuid_content:
                                            original_variants_list.append(non_skuid_content)
                                            print(f"⚠️ SKUID块{i+1}未找到Variants标签，使用非SKUID内容: '{non_skuid_content}'")
                                        else:
                                            original_variants_list.append('')
                                            print(f"❌ SKUID块{i+1}无有效内容")
                                
                                if len(original_variants_list) == len(sku_blocks):
                                    size_list = original_variants_list
                                    print(f"✅ 原始Variants提取成功，获得 {len(original_variants_list)} 个尺寸: {original_variants_list}")
                                else:
                                    # 回退逻辑：如果提取失败，使用原始B列内容
                                    print(f"❌ 原始Variants提取失败或数量不匹配，使用原始B列内容作为回退")
                                    size_list = [b_content] * len(sku_blocks)
                                
                                multi_sku_rows.append((row_idx, skuid_matches, size_list, sku_blocks))
                            elif is_multi_line and len(skuid_matches) <= 1:
                                # 边界情况：单个SKUID但有多个关键词的情况
                                print(f"检测到单SKUID多关键词行: 第{row_idx}行，SKUID数量: {len(skuid_matches)}")
                                print(f"关键词统计: {keyword_counts}")
                                # 这种情况可能是格式错误或特殊情况，记录但不拆分
                                print("警告：检测到多关键词但SKUID数量<=1的情况，跳过拆分")
                        
                        # 检测简单的空格分隔SKU（兼容原有格式）- 仅在没有SKUID:格式时处理
                        elif not is_multi_line and ('\n' in b_content or ' ' in b_content):
                            sku_list = [sku.strip() for sku in re.split(r'[\s\n\t]+', b_content) if sku.strip()]
                            if len(sku_list) > 1:
                                print(f"检测到简单多SKU行: 第{row_idx}行，包含{len(sku_list)}个SKU: {sku_list}")
                                is_multi_line = True
                                skuid_matches = sku_list  # 对于简单格式，直接使用sku_list
                                
                                # 优化简单格式的尺寸提取：使用统一的尺寸提取方法
                                if size_column_index:
                                    b_cell = worksheet.cell(row=row_idx, column=size_column_index)
                                    if b_cell.value:
                                        b_content_size = str(b_cell.value).strip()
                                        print(f"🔧 [简单格式] 使用统一尺寸提取方法处理B列: {b_content_size}")
                                        
                                        # 首先尝试按空格或换行分割B列尺寸
                                        size_parts = [s.strip() for s in re.split(r'[\s\n\t]+', b_content_size) if s.strip()]
                                        if len(size_parts) == len(sku_list):
                                            # 对每个尺寸部分使用统一提取方法
                                            processed_sizes = []
                                            for size_part in size_parts:
                                                extracted = self._extract_single_size(size_part)
                                                processed_sizes.append(extracted if extracted else size_part)
                                            size_list = processed_sizes
                                            print(f"✅ [简单格式] B列尺寸数量匹配SKU数量，使用处理后的尺寸: {size_list}")
                                        else:
                                            # 如果数量不匹配，尝试统一提取整个B列内容
                                            extracted_size = self._extract_single_size(b_content_size)
                                            final_size = extracted_size if extracted_size else b_content_size
                                            size_list = [final_size] * len(sku_list)
                                            print(f"✅ [简单格式] B列尺寸数量不匹配，所有SKU使用统一处理后的尺寸: {final_size}")
                                    else:
                                        print("❌ [简单格式] 简单多SKU行B列为空")
                                        size_list = [''] * len(sku_list)
                                
                                multi_sku_rows.append((row_idx, skuid_matches, size_list, []))  # 简单格式sku_blocks为空
                        
                        # 调试信息：记录判断结果
                        if is_multi_line:
                            print(f"第{row_idx}行判断为多行，SKU数量: {len(skuid_matches)}")
                        else:
                            print(f"第{row_idx}行判断为单行")
            
            # 从后往前处理多SKU行，避免行号变化影响
            for row_data in reversed(multi_sku_rows):
                if len(row_data) == 4:  # 新格式：(row_idx, skuid_list, size_list, sku_blocks)
                    row_idx, skuid_list, size_list, sku_blocks = row_data
                    self.split_multi_sku_row_advanced(worksheet, row_idx, skuid_list, size_list, sku_blocks, size_column_index, processing_column_index, material_column_index, quantity_column_index, order_no_column_index, sku_column_index)
                else:  # 旧格式：(row_idx, sku_list)
                    row_idx, sku_list = row_data
                    self.split_multi_sku_row(worksheet, row_idx, sku_list, size_column_index, processing_column_index, material_column_index, quantity_column_index, order_no_column_index, sku_column_index)
            
            # 所有多SKU行拆分完成后，统一执行合并操作
            if multi_sku_rows:
                print(f"所有多SKU行拆分完成，开始统一合并操作...")
                self.merge_cells_by_i_column(worksheet, order_no_column_index)
                print(f"合并操作完成")
            
            for row_idx in range(3, worksheet.max_row + 1):
                # 检查是否是空行
                is_empty_row = True
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        is_empty_row = False
                        break
                
                # 跳过空行
                if is_empty_row:
                    continue
                
                # 设置正文内容行行高为79
                worksheet.row_dimensions[row_idx].height = 79
                
                # 计算F列数量总和
                if quantity_column_index:
                    quantity_cell = worksheet.cell(row=row_idx, column=quantity_column_index)
                    if quantity_cell.value:
                        try:
                            quantity = float(quantity_cell.value)
                            total_quantity += quantity
                            # 将数量显示为整数格式
                            quantity_cell.value = int(quantity)
                            # 设置产品数量字体为宋体、36号
                            quantity_cell.font = Font(name='宋体', size=36)
                            # 如果数量大于1，设置为红色
                            if int(quantity) > 1:
                                quantity_cell.font = Font(name='宋体', size=36, color='FF0000')
                        except (ValueError, TypeError):
                            pass  # 忽略无法转换为数字的值
                
                # 自动填充序号（A列）
                if sequence_column_index == 1 or not sequence_column_index:  # A列是第1列
                    sequence_cell = worksheet.cell(row=row_idx, column=1)
                    from openpyxl.cell.cell import MergedCell
                    # 如果当前单元格为合并单元格（非起始单元格），则跳过写入以避免“read-only”异常
                    if isinstance(sequence_cell, MergedCell):
                        pass  # 已在合并区域内，跳过写入
                    else:
                        # 检查是否是新增行（标记为跳过序号）
                        if sequence_cell.value == "SKIP_SEQUENCE":
                            sequence_cell.value = None  # 清空标记，不分配序号，不递增计数器保持连续性
                        else:
                            sequence_cell.value = sequence_number
                            # 设置序号字体为宋体、16号
                            sequence_cell.font = Font(name='宋体', size=16)
                            sequence_number += 1
                
                # 应用尺寸映射
                if self.size_mapping and self.size_column and size_column_index:
                    cell = worksheet.cell(row=row_idx, column=size_column_index)
                    if cell.value:
                        # 使用统一尺寸映射函数进行处理
                        cell.value = self._map_size_with_rules(str(cell.value))
                        mapped = True  # 标记为已映射，禁用旧版尺寸逻辑
                        # 旧尺寸映射逻辑已移除
                        
                        # ↓↓↓ 其余旧版逻辑被统一方法替代，已注释保留以便回溯 ↓↓↓
                        if not mapped:
                            filename_standard_size = self.normalize_size_to_standard(size_str)
                            if filename_standard_size:
                                # 尝试完整标准化匹配
                                for original_size, new_size in self.size_mapping.items():
                                    standard_size = self.normalize_size_to_standard(original_size)
                                    if standard_size and filename_standard_size == standard_size:
                                        cell.value = new_size
                                        mapped = True
                                        print(f"✅ 标准化完整匹配成功: '{size_str}' (标准化为 '{filename_standard_size}') -> '{new_size}'")
                                        break
                                
                                # 尝试基础尺寸匹配（同时比较宽高）
                                if not mapped:
                                    w1, h1 = self._parse_standard_size(filename_standard_size)
                                    if w1 and h1:
                                        for original_size, new_size in self.size_mapping.items():
                                            standard_size = self.normalize_size_to_standard(original_size)
                                            if standard_size:
                                                w2, h2 = self._parse_standard_size(standard_size)
                                                if self._is_basic_size_match(w1, h1, w2, h2):
                                                    cell.value = new_size
                                                    mapped = True
                                                    print(f"✅ 基础尺寸双维匹配成功: '{size_str}' ({w1}x{h1}) ≈ {w2}x{h2} -> '{new_size}'")
                                                    break
                        
                        # 第三优先级：相似度匹配（使用现有的calculate_size_similarity函数）
                        if not mapped:
                            best_match = None
                            best_similarity = 0
                            for original_size, new_size in self.size_mapping.items():
                                similarity = self.calculate_size_similarity(size_str, original_size)
                                if similarity > best_similarity and similarity >= 0.9:  # 设置较高的相似度阈值
                                    best_similarity = similarity
                                    best_match = (original_size, new_size)
                            
                            if best_match:
                                cell.value = best_match[1]
                                mapped = True
                                print(f"✅ 相似度匹配成功: '{size_str}' -> '{best_match[1]}' (相似度: {best_similarity:.3f})")
                        
                        # 最后：如果没有找到映射，使用标准化结果或备用逻辑
                        if not mapped:
                            filename_standard_size = self.normalize_size_to_standard(size_str)
                            if filename_standard_size:
                                # 优先使用新的标准化结果，统一维度分隔符为 *
                                cell.value = filename_standard_size.replace(' x ', '*').replace('×', '*')
                                print(f"⚠️ 无映射匹配，使用标准化结果: '{size_str}' -> '{cell.value}'")
                            else:
                                # 尝试从原始内容中提取variants并标准化
                                variants_content = self.extract_variants_content(size_str)
                                if variants_content:
                                    vc_std = self.normalize_size_to_standard(variants_content)
                                    if vc_std:
                                        cell.value = vc_std.replace(' x ', '*').replace('×', '*')
                                        print(f"⚠️ 无映射匹配，使用variants标准化结果: '{size_str}' -> '{cell.value}'")
                                    else:
                                        cell.value = variants_content
                                        print(f"⚠️ 无映射匹配，使用variants原始结果: '{size_str}' -> '{cell.value}'")
                                else:
                                    # 最后才使用旧的process_variants_content作为备选
                                    auto_converted = self.process_variants_content(size_str)
                                    if auto_converted and auto_converted != size_str:
                                        cell.value = auto_converted.replace(' x ', '*').replace('×', '*')
                                        print(f"⚠️ 无映射匹配，使用旧逻辑结果: '{size_str}' -> '{cell.value}'")
                                    else:
                                        print(f"❌ 无法处理尺寸: '{size_str}'，保持原样")
                    # 设置尺寸字体为宋体、18号
                    cell.font = Font(name='宋体', size=18)
                
                # 为订单号列设置字体
                if order_no_column_index:
                    order_no_cell = worksheet.cell(row=row_idx, column=order_no_column_index)
                    from openpyxl.cell.cell import MergedCell
                    # 如果是合并单元格的非起始位置，直接跳过设置字体/样式
                    if isinstance(order_no_cell, MergedCell):
                        pass
                    elif order_no_cell.value:
                        # 设置订单号字体为宋体、18号
                        order_no_cell.font = Font(name='宋体', size=18)
                
                # 应用加工方式和材质配置
                if sku_column_index:
                    sku_cell = worksheet.cell(row=row_idx, column=sku_column_index)
                    if sku_cell.value:
                        sku = str(sku_cell.value).strip()
                        # 设置SKU字体为宋体、18号
                        sku_cell.font = Font(name='宋体', size=18)
                        
                        processing, material = self.get_processing_info_by_sku(sku)
                        
                        # 填充加工方式
                        if processing and processing_column_index:
                            processing_cell = worksheet.cell(row=row_idx, column=processing_column_index)
                            processing_cell.value = processing
                            # 设置加工方式字体为宋体、24号
                            processing_cell.font = Font(name='宋体', size=24)
                        
                        # 填充材质
                        if material and material_column_index:
                            material_cell = worksheet.cell(row=row_idx, column=material_column_index)
                            material_cell.value = material
                            # 设置材质字体为宋体、24号、红色
                            material_cell.font = Font(name='宋体', size=24, color='FF0000')

                        # 二次尺寸处理：仅在用户手动选择了预设时触发（会话内所有相关行均按预设补充处理）
                        try:
                            secondary_flag = getattr(self, '_secondary_processing_triggered_manually', False)
                            print(f"调试：二次处理标志 = {secondary_flag}, 尺寸列索引 = {size_column_index}")
                            if size_column_index and secondary_flag:
                                size_cell2 = worksheet.cell(row=row_idx, column=size_column_index)
                                original_size = str(size_cell2.value) if size_cell2.value else ""
                                print(f"调试：行{row_idx} 原始尺寸 = '{original_size}', 加工方式 = '{processing}'")
                                if size_cell2.value:
                                    # 判断产品类型，只对窗帘和床上三件套进行二次处理
                                    current_preset = getattr(self, 'last_selected_preset', '')
                                    should_process = self._should_apply_secondary_processing(processing, current_preset)
                                    print(f"调试：产品类型判断 - 是否需要二次处理: {should_process}")
                                    
                                    if should_process:
                                        new_size = self.apply_secondary_size_processing(str(size_cell2.value), current_preset)
                                        print(f"调试：二次处理结果 = '{new_size}'")
                                        if new_size:
                                            size_cell2.value = new_size
                                            size_cell2.font = Font(name='宋体', size=18)
                                            print(f"调试：已更新尺寸为 '{new_size}'")
                                    else:
                                        print(f"调试：画或其他产品类型，保留原格式 '{original_size}'")
                        except Exception as e:
                            print(f"二次尺寸处理集成出错: {e}")
            
            # 在第一行的F-H列填入总数量并合并
            if quantity_column_index and total_quantity > 0:
                # 合并F-H列（第6列到第8列）
                worksheet.merge_cells('F1:H1')
                # F列填入"共n件"格式，显示为整数
                cell = worksheet.cell(row=1, column=quantity_column_index)
                cell.value = f"共{int(total_quantity)}件"
                # 设置字体为宋体、20号、红色、居中对齐
                cell.font = Font(name='宋体', size=20, color='FF0000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 为A-H列的所有内容添加边框
            from openpyxl.styles import Border, Side
            
            # 定义边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为A-H列的所有有内容的单元格添加边框
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, 9):  # A-H列（1-8列）
                    cell = worksheet.cell(row=row, column=col)
                    # 只为有内容的单元格或在数据范围内的单元格添加边框
                    if cell.value is not None or row <= worksheet.max_row:
                        cell.border = thin_border
            
            # 统一设置行高（在边框设置之后，避免被覆盖）
            print("🔧 开始统一设置行高...")
            # 设置第一行行高为77
            worksheet.row_dimensions[1].height = 77
            print(f"✅ 第1行高度设置为77")
            
            # 设置第二行行高为44
            worksheet.row_dimensions[2].height = 44
            print(f"✅ 第2行高度设置为44")
            
            # 设置所有数据行（第三行开始）行高为79
            for row_idx in range(3, worksheet.max_row + 1):
                # 检查是否是空行
                is_empty_row = True
                for col_idx in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None and str(cell_value).strip():
                        is_empty_row = False
                        break
                
                # 只为非空行设置行高
                if not is_empty_row:
                    worksheet.row_dimensions[row_idx].height = 79
                    print(f"✅ 第{row_idx}行高度设置为79")
            
            print("🎉 所有行高设置完成！")
            
            # 🔧 导出前清空I列的数字标记
            print("🔧 开始清空I列的数字标记...")
            i_column_index = 9  # I列是第9列
            cleared_count = 0
            
            for row_idx in range(1, worksheet.max_row + 1):
                i_cell = worksheet.cell(row=row_idx, column=i_column_index)
                if i_cell.value is not None:
                    i_cell.value = None  # 清空单元格内容
                    cleared_count += 1
            
            print(f"✅ I列数字标记清空完成！共清空了 {cleared_count} 个单元格")
            
            # 保存到新文件，保留所有图片对象
            workbook.save(export_path)
            
            # 记录导出到每日时间轴
            if custom_name:
                self.record_to_timeline(custom_name, 'excel_exports')
            
            # 提取SKU数据并去除序号尾缀，准备自动搜索
            sku_list = []
            if sku_column_index:
                for row_idx in range(3, worksheet.max_row + 1):  # 现在从第三行开始，因为插入了新行
                    # 检查是否是空行
                    is_empty_row = True
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value is not None and str(cell_value).strip():
                            is_empty_row = False
                            break
                    
                    # 跳过空行
                    if is_empty_row:
                        continue
                    
                    sku_cell = worksheet.cell(row=row_idx, column=sku_column_index)
                    if sku_cell.value:
                        sku = str(sku_cell.value).strip()
                        # 去除序号尾缀（如 -1, -2, -3 等）
                        processed_sku = re.sub(r'-\d+$', '', sku)
                        if processed_sku and processed_sku not in sku_list:
                            sku_list.append(processed_sku)
            
            workbook.close()
            
            # 统计处理信息
            processed_info = []
            processed_info.append("序号自动填充")
            if self.size_mapping and size_column_index:
                processed_info.append("尺寸映射")
            if processing_column_index or material_column_index:
                processed_info.append("加工方式和材质")
            
            info_text = "、".join(processed_info) if processed_info else "数据"
            
            # 静默处理完成，不显示弹窗
            print(f"Excel处理完成：{info_text}，导出到 {export_path}，处理了 {sequence_number - 1} 行数据")
            
            # 多表格处理时不修改搜索框，只静默处理
            if sku_list:
                # 静默处理完成，不修改搜索框
                print(f"已提取 {len(sku_list)} 个SKU并去除序号尾缀，多表格处理中不自动搜索")
            
            # 如果需要打开文件
            if open_after_export:
                try:
                    import subprocess
                    subprocess.run(['start', '', export_path], shell=True, check=True)
                except Exception as e:
                    print(f"无法打开文件 {export_path}: {str(e)}")
            
            # 显示面单名称提示弹窗（仅在单表格处理时显示）
            if not getattr(self, '_is_multi_table_processing', False):
                try:
                    # 导入面单名称弹窗组件
                    from waybill_name_dialog import WaybillNameDialog
                    
                    # 计算数据行数（sequence_number - 1 是最终的数据行数）
                    data_row_count = sequence_number - 1
                    
                    # 检查是否已有全局弹窗实例
                    if self.waybill_dialog is None:
                        # 创建新的面单名称弹窗
                        self.waybill_dialog = WaybillNameDialog(self.root, self)  # 直接传入organizer_instance
                        self.waybill_dialog.show_single_waybill(export_name_for_waybill, data_row_count)
                    else:
                        # 向已存在的弹窗添加新面单
                        self.waybill_dialog.add_waybill_to_existing(export_name_for_waybill, data_row_count)
                    
                except Exception as e:
                    print(f"显示面单名称弹窗时出错: {e}")
            else:
                # 多表格处理时，收集面单名称到列表中
                if not hasattr(self, '_waybill_names_collection'):
                    self._waybill_names_collection = []
                
                try:
                    from waybill_name_dialog import WaybillNameDialog
                    
                    # 计算数据行数（sequence_number - 1 是最终的数据行数）
                    data_row_count = sequence_number - 1
                    
                    # 使用全局弹窗实例进行面单名称转换
                    if self.waybill_dialog is None:
                        self.waybill_dialog = WaybillNameDialog(self.root, self)  # 直接传入organizer_instance
                    
                    waybill_name = self.waybill_dialog.convert_export_name_to_waybill(export_name_for_waybill, data_row_count)
                    self._waybill_names_collection.append(waybill_name)
                    print(f"📋 收集面单名称: {waybill_name}")
                except Exception as e:
                    print(f"收集面单名称时出错: {e}")
            
        except Exception as e:
            messagebox.showerror("导出失败", f"导出处理后的Excel文件时出错：{str(e)}")

    def show_size_processing_dialog(self, sku_count):
        """显示尺寸处理选择弹窗，包含三个选项"""
        dialog = tk.Toplevel(self.root)
        dialog.title("整理尺寸")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        # dialog.attributes('-topmost', True)  # 移除置顶，避免遮挡文件选择框
        dialog.focus_force()
        
        # 立即隐藏窗口，避免在左上角显示
        dialog.withdraw()
        
        # 居中显示 - 先更新布局但窗口仍然隐藏
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 设置好位置后再显示窗口，避免移动效果
        dialog.deiconify()
        
        # 图标和消息
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 消息文本
        message_label = ttk.Label(main_frame, 
                                text=f"检测到Excel文件包含尺寸数据（{sku_count} 个SKU）。\n\n是否整理尺寸并另存到桌面？",
                                font=('Microsoft YaHei', 10),
                                justify=tk.CENTER)
        message_label.pack(pady=(0, 20))
        
        # 按钮框架
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        result = {"value": None}
        
        def on_process_and_open():
            result["value"] = "process_and_open"
            dialog.destroy()
            
        def on_yes():
            result["value"] = "yes"
            dialog.destroy()
            
        def on_no():
            result["value"] = "no"
            dialog.destroy()
        
        def on_key_press(event):
            if event.keysym.lower() == 'y':
                on_yes()
            elif event.keysym.lower() == 'n':
                on_no()
            elif event.keysym == 'Return':
                on_process_and_open()
            elif event.keysym == 'Escape':
                on_no()
        
        # 按钮样式
        button_style = {'width': 12, 'padding': (5, 8)}
        process_button_style = {'width': 16, 'padding': (5, 8)}  # 处理按钮使用更宽的样式
        
        # 处理完打开表格按钮
        process_open_btn = ttk.Button(button_frame, text="是（处理完打开表格）",
                                    command=on_process_and_open, **process_button_style)
        process_open_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # 添加弹性空间，将是和否按钮推向右边
        spacer_frame = ttk.Frame(button_frame)
        spacer_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # 是按钮
        yes_btn = ttk.Button(button_frame, text="✅ 是(Y)", 
                           command=on_yes, **button_style)
        yes_btn.pack(side=tk.RIGHT, padx=(10, 0))
        
        # 否按钮
        no_btn = ttk.Button(button_frame, text="❌ 否(N)", 
                          command=on_no, **button_style)
        no_btn.pack(side=tk.RIGHT, padx=(10, 0))
        
        # 键盘快捷键
        dialog.bind('<Key>', on_key_press)
        dialog.focus_set()
        
        # 等待用户选择
        dialog.wait_window()
        return result["value"]
    
    def process_size_mapping(self, df):
        try:
            # 直接返回原始数据，不弹出尺寸映射对话框
            # 符合预设的会自动匹配，不符合的不做处理
            return df
            
        except Exception as e:
            messagebox.showerror("错误", f"处理尺寸映射时出错：{str(e)}")
            return df
    
    def save_size_mapping(self, dialog):
        """保存尺寸映射配置"""
        try:
            # 更新映射规则
            for size, entry_var in self.mapping_entries.items():
                mapping_value = entry_var.get().strip()
                if mapping_value:
                    self.size_mapping[size] = mapping_value
            
            # 保存到配置文件
            self.save_config()
            
            dialog.destroy()
            messagebox.showinfo("成功", "尺寸映射配置已保存！")
            
        except Exception as e:
            messagebox.showerror("错误", f"保存配置时出错：{str(e)}")
    
    def extract_variants_content(self, text):
        """从文本中提取Variants内容 - 支持多SKUID处理"""
        try:
            if not text:
                return None
            
            text = str(text).strip()
            
            # 检查是否包含多个SKUID
            skuid_matches = re.findall(r'SKUID:', text)
            
            if len(skuid_matches) > 1:
                # 多SKUID情况：按SKUID分割并分别处理
                skuid_blocks = re.split(r'(?=SKUID:)', text)
                skuid_blocks = [block.strip() for block in skuid_blocks if block.strip()]
                
                results = []
                for block in skuid_blocks:
                    # 从每个块中提取Variants内容
                    variants_pattern = r'Variants:\s*(.*?)(?:\s*Warehouse:|$)'
                    match = re.search(variants_pattern, block, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                    
                    if match:
                        variants_content = match.group(1).strip()
                        # 移除可能的颜色信息（如 "/ White", "/ Red"等），但保留数字/Set格式（如 "2Pcs/Set"）
                        if "Pcs/Set" not in variants_content:
                            variants_content = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff\s]+(?=\s*(?:\n|$))', '', variants_content)
                        variants_content = variants_content.strip() if variants_content else None
                        
                        if variants_content:
                            # 进一步处理提取的内容
                            processed_content = self.process_variants_content(variants_content)
                            results.append(processed_content)
                
                # 返回多行结果
                if results:
                    return '\n'.join(results)
                else:
                    return None
            else:
                # 单SKUID情况：使用原有逻辑
                variants_pattern = r'Variants:\s*(.*?)(?:\s*Warehouse:|$)'
                match = re.search(variants_pattern, text, re.IGNORECASE | re.MULTILINE | re.DOTALL)
                
                if match:
                    variants_content = match.group(1).strip()
                    # 移除可能的颜色信息（如 "/ White", "/ Red"等），但保留数字/Set格式（如 "2Pcs/Set"）
                    if "Pcs/Set" not in variants_content:
                        variants_content = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff\s]+(?=\s*(?:\n|$))', '', variants_content)
                    variants_content = variants_content.strip() if variants_content else None
                    
                    if variants_content:
                        # 进一步处理提取的内容
                        processed_content = self.process_variants_content(variants_content)
                        return processed_content
                    
                    return variants_content
            
            return None
            
        except Exception as e:
            print(f"提取Variants内容时出错: {e}")
            return None
    
    def process_variants_content(self, content):
        """进一步处理Variants内容，处理cm和英寸的转换，支持新的格式"""
        try:
            if not content:
                return content
            
            # 1. 处理西班牙语pulgadas格式
            pulgadas_result = self.handle_spanish_pulgadas(content)
            if pulgadas_result:
                return pulgadas_result
            
            # 2. 处理cm格式中间分隔符问题
            cm_slash_result = self.handle_cm_slash_format(content)
            if cm_slash_result:
                return cm_slash_result
            
            # 3. 处理混合英寸格式
            mixed_inch_result = self.handle_mixed_inch_format(content)
            if mixed_inch_result:
                return mixed_inch_result
            
            # 4. 检查是否包含cm - 如果包含cm，提取cm内容和数量
            if 'cm' in content.lower():
                return self.extract_cm_content(content)
            
            # 5. 检查是否只包含英寸 - 如果只有英寸，进行转换
            elif 'inch' in content.lower() or '"' in content or "''" in content:
                return self.convert_inches_to_cm(content)
            
            # 其他情况直接返回原内容
            return content
            
        except Exception as e:
            print(f"处理Variants内容时出错: {e}")
            return content
    
    def extract_cm_content(self, content):
        """从内容中提取cm尺寸和数量信息"""
        try:
            # 首先尝试提取数量信息
            quantity = None
            # 查找各种数量表达格式
            quantity_patterns = [
                r'\*\s*([０-９\d]+)(?:\s|$|个|装)',  # *2, *２, *2个装 (最常见的格式，优先匹配)
                r'[×xX]\s*([０-９\d]+)(?:\s|$|个|装)',  # ×2, x2, X2, ×２
                r'([０-９\d]+)\s*Pcs?/Set',  # 2Pcs/Set, ２Pcs/Set
                r'([０-９\d]+)\s*pcs?',      # 2pcs, ２pcs
                r'([０-９\d]+)\s*PC',        # 2PC, ２PC
                r'([０-９\d]+)\s*PCS',       # 2PCS, ２PCS
                r'Set\s+Of\s+([０-９\d]+)',  # Set Of 2, Set Of ２
                r'([０-９\d]+)\s+pieces?',   # 2 pieces, ２ pieces
                r'([０-９\d]+)\s*个装',       # 2个装, ２个装
                r'\*\s*(一|二|三|四|五|六|七|八|九|十)(?:\s|$)',  # 中文数字
            ]
            
            for pattern in quantity_patterns:
                match = re.search(pattern, content, re.IGNORECASE)
                if match:
                    quantity_raw = match.group(1)
                    # 转换全角数字和中文数字为半角数字
                    quantity = self._normalize_quantity(quantity_raw)
                    break
            
            # 匹配包含cm的尺寸格式，支持多种英寸单位表示和分隔符
            # 支持格式：
            # 1. 30cm/11.81inch*60cm/23.62inc*3 (支持inc缩写)
            # 2. 85 cm/33.46 pulgadas * 55cm / 21.65in (支持pulgadas和空格)
            # 3. 90 cm/35.43 pulgadas x 30 cm/11.81 pulgadas x 2 (支持x分隔符)
            
            # 模式1：标准cm格式 (cm在前) - 增强版，支持更多单位和分隔符
            cm_pattern1 = r'(\d+(?:\.\d+)?)\s*cm\s*(?:/\s*[\d.]+\s*(?:in|inch|inches|inc|pulgadas))?\s*[*×xX]\s*(\d+(?:\.\d+)?)\s*cm\s*(?:/\s*[\d.]+\s*(?:in|inch|inches|inc|pulgadas))?'
            match1 = re.search(cm_pattern1, content, re.IGNORECASE)
            
            # 模式2：英寸在前的格式 (W45inch*H106inch/115cm*270cm)
            cm_pattern2 = r'W(\d+(?:\.\d+)?)inch\*H(\d+(?:\.\d+)?)inch/(\d+(?:\.\d+)?)cm\*(\d+(?:\.\d+)?)cm'
            match2 = re.search(cm_pattern2, content, re.IGNORECASE)
            
            # 模式3：简单的英寸转cm格式
            cm_pattern3 = r'(\d+(?:\.\d+)?)\s*(?:inch|inches|inc)\s*[^/]*/(\d+(?:\.\d+)?)\s*cm\s*[*×xX]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches|inc)\s*[^/]*/(\d+(?:\.\d+)?)\s*cm'
            match3 = re.search(cm_pattern3, content, re.IGNORECASE)
            
            # 模式4：处理pulgadas格式 - 90 cm/35.43 pulgadas x 30 cm/11.81 pulgadas x 2
            cm_pattern4 = r'(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*pulgadas\s*[xX]\s*(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*pulgadas\s*[xX]\s*(\d+)'
            match4 = re.search(cm_pattern4, content, re.IGNORECASE)
            
            width_cm = None
            height_cm = None
            
            if match4:  # pulgadas x格式
                width_cm = match4.group(1)
                height_cm = match4.group(2)
                if not quantity:  # 如果之前没找到数量，使用这里的
                    quantity = int(match4.group(3))
            elif match2:  # W45inch*H106inch/115cm*270cm 格式
                width_cm = match2.group(3)   # 115
                height_cm = match2.group(4)  # 270
            elif match1:  # 标准cm格式
                width_cm = match1.group(1)
                height_cm = match1.group(2)
            elif match3:  # 简单英寸转cm格式
                width_cm = match3.group(2)
                height_cm = match3.group(4)
            
            if width_cm and height_cm:
                if quantity:
                    return f"{width_cm}cm*{height_cm}cm*{quantity}"
                else:
                    return f"{width_cm}cm*{height_cm}cm"
            
            # 如果没有匹配到标准格式，尝试简单提取所有cm相关内容
            cm_simple_pattern = r'(\d+(?:\.\d+)?cm[^/]*(?:\*\d+(?:\.\d+)?cm[^/]*)*(?:\*\d+)?)'
            simple_match = re.search(cm_simple_pattern, content, re.IGNORECASE)
            if simple_match:
                extracted = simple_match.group(1)
                if quantity and quantity not in extracted:
                    return f"{extracted}*{quantity}"
                return extracted
            
            return content
            
        except Exception as e:
            print(f"提取cm内容时出错: {e}")
            return content
    
    def convert_inches_to_cm(self, content):
        """将英寸内容转换为厘米，如果已经是厘米则直接返回"""
        try:
            # 优先处理复杂的混合英寸格式（如 Set Of 2, 2pcs 等）
            mixed_result = self.handle_mixed_inch_format(content)
            if mixed_result:
                return mixed_result
            
            # 检查是否已经包含厘米格式，如果是则直接返回厘米值
            cm_patterns = [
                # 厘米格式（带数量）- 85cm*55cm*3 或 85cm/33.46in*55cm/21.65in*3
                r'(\d+(?:\.\d+)?)\s*cm(?:/[\d.]+(?:in|inch|pulgadas|"|\'\')?)?[*×x](\d+(?:\.\d+)?)\s*cm(?:/[\d.]+(?:in|inch|pulgadas|"|\'\')?)?[*×x](\d+)',
                
                # 厘米格式（不带数量）- 85cm*55cm 或 85cm/33.46in*55cm/21.65in
                r'(\d+(?:\.\d+)?)\s*cm(?:/[\d.]+(?:in|inch|pulgadas|"|\'\')?)?[*×x](\d+(?:\.\d+)?)\s*cm(?:/[\d.]+(?:in|inch|pulgadas|"|\'\')?)?(?![*×x]\s*\d)'
            ]
            
            # 先检查厘米格式
            for pattern in cm_patterns:
                match = re.search(pattern, content, re.IGNORECASE)
                if match:
                    width = int(float(match.group(1)))
                    height = int(float(match.group(2)))
                    quantity = match.group(3) if len(match.groups()) >= 3 and match.group(3) else None
                    
                    if quantity:
                        return f"{width}cm*{height}cm*{quantity}"
                    else:
                        return f"{width}cm*{height}cm"
            
            # 只有在没有厘米格式时，才处理英寸格式
            inch_patterns = [
                # 纯英寸格式（带数量）- 52inch*84inch*2
                (r'(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')\s*[*×x]\s*(\d+)', True),
                
                # 纯英寸格式（不带数量）- 52inch*84inch
                (r'(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')\s*(?![*×x]\s*\d)', False),
                
                # W*H英寸格式（带数量）
                (r'W(\d+(?:\.\d+)?)\s*[*×x]\s*H(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')?\s+(\d+)\s*pcs', True),
                
                # W*H英寸格式（不带数量）
                (r'W(\d+(?:\.\d+)?)\s*[*×x]\s*H(\d+(?:\.\d+)?)\s*(?:in|inch|"|\'\')?\s*$', False)
            ]
            
            for pattern, has_quantity in inch_patterns:
                match = re.search(pattern, content, re.IGNORECASE)
                if match:
                    width = float(match.group(1))
                    height = float(match.group(2))
                    quantity = match.group(3) if has_quantity and len(match.groups()) >= 3 and match.group(3) else None
                    
                    # 英寸格式，转换为厘米 (1英寸 = 2.54厘米)
                    width_result = int(width * 2.54)
                    height_result = int(height * 2.54)
                    
                    if quantity:
                        return f"{width_result}cm*{height_result}cm*{quantity}"
                    else:
                        return f"{width_result}cm*{height_result}cm"
            
            return content
            
        except Exception as e:
            print(f"转换英寸到厘米时出错: {e}")
            return content

    def handle_spanish_pulgadas(self, content):
        """处理西班牙语pulgadas格式"""
        try:
            # 匹配格式：51,96 pulgadas de ancho x 84 pulgadas x 2 piezas
            # 或者：51.96 pulgadas de ancho x 84 pulgadas x 2 piezas
            pulgadas_pattern = r'(\d+[,.]?\d*)\s+pulgadas\s+de\s+ancho\s+x\s+(\d+)\s+pulgadas\s+x\s+(\d+)\s+piezas'
            match = re.search(pulgadas_pattern, content, re.IGNORECASE)
            
            if match:
                # 提取数字，处理逗号分隔的小数
                width_str = match.group(1).replace(',', '.')
                width = float(width_str)
                height = float(match.group(2))
                quantity = int(match.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并四舍五入取整数
                width_cm = round(width * 2.54)
                height_cm = round(height * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            return None
            
        except Exception as e:
            print(f"处理西班牙语pulgadas格式时出错: {e}")
            return None

    def handle_cm_slash_format(self, content):
        """处理cm格式中间分隔符问题"""
        try:
            # 匹配格式：100*220cm*2/39.37*86.61inch*2
            # 重点是识别前半部分的100*220cm*2为连续整体
            cm_slash_pattern = r'(\d+)\*(\d+)cm\*(\d+)/[\d.*]+inch\*\d+'
            match = re.search(cm_slash_pattern, content, re.IGNORECASE)
            
            if match:
                width = int(match.group(1))
                height = int(match.group(2))
                quantity = int(match.group(3))
                
                result = f"{width}cm*{height}cm*{quantity}"
                return result
            
            return None
            
        except Exception as e:
            print(f"处理cm分隔符格式时出错: {e}")
            return None

    def handle_mixed_inch_format(self, content):
        """处理混合英寸格式，支持多种数量表达方式"""
        try:
            # 模式1: 52"X 63"(51.97inch*62.99inch) Set Of 2 (带括号)
            pattern1 = r'(\d+)"?\s*[Xx]\s*(\d+)"?\s*\([^)]*\)\s*Set\s+Of\s+(\d+)'
            match1 = re.search(pattern1, content, re.IGNORECASE)
            
            if match1:
                width_inch = int(match1.group(1))
                height_inch = int(match1.group(2))
                quantity = int(match1.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并取整数
                width_cm = int(width_inch * 2.54)
                height_cm = int(height_inch * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            # 模式2: W39.5in/inch x H87in/inch X 2PC/PCS (忽略W和H前缀，识别X 2PC/PCS)
            pattern2 = r'[WwHh]?(\d+(?:\.\d+)?)\s*(?:in|inch)\s*[xX]\s*[WwHh]?(\d+(?:\.\d+)?)\s*(?:in|inch)\s+[Xx]\s+(\d+)(?:PC|PCS|Pcs|pcs)?'
            match2 = re.search(pattern2, content, re.IGNORECASE)
            
            if match2:
                width_inch = float(match2.group(1))
                height_inch = float(match2.group(2))
                quantity = int(match2.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并取整数
                width_cm = int(width_inch * 2.54)
                height_cm = int(height_inch * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            # 模式3: 通用的 "数字 + 单位 + X/x + 数字 + 单位 + 数量词" 格式
            # 支持更多变体如: 39.5"x87" X 2, 39.5inchx87inch X 2PCS, 等
            pattern3 = r'[WwHh]?(\d+(?:\.\d+)?)(?:in|inch|"|\'\')?\s*[xX×*]\s*[WwHh]?(\d+(?:\.\d+)?)(?:in|inch|"|\'\')?\s+[Xx]\s+(\d+)(?:PC|PCS|Pcs|pcs|PIECES?|pieces?)?'
            match3 = re.search(pattern3, content, re.IGNORECASE)
            
            if match3:
                width_inch = float(match3.group(1))
                height_inch = float(match3.group(2))
                quantity = int(match3.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并取整数
                width_cm = int(width_inch * 2.54)
                height_cm = int(height_inch * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            # 模式4: 简单的 Set Of 格式，不需要括号
            pattern4 = r'(\d+)"?\s*[Xx]\s*(\d+)"?\s*Set\s+Of\s+(\d+)'
            match4 = re.search(pattern4, content, re.IGNORECASE)
            
            if match4:
                width_inch = int(match4.group(1))
                height_inch = int(match4.group(2))
                quantity = int(match4.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并取整数
                width_cm = int(width_inch * 2.54)
                height_cm = int(height_inch * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            # 模式5: 新增 - 处理 52"X 84"(51.97inch*83.86inch) 2pcs 格式
            pattern5 = r'(\d+)"?\s*[Xx]\s*(\d+)"?\s*\([^)]*\)\s*(\d+)\s*pcs?'
            match5 = re.search(pattern5, content, re.IGNORECASE)
            
            if match5:
                width_inch = int(match5.group(1))
                height_inch = int(match5.group(2))
                quantity = int(match5.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并取整数
                width_cm = int(width_inch * 2.54)
                height_cm = int(height_inch * 2.54)
                
                result = f"{width_cm}cm*{height_cm}cm*{quantity}"
                return result
            
            return None
            
        except Exception as e:
            print(f"处理混合英寸格式时出错: {e}")
            return None

    def _parse_standard_size(self, standard_size_str):
        """将标准化尺寸字符串解析为 (width_cm:int, height_cm:int)。失败返回 (None, None)。"""
        if not standard_size_str:
            return (None, None)
        # 仅取前两段，忽略数量后缀
        parts = standard_size_str.split('*')[:2]
        if len(parts) < 2:
            return (None, None)
        try:
            w = re.sub(r"[^0-9]", "", parts[0])
            h = re.sub(r"[^0-9]", "", parts[1])
            if not w or not h:
                return (None, None)
            return (int(w), int(h))
        except Exception:
            return (None, None)

    def _is_basic_size_match(self, w1, h1, w2, h2):
        """双维比较，允许小范围误差。"""
        if w1 is None or h1 is None or w2 is None or h2 is None:
            return False
        return (abs(w1 - w2) <= self.BASIC_SIZE_TOLERANCE_CM and
                abs(h1 - h2) <= self.BASIC_SIZE_TOLERANCE_CM)

    # ------------------------------------------------------------------
    # 🆕 统一尺寸映射入口，供各处调用
    # ------------------------------------------------------------------
    def _map_size_with_rules(self, size_str):
        """根据配置 self.size_mapping 对尺寸文本进行映射并返回结果。
        流程：精确匹配 → 标准化完整匹配 → 双维基础匹配 → 相似度匹配 → 回退逻辑"""
        try:
            if not size_str:
                return size_str

            original_value = str(size_str)
            # 若无映射表，直接返回标准化或原值
            if not getattr(self, 'size_mapping', None):
                return self.normalize_size_to_standard(original_value) or original_value

            # 1️⃣ 精确匹配
            for original_size, new_size in self.size_mapping.items():
                if original_value.strip() == original_size.strip():
                    print(f"✅ 精确匹配成功: '{original_value}' -> '{new_size}'")
                    return new_size

            # 2️⃣ 标准化完整匹配
            filename_standard_size = self.normalize_size_to_standard(original_value)
            if filename_standard_size:
                for original_size, new_size in self.size_mapping.items():
                    standard_size = self.normalize_size_to_standard(original_size)
                    if standard_size and filename_standard_size == standard_size:
                        print(f"✅ 标准化完整匹配成功: '{original_value}' -> '{new_size}'")
                        return new_size

                # 3️⃣ 基础尺寸双维匹配
                w1, h1 = self._parse_standard_size(filename_standard_size)
                if w1 and h1:
                    for original_size, new_size in self.size_mapping.items():
                        standard_size = self.normalize_size_to_standard(original_size)
                        if standard_size:
                            w2, h2 = self._parse_standard_size(standard_size)
                            if self._is_basic_size_match(w1, h1, w2, h2):
                                print(f"✅ 基础尺寸双维匹配成功: '{original_value}' -> '{new_size}'")
                                return new_size

            # 4️⃣ 相似度匹配
            best_match = None
            best_similarity = 0
            for original_size, new_size in self.size_mapping.items():
                similarity = self.calculate_size_similarity(original_value, original_size)
                if similarity > best_similarity and similarity >= 0.9:
                    best_similarity = similarity
                    best_match = new_size
            if best_match:
                print(f"✅ 相似度匹配成功: '{original_value}' -> '{best_match}' (相似度: {best_similarity:.3f})")
                return best_match

            # 5️⃣ 回退逻辑
            if filename_standard_size:
                return filename_standard_size.replace(' x ', '*').replace('×', '*')

            variants_content = self.extract_variants_content(original_value)
            if variants_content:
                vc_std = self.normalize_size_to_standard(variants_content)
                if vc_std:
                    return vc_std.replace(' x ', '*').replace('×', '*')
                else:
                    return variants_content

            auto_converted = self.process_variants_content(original_value)
            if auto_converted and auto_converted != original_value:
                return auto_converted.replace(' x ', '*').replace('×', '*')

            return original_value
        except Exception as e:
            print(f"_map_size_with_rules error: {e}")
            return size_str

    def normalize_size_to_standard(self, size_text):
        """将各种尺寸格式标准化为统一格式 (width x height)，支持从复杂文本中提取，支持 + 复合尺寸"""
        try:
            if not size_text:
                return None
            
            size_text = str(size_text).strip()
            
            # 处理 + 复合尺寸 (例如: A+B)
            if '+' in size_text:
                parts = size_text.split('+')
                normalized_parts = []
                for part in parts:
                    norm = self._normalize_single_size(part.strip())
                    if norm:
                        normalized_parts.append(norm)
                    else:
                        normalized_parts.append(part.strip())
                return '+'.join(normalized_parts)
            
            return self._normalize_single_size(size_text)
            
        except Exception as e:
            print(f"标准化尺寸时出错: {e}")
            return size_text

    def _normalize_single_size(self, size_text):
        """内部方法：标准化单个尺寸单元"""
        try:
            # 1. 如果包含 Variants: 信息，优先从中提取完整内容
            variants_match = re.search(r'Variants:\s*(.+?)(?:\s+Warehouse:|$)', size_text, re.IGNORECASE | re.DOTALL)
            if variants_match:
                variants_text = variants_match.group(1).strip()
                size_in_variants = self._extract_size_from_text(variants_text)
                if size_in_variants:
                    return size_in_variants
                else:
                    print(f"🔄 无法识别尺寸格式，返回原始Variants内容: {variants_text}")
                    return variants_text
            
            # 2. 尝试从整个文本中提取尺寸
            extracted_size = self._extract_size_from_text(size_text)
            if extracted_size:
                return extracted_size
            
            # 3. 如果完全无法识别，尝试回退逻辑
            variants_fallback_match = re.search(r'Variants:\s*([^W]*?)(?=\s*Warehouse:|$)', size_text)
            if variants_fallback_match:
                variants_content = variants_fallback_match.group(1).strip()
                return variants_content
            
            return size_text
        except Exception:
            return size_text
    
    def _extract_size_components(self, text):
        """组件化解析：从文本中提取尺寸相关的各种组件"""
        components = {
            'numbers': [],          # 数字列表
            'units': [],           # 单位列表 (cm, inch, inches, pulgadas, etc.)
            'separators': [],      # 分隔符 (x, ×, *, etc.)
            'brackets': [],        # 括号内容
            'language_markers': [], # 语言标识 (Ancho, Alto, Width, Height, etc.)
            'quantities': [],      # 数量词 (piezas, pieces, pcs, *2, x2, etc.)
            'colors': []           # 颜色信息
        }
        
        # 提取括号内容（最高优先级）
        bracket_patterns = [
            r'\(\s*(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)\s*cm\s*\)',  # (113x230cm)
            r'\(\s*(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)\s*\)',       # (113x230)
        ]
        
        for pattern in bracket_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                components['brackets'].append({
                    'width': float(match.group(1)),
                    'height': float(match.group(2)),
                    'unit': 'cm' if 'cm' in match.group(0).lower() else 'unknown',
                    'full_match': match.group(0),
                    'priority': 10  # 最高优先级
                })
        
        # 提取数量信息
        quantity_patterns = [
            (r'(\d+)\s*piezas?\b', 'spanish'),
            (r'(\d+)\s*pieces?\b', 'english'),
            (r'(\d+)\s*pcs?\b', 'abbreviation'),
            (r'\*\s*(\d+)(?:\s*$|\s+)', 'multiplier'),
            (r'x\s*(\d+)(?:\s*$|\s+)', 'multiplier'),
        ]
        
        for pattern, qtype in quantity_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                quantity = int(match.group(1))
                if 1 <= quantity <= 10:  # 合理范围
                    components['quantities'].append({
                        'value': quantity,
                        'type': qtype,
                        'full_match': match.group(0),
                        'position': match.start()
                    })
        
        # 提取语言标识的尺寸格式
        language_patterns = [
            # 西班牙语格式
            (r'Ancho\s+(\d+(?:\.\d+)?)\s*cm\s*[*×x]\s*Alto\s+(\d+(?:\.\d+)?)\s*cm', 'spanish_cm'),
            # 英语格式
            (r'Width\s+(\d+(?:\.\d+)?)\s*(?:inches?|cm)\s*[*×x]\s*Height\s+(\d+(?:\.\d+)?)\s*(?:inches?|cm)', 'english_labeled'),
            # W/H格式
            (r'W(\d+(?:\.\d+)?)\s*(?:inch|cm)\s*[*×x]\s*H(\d+(?:\.\d+)?)\s*(?:inch|cm)', 'wh_format'),
        ]
        
        for pattern, ltype in language_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                # 检测单位
                unit = 'cm' if 'cm' in match.group(0).lower() else ('inch' if 'inch' in match.group(0).lower() else 'unknown')
                components['language_markers'].append({
                    'width': float(match.group(1)),
                    'height': float(match.group(2)),
                    'unit': unit,
                    'type': ltype,
                    'full_match': match.group(0),
                    'priority': 8
                })
        
        # 提取标准数字+单位格式
        standard_patterns = [
            # 新增：对单单位简写格式（数字*数字 单位）的高优先级识别支持 (计划书 1.8 核心改进)
            (r'(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:cm|inch|inches|in|pulgadas|inc)\b', 'suffix_unit', 9.5),
            (r'(\d+(?:\.\d+)?)\s*cm\s*[*×x]\s*(\d+(?:\.\d+)?)\s*cm', 'standard_cm', 9),
            (r'(\d+(?:\.\d+)?)\s*(?:inch|inches|pulgadas)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches|pulgadas)', 'standard_inches', 7),
            (r'(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:\'\'|")?', 'quote_inches', 6),
            # 添加混合cm/inches格式支持
            (r'(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*(?:inch|inches)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*(?:inch|inches)', 'mixed_cm_inches', 8),
            (r'(\d+(?:\.\d+)?)\s*(?:inch|inches)/[\d.]+\s*cm\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)/[\d.]+\s*cm', 'mixed_inches_cm', 8),
            # 添加混合cm/pulgadas格式支持
            (r'(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*pulgadas\s*[*×x]\s*(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*pulgadas', 'mixed_cm_pulgadas', 8),
        ]
        
        for pattern, stype, priority in standard_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                # 对于混合格式，优先使用cm作为单位
                if stype in ['mixed_cm_inches', 'mixed_inches_cm', 'mixed_cm_pulgadas']:
                    unit = 'cm'
                else:
                    unit = 'cm' if 'cm' in match.group(0).lower() else 'inch'
                components['units'].append({
                    'width': float(match.group(1)),
                    'height': float(match.group(2)),
                    'unit': unit,
                    'type': stype,
                    'full_match': match.group(0),
                    'priority': priority
                })
        
        return components

    def _extract_size_from_text(self, text):
        """组件化尺寸提取 - 新版本，智能处理各种格式组合"""
        try:
            if not text:
                return None
            
            original_text = text.strip()
            print(f"🔍 [新版] 处理文本: {original_text}")
            
            # 第一步：组件提取
            components = self._extract_size_components(text)
            print(f"🧩 提取的组件: {components}")
            
            # 第二步：智能优先级处理
            size_candidates = []
            
            # 优先级1: 括号内容（最可靠）
            if components['brackets']:
                for bracket in components['brackets']:
                    if bracket['unit'] == 'cm':
                        size_candidates.append({
                            'width': bracket['width'],
                            'height': bracket['height'],
                            'unit': 'cm',
                            'priority': 10,
                            'source': 'brackets_cm'
                        })
                        print(f"🎯 发现括号厘米格式: {bracket['width']}x{bracket['height']}cm")
            
            # 优先级2: 语言标识格式
            if components['language_markers']:
                for marker in components['language_markers']:
                    size_candidates.append({
                        'width': marker['width'],
                        'height': marker['height'],
                        'unit': marker['unit'],
                        'priority': marker['priority'],
                        'source': f"language_{marker['type']}"
                    })
                    print(f"🌐 发现语言标识格式: {marker['type']}")
            
            # 优先级3: 标准单位格式
            if components['units']:
                for unit_info in components['units']:
                    size_candidates.append({
                        'width': unit_info['width'],
                        'height': unit_info['height'],
                        'unit': unit_info['unit'],
                        'priority': unit_info['priority'],
                        'source': f"standard_{unit_info['type']}"
                    })
                    print(f"📏 发现标准单位格式: {unit_info['type']}")
            
            # 第三步：选择最佳候选
            if size_candidates:
                # 按优先级排序，选择最高优先级的
                best_candidate = max(size_candidates, key=lambda x: x['priority'])
                print(f"🏆 选择最佳候选: {best_candidate}")
                
                # 单位转换
                if best_candidate['unit'] == 'inch':
                    width_cm = round(best_candidate['width'] * 2.54)
                    height_cm = round(best_candidate['height'] * 2.54)
                else:
                    width = best_candidate['width']
                    height = best_candidate['height']
                    width_cm = int(width) if width == int(width) else width
                    height_cm = int(height) if height == int(height) else height
                
                # 第四步：处理数量后缀
                quantity_value = 1
                if components['quantities']:
                    # 选择最合适的数量
                    quantity = max(components['quantities'], key=lambda x: x['value'] if x['type'] != 'multiplier' else x['value'] * 2)
                    quantity_value = quantity['value']
                    print(f"📊 发现数量: {quantity_value}")
                
                # 格式化输出
                if isinstance(width_cm, float):
                    width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                else:
                    width_str = str(width_cm)
                
                if isinstance(height_cm, float):
                    height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                else:
                    height_str = str(height_cm)
                
                # 所有产品在第一次处理时都返回CM格式，让二次处理来决定最终格式
                quantity_suffix = f"*{quantity_value}" if quantity_value > 1 else ""
                result = f"{width_str}CM*{height_str}CM{quantity_suffix}"
                
                print(f"✅ [新版] 尺寸提取成功: {original_text} -> {result} (来源: {best_candidate['source']})")
                return result
            
            # 第五步：回退逻辑 - 智能处理复杂格式
            print(f"❌ [新版] 未找到匹配的尺寸格式，使用智能回退逻辑")
            
            # 尝试智能解析复杂混合格式
            fallback_result = self._smart_fallback_parsing(original_text)
            if fallback_result:
                print(f"✅ 智能回退成功: {original_text} -> {fallback_result}")
                return fallback_result
            
            # 最后的回退：提取Variants内容
            variants_match = re.search(r'Variants:\s*([^W]*?)(?=\s*Warehouse:|$)', original_text)
            if variants_match:
                variants_content = variants_match.group(1).strip()
                print(f"🔄 最终回退: 提取Variants内容: {variants_content}")
                return variants_content
            return original_text
            
        except Exception as e:
            print(f"[新版] 从文本提取尺寸时出错: {e}")
            return original_text if 'original_text' in locals() else text

    def _smart_fallback_parsing(self, text):
        """智能回退解析 - 处理复杂混合格式"""
        try:
            # 提取Variants内容
            variants_match = re.search(r'Variants:\s*([^\n]+)', text)
            if variants_match:
                variants_text = variants_match.group(1).strip()
            else:
                variants_text = text
            
            # 移除颜色信息
            variants_text = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff\s]+(?=\s*$)', '', variants_text)
            
            # 新增：处理用户案例1 - Green / 30cm/11.81inch*60cm/23.62inc*3
            # 匹配格式：30cm/11.81inch*60cm/23.62inc*3 (支持inc缩写)
            case1_pattern = r'(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*(?:inch|inc)\s*\*\s*(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*(?:inch|inc)\s*\*\s*(\d+)'
            case1_match = re.search(case1_pattern, variants_text, re.IGNORECASE)
            if case1_match:
                width_cm = float(case1_match.group(1))
                height_cm = float(case1_match.group(2))
                quantity = int(case1_match.group(3))
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                return f"{width_str}CM*{height_str}CM*{quantity}"
            
            # 新增：处理用户案例2 - 85 cm/33.46 pulgadas * 55cm / 21.65in
            # 匹配格式：85 cm/33.46 pulgadas * 55cm / 21.65in (支持空格和不同单位)
            case2_pattern = r'(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*pulgadas\s*\*\s*(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*(?:in|inch)'
            case2_match = re.search(case2_pattern, variants_text, re.IGNORECASE)
            if case2_match:
                width_cm = float(case2_match.group(1))
                height_cm = float(case2_match.group(2))
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理用户案例3 - 90 cm/35.43 pulgadas x 30 cm/11.81 pulgadas x 2 / Rojo
            # 匹配格式：90 cm/35.43 pulgadas x 30 cm/11.81 pulgadas x 2 (支持x分隔符)
            case3_pattern = r'(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*pulgadas\s*[xX]\s*(\d+(?:\.\d+)?)\s*cm\s*/\s*[\d.]+\s*pulgadas\s*[xX]\s*(\d+)'
            case3_match = re.search(case3_pattern, variants_text, re.IGNORECASE)
            if case3_match:
                width_cm = float(case3_match.group(1))
                height_cm = float(case3_match.group(2))
                quantity = int(case3_match.group(3))
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                return f"{width_str}CM*{height_str}CM*{quantity}"
            
            # 优先处理西班牙语pulgadas格式: "51,96 pulgadas de ancho x 84 pulgadas x 2 piezas"
            pulgadas_pattern = r'(\d+[,.]?\d*)\s+pulgadas\s+de\s+ancho\s+x\s+(\d+)\s+pulgadas\s+x\s+(\d+)\s+piezas'
            pulgadas_match = re.search(pulgadas_pattern, variants_text, re.IGNORECASE)
            if pulgadas_match:
                # 提取数字，处理逗号分隔的小数
                width_str = pulgadas_match.group(1).replace(',', '.')
                width = float(width_str)
                height = float(pulgadas_match.group(2))
                quantity = int(pulgadas_match.group(3))
                
                # 转换为厘米 (1英寸 = 2.54厘米) 并四舍五入取整数
                width_cm = round(width * 2.54)
                height_cm = round(height * 2.54)
                
                return f"{width_cm}CM*{height_cm}CM*{quantity}"
            
            # 处理cm格式中间分隔符问题: "100*220cm*2/39.37*86.61inch*2"
            cm_slash_pattern = r'(\d+)\*(\d+)cm\*(\d+)/[\d.*]+inch\*\d+'
            cm_slash_match = re.search(cm_slash_pattern, variants_text, re.IGNORECASE)
            if cm_slash_match:
                width = int(cm_slash_match.group(1))
                height = int(cm_slash_match.group(2))
                quantity = int(cm_slash_match.group(3))
                
                return f"{width}CM*{height}CM*{quantity}"
            
            # 尝试解析复杂混合格式: "25 cm/9.84 inches * 60 cm/23.62 inches * 3"
            complex_mixed_pattern = r'(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*(?:inch|inches)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*(?:inch|inches)(?:\s*[*×x]\s*(\d+))?'
            match = re.search(complex_mixed_pattern, variants_text, re.IGNORECASE)
            if match:
                width = float(match.group(1))
                height = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                # 格式化输出
                width_str = f"{width:.1f}" if width != int(width) else f"{int(width)}"
                height_str = f"{height:.1f}" if height != int(height) else f"{int(height)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 尝试解析其他复杂格式
            # 格式: "19.68 inches/50 cm * 7.87 inches/20 cm * 3"
            inches_cm_pattern = r'([\d.]+)\s*(?:inch|inches)/(\d+(?:\.\d+)?)\s*cm\s*[*×x]\s*([\d.]+)\s*(?:inch|inches)/(\d+(?:\.\d+)?)\s*cm(?:\s*[*×x]\s*(\d+))?'
            match = re.search(inches_cm_pattern, variants_text, re.IGNORECASE)
            if match:
                width_cm = float(match.group(2))
                height_cm = float(match.group(4))
                quantity = int(match.group(5)) if match.group(5) else 1
                
                # 格式化数字：如果四舍五入后是整数则不显示小数点，否则显示一位小数
                width_str = f"{int(round(width_cm))}" if round(width_cm, 1) == round(width_cm) else f"{width_cm:.1f}"
                height_str = f"{int(round(height_cm))}" if round(height_cm, 1) == round(height_cm) else f"{height_cm:.1f}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理三维英寸格式: "15.16 * 0.79 * 78.74 inches"
            # 根据用户说明：宽度15.16英寸，包裹中的产品数量0.79，高度78.74英寸
            three_dim_inches_pattern = r'(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)'
            match = re.search(three_dim_inches_pattern, variants_text, re.IGNORECASE)
            if match:
                dim1 = float(match.group(1))  # 宽度
                dim2 = float(match.group(2))  # 包裹中的产品数量
                dim3 = float(match.group(3))  # 高度
                
                # 取宽度和高度
                width_inches = dim1
                height_inches = dim3
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                # 将中间的数量值转换为整数数量（0.79 * 2.54 ≈ 2）
                quantity = round(dim2 * 2.54)
                
                # 格式化数字：如果四舍五入后是整数则不显示小数点，否则显示一位小数
                width_str = f"{int(round(width_cm))}" if round(width_cm, 1) == round(width_cm) else f"{width_cm:.1f}"
                height_str = f"{int(round(height_cm))}" if round(height_cm, 1) == round(height_cm) else f"{height_cm:.1f}"
                
                # 始终包含数量，即使是1
                return f"{width_str}CM*{height_str}CM*{quantity}"
            
            # 新增：处理cm/in缩写格式: "60cm/23.62in*60cm/23.62in*1"
            cm_in_pattern = r'(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*in\s*[*×x]\s*(\d+(?:\.\d+)?)\s*cm/[\d.]+\s*in(?:\s*[*×x]\s*(\d+))?'
            match = re.search(cm_in_pattern, variants_text, re.IGNORECASE)
            if match:
                width = float(match.group(1))
                height = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                width_str = f"{width:.1f}" if width != int(width) else f"{int(width)}"
                height_str = f"{height:.1f}" if height != int(height) else f"{int(height)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理纯英寸格式: "25.59 inches * 23.03 inches"
            pure_inches_pattern = r'(\d+(?:\.\d+)?)\s*(?:inch|inches)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)(?:\s*/\s*(\d+))?'
            match = re.search(pure_inches_pattern, variants_text, re.IGNORECASE)
            if match:
                width_inches = float(match.group(1))
                height_inches = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理带数量的英寸格式: "25 * 23 inches / 1"
            inches_with_quantity_pattern = r'(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)\s*/\s*(\d+)'
            match = re.search(inches_with_quantity_pattern, variants_text, re.IGNORECASE)
            if match:
                width_inches = float(match.group(1))
                height_inches = float(match.group(2))
                quantity = int(match.group(3))
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理X格式的英寸: "30.3X78.7 inches", "37.4X82.6 inch"
            x_inches_pattern = r'(\d+(?:\.\d+)?)\s*[Xx]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)(?:\s*/\s*(\d+))?'
            match = re.search(x_inches_pattern, variants_text, re.IGNORECASE)
            if match:
                width_inches = float(match.group(1))
                height_inches = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理x格式的英寸: "23.03 x 25.59 inches"
            x_space_inches_pattern = r'(\d+(?:\.\d+)?)\s+x\s+(\d+(?:\.\d+)?)\s*(?:inch|inches)(?:\s*/\s*(\d+))?'
            match = re.search(x_space_inches_pattern, variants_text, re.IGNORECASE)
            if match:
                width_inches = float(match.group(1))
                height_inches = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理带pieces的英寸格式: "11.81 * 70.8 inches * 2 pieces"
            pieces_pattern = r'(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:inch|inches)\s*[*×x]\s*(\d+)\s*pieces?'
            match = re.search(pieces_pattern, variants_text, re.IGNORECASE)
            if match:
                width_inches = float(match.group(1))
                height_inches = float(match.group(2))
                quantity = int(match.group(3))
                
                # 转换为厘米
                width_cm = width_inches * 2.54
                height_cm = height_inches * 2.54
                
                width_str = f"{width_cm:.1f}" if width_cm != int(width_cm) else f"{int(width_cm)}"
                height_str = f"{height_cm:.1f}" if height_cm != int(height_cm) else f"{int(height_cm)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            # 新增：处理纯数字格式: "65*58.5" (改进负向前瞻，检测整个字符串中的单位标识符)
            pure_numbers_pattern = r'(\d+(?:\.\d+)?)\s*[*×x]\s*(\d+(?:\.\d+)?)(?:\s*[*×x]\s*(\d+))?(?!.*(?:cm|inch|inches|in))'
            match = re.search(pure_numbers_pattern, variants_text, re.IGNORECASE)
            if match:
                width = float(match.group(1))
                height = float(match.group(2))
                quantity = int(match.group(3)) if match.group(3) else 1
                
                width_str = f"{width:.1f}" if width != int(width) else f"{int(width)}"
                height_str = f"{height:.1f}" if height != int(height) else f"{int(height)}"
                
                if quantity > 1:
                    return f"{width_str}CM*{height_str}CM*{quantity}"
                else:
                    return f"{width_str}CM*{height_str}CM"
            
            return None
            
        except Exception as e:
            print(f"智能回退解析出错: {e}")
            return None

    def auto_convert_inches_to_cm(self, size_text):
        """自动将英寸尺寸转换为厘米格式"""
        try:
            if not size_text:
                return None
            
            size_text = str(size_text).strip()
            print(f"自动转换输入: {size_text}")
            
            # 从复杂文本中提取Variants信息
            variants_match = re.search(r'Variants:\s*([^\n]+)', size_text)
            if variants_match:
                size_text = variants_match.group(1).strip()
                # 移除颜色信息 (如 "/ Multicolor", "/ Verde", "/ Red")
                size_text = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff]+\s*$', '', size_text)
                print(f"提取Variants信息: {size_text}")
            
            # 处理倍数模式 (如 "*2", "*3")
            multiplier = 1
            multiplier_patterns = [
                r'\*\s*(\d+)\s*$',  # 末尾的 *2, *3 等
            ]
            
            for pattern in multiplier_patterns:
                match = re.search(pattern, size_text)
                if match:
                    multiplier = int(match.group(1))
                    size_text = re.sub(pattern, '', size_text).strip()
                    print(f"发现倍数: {multiplier}")
                    break
            
            # 英寸双引号格式 (如 93.16''*92.16'')
            double_quote_pattern = r"(\d+(?:\.\d+)?)''\s*[*×x]\s*(\d+(?:\.\d+)?)''"
            match = re.search(double_quote_pattern, size_text)
            if match:
                width_inch = float(match.group(1))
                height_inch = float(match.group(2))
                width_cm = round(width_inch * 2.54)
                height_cm = round(height_inch * 2.54)
                result = f"{width_cm}CM*{height_cm}CM"
                if multiplier > 1:
                    result += f"*{multiplier}"
                print(f"英寸双引号转换: {width_inch}''*{height_inch}'' -> {result}")
                return result
            
            # 标准混合格式 (如 85cm/33.46in*55cm/21.65in)
            mixed_pattern = r'(\d+(?:\.\d+)?)cm/(\d+(?:\.\d+)?)in\s*[*×x]\s*(\d+(?:\.\d+)?)cm/(\d+(?:\.\d+)?)in'
            match = re.search(mixed_pattern, size_text)
            if match:
                width_cm = round(float(match.group(1)))
                height_cm = round(float(match.group(3)))
                result = f"{width_cm}CM*{height_cm}CM"
                if multiplier > 1:
                    result += f"*{multiplier}"
                print(f"混合格式转换: {result}")
                return result
            
            # 纯CM格式 (如 85cm*55cm)
            cm_pattern = r'(\d+(?:\.\d+)?)cm\s*[*×x]\s*(\d+(?:\.\d+)?)cm'
            match = re.search(cm_pattern, size_text)
            if match:
                width_cm = round(float(match.group(1)))
                height_cm = round(float(match.group(2)))
                result = f"{width_cm}CM*{height_cm}CM"
                if multiplier > 1:
                    result += f"*{multiplier}"
                print(f"纯CM格式: {result}")
                return result
            
            # 纯英寸格式 (如 33.46in*21.65in)
            inch_pattern = r'(\d+(?:\.\d+)?)in\s*[*×x]\s*(\d+(?:\.\d+)?)in'
            match = re.search(inch_pattern, size_text)
            if match:
                width_inch = float(match.group(1))
                height_inch = float(match.group(2))
                width_cm = round(width_inch * 2.54)
                height_cm = round(height_inch * 2.54)
                result = f"{width_cm}CM*{height_cm}CM"
                if multiplier > 1:
                    result += f"*{multiplier}"
                print(f"纯英寸转换: {width_inch}in*{height_inch}in -> {result}")
                return result
            
            # 处理西班牙语单位 pulgadas (英寸)
            pulgadas_pattern = r'(\d+(?:\.\d+)?)cm/(\d+(?:\.\d+)?)pulgadas\s*[*×x]\s*(\d+(?:\.\d+)?)cm/(\d+(?:\.\d+)?)pulgadas'
            match = re.search(pulgadas_pattern, size_text)
            if match:
                width_cm = round(float(match.group(1)))
                height_cm = round(float(match.group(3)))
                result = f"{width_cm}CM*{height_cm}CM"
                if multiplier > 1:
                    result += f"*{multiplier}"
                print(f"西班牙语格式转换: {result}")
                return result
            
            print(f"无法识别的尺寸格式: {size_text}")
            return None
            
        except Exception as e:
            print(f"自动转换尺寸时出错: {e}")
            return None

    def _check_numerical_exactness(self, size1, size2):
        """检查两个尺寸的数字是否完全相同（忽略单位和标点符号差异）"""
        try:
            # 提取纯数字
            def extract_numbers(size_str):
                # 移除所有非数字字符，只保留数字和小数点
                clean_str = re.sub(r'[^\d.]', ' ', size_str)
                # 提取所有数字
                numbers = re.findall(r'\d+(?:\.\d+)?', clean_str)
                return [float(num) for num in numbers]
            
            nums1 = extract_numbers(size1)
            nums2 = extract_numbers(size2)
            
            # 数字个数必须相同
            if len(nums1) != len(nums2):
                return False
            
            # 每个数字必须完全相同
            for n1, n2 in zip(nums1, nums2):
                if abs(n1 - n2) > 0.001:  # 允许极小的浮点误差
                    return False
            
            return True
            
        except Exception:
            return False

    def calculate_size_similarity(self, size1, size2, tolerance=0.1):
        """计算两个尺寸的相似度，支持容错匹配和英寸厘米转换"""
        try:
            if not size1 or not size2:
                return 0.0
            
            # 首先检查数字是否完全相同
            if self._check_numerical_exactness(size1, size2):
                return 1.0  # 数字完全相同，返回最高相似度
            
            # 如果数字不完全相同，则不进行相似度匹配，直接返回0
            # 这确保了数字必须绝对准确，只允许单位和标点符号的差异
            return 0.0
            
        except Exception as e:
            print(f"计算尺寸相似度时出错: {e}")
            return 0.0

    def _normalize_quantity(self, quantity_str):
        """标准化数量字符串，支持全角数字和中文数字"""
        try:
            if not quantity_str:
                return None
            
            # 中文数字映射
            chinese_numbers = {
                '一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                '六': '6', '七': '7', '八': '8', '九': '9', '十': '10'
            }
            
            # 如果是中文数字，直接转换
            if quantity_str in chinese_numbers:
                return chinese_numbers[quantity_str]
            
            # 转换全角数字为半角数字
            full_to_half = str.maketrans('０１２３４５６７８９', '0123456789')
            normalized = quantity_str.translate(full_to_half)
            
            # 验证是否为有效数字
            if normalized.isdigit():
                return normalized
            
            return None
        except Exception:
            return None

    def _normalize_qty_suffix(self, qty_text):
        try:
            if not qty_text:
                return ""
            text = str(qty_text).strip()
            patterns = [
                r'\*\s*(\d+)\b',
                r'[×xX]\s*(\d+)\b',
                r'(?:Set\s*Of|Set\s*of|Pack\s*of)\s*(\d+)\b',
                r'(\d+)\s*(?:pcs?|pieces?|pc|PC|PCS|Pcs|Pieces)\b',
            ]
            for pat in patterns:
                m = re.search(pat, text, re.IGNORECASE)
                if m:
                    n = int(m.group(1))
                    if n > 0:
                        return f"*{n}"
            return ""
        except Exception:
            return ""

    def _extract_original_size_from_filename(self, filename):
        """从文件名中提取原始尺寸字符串（不进行标准化）"""
        try:
            if not filename:
                return None
            
            filename = str(filename).strip()
            
            # 定义各种尺寸模式，按优先级排序
            size_patterns = [
                # 英寸双引号格式
                r'(\d+(?:\.\d+)?\s*(?:\'\'|")\s*[*×xX]\s*\d+(?:\.\d+)?(?:\s*(?:\'\'|"))?(?:\s*[*×xX]\s*\d+)?)',
                # 英寸格式
                r'(\d+(?:\.\d+)?\s*(?:inch|inches|in)\s*[*×xX]\s*\d+(?:\.\d+)?\s*(?:inch|inches|in)(?:\s*[*×xX]\s*\d+)?)',
                # 西班牙语英寸格式
                r'(\d+(?:\.\d+)?\s+(?:pulgadas|Pulgadas)\s+[*×xX]\s+\d+(?:\.\d+)?\s+(?:pulgadas|Pulgadas)(?:\s*[*×xX]\s*\d+)?)',
                # 厘米格式
                r'(\d+(?:\.\d+)?\s*cm\s*[*×xX]\s*\d+(?:\.\d+)?\s*cm(?:\s*[*×xX]\s*\d+)?)',
                # W/H格式
                r'(W\d+(?:\.\d+)?\s*(?:in|inch|inches)\s*[*×xX]\s*H\d+(?:\.\d+)?\s*(?:in|inch|inches)(?:\s*[*×xX]\s*\d+)?)',
                # 纯数字格式（最后匹配，避免误匹配）
                r'(\d+(?:\.\d+)?\s*[*×xX]\s*\d+(?:\.\d+)?(?:\s*[*×xX]\s*\d+)?)(?!\s*(?:inch|inches|cm|\'\'|"|pulgadas|in))',
            ]
            
            for pattern in size_patterns:
                match = re.search(pattern, filename, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
            
            return None
            
        except Exception as e:
            print(f"提取原始尺寸时出错: {e}")
            return None

    def _extract_dimension_numbers(self, text):
        """从文本中提取纯数字维度结构，用于模糊匹配"""
        try:
            if not text:
                return None
            
            text = str(text).strip()
            
            # 提取所有数字（包括小数）
            numbers = re.findall(r'\d+(?:\.\d+)?', text)
            
            if len(numbers) >= 2:
                # 检查是否有第三个数字（数量）
                if len(numbers) >= 3:
                    # 验证第三个数字是否可能是数量（通常较小，如1-10）
                    third_num = float(numbers[2])
                    if third_num <= 20:  # 假设数量不超过20
                        return f"{numbers[0]}*{numbers[1]}*{numbers[2]}"
                
                # 只有两个维度
                return f"{numbers[0]}*{numbers[1]}"
            
            return None
            
        except Exception as e:
            print(f"提取数字维度时出错: {e}")
            return None

    def _check_unit_compatibility(self, text1, text2):
        """检查两个文本的单位兼容性"""
        try:
            # 检查是否都不包含cm单位
            has_cm1 = bool(re.search(r'\bcm\b', text1, re.IGNORECASE))
            has_cm2 = bool(re.search(r'\bcm\b', text2, re.IGNORECASE))
            
            # 检查是否都包含英寸相关单位
            inch_patterns = r'\b(?:inch|inches|in|pulgadas|pulgada|\'\'|")\b'
            has_inch1 = bool(re.search(inch_patterns, text1, re.IGNORECASE))
            has_inch2 = bool(re.search(inch_patterns, text2, re.IGNORECASE))
            
            # 如果都不含cm，或者都含英寸单位，则兼容
            if (not has_cm1 and not has_cm2) or (has_inch1 and has_inch2):
                return True
            
            # 如果一个有cm，另一个没有，则不兼容
            if has_cm1 != has_cm2:
                return False
            
            return True
            
        except Exception:
            return True  # 默认兼容

    def apply_size_mapping(self, filename):
        """应用尺寸映射规则替换文件名中的尺寸数据 - 优先精确匹配"""
        try:
            if not self.size_mapping:
                return filename
            
            print(f"🔍 开始尺寸映射处理: {filename}")
            
            # 从文件名中提取原始尺寸字符串（不标准化）
            original_size_in_filename = self._extract_original_size_from_filename(filename)
            
            # 从文件名中提取并标准化尺寸
            filename_standard_size = self.normalize_size_to_standard(filename)
            if not filename_standard_size:
                print(f"❌ 无法从文件名中提取尺寸")
                return filename
            
            print(f"📏 提取到的原始尺寸: {original_size_in_filename}")
            print(f"📏 提取到的标准化尺寸: {filename_standard_size}")
            
            # 优先级1: 精确匹配原始尺寸字符串
            best_match = None
            best_similarity = 0.0
            best_new_size = None
            
            # 首先尝试精确匹配原始尺寸字符串
            if original_size_in_filename:
                for original_size, new_size in self.size_mapping.items():
                    if original_size_in_filename.lower() == original_size.lower():
                        print(f"✅ 原始尺寸精确匹配: {original_size} -> {new_size}")
                        best_match = original_size
                        best_similarity = 1.0
                        best_new_size = new_size
                        break
            
            # 优先级2: 标准化后的精确匹配
            if not best_match:
                for original_size, new_size in self.size_mapping.items():
                    standard_size = self.normalize_size_to_standard(original_size)
                    if standard_size and filename_standard_size == standard_size:
                        print(f"✅ 标准化精确匹配: {original_size} -> {new_size}")
                        best_match = original_size
                        best_similarity = 1.0
                        best_new_size = new_size
                        break
            
            # 优先级3: 数字结构匹配（用于处理不同标点符号和干扰内容）
            if not best_match:
                filename_numbers = self._extract_dimension_numbers(filename)
                if filename_numbers:
                    print(f"📊 提取到的数字结构: {filename_numbers}")
                    
                    for original_size, new_size in self.size_mapping.items():
                        # 检查单位兼容性
                        if not self._check_unit_compatibility(filename, original_size):
                            print(f"⚠️ 单位不兼容，跳过: {original_size}")
                            continue
                        
                        # 提取映射规则的数字结构
                        mapping_numbers = self._extract_dimension_numbers(original_size)
                        if mapping_numbers and filename_numbers == mapping_numbers:
                            print(f"✅ 数字结构匹配: {filename_numbers} -> {original_size} -> {new_size}")
                            best_match = original_size
                            best_similarity = 0.95  # 高于相似度匹配但低于精确匹配
                            best_new_size = new_size
                            break
            
            # 优先级4: 相似度匹配（仅在没有精确匹配时）
            if not best_match:
                for original_size, new_size in self.size_mapping.items():
                    standard_size = self.normalize_size_to_standard(original_size)
                    if standard_size:
                        # 计算相似度
                        similarity = self.calculate_size_similarity(filename_standard_size, standard_size)
                        print(f"🔍 相似度计算: {filename_standard_size} vs {standard_size} = {similarity:.3f}")
                        # 只接受高精度匹配 (≥0.9)，确保尺寸准确性
                        if similarity > best_similarity and similarity >= 0.9:
                            best_match = original_size
                            best_similarity = similarity
                            best_new_size = new_size
                            print(f"🎯 找到高精度匹配: {original_size} (相似度: {similarity:.3f})")
                        elif similarity >= 0.8:
                            print(f"⚠️ 中等相似度但未达到高精度要求: {original_size} (相似度: {similarity:.3f}) - 跳过")
            
            # 如果找到匹配（精确或模糊），进行替换
            if best_match and best_new_size:
                print(f"🔄 应用映射: {best_match} -> {best_new_size} (相似度: {best_similarity:.3f})")
                print(f"📋 映射详情: 原始尺寸={filename_standard_size}, 匹配规则={best_match}, 替换值={best_new_size}")
                
                # 找到文件名中的尺寸部分并替换
                # 优化：在尺寸映射时保留原始数量尾缀（如 *2 / Set Of 2 / 2pcs）
                qty_tail = r'(?P<qty_suffix>\s*(?:[*×xX]\s*\d+|X\s*\d+(?:\s*(?:pcs|Pcs|PCS|pieces|Pieces|pc|PC))?|(?:Set\s*Of|Set\s*of|Pack\s*of)\s*\d+|\d+\s*(?:pcs|Pcs|PCS|pieces|Pieces|pc|PC))\b)?'
                size_patterns = [
                    re.compile(r'(?P<size>\d+(?:\.\d+)?\s*(?:\'\'|")\s*[*×xX]\s*\d+(?:\.\d+)?(?:\s*(?:\'\'|"))?)' + qty_tail, re.IGNORECASE),
                    re.compile(r'(?P<size>\d+(?:\.\d+)?\s*(?:inch|inches|in)\s*[*×xX]\s*\d+(?:\.\d+)?\s*(?:inch|inches|in))' + qty_tail, re.IGNORECASE),
                    re.compile(r'(?P<size>\d+(?:\.\d+)?\s+(?:inch|inches|pulgadas|in)\s+[*×xX]\s+\d+(?:\.\d+)?\s+(?:inch|inches|pulgadas|in))' + qty_tail, re.IGNORECASE),
                    re.compile(r'(?P<size>\d+(?:\.\d+)?\s*cm\s*[*×xX]\s*\d+(?:\.\d+)?\s*cm)' + qty_tail, re.IGNORECASE),
                    re.compile(r'(?P<size>\d+(?:\.\d+)?\s*[*×xX]\s*\d+(?:\.\d+)?)(?!\s*(?:inch|inches|cm|\'\'|"|pulgadas|in))' + qty_tail, re.IGNORECASE),
                    re.compile(r'(?P<size>W\d+(?:\.\d+)?\s*(?:in|inch|inches)\s*[*×xX]\s*H\d+(?:\.\d+)?\s*(?:in|inch|inches))' + qty_tail, re.IGNORECASE),
                ]
                
                result_filename = filename
                for pattern in size_patterns:
                    m = pattern.search(result_filename)
                    if m:
                        qty_suffix = m.group('qty_suffix') or ''
                        normalized_qty = self._normalize_qty_suffix(qty_suffix)
                        # 若映射值已含数量信息（*K），则不重复追加
                        if normalized_qty and re.search(r'\*\s*\d+\b', best_new_size):
                            final_replacement = best_new_size
                        else:
                            final_replacement = best_new_size + (normalized_qty if normalized_qty else '')
                        result_filename = pattern.sub(final_replacement, result_filename, count=1)
                        print(f"✅ 替换成功: {filename} -> {result_filename}")
                        return result_filename
                
                print(f"⚠️ 未找到可替换的尺寸模式")
                return filename
            else:
                print(f"❌ 未找到匹配的尺寸映射规则")
                return filename
            
        except Exception as e:
            print(f"应用尺寸映射时出错: {e}")
            return filename
    
    def export_size_mapping(self):
        """导出尺寸映射配置到文件"""
        try:
            if not self.size_mapping:
                messagebox.showwarning("警告", "当前没有尺寸映射配置可导出！")
                return
            
            # 选择保存文件
            file_path = filedialog.asksaveasfilename(
                title="导出尺寸映射配置",
                defaultextension=".json",
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
                initialfilename="size_mapping_config.json"
            )
            
            if file_path:
                # 创建配置数据
                config_data = {
                    "version": "1.9",
                    "description": "Y2订单辅助工具 - 尺寸映射配置",
                    "size_mapping": self.size_mapping,
                    "export_time": time.strftime("%Y-%m-%d %H:%M:%S")
                }
                
                # 保存到文件
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(config_data, f, ensure_ascii=False, indent=2)
                
                messagebox.showinfo("成功", f"尺寸映射配置已导出到：\n{file_path}")
                
        except Exception as e:
            messagebox.showerror("错误", f"导出配置失败：{str(e)}")
            print(f"导出尺寸映射配置时出错: {e}")
    
    def import_size_mapping(self, dialog):
        """从文件导入尺寸映射配置"""
        try:
            # 临时释放设置窗口的grab_set，避免文件对话框被遮挡
            dialog.grab_release()
            
            # 选择导入文件
            file_path = filedialog.askopenfilename(
                title="导入尺寸映射配置",
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
            )
            
            # 恢复设置窗口的grab_set
            dialog.grab_set()
            
            if file_path:
                # 读取配置文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                
                # 验证配置格式
                if "size_mapping" not in config_data:
                    messagebox.showerror("错误", "配置文件格式不正确，缺少size_mapping字段！")
                    return
                
                imported_mapping = config_data["size_mapping"]
                if not isinstance(imported_mapping, dict):
                    messagebox.showerror("错误", "配置文件格式不正确，size_mapping必须是字典格式！")
                    return
                
                # 确认导入
                result = messagebox.askyesno(
                    "确认导入", 
                    f"将导入 {len(imported_mapping)} 条尺寸映射规则。\n\n这将覆盖当前的配置，是否继续？"
                )
                
                if result:
                    # 更新映射配置
                    self.size_mapping.update(imported_mapping)
                    
                    # 更新对话框中的输入框
                    for size, new_value in imported_mapping.items():
                        if size in self.mapping_entries:
                            self.mapping_entries[size].set(new_value)
                    
                    messagebox.showinfo("成功", f"已成功导入 {len(imported_mapping)} 条尺寸映射规则！")
                
        except json.JSONDecodeError:
            messagebox.showerror("错误", "配置文件格式错误，请选择有效的JSON文件！")
        except Exception as e:
            messagebox.showerror("错误", f"导入配置失败：{str(e)}")
            print(f"导入尺寸映射配置时出错: {e}")
    
    def open_settings_page(self):
        """打开设置页面"""
        # 在打开设置页面前重新加载配置，确保显示最新的配置值
        # apply_geometry=False 防止主窗口被重置到旧位置
        self.load_config(apply_geometry=False)
        
        settings_window = tk.Toplevel(self.root)
        settings_window.attributes('-alpha', 0.0)  # 初始设为全透明，防止空白闪烁
        settings_window.title("设置")
        
        # 从配置中加载设置窗口大小，如果没有则使用默认值
        settings_geometry = getattr(self, 'settings_window_geometry', "700x650")
        settings_window.geometry(settings_geometry)
        settings_window.resizable(True, True)
        settings_window.transient(self.root)
        settings_window.grab_set()
        # settings_window.attributes('-topmost', True)  # 移除设置窗口置顶，避免遮挡文件选择框
        
        # 设置窗口图标
        try:
            settings_window.iconbitmap(self.icon_path)
        except FileNotFoundError:
            print(f"图标文件未找到: {self.icon_path}")
        except Exception as e:
            print(f"设置窗口图标时出错: {e}")
        
        # 主框架，减少内边距
        main_frame = ttk.Frame(settings_window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题，减小字体和间距
        title_label = ttk.Label(main_frame, text="⚙ 设置", 
                               font=('Microsoft YaHei UI', 14, 'bold'))
        title_label.pack(anchor=tk.W, pady=(0, 15))
        
        # 创建多标签页控件
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # 标签页1：尺寸映射配置，减少内边距
        size_mapping_frame = ttk.Frame(notebook, padding="10")
        notebook.add(size_mapping_frame, text="📏 尺寸映射")
        
        # 标签页2：加工方式和材质设置，减少内边距
        processing_frame = ttk.Frame(notebook, padding="10")
        notebook.add(processing_frame, text="🔧 加工方式与材质")
        
        # 标签页3：导出路径设置
        export_path_frame = ttk.Frame(notebook, padding="10")
        notebook.add(export_path_frame, text="📁 导出路径")
        
        # 标签页4：API配置
        api_config_frame = ttk.Frame(notebook, padding="10")
        notebook.add(api_config_frame, text="🔑 API配置")
        
        # 标签页5：物流与命名设置
        naming_frame = ttk.Frame(notebook, padding="10")
        notebook.add(naming_frame, text="📦 物流与命名")
        
        # 标签页6：作者信息
        author_info_frame = ttk.Frame(notebook, padding="10")
        notebook.add(author_info_frame, text="👤 作者信息")
        
        # 创建尺寸映射标签页内容
        self.create_size_mapping_tab(size_mapping_frame)
        
        # 创建加工方式和材质标签页内容
        self.create_processing_tab(processing_frame)
        
        # 创建导出路径标签页内容
        self.create_export_path_tab(export_path_frame)
        
        # 创建API配置标签页内容
        self.create_api_config_tab(api_config_frame)
        
        # 创建物流与命名标签页内容
        self.create_naming_config_tab(naming_frame)
        
        # 创建作者信息标签页内容
        self.create_author_info_tab(author_info_frame)
        
        # 相对于主窗口居中显示
        settings_window.update_idletasks()
        
        # 确保主窗口坐标是最新的
        self.root.update_idletasks()
        
        # 获取主窗口的位置和尺寸
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        # 如果主窗口处于最小化或尚未完全渲染，使用默认尺寸
        if main_width <= 1: main_width = 1320
        if main_height <= 1: main_height = 800
        
        # 获取设置窗口的尺寸（使用请求尺寸，因为可能尚未完全渲染）
        settings_width = settings_window.winfo_reqwidth()
        settings_height = settings_window.winfo_reqheight()
        
        # 如果 winfo_width 已经有值，优先使用
        if settings_window.winfo_width() > 1:
            settings_width = settings_window.winfo_width()
        if settings_window.winfo_height() > 1:
            settings_height = settings_window.winfo_height()
            
        # 计算居中位置
        x = main_x + (main_width - settings_width) // 2
        y = main_y + (main_height - settings_height) // 2
        
        # 确保窗口不会超出屏幕边界
        screen_width = settings_window.winfo_screenwidth()
        screen_height = settings_window.winfo_screenheight()
        
        if x < 0:
            x = 0
        elif x + settings_width > screen_width:
            x = screen_width - settings_width
            
        if y < 0:
            y = 0
        elif y + settings_height > screen_height:
            y = screen_height - settings_height
        
        # 设置最终位置和尺寸
        settings_window.geometry(f"{settings_width}x{settings_height}+{x}+{y}")
        
        # 渐入显示动画 (计划书 1.8 交互微调)
        def animate_fade_in(current_alpha=0.0):
            if current_alpha < 1.0:
                new_alpha = min(1.0, current_alpha + 0.15)
                settings_window.attributes('-alpha', new_alpha)
                settings_window.after(15, lambda: animate_fade_in(new_alpha))
        
        # 启动动画
        animate_fade_in()
        
        # 保存设置窗口引用，用于配置加载
        self.settings_window = settings_window
        
        # 绑定窗口关闭事件
        settings_window.protocol("WM_DELETE_WINDOW", lambda: self.close_settings_window(settings_window))
    
    def create_size_mapping_tab(self, parent_frame):
        """创建尺寸映射标签页内容"""
        # 创建主容器
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # 顶部说明区域
        info_frame = ttk.Frame(main_container)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 图标和标题
        title_frame = ttk.Frame(info_frame)
        title_frame.pack(fill=tk.X, pady=(0, 5))
        
        title_label = ttk.Label(title_frame, text="📏 尺寸映射配置", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(side=tk.LEFT)
        
        # 说明文字
        info_label = ttk.Label(info_frame, 
                              text="💡 在下方表格中设置尺寸映射规则，左列为原始尺寸，右列为对应尺寸",
                              font=('Microsoft YaHei UI', 9),
                              foreground='#666666')
        info_label.pack(anchor=tk.W)
        
        # 分隔线
        separator = ttk.Separator(main_container, orient='horizontal')
        separator.pack(fill=tk.X, pady=(0, 8))
        
        # 表格区域
        table_container = ttk.Frame(main_container)
        table_container.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # 表格标题区域
        header_container = ttk.Frame(table_container)
        header_container.pack(fill=tk.X, pady=(0, 5))
        
        # 美化的表格标题
        header_frame = ttk.Frame(header_container)
        header_frame.pack(fill=tk.X)
        
        # 原始尺寸标题
        original_header = ttk.Frame(header_frame)
        original_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(original_header, text="📋 原始尺寸", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 箭头
        arrow_label = ttk.Label(header_frame, text="➡️", 
                               font=('Microsoft YaHei UI', 10))
        arrow_label.pack(side=tk.LEFT, padx=(0, 8))
        
        # 对应尺寸标题
        mapped_header = ttk.Frame(header_frame)
        mapped_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(mapped_header, text="🎯 对应尺寸", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 操作标题
        action_header = ttk.Frame(header_frame)
        action_header.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(action_header, text="⚙️ 操作", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack()
        
        # 滚动区域
        scroll_container = ttk.Frame(table_container)
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(scroll_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建画布和内部框架用于滚动
        canvas = tk.Canvas(scroll_container, yscrollcommand=scrollbar.set,
                          highlightthickness=0, bg='#f8f9fa')
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=canvas.yview)
        
        # 内部滚动框架
        scroll_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        
        # 存储映射条目的容器
        self.mapping_entries_frame = ttk.Frame(scroll_frame)
        self.mapping_entries_frame.pack(fill=tk.X, padx=5, pady=3)
        
        # 存储当前的映射输入框
        self.current_mapping_entries = []
        
        # 添加新条目按钮区域
        add_button_frame = ttk.Frame(scroll_frame)
        add_button_frame.pack(fill=tk.X, pady=(8, 5))
        
        add_btn = ttk.Button(add_button_frame, text="➕ 添加新映射", 
                            command=self.add_mapping_entry,
                            style='Accent.TButton')
        add_btn.pack()
        
        # 更新滚动区域
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scroll_frame.bind("<Configure>", configure_scroll_region)
        
        # 绑定鼠标滚轮事件到设置页面的Canvas
        self.bind_mousewheel(canvas)
        self.bind_mousewheel(scroll_frame)
        
        # 确保Canvas能够接收焦点和滚轮事件
        canvas.focus_set()
        
        # 底部按钮区域
        button_container = ttk.Frame(main_container)
        button_container.pack(fill=tk.X, pady=(8, 0))
        
        # 左侧按钮组
        left_buttons = ttk.Frame(button_container)
        left_buttons.pack(side=tk.LEFT)
        
        # 导出配置按钮
        export_btn = ttk.Button(left_buttons, text="📤 导出配置", 
                               command=self.export_size_mapping_excel)
        export_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # 导入配置按钮
        import_btn = ttk.Button(left_buttons, text="📥 导入配置", 
                               command=self.import_size_mapping_excel)
        import_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # 右侧按钮组
        right_buttons = ttk.Frame(button_container)
        right_buttons.pack(side=tk.RIGHT)
        
        # 清空配置按钮
        clear_btn = ttk.Button(right_buttons, text="🗑️ 清空配置", 
                              command=self.clear_size_mapping)
        clear_btn.pack(side=tk.RIGHT, padx=(8, 0))
        
        # 保存配置按钮
        save_btn = ttk.Button(right_buttons, text="💾 保存配置", 
                             command=self.save_settings_mapping,
                             style='Accent.TButton')
        save_btn.pack(side=tk.RIGHT, padx=(8, 0))
        
        # 加载现有配置
        self.load_existing_mappings()
    
    def create_processing_tab(self, parent_frame):
        """创建加工方式和材质设置标签页内容"""
        # 创建主容器
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        
        # 顶部说明区域
        info_frame = ttk.Frame(main_container)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 图标和标题
        title_frame = ttk.Frame(info_frame)
        title_frame.pack(fill=tk.X, pady=(0, 5))
        
        title_label = ttk.Label(title_frame, text="⚙️ 加工方式与材质配置", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(side=tk.LEFT)
        
        # 说明文字
        info_label = ttk.Label(info_frame, 
                              text="💡 根据SKU前缀设置对应的加工方式和材质信息，支持自动识别和匹配",
                              font=('Microsoft YaHei UI', 9),
                              foreground='#666666')
        info_label.pack(anchor=tk.W)
        
        # 分隔线
        separator = ttk.Separator(main_container, orient='horizontal')
        separator.pack(fill=tk.X, pady=(0, 8))
        
        # 表格区域
        table_container = ttk.Frame(main_container)
        table_container.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        # 表格标题区域
        header_container = ttk.Frame(table_container)
        header_container.pack(fill=tk.X, pady=(0, 5))
        
        # 美化的表格标题
        header_frame = ttk.Frame(header_container)
        header_frame.pack(fill=tk.X)
        
        # 序号标题
        seq_header = ttk.Frame(header_frame)
        seq_header.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(seq_header, text="🔢 序号", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack()
        
        # SKU前缀标题
        sku_header = ttk.Frame(header_frame)
        sku_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(sku_header, text="🏷️ SKU前缀", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 产品名称标题
        product_header = ttk.Frame(header_frame)
        product_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(product_header, text="📦 产品名称", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 加工方式标题
        process_header = ttk.Frame(header_frame)
        process_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(process_header, text="🔧 加工方式", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 材质标题
        material_header = ttk.Frame(header_frame)
        material_header.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Label(material_header, text="🧱 材质", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack(anchor=tk.W)
        
        # 操作标题
        action_header = ttk.Frame(header_frame)
        action_header.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(action_header, text="⚙️ 操作", 
                 font=('Microsoft YaHei UI', 10, 'bold'),
                 foreground='#2c3e50').pack()
        
        # 滚动区域
        scroll_container = ttk.Frame(table_container)
        scroll_container.pack(fill=tk.BOTH, expand=True)
        
        # 创建滚动条
        scrollbar = ttk.Scrollbar(scroll_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 创建画布和内部框架用于滚动
        canvas = tk.Canvas(scroll_container, yscrollcommand=scrollbar.set,
                          highlightthickness=0, bg='#f8f9fa')
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=canvas.yview)
        
        # 内部滚动框架
        scroll_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        
        # 存储加工方式条目的容器
        self.processing_entries_frame = ttk.Frame(scroll_frame)
        self.processing_entries_frame.pack(fill=tk.X, padx=5, pady=3)
        
        # 存储当前的加工方式输入框
        self.current_processing_entries = []
        
        # 添加新条目按钮区域
        add_button_frame = ttk.Frame(scroll_frame)
        add_button_frame.pack(fill=tk.X, pady=(8, 5))
        
        add_processing_btn = ttk.Button(add_button_frame, text="➕ 添加新配置", 
                                       command=self.add_processing_entry,
                                       style='Accent.TButton')
        add_processing_btn.pack()
        
        # 更新滚动区域
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scroll_frame.bind("<Configure>", configure_scroll_region)
        
        # 绑定鼠标滚轮事件
        self.bind_mousewheel(canvas)
        self.bind_mousewheel(scroll_frame)
        canvas.focus_set()
        
        # 底部按钮区域
        button_container = ttk.Frame(main_container)
        button_container.pack(fill=tk.X, pady=(8, 0))
        
        # 左侧按钮组
        left_buttons = ttk.Frame(button_container)
        left_buttons.pack(side=tk.LEFT)
        
        # 导出配置按钮
        export_processing_btn = ttk.Button(left_buttons, text="📤 导出配置", 
                                          command=self.export_processing_config)
        export_processing_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # 导入配置按钮
        import_processing_btn = ttk.Button(left_buttons, text="📥 导入配置", 
                                          command=self.import_processing_config)
        import_processing_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # 右侧按钮组
        right_buttons = ttk.Frame(button_container)
        right_buttons.pack(side=tk.RIGHT)
        
        # 清空配置按钮
        clear_processing_btn = ttk.Button(right_buttons, text="🗑️ 清空配置", 
                                         command=self.clear_processing_config)
        clear_processing_btn.pack(side=tk.RIGHT, padx=(8, 0))
        
        # 保存配置按钮
        save_processing_btn = ttk.Button(right_buttons, text="💾 保存配置", 
                                        command=self.save_processing_config,
                                        style='Accent.TButton')
        save_processing_btn.pack(side=tk.RIGHT, padx=(8, 0))
        
        # 加载现有配置
        self.load_existing_processing_config()
        
        # 如果程序启动时保存了配置数据，现在加载它
        if hasattr(self, 'saved_processing_config') and self.saved_processing_config:
            self.load_saved_processing_config()
    
    def create_export_path_tab(self, parent_frame):
        """创建导出路径配置标签页"""
        # 主标题
        title_label = ttk.Label(parent_frame, text="📁 导出路径配置", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(pady=(0, 15))
        
        # Excel导出路径配置
        excel_frame = ttk.LabelFrame(parent_frame, text="📊 Excel表格导出路径", padding="15")
        excel_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Excel路径显示和选择
        excel_path_frame = ttk.Frame(excel_frame)
        excel_path_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(excel_path_frame, text="当前路径：").pack(side=tk.LEFT)
        
        # 确保显示正确的路径，如果为空则显示默认桌面路径
        display_excel_path = self.excel_export_path if self.excel_export_path else os.path.join(os.path.expanduser("~"), "Desktop")
        self.excel_path_var = tk.StringVar(value=display_excel_path)
        excel_path_entry = ttk.Entry(excel_path_frame, textvariable=self.excel_path_var, 
                                    state='readonly', width=35)
        excel_path_entry.pack(side=tk.LEFT, padx=(5, 10), fill=tk.X, expand=True)
        
        # Excel重置按钮
        excel_reset_btn = ttk.Button(excel_path_frame, text="🔄 重置为桌面", 
                                    command=self.reset_excel_export_path)
        excel_reset_btn.pack(side=tk.RIGHT)
        
        excel_browse_btn = ttk.Button(excel_path_frame, text="浏览...", 
                                     command=self.browse_excel_export_path)
        excel_browse_btn.pack(side=tk.RIGHT, padx=(0, 5))
        
        # 图片整理导出路径配置
        image_frame = ttk.LabelFrame(parent_frame, text="🖼️ 图片整理导出路径", padding="15")
        image_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 图片路径显示和选择
        image_path_frame = ttk.Frame(image_frame)
        image_path_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(image_path_frame, text="当前路径：").pack(side=tk.LEFT)
        
        # 确保显示正确的路径，如果为空则显示默认桌面路径
        display_image_path = self.image_export_path if self.image_export_path else os.path.join(os.path.expanduser("~"), "Desktop")
        self.image_path_var = tk.StringVar(value=display_image_path)
        image_path_entry = ttk.Entry(image_path_frame, textvariable=self.image_path_var, 
                                    state='readonly', width=35)
        image_path_entry.pack(side=tk.LEFT, padx=(5, 10), fill=tk.X, expand=True)
        
        # 图片重置按钮
        image_reset_btn = ttk.Button(image_path_frame, text="🔄 重置为桌面", 
                                    command=self.reset_image_export_path)
        image_reset_btn.pack(side=tk.RIGHT)
        
        image_browse_btn = ttk.Button(image_path_frame, text="浏览...", 
                                     command=self.browse_image_export_path)
        image_browse_btn.pack(side=tk.RIGHT, padx=(0, 5))
        
        # 高清图片保存路径配置
        upscale_frame = ttk.LabelFrame(parent_frame, text="🎨 高清图片保存路径", padding="15")
        upscale_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 高清图片路径显示和选择
        upscale_path_frame = ttk.Frame(upscale_frame)
        upscale_path_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(upscale_path_frame, text="当前路径：").pack(side=tk.LEFT)
        
        self.upscale_path_var = tk.StringVar(value=self.upscale_export_path if self.upscale_export_path else "源文件位置")
        upscale_path_entry = ttk.Entry(upscale_path_frame, textvariable=self.upscale_path_var, 
                                      state='readonly', width=35)
        upscale_path_entry.pack(side=tk.LEFT, padx=(5, 10), fill=tk.X, expand=True)
        
        # 高清图片重置按钮
        upscale_reset_btn = ttk.Button(upscale_path_frame, text="🔄 重置为源文件位置", 
                                      command=self.reset_upscale_export_path)
        upscale_reset_btn.pack(side=tk.RIGHT)
        
        upscale_browse_btn = ttk.Button(upscale_path_frame, text="浏览...", 
                                       command=self.browse_upscale_export_path)
        upscale_browse_btn.pack(side=tk.RIGHT, padx=(0, 5))
        
        # 高清图片重命名规则配置
        rename_frame = ttk.Frame(upscale_frame)
        rename_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(rename_frame, text="重命名规则：").pack(side=tk.LEFT)
        
        self.upscale_rename_var = tk.StringVar(value=self.upscale_config.get('rename_rule', '（高清）'))
        rename_entry = ttk.Entry(rename_frame, textvariable=self.upscale_rename_var, 
                                width=20, font=('Microsoft YaHei UI', 9))
        rename_entry.pack(side=tk.LEFT, padx=(5, 10))
        
        # 绑定变化事件
        self.upscale_rename_var.trace('w', self.on_upscale_rename_change)
        
        # 重命名规则说明
        rename_info = ttk.Label(rename_frame, 
                               text="💡 在文件名后添加的标识，如：图片（高清）.jpg", 
                               font=('Microsoft YaHei UI', 8), 
                               foreground='#666666')
        rename_info.pack(side=tk.LEFT, padx=(10, 0))
        
        # 说明文字
        upscale_info_label = ttk.Label(upscale_frame, 
                                      text="💡 提示：选择\"源文件位置\"时，高清图片将保存在原图片相同目录下", 
                                      font=('Microsoft YaHei UI', 9), 
                                      foreground='#666666')
        upscale_info_label.pack(pady=(5, 0))
    
    def create_naming_config_tab(self, parent_frame):
        """创建物流与命名标签页内容"""
        # 主滚动容器
        canvas = tk.Canvas(parent_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 1. 名称缩写管理
        identity_frame = ttk.LabelFrame(scrollable_frame, text="🏢 名称缩写管理", padding="15")
        identity_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

        # 缩写列表显示
        self.abbrev_listbox = tk.Listbox(identity_frame, height=3, font=('Microsoft YaHei UI', 9))
        self.abbrev_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 填充列表 (兼容旧版单字符串格式)
        abbrev_data = self.naming_center.get('business_abbreviations', [])
        if not abbrev_data and 'business_abbreviation' in self.naming_center:
            abbrev_data = [self.naming_center['business_abbreviation']]
            self.naming_center['business_abbreviations'] = abbrev_data
            
        for abbrev in abbrev_data:
            self.abbrev_listbox.insert(tk.END, abbrev)

        abbrev_btns = ttk.Frame(identity_frame)
        abbrev_btns.pack(side=tk.LEFT, padx=(10, 0))
        
        def add_abbrev():
            import tkinter.simpledialog as sd
            new_a = sd.askstring("添加缩写", "请输入新的名称缩写 (如: CHX):", parent=self.root)
            if new_a and new_a.strip():
                new_a = new_a.strip().upper()
                if 'business_abbreviations' not in self.naming_center:
                    self.naming_center['business_abbreviations'] = []
                if new_a not in self.naming_center['business_abbreviations']:
                    self.naming_center['business_abbreviations'].append(new_a)
                    self.abbrev_listbox.insert(tk.END, new_a)
                    self.save_all_settings_and_feedback()

        def remove_abbrev():
            selection = self.abbrev_listbox.curselection()
            if selection:
                idx = selection[0]
                val = self.abbrev_listbox.get(idx)
                self.naming_center['business_abbreviations'].remove(val)
                self.abbrev_listbox.delete(idx)
                self.save_all_settings_and_feedback()

        def edit_abbrev():
            selection = self.abbrev_listbox.curselection()
            if selection:
                idx = selection[0]
                old_val = self.abbrev_listbox.get(idx)
                import tkinter.simpledialog as sd
                new_a = sd.askstring("修改缩写", f"修改名称缩写 (原名: {old_val}):", 
                                   initialvalue=old_val, parent=self.root)
                if new_a and new_a.strip() and new_a.strip().upper() != old_val:
                    new_a = new_a.strip().upper()
                    a_list = self.naming_center['business_abbreviations']
                    if old_val in a_list:
                        a_idx = a_list.index(old_val)
                        a_list[a_idx] = new_a
                    
                    self.abbrev_listbox.delete(idx)
                    self.abbrev_listbox.insert(idx, new_a)
                    self.abbrev_listbox.selection_set(idx)
                    self.save_all_settings_and_feedback()

        ttk.Button(abbrev_btns, text="添加", command=add_abbrev).pack(fill=tk.X, pady=2)
        ttk.Button(abbrev_btns, text="修改", command=edit_abbrev).pack(fill=tk.X, pady=2)
        ttk.Button(abbrev_btns, text="删除", command=remove_abbrev).pack(fill=tk.X, pady=2)

        # 2. 物流服务商管理
        logistics_frame = ttk.LabelFrame(scrollable_frame, text="🚚 物流服务商列表", padding="15")
        logistics_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

        # 服务商列表显示
        self.logistics_listbox = tk.Listbox(logistics_frame, height=5, font=('Microsoft YaHei UI', 9))
        self.logistics_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 填充列表
        for provider in self.naming_center.get('logistics_providers', []):
            self.logistics_listbox.insert(tk.END, provider)

        logistics_btns = ttk.Frame(logistics_frame)
        logistics_btns.pack(side=tk.LEFT, padx=(10, 0))
        
        def add_provider():
            import tkinter.simpledialog as sd
            new_p = sd.askstring("添加服务商", "请输入新的物流服务商名称 (如: Y3顺丰):", parent=self.root)
            if new_p and new_p.strip():
                if new_p not in self.naming_center['logistics_providers']:
                    self.naming_center['logistics_providers'].append(new_p)
                    self.logistics_listbox.insert(tk.END, new_p)
                    self.save_all_settings_and_feedback()

        def remove_provider():
            selection = self.logistics_listbox.curselection()
            if selection:
                idx = selection[0]
                val = self.logistics_listbox.get(idx)
                self.naming_center['logistics_providers'].remove(val)
                self.logistics_listbox.delete(idx)
                self.save_all_settings_and_feedback()

        def edit_provider():
            selection = self.logistics_listbox.curselection()
            if selection:
                idx = selection[0]
                old_val = self.logistics_listbox.get(idx)
                import tkinter.simpledialog as sd
                new_p = sd.askstring("修改服务商", f"修改服务商名称 (原名: {old_val}):", 
                                   initialvalue=old_val, parent=self.root)
                if new_p and new_p.strip() and new_p != old_val:
                    # 更新数据列表
                    p_list = self.naming_center['logistics_providers']
                    if old_val in p_list:
                        p_idx = p_list.index(old_val)
                        p_list[p_idx] = new_p
                    
                    # 更新UI列表
                    self.logistics_listbox.delete(idx)
                    self.logistics_listbox.insert(idx, new_p)
                    self.logistics_listbox.selection_set(idx)
                    self.save_all_settings_and_feedback()

        ttk.Button(logistics_btns, text="添加", command=add_provider).pack(fill=tk.X, pady=2)
        ttk.Button(logistics_btns, text="修改", command=edit_provider).pack(fill=tk.X, pady=2)
        ttk.Button(logistics_btns, text="删除", command=remove_provider).pack(fill=tk.X, pady=2)

        # 3. 自定义后缀管理
        suffix_mgmt_frame = ttk.LabelFrame(scrollable_frame, text="🚪 所属部门尾缀列表", padding="15")
        suffix_mgmt_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

        self.suffix_listbox = tk.Listbox(suffix_mgmt_frame, height=3, font=('Microsoft YaHei UI', 9))
        self.suffix_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        for s in self.naming_center.get('custom_suffixes', []):
            self.suffix_listbox.insert(tk.END, s)

        suffix_btns = ttk.Frame(suffix_mgmt_frame)
        suffix_btns.pack(side=tk.LEFT, padx=(10, 0))
        
        def add_suffix():
            import tkinter.simpledialog as sd
            new_s = sd.askstring("添加后缀", "请输入新的后缀 (如: 设计师):", parent=self.root)
            if new_s and new_s.strip():
                if new_s not in self.naming_center['custom_suffixes']:
                    self.naming_center['custom_suffixes'].append(new_s)
                    self.suffix_listbox.insert(tk.END, new_s)
                    self.save_all_settings_and_feedback()

        def remove_suffix():
            selection = self.suffix_listbox.curselection()
            if selection:
                idx = selection[0]
                val = self.suffix_listbox.get(idx)
                self.naming_center['custom_suffixes'].remove(val)
                self.suffix_listbox.delete(idx)
                self.save_all_settings_and_feedback()

        def edit_suffix():
            selection = self.suffix_listbox.curselection()
            if selection:
                idx = selection[0]
                old_val = self.suffix_listbox.get(idx)
                import tkinter.simpledialog as sd
                new_s = sd.askstring("修改后缀", f"修改后缀名称 (原名: {old_val}):", 
                                   initialvalue=old_val, parent=self.root)
                if new_s and new_s.strip() and new_s != old_val:
                    # 更新数据列表
                    s_list = self.naming_center['custom_suffixes']
                    if old_val in s_list:
                        s_idx = s_list.index(old_val)
                        s_list[s_idx] = new_s
                    
                    # 更新UI列表
                    self.suffix_listbox.delete(idx)
                    self.suffix_listbox.insert(idx, new_s)
                    self.suffix_listbox.selection_set(idx)
                    self.save_all_settings_and_feedback()

        ttk.Button(suffix_btns, text="添加", command=add_suffix).pack(fill=tk.X, pady=2)
        ttk.Button(suffix_btns, text="修改", command=edit_suffix).pack(fill=tk.X, pady=2)
        ttk.Button(suffix_btns, text="删除", command=remove_suffix).pack(fill=tk.X, pady=2)

        # 4. 命名模块排序
        template_frame = ttk.LabelFrame(scrollable_frame, text="🧩 导出文件名模块排序", padding="15")
        template_frame.pack(fill=tk.X, pady=(0, 15), padx=5)
        
        ttk.Label(template_frame, text="模板格式：", font=('Microsoft YaHei UI', 9, 'bold')).pack(anchor=tk.W)
        self.template_var = tk.StringVar(value=self.naming_center.get('naming_template', '{prefix}-{date}-{provider}-{product}-{suffix}'))
        template_entry = ttk.Entry(template_frame, textvariable=self.template_var)
        template_entry.pack(fill=tk.X, pady=5)
        
        hint_text = "可用模块：{prefix} (缩写), {date} (日期), {provider} (物流), {product} (产品), {suffix} (后缀)\n示例：{prefix}-{date}-{provider}-{product}-{suffix}"
        ttk.Label(template_frame, text=hint_text, foreground="#666666", justify=tk.LEFT).pack(anchor=tk.W)

        # 4. 其他配置
        other_frame = ttk.LabelFrame(scrollable_frame, text="⚙️ 其他命名设置", padding="15")
        other_frame.pack(fill=tk.X, pady=(0, 15), padx=5)
        
        self.auto_save_new_var = tk.BooleanVar(value=self.naming_center.get('auto_save_new_items', True))
        ttk.Checkbutton(other_frame, text="自动保存弹窗中手动输入的服务商到预设列表", variable=self.auto_save_new_var).pack(anchor=tk.W)

        # 布局滚动条
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        return parent_frame

    def create_api_config_tab(self, parent_frame):
        """创建API配置标签页"""
        # 主标题
        title_label = ttk.Label(parent_frame, text="🔑 API配置", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(pady=(0, 15))
        
        # BigJPG API配置
        api_frame = ttk.LabelFrame(parent_frame, text="🎨 BigJPG API配置", padding="15")
        api_frame.pack(fill=tk.X, pady=(0, 15))
        
        # API密钥输入
        api_key_frame = ttk.Frame(api_frame)
        api_key_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(api_key_frame, text="API密钥：").pack(side=tk.LEFT)
        
        # 每次打开设置页面时重新创建API密钥变量，确保读取最新值
        self.api_key_var = tk.StringVar(value=getattr(self, 'bigjpg_api_key', ''))
        
        # 创建API密钥变量（如果不存在）
        if False:  # 禁用原有的条件检查
            self.api_key_var = tk.StringVar(value=getattr(self, 'bigjpg_api_key', ''))
        
        api_key_entry = ttk.Entry(api_key_frame, textvariable=self.api_key_var, 
                                 show="*", width=50)
        api_key_entry.pack(side=tk.LEFT, padx=(5, 10), fill=tk.X, expand=True)
        
        # 保存API密钥按钮
        save_api_btn = ttk.Button(api_key_frame, text="保存", 
                                 command=self.save_api_key)
        save_api_btn.pack(side=tk.RIGHT)
        
        # ImgBB API配置 (计划书 1.8 核心改进)
        imgbb_frame = ttk.LabelFrame(parent_frame, text="🖼️ ImgBB 图床 API配置", padding="15")
        imgbb_frame.pack(fill=tk.X, pady=(0, 15))
        
        # ImgBB API密钥输入
        imgbb_key_frame = ttk.Frame(imgbb_frame)
        imgbb_key_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(imgbb_key_frame, text="ImgBB 密钥：").pack(side=tk.LEFT)
        
        self.imgbb_key_var = tk.StringVar(value=getattr(self, 'imgbb_api_key', '5d170edec64cef15aefa2540d93724cc'))
        
        imgbb_key_entry = ttk.Entry(imgbb_key_frame, textvariable=self.imgbb_key_var, 
                                   show="*", width=50)
        imgbb_key_entry.pack(side=tk.LEFT, padx=(5, 10), fill=tk.X, expand=True)
        
        # 保存ImgBB密钥按钮
        save_imgbb_btn = ttk.Button(imgbb_key_frame, text="保存", 
                                   command=self.save_all_settings_and_feedback)
        save_imgbb_btn.pack(side=tk.RIGHT)
        
        # 高清处理图片的默认配置
        default_config_frame = ttk.LabelFrame(parent_frame, text="🎯 高清处理图片的默认配置", padding="15")
        default_config_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 目标尺寸配置
        target_size_frame = ttk.Frame(default_config_frame)
        target_size_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(target_size_frame, text="目标尺寸：").pack(side=tk.LEFT)
        
        # 目标宽度 - 每次打开设置页面时重新读取配置
        self.target_width_var = tk.StringVar(value=str(self.smart_upscale_config.get('target_width', 8000)))
        width_entry = ttk.Entry(target_size_frame, textvariable=self.target_width_var, width=8)
        width_entry.pack(side=tk.LEFT, padx=(5, 2))
        
        ttk.Label(target_size_frame, text="×").pack(side=tk.LEFT, padx=(2, 2))
        
        # 目标高度 - 每次打开设置页面时重新读取配置
        self.target_height_var = tk.StringVar(value=str(self.smart_upscale_config.get('target_height', 8000)))
        height_entry = ttk.Entry(target_size_frame, textvariable=self.target_height_var, width=8)
        height_entry.pack(side=tk.LEFT, padx=(2, 5))
        
        ttk.Label(target_size_frame, text="像素").pack(side=tk.LEFT, padx=(5, 10))
        
        # 显示参考信息
        size_info = ttk.Label(target_size_frame, text="(8.0K × 8.0K)", 
                             font=('Microsoft YaHei UI', 9), foreground='#666666')
        size_info.pack(side=tk.LEFT, padx=(10, 0))
        
        # 默认图片类型配置
        type_frame = ttk.Frame(default_config_frame)
        type_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(type_frame, text="默认图片类型：").pack(side=tk.LEFT)
        
        # 图片类型下拉框 - 每次打开设置页面时重新读取配置
        self.default_style_var = tk.StringVar(value=self.upscale_config.get('style', 'art'))
        style_combo = ttk.Combobox(type_frame, textvariable=self.default_style_var, 
                                  values=['art', 'photo'], state='readonly', width=15)
        style_combo.pack(side=tk.LEFT, padx=(5, 10))
        
        ttk.Label(type_frame, text="(art=卡通/插画, photo=照片)").pack(side=tk.LEFT, padx=(10, 0))
        
        # 默认降噪程度配置
        noise_frame = ttk.Frame(default_config_frame)
        noise_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(noise_frame, text="默认降噪程度：").pack(side=tk.LEFT)
        
        # 降噪程度下拉框 - 每次打开设置页面时重新读取配置
        self.default_noise_var = tk.StringVar(value=self.upscale_config.get('noise', '0'))
        noise_combo = ttk.Combobox(noise_frame, textvariable=self.default_noise_var,
                                  values=['-1', '0', '1', '2', '3'], state='readonly', width=8)
        noise_combo.pack(side=tk.LEFT, padx=(5, 10))
        
        ttk.Label(noise_frame, text="(-1=无, 0=低, 1=中, 2=高, 3=最高)").pack(side=tk.LEFT, padx=(10, 0))
        
        # 智能模式配置
        smart_frame = ttk.Frame(default_config_frame)
        smart_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 启用智能算放大倍数 - 每次打开设置页面时重新读取配置
        self.smart_upscale_enabled_var = tk.BooleanVar(value=self.smart_upscale_config.get('enabled', True))
        smart_enable_cb = ttk.Checkbutton(smart_frame, text="自动算放大倍数", 
                                         variable=self.smart_upscale_enabled_var,
                                         command=self.on_smart_upscale_enable_change)
        smart_enable_cb.pack(side=tk.LEFT, padx=(0, 20))
        
        # 跳过已达标图片 - 每次打开设置页面时重新读取配置
        self.skip_qualified_var = tk.BooleanVar(value=self.smart_upscale_config.get('skip_qualified', True))
        skip_qualified_cb = ttk.Checkbutton(smart_frame, text="跳过已达标图片", 
                                           variable=self.skip_qualified_var,
                                           command=self.on_skip_qualified_change)
        skip_qualified_cb.pack(side=tk.LEFT, padx=(0, 20))
        
        # 显示处理预览
        self.show_preview_var = tk.BooleanVar(value=True)
        show_preview_cb = ttk.Checkbutton(smart_frame, text="显示处理预览", 
                                         variable=self.show_preview_var)
        show_preview_cb.pack(side=tk.LEFT)
        
        # 移除超时配置区域 - 用户手动控制，不需要系统自动超时关闭
        
        # 绑定自动保存事件
        self.bind_auto_save_events()
        
        # 获取API密钥说明
        api_help_frame = ttk.Frame(api_frame)
        api_help_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 移除API配置说明文字以节省空间

    
    def create_author_info_tab(self, parent_frame):
        """创建作者信息标签页"""
        # 主标题
        title_label = ttk.Label(parent_frame, text="👤 作者信息", 
                               font=('Microsoft YaHei UI', 12, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 作者信息框架
        author_frame = ttk.LabelFrame(parent_frame, text="📝 开发者信息", padding="20")
        author_frame.pack(fill=tk.X, pady=(0, 20))
        
        # 软件名称
        app_name_label = ttk.Label(author_frame, text="Y2订单处理辅助工具", 
                                  font=('Microsoft YaHei UI', 14, 'bold'),
                                  foreground='#2E86AB')
        app_name_label.pack(pady=(0, 10))
        
        # 版本信息
        version_label = ttk.Label(author_frame, text="版本：1.9", 
                                 font=('Microsoft YaHei UI', 10),
                                 foreground='#666666')
        version_label.pack(pady=(0, 15))
        
        # 作者信息
        author_info_text = ("👨‍💻 开发者：陈泓旭\n"
                           "📧 联系邮箱：chx_wy@qq.com\n"
                           )
        
        author_info_label = ttk.Label(author_frame, text=author_info_text, 
                                     font=('Microsoft YaHei UI', 10),
                                     foreground='#333333')
        author_info_label.pack(pady=(0, 15))
        
        # 检查更新按钮
        update_btn = ttk.Button(author_frame, text="🔍 检查更新", 
                               command=self.check_for_updates)
        update_btn.pack(pady=(0, 10))
        
        # 自动搜索功能配置
        auto_search_frame = ttk.LabelFrame(parent_frame, text="🔍 自动搜索功能", padding="15")
        auto_search_frame.pack(fill=tk.X, pady=(20, 0))
        
        # 自动搜索开关
        auto_search_switch_frame = ttk.Frame(auto_search_frame)
        auto_search_switch_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 创建自动搜索开关变量（如果不存在）
        if not hasattr(self, 'auto_search_var'):
            self.auto_search_var = tk.BooleanVar(value=getattr(self, 'auto_search_enabled', False))
        
        auto_search_checkbox = ttk.Checkbutton(auto_search_switch_frame, 
                                              text="启用表格拖入后自动搜索图片", 
                                              variable=self.auto_search_var,
                                              command=self.on_auto_search_toggle)
        auto_search_checkbox.pack(side=tk.LEFT)
        
        # 自动搜索功能说明
        auto_search_info_label = ttk.Label(auto_search_frame, 
                                          text="💡 启用后，拖入表格并填入SKU、去除尾缀序号后会自动开始搜索图片\n💡 禁用后，需要手动点击\"开始搜索\"按钮来搜索图片", 
                                          font=('Microsoft YaHei UI', 9), 
                                          foreground='#666666')
        auto_search_info_label.pack(anchor=tk.W, pady=(5, 0))
    
    def save_api_key(self):
        """保存API密钥"""
        api_key = self.api_key_var.get().strip()
        if api_key:
            self.bigjpg_api_key = api_key
            self.save_config()
            messagebox.showinfo("成功", "API密钥已保存！")
        else:
            messagebox.showwarning("警告", "请输入有效的API密钥！")
    
    def bind_auto_save_events(self):
        """绑定自动保存事件到所有配置控件"""
        # 为所有配置变量绑定自动保存事件
        if hasattr(self, 'api_key_var'):
            self.api_key_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'target_width_var'):
            self.target_width_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'target_height_var'):
            self.target_height_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'default_style_var'):
            self.default_style_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'default_noise_var'):
            self.default_noise_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'smart_upscale_enabled_var'):
            self.smart_upscale_enabled_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'skip_qualified_var'):
            self.skip_qualified_var.trace('w', lambda *args: self.auto_save_config())
        if hasattr(self, 'show_preview_var'):
            self.show_preview_var.trace('w', lambda *args: self.auto_save_config())
        # 移除超时配置变量的绑定 - 不再需要超时机制
    
    def auto_save_config(self):
        """自动保存配置（静默保存，无提示）"""
        try:
            # 验证数值输入
            if hasattr(self, 'target_width_var') and self.target_width_var.get():
                target_width = int(self.target_width_var.get())
                if target_width <= 0:
                    return
            if hasattr(self, 'target_height_var') and self.target_height_var.get():
                target_height = int(self.target_height_var.get())
                if target_height <= 0:
                    return
            # 移除超时变量的验证 - 不再需要超时机制
            
            # 静默保存配置
            self.save_smart_upscale_config_silent()
            
        except (ValueError, AttributeError):
            # 输入无效时不保存，避免错误提示
            pass
    
    def save_smart_upscale_config_silent(self):
        """静默保存智能高清处理配置（无消息提示）"""
        try:
            # 验证输入
            target_width = int(self.target_width_var.get())
            target_height = int(self.target_height_var.get())
            # 移除超时变量的获取 - 不再需要超时机制
            
            if target_width <= 0 or target_height <= 0:
                return
            
            # 更新智能高清处理配置
            self.smart_upscale_config['target_width'] = target_width
            self.smart_upscale_config['target_height'] = target_height
            self.smart_upscale_config['enabled'] = self.smart_upscale_enabled_var.get()
            self.smart_upscale_config['skip_qualified'] = self.skip_qualified_var.get()
            # 移除超时配置的保存 - 不再需要超时机制
            
            # 更新API配置
            self.bigjpg_api_key = self.api_key_var.get()
            
            # 更新高清处理配置
            self.upscale_config['style'] = self.default_style_var.get()
            self.upscale_config['noise'] = self.default_noise_var.get()
            # 移除超时配置的保存 - 不再需要超时机制
            
            # 保存配置（静默）
            self.save_config()
            
        except (ValueError, AttributeError):
            # 静默处理错误，不显示提示
            pass

    def on_smart_upscale_enable_change(self):
        """智能放大启用状态改变时的回调"""
        # 这里可以添加额外的逻辑，比如启用/禁用相关控件
        pass
    
    def on_skip_qualified_change(self):
        """跳过已达标图片选项改变时的回调"""
        # 这里可以添加额外的逻辑
        pass
    
    def browse_excel_export_path(self):
        """浏览选择Excel导出路径"""
        folder_path = filedialog.askdirectory(
            title="选择Excel导出路径",
            initialdir=self.excel_export_path
        )
        if folder_path:
            self.excel_export_path = folder_path
            self.excel_path_var.set(folder_path)
            self.save_config()
    
    def browse_image_export_path(self):
        """浏览选择图片导出路径"""
        folder_path = filedialog.askdirectory(
            title="选择图片整理导出路径",
            initialdir=self.image_export_path
        )
        if folder_path:
            self.image_export_path = folder_path
            self.image_path_var.set(folder_path)
            self.save_config()
    
    def reset_excel_export_path(self):
        """重置Excel导出路径为桌面"""
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.excel_export_path = desktop_path
        self.excel_path_var.set(desktop_path)
        self.save_config()
        messagebox.showinfo("重置成功", "Excel导出路径已重置为桌面")
    
    def reset_image_export_path(self):
        """重置图片导出路径为桌面"""
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.image_export_path = desktop_path
        self.image_path_var.set(desktop_path)
        self.save_config()
        messagebox.showinfo("重置成功", "图片整理导出路径已重置为桌面")
    
    def validate_and_reset_export_path(self, path_type):
        """
        验证导出路径是否存在，如果不存在则自动重置为桌面
        
        Args:
            path_type: 路径类型，'excel' 或 'image'
            
        Returns:
            str: 有效的导出路径
        """
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        if path_type == 'excel':
            current_path = self.excel_export_path
            if not current_path or not os.path.exists(current_path) or not os.path.isdir(current_path):
                self.excel_export_path = desktop_path
                if hasattr(self, 'excel_path_var'):
                    self.excel_path_var.set(desktop_path)
                self.save_config()
                print(f"Excel导出路径不存在，已自动重置为桌面: {desktop_path}")
                return desktop_path
            return current_path
            
        elif path_type == 'image':
            current_path = self.image_export_path
            if not current_path or not os.path.exists(current_path) or not os.path.isdir(current_path):
                self.image_export_path = desktop_path
                if hasattr(self, 'image_path_var'):
                    self.image_path_var.set(desktop_path)
                self.save_config()
                print(f"图片导出路径不存在，已自动重置为桌面: {desktop_path}")
                return desktop_path
            return current_path
            
        return desktop_path
    
    def browse_upscale_export_path(self):
        """浏览选择高清图片保存路径"""
        folder_path = filedialog.askdirectory(title="选择高清图片保存路径")
        if folder_path:
            self.upscale_export_path = folder_path
            if hasattr(self, 'upscale_path_var'):
                self.upscale_path_var.set(folder_path)
            self.save_config()
    
    def reset_upscale_export_path(self):
        """重置高清图片保存路径为源文件位置"""
        self.upscale_export_path = ""
        if hasattr(self, 'upscale_path_var'):
            self.upscale_path_var.set("源文件位置")
        self.save_config()
        messagebox.showinfo("重置成功", "高清图片保存路径已重置为源文件位置")
    
    def add_processing_entry(self):
        """添加新的加工方式配置条目"""
        entry_frame = ttk.Frame(self.processing_entries_frame)
        entry_frame.pack(fill=tk.X, pady=2)
        
        # 序号标签
        seq_num = len(self.current_processing_entries) + 1
        seq_label = ttk.Label(entry_frame, text=f"{seq_num}.", width=3)
        seq_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # SKU前缀输入框
        prefix_entry = ttk.Entry(entry_frame, width=12)
        prefix_entry.pack(side=tk.LEFT, padx=(0, 8))
        
        # 产品名称输入框
        product_name_entry = ttk.Entry(entry_frame, width=12)
        product_name_entry.pack(side=tk.LEFT, padx=(0, 8))
        
        # 加工方式输入框
        processing_entry = ttk.Entry(entry_frame, width=12)
        processing_entry.pack(side=tk.LEFT, padx=(0, 8))
        
        # 材质输入框
        material_entry = ttk.Entry(entry_frame, width=12)
        material_entry.pack(side=tk.LEFT, padx=(0, 8))
        
        # 删除按钮
        delete_btn = ttk.Button(entry_frame, text="🗑", width=3,
                               command=lambda: self.delete_processing_entry(entry_frame))
        delete_btn.pack(side=tk.LEFT, padx=(10, 0))
        
        # 存储条目信息
        entry_info = {
            'frame': entry_frame,
            'seq_label': seq_label,
            'prefix_entry': prefix_entry,
            'product_name_entry': product_name_entry,
            'processing_entry': processing_entry,
            'material_entry': material_entry,
            'delete_btn': delete_btn
        }
        
        self.current_processing_entries.append(entry_info)
        
        # 更新序号
        self.update_processing_sequence_numbers()
    
    def delete_processing_entry(self, entry_frame):
        """删除加工方式配置条目"""
        # 找到要删除的条目
        for i, entry_info in enumerate(self.current_processing_entries):
            if entry_info['frame'] == entry_frame:
                entry_frame.destroy()
                self.current_processing_entries.pop(i)
                break
        
        # 更新序号
        self.update_processing_sequence_numbers()
    
    def update_processing_sequence_numbers(self):
        """更新加工方式配置条目的序号"""
        for i, entry_info in enumerate(self.current_processing_entries):
            entry_info['seq_label'].config(text=f"{i + 1}.")
    
    def save_processing_config(self):
        """保存加工方式配置"""
        try:
            config = {}
            for entry_info in self.current_processing_entries:
                prefix = entry_info['prefix_entry'].get().strip()
                product_name = entry_info['product_name_entry'].get().strip()
                processing = entry_info['processing_entry'].get().strip()
                material = entry_info['material_entry'].get().strip()
                
                if prefix and (product_name or processing or material):
                    config[prefix] = {
                        'product_name': product_name,
                        'processing': processing,
                        'material': material
                    }
            
            # 保存到文件，添加备份机制 - 使用应用程序目录
            app_dir = get_app_directory()
            config_file = os.path.join(app_dir, 'processing_config.json')
            backup_file = os.path.join(app_dir, 'processing_config.json.bak')
            
            # 如果原文件存在，先备份
            if os.path.exists(config_file):
                try:
                    shutil.copy2(config_file, backup_file)
                except Exception as e:
                    print(f"创建备份文件失败：{str(e)}")
            
            # 保存新配置
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            messagebox.showinfo("成功", "加工方式配置已保存！")
            
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败：{str(e)}")
            # 如果保存失败，尝试从备份恢复
            self.restore_config_from_backup()
    
    def load_existing_processing_config(self):
        """加载现有的加工方式配置"""
        try:
            app_dir = get_app_directory()
            config_file = os.path.join(app_dir, 'processing_config.json')
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # 清空现有条目
                for entry_info in self.current_processing_entries:
                    entry_info['frame'].destroy()
                self.current_processing_entries.clear()
                
                # 添加配置条目
                for prefix, info in config.items():
                    self.add_processing_entry()
                    # 填充数据
                    entry_info = self.current_processing_entries[-1]
                    entry_info['prefix_entry'].insert(0, prefix)
                    entry_info['product_name_entry'].insert(0, info.get('product_name', ''))
                    entry_info['processing_entry'].insert(0, info.get('processing', ''))
                    entry_info['material_entry'].insert(0, info.get('material', ''))
                    
        except Exception as e:
            print(f"加载加工方式配置失败：{str(e)}")
    
    def auto_load_processing_config(self):
        """程序启动时自动加载加工方式配置（静默加载，不显示消息框）"""
        try:
            # 使用应用程序目录，确保在打包后的exe中也能正确找到配置文件
            config_file = os.path.join(get_app_directory(), 'processing_config.json')
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # 如果配置不为空，则加载配置
                if config:
                    # 确保current_processing_entries已初始化
                    if not hasattr(self, 'current_processing_entries'):
                        self.current_processing_entries = []
                    
                    # 清空现有条目（如果有的话）
                    for entry_info in self.current_processing_entries:
                        if 'frame' in entry_info:
                            entry_info['frame'].destroy()
                    self.current_processing_entries.clear()
                    
                    # 添加配置条目（仅在设置窗口存在时）
                    if hasattr(self, 'settings_window') and self.settings_window and self.settings_window.winfo_exists():
                        for prefix, info in config.items():
                            self.add_processing_entry()
                            # 填充数据
                            entry_info = self.current_processing_entries[-1]
                            entry_info['prefix_entry'].insert(0, prefix)
                            entry_info['product_name_entry'].insert(0, info.get('product_name', ''))
                            entry_info['processing_entry'].insert(0, info.get('processing', ''))
                            entry_info['material_entry'].insert(0, info.get('material', ''))
                    else:
                        # 如果设置窗口不存在，只保存配置数据供后续使用
                        self.saved_processing_config = config
                    
        except Exception as e:
            print(f"自动加载加工方式配置失败：{str(e)}")
    
    def load_saved_processing_config(self):
        """加载程序启动时保存的配置数据"""
        try:
            if hasattr(self, 'saved_processing_config') and self.saved_processing_config:
                config = self.saved_processing_config
                
                # 清空现有条目
                for entry_info in self.current_processing_entries:
                    if 'frame' in entry_info:
                        entry_info['frame'].destroy()
                self.current_processing_entries.clear()
                
                # 添加配置条目
                for prefix, info in config.items():
                    self.add_processing_entry()
                    # 填充数据
                    entry_info = self.current_processing_entries[-1]
                    entry_info['prefix_entry'].insert(0, prefix)
                    entry_info['product_name_entry'].insert(0, info.get('product_name', ''))
                    entry_info['processing_entry'].insert(0, info.get('processing', ''))
                    entry_info['material_entry'].insert(0, info.get('material', ''))
                
                # 清除保存的配置数据
                delattr(self, 'saved_processing_config')
                    
        except Exception as e:
            print(f"加载保存的配置数据失败：{str(e)}")
    
    def save_all_settings_and_feedback(self):
        """保存所有设置并提供反馈 (计划书 1.8 交互升级)"""
        try:
            # 1. 保存尺寸映射 (静默模式)
            self.save_settings_mapping(silent=True)
            
            # 2. 保存API密钥
            if hasattr(self, 'api_key_var'):
                self.bigjpg_api_key = self.api_key_var.get().strip()
            if hasattr(self, 'imgbb_key_var'):
                self.imgbb_api_key = self.imgbb_key_var.get().strip()
                
            # 3. 保存高清处理配置
            if hasattr(self, 'target_width_var') and hasattr(self, 'target_height_var'):
                try:
                    self.smart_upscale_config['target_width'] = int(self.target_width_var.get())
                    self.smart_upscale_config['target_height'] = int(self.target_height_var.get())
                except ValueError:
                    pass
            
            if hasattr(self, 'default_style_var'):
                self.upscale_config['style'] = self.default_style_var.get()
            if hasattr(self, 'default_noise_var'):
                self.upscale_config['noise'] = self.default_noise_var.get()
                
            # 4. 保存命名配置
            if hasattr(self, 'template_var'):
                self.naming_center['naming_template'] = self.template_var.get().strip()
            if hasattr(self, 'auto_save_new_var'):
                self.naming_center['auto_save_new_items'] = self.auto_save_new_var.get()
                
            # 5. 执行持久化
            self.save_config()
            
            # 6. 提供视觉反馈
            if hasattr(self, 'settings_status_var'):
                self.settings_status_var.set("✅ 所有设置已成功保存！")
                # 3秒后清除反馈
                self.root.after(3000, lambda: self.settings_status_var.set(""))
                
        except Exception as e:
            print(f"保存所有设置时出错: {e}")
            if hasattr(self, 'settings_status_var'):
                self.settings_status_var.set(f"❌ 保存失败: {str(e)}")

    def close_settings_window(self, settings_window):
        """关闭设置窗口时自动保存所有配置"""
        try:
            # 自动调用全局保存逻辑（包含 API 密钥、尺寸映射等）
            self.save_all_settings_and_feedback()
            
            # 保存设置窗口的大小
            self.settings_window_geometry = settings_window.geometry()
            
            # 自动保存加工方式配置
            self.save_processing_config_silent()
            
            # 清除设置窗口引用
            if hasattr(self, 'settings_window'):
                self.settings_window = None
            
            # 关闭窗口
            settings_window.destroy()
            
        except Exception as e:
            print(f"关闭设置窗口时出错：{str(e)}")
            # 即使出错也要关闭窗口
            settings_window.destroy()
    
    def save_processing_config_silent(self):
        """静默保存加工方式配置（不显示消息框）"""
        try:
            config = {}
            for entry_info in self.current_processing_entries:
                prefix = entry_info['prefix_entry'].get().strip()
                processing = entry_info['processing_entry'].get().strip()
                material = entry_info['material_entry'].get().strip()
                product_name = entry_info['product_name_entry'].get().strip()
                
                if prefix and (processing or material):
                    config[prefix] = {
                        'processing': processing,
                        'material': material,
                        'product_name': product_name
                    }
            
            # 保存到文件，添加备份机制 - 使用应用程序目录
            config_file = os.path.join(get_app_directory(), 'processing_config.json')
            backup_file = os.path.join(get_app_directory(), 'processing_config.json.bak')
            
            # 如果原文件存在，先备份
            if os.path.exists(config_file):
                try:
                    shutil.copy2(config_file, backup_file)
                except Exception as e:
                    print(f"创建备份文件失败：{str(e)}")
            
            # 保存新配置
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            print("加工方式配置已自动保存")
            
        except Exception as e:
            print(f"自动保存配置失败：{str(e)}")
            # 如果保存失败，尝试从备份恢复
            self.restore_config_from_backup()
    
    def restore_config_from_backup(self):
        """从备份文件恢复配置"""
        try:
            config_file = os.path.join(get_app_directory(), 'processing_config.json')
            backup_file = os.path.join(get_app_directory(), 'processing_config.json.bak')
            
            if os.path.exists(backup_file):
                shutil.copy2(backup_file, config_file)
                print("已从备份文件恢复配置")
            else:
                print("备份文件不存在，无法恢复")
                
        except Exception as e:
            print(f"从备份恢复配置失败：{str(e)}")
    
    def clear_processing_config(self):
        """清空加工方式配置"""
        if messagebox.askyesno("确认", "确定要清空所有加工方式配置吗？"):
            # 销毁所有条目框架
            for entry_info in self.current_processing_entries:
                entry_info['frame'].destroy()
            # 清空条目列表
            self.current_processing_entries.clear()
            # 保存空配置到文件
            self.save_processing_config()
            messagebox.showinfo("成功", "加工方式配置已清空！")
    
    def export_processing_config(self):
        """导出加工方式配置到Excel"""
        try:
            # 收集配置数据
            data = []
            for i, entry_info in enumerate(self.current_processing_entries, 1):
                prefix = entry_info['prefix_entry'].get().strip()
                processing = entry_info['processing_entry'].get().strip()
                material = entry_info['material_entry'].get().strip()
                
                if prefix or processing or material:
                    data.append({
                        '序号': i,
                        'SKU前缀': prefix,
                        '加工方式': processing,
                        '材质': material
                    })
            
            if not data:
                messagebox.showwarning("警告", "没有配置数据可导出！")
                return
            
            # 临时释放设置窗口的grab_set，避免文件对话框被遮挡
            if hasattr(self, 'settings_window') and self.settings_window:
                self.settings_window.grab_release()
            
            # 选择保存位置
            file_path = filedialog.asksaveasfilename(
                title="导出加工方式配置",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            # 恢复设置窗口的grab_set
            if hasattr(self, 'settings_window') and self.settings_window:
                self.settings_window.grab_set()
            
            if file_path:
                pd = _import_pandas()
                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False, engine='openpyxl')
                messagebox.showinfo("成功", f"配置已导出到：{file_path}")
                
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")
    
    def import_processing_config(self):
        """从Excel导入加工方式配置"""
        try:
            # 临时释放设置窗口的grab_set，确保文件对话框能正常显示
            settings_window = None
            if hasattr(self, 'settings_window') and self.settings_window:
                settings_window = self.settings_window
                settings_window.grab_release()
            
            file_path = filedialog.askopenfilename(
                title="导入加工方式配置",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            # 恢复设置窗口的grab_set
            if settings_window:
                settings_window.grab_set()
            
            if file_path:
                pd = _import_pandas()
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # 清空现有配置
                self.clear_processing_config()
                
                # 导入数据
                for _, row in df.iterrows():
                    self.add_processing_entry()
                    entry_info = self.current_processing_entries[-1]
                    
                    # 填充数据
                    if 'SKU前缀' in row and pd.notna(row['SKU前缀']):
                        entry_info['prefix_entry'].insert(0, str(row['SKU前缀']))
                    if '加工方式' in row and pd.notna(row['加工方式']):
                        entry_info['processing_entry'].insert(0, str(row['加工方式']))
                    if '材质' in row and pd.notna(row['材质']):
                        entry_info['material_entry'].insert(0, str(row['材质']))
                
                messagebox.showinfo("成功", "配置导入完成！")
                
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")
    
    def check_unmatched_skus(self, sku_list):
        """检测未匹配到加工方式配置的SKU"""
        try:
            config_file = os.path.join(get_app_directory(), 'processing_config.json')
            if not os.path.exists(config_file):
                # 如果配置文件不存在，所有SKU都是未匹配的
                print(f"🚨 配置文件不存在: {config_file}")
                return list(set(sku_list))
            
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            print(f"🔍 check_unmatched_skus 开始检查:")
            print(f"  📋 SKU列表: {sku_list}")
            print(f"  ⚙️ 配置前缀: {list(config.keys())}")
            
            unmatched_skus = []
            for sku in sku_list:
                matched = False
                sku_prefix = self.extract_sku_prefix(sku)
                print(f"  🔍 检查SKU '{sku}' -> 提取前缀: '{sku_prefix}'")
                
                for prefix in config.keys():
                    if sku_prefix.upper() == prefix.upper():
                        print(f"    ✅ 匹配成功: '{sku_prefix}' == '{prefix}'")
                        matched = True
                        break
                    else:
                        print(f"    ❌ 不匹配: '{sku_prefix}' != '{prefix}'")
                
                if not matched:
                    print(f"  🚨 SKU '{sku}' 未匹配到任何前缀")
                    unmatched_skus.append(sku)
                else:
                    print(f"  ✅ SKU '{sku}' 已匹配")
            
            print(f"📊 检查结果: 未匹配SKU = {unmatched_skus}")
            # 去重并返回
            return list(set(unmatched_skus))
            
        except Exception as e:
            print(f"检测未匹配SKU失败：{str(e)}")
            return list(set(sku_list))

    def _collect_presets_for_unmatched(self, unmatched_skus):
        """收集与未匹配SKU相关的预设(名称匹配/加工材质)用于弹窗推荐"""
        try:
            print(f"调试：开始收集预设，未匹配SKU: {unmatched_skus}")
            
            # 加工方式与材质配置
            processing_config = {}
            p_file = os.path.join(get_app_directory(), 'processing_config.json')
            if os.path.exists(p_file):
                try:
                    with open(p_file, 'r', encoding='utf-8') as f:
                        processing_config = json.load(f) or {}
                    print(f"调试：成功加载processing_config: {processing_config}")
                except Exception as e:
                    print(f"调试：加载processing_config失败: {e}")
                    processing_config = {}

            def matches_prefix(sku, prefix):
                sku_prefix = self.extract_sku_prefix(sku)
                result = sku_prefix.upper() == str(prefix).upper()
                print(f"调试：SKU '{sku}' 前缀 '{sku_prefix}' 匹配前缀 '{prefix}': {result}")
                return result

            all_prefixes = set(list(processing_config.keys()))
            print(f"调试：所有可用前缀: {all_prefixes}")
            
            suggestions = []
            matched_suggestions = []
            all_suggestions = []
            
            # 构建所有预设选项
            for prefix in all_prefixes:
                proc_info = processing_config.get(prefix, {}) or {}
                processing = proc_info.get('processing', '')
                material = proc_info.get('material', '')
                product_name = proc_info.get('product_name', '')
                
                base = product_name if product_name else str(prefix)
                details = [v for v in [processing, material] if v]
                display = base + (f"（{' / '.join(details)}）" if details else '')
                
                suggestion = {
                    'prefix': prefix,
                    'display': display,
                    'processing': processing,
                    'material': material,
                    'export_name': product_name
                }
                
                all_suggestions.append(suggestion)
                
                # 检查是否与未匹配SKU相关
                if any(matches_prefix(sku, prefix) for sku in unmatched_skus):
                    matched_suggestions.append(suggestion)
                    print(f"调试：匹配的预设: {suggestion['display']}")

            print(f"调试：匹配的预设数量: {len(matched_suggestions)}")
            print(f"调试：所有预设数量: {len(all_suggestions)}")

            # 优先返回匹配的预设，如果没有匹配的则返回所有预设
            # 这样确保用户总是能看到预设选项，特别是在连续处理表格时
            if matched_suggestions:
                suggestions = matched_suggestions
                print("调试：使用匹配的预设")
            else:
                suggestions = all_suggestions
                print("调试：使用所有预设（回退模式）")
            
            print(f"调试：最终返回预设数量: {len(suggestions)}")
            for i, s in enumerate(suggestions):
                print(f"调试：预设 {i+1}: {s['display']}")
                
            return suggestions
            
        except Exception as e:
            print(f"收集预设选项失败：{str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def show_manual_processing_dialog(self, unmatched_skus, all_skus, df, excel_path):
        """显示手动填写加工方式与材质的弹窗"""
        
        def apply_and_continue():
            """应用设置并继续处理"""
            processing = processing_var.get().strip()
            material = material_var.get().strip()
            
            # 获取选择的预设
            try:
                selected_preset_display = preset_var.get().strip() if 'preset_var' in locals() else ''
            except Exception:
                selected_preset_display = ''
            print(f"调试：选择的预设 = '{selected_preset_display}'")
            
            # 保存用户输入到配置中，包括预设选择
            self.last_manual_processing = {
                'processing': processing,
                'material': material
            }
            # 保存预设选择
            if selected_preset_display:
                self.last_selected_preset = selected_preset_display
                print(f"调试：保存预设选择到配置: {selected_preset_display}")
            
            self.save_config()
            
            # 创建特定SKU的临时映射
            self.temp_sku_mapping = {}
            for sku in unmatched_skus:
                self.temp_sku_mapping[sku] = {
                    'processing': processing,
                    'material': material
                }
            print(f"调试：创建temp_sku_mapping，包含{len(self.temp_sku_mapping)}个SKU:")
            for sku, info in self.temp_sku_mapping.items():
                print(f"  - {sku}: 加工方式='{info['processing']}', 材质='{info['material']}'")
            print(f"调试：temp_sku_mapping对象ID = {id(self.temp_sku_mapping)}")
            
            # 标记对话框已关闭
            self._current_dialog_active = False
            dialog.destroy()

            # 仅在用户选择了预设时触发二次尺寸处理；并记录本次涉及的未匹配SKU
            if selected_preset_display:
                self._secondary_processing_triggered_manually = True
                try:
                    self._manual_preset_skus = set(unmatched_skus)
                except Exception:
                    self._manual_preset_skus = set()
                print(f"调试：设置二次处理标志为True，涉及SKU数量 = {len(getattr(self, '_manual_preset_skus', set()))}")
            else:
                self._secondary_processing_triggered_manually = False
                self._manual_preset_skus = set()

            # 继续处理Excel文件
            if hasattr(self, '_is_multi_table_processing') and self._is_multi_table_processing:
                self.continue_excel_processing_multi(all_skus, df, excel_path)
                # 在多表格处理模式下，继续处理队列
                self.continue_queue_processing()
            else:
                self.continue_excel_processing(all_skus, df, excel_path)

        def cancel_processing():
            """取消处理，直接继续处理表格"""
            # 清除任何临时设置
            if hasattr(self, 'temp_processing_info'):
                delattr(self, 'temp_processing_info')
            if hasattr(self, 'temp_sku_mapping'):
                delattr(self, 'temp_sku_mapping')
            # 清除二次尺寸处理触发标记与集合
            self._secondary_processing_triggered_manually = False
            self._manual_preset_skus = set()
            
            # 标记对话框已关闭
            self._current_dialog_active = False
            dialog.destroy()
            
            # 直接继续处理Excel文件，不设置任何临时信息
            if hasattr(self, '_is_multi_table_processing') and self._is_multi_table_processing:
                self.continue_excel_processing_multi(all_skus, df, excel_path)
                # 在多表格处理模式下，继续处理队列
                self.continue_queue_processing()
            else:
                self.continue_excel_processing(all_skus, df, excel_path)
        
        dialog = tk.Toplevel(self.root)
        dialog.title("手动设置加工方式与材质")
        dialog.geometry("500x550")
        dialog.resizable(True, True)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.attributes('-topmost', True)
        
        # 立即隐藏窗口，避免在左上角显示
        dialog.withdraw()
        
        # 居中显示 - 先更新布局但窗口仍然隐藏
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 设置好位置后再显示窗口，避免移动效果
        dialog.deiconify()
        
        # 主框架
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="🔧 检测到未配置的SKU前缀", 
                               font=('Microsoft YaHei UI', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # 说明文字
        info_text = f"有 {len(unmatched_skus)} 个SKU未匹配，请设置加工方式与材质"
        info_label = ttk.Label(main_frame, text=info_text, 
                              font=('Microsoft YaHei UI', 9))
        info_label.pack(pady=(0, 8))
        
        # 未匹配SKU显示区域
        sku_frame = ttk.LabelFrame(main_frame, text="未匹配的SKU", padding="10")
        sku_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # SKU列表显示
        sku_text = tk.Text(sku_frame, height=9, width=50, wrap=tk.WORD, 
                          font=('Consolas', 9), state=tk.DISABLED)
        sku_scrollbar = ttk.Scrollbar(sku_frame, orient=tk.VERTICAL, command=sku_text.yview)
        sku_text.configure(yscrollcommand=sku_scrollbar.set)
        
        sku_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sku_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 显示未匹配的SKU
        sku_text.config(state=tk.NORMAL)
        sku_text.insert(tk.END, '\n'.join(unmatched_skus))
        sku_text.config(state=tk.DISABLED)

        # 推荐预设区域（如果存在）
        presets = self._collect_presets_for_unmatched(unmatched_skus)
        preset_map = {p['display']: p for p in presets}
        
        # 初始化预设变量（无论是否有预设都要初始化）
        preset_var = tk.StringVar()
        
        if presets:
            preset_frame = ttk.LabelFrame(main_frame, text="选择预设配置", padding="10")
            preset_frame.pack(fill=tk.X, pady=(0, 10))
            ttk.Label(preset_frame, text="选择预设：", font=('Microsoft YaHei UI', 10)).grid(row=0, column=0, sticky=tk.W)
            preset_combo = ttk.Combobox(preset_frame, textvariable=preset_var, state='readonly',
                                        values=list(preset_map.keys()), width=50)
            preset_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))
            preset_frame.columnconfigure(1, weight=1)
            
            # 设置上次选择的预设
            if hasattr(self, 'last_selected_preset') and self.last_selected_preset:
                if self.last_selected_preset in preset_map:
                    preset_var.set(self.last_selected_preset)

        # 输入区域
        input_frame = ttk.LabelFrame(main_frame, text="设置加工方式与材质", padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 加工方式输入
        ttk.Label(input_frame, text="🔧 加工方式：", font=('Microsoft YaHei UI', 10)).grid(
            row=0, column=0, sticky=tk.W, pady=(0, 5))
        processing_var = tk.StringVar()
        # 设置上次保存的加工方式
        if hasattr(self, 'last_manual_processing') and self.last_manual_processing['processing']:
            processing_var.set(self.last_manual_processing['processing'])
        processing_entry = ttk.Entry(input_frame, textvariable=processing_var, 
                                   font=('Microsoft YaHei UI', 10), width=30)
        processing_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 5), padx=(10, 0))
        
        # 材质输入
        ttk.Label(input_frame, text="🧱 材质：", font=('Microsoft YaHei UI', 10)).grid(
            row=1, column=0, sticky=tk.W, pady=(5, 0))
        material_var = tk.StringVar()
        # 设置上次保存的材质
        if hasattr(self, 'last_manual_processing') and self.last_manual_processing['material']:
            material_var.set(self.last_manual_processing['material'])
        material_entry = ttk.Entry(input_frame, textvariable=material_var, 
                                 font=('Microsoft YaHei UI', 10), width=30)
        material_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(5, 0), padx=(10, 0))

        input_frame.columnconfigure(1, weight=1)

        # 选择预设后自动填充并记忆导出名称
        def on_preset_selected(event=None):
            if not presets:
                return
            selected = preset_var.get()
            preset = preset_map.get(selected)
            if not preset:
                return
            
            # 保存用户选择的预设
            self.last_selected_preset = selected
            try:
                self.save_config()
            except Exception:
                pass
            
            # 无论预设值是否为空，都设置到输入框（空值会清空输入框）
            processing_var.set(preset.get('processing', ''))
            material_var.set(preset.get('material', ''))
            # 名称处理：不直接使用“名称匹配的原名”，而是按普通匹配模式仅替换产品类型部分
            try:
                target_product = (preset.get('export_name') or '').strip()
                if target_product:
                    # 获取当前的基础名称
                    current_name = getattr(self, 'last_excel_export_name', None) or "已整理尺寸表格"
                    
                    # 检查当前名称是否已经包含目标产品类型
                    # 移除序列号后检查基础名称
                    base_name_without_sequence = self._remove_sequence_number(current_name)
                    
                    # 检查基础名称是否已经包含目标产品类型
                    if target_product in base_name_without_sequence:
                        print(f"  ℹ️ 当前名称已包含目标产品类型 '{target_product}'，跳过替换")
                        # 如果已经包含目标产品类型，不需要再次替换
                        # 但仍需要更新last_excel_export_name为干净的基础名称
                        self.last_excel_export_name = base_name_without_sequence
                    else:
                        # 如果不包含目标产品类型，进行替换
                        print(f"  🔄 当前名称不包含目标产品类型 '{target_product}'，进行替换")
                        prefix_for_reason = [preset.get('prefix')] if preset.get('prefix') else []
                        modified_name, _ = self._flexible_name_replacement(base_name_without_sequence, target_product, prefix_for_reason)
                        self.last_excel_export_name = modified_name
                    
                    try:
                        self.save_config()
                    except Exception:
                        pass
            except Exception:
                # 如果替换逻辑出错，保持现状，不影响用户继续使用
                pass

        if presets:
            preset_combo.bind('<<ComboboxSelected>>', on_preset_selected)
            # 如果已经设置了预设值，手动触发一次联动更新
            if preset_var.get():
                on_preset_selected()
        
        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        # 按钮
        ttk.Button(button_frame, text="✅ 应用并继续", 
                  command=apply_and_continue, style="Success.TButton").pack(
                  side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="❌ 跳过并继续", 
                  command=cancel_processing).pack(side=tk.RIGHT)
        
        # 设置焦点
        processing_entry.focus()

    def continue_excel_processing(self, sku_list, df, excel_path):
        """使用临时设置继续处理Excel文件"""
        try:
            # 显示进度
            self.show_progress()
            self.update_progress(90, "应用临时设置并处理...")
            
            # 检查搜索框是否已经包含相同的SKU内容，避免重复填入
            current_content = self.search_entry.get('1.0', tk.END).strip()
            expected_content = '\n'.join(sku_list)
            
            # 只有当搜索框内容与期望内容不同时才重新填入
            if current_content != expected_content:
                # 清空搜索框并填入SKU数据
                self.search_entry.delete('1.0', tk.END)
                self.search_entry.insert('1.0', expected_content)
                self.search_entry.config(fg='#212529')
                
                # 自动去除序号尾缀，确保搜索框中的SKU是无尾缀的
                self.remove_suffix()
            
            self.update_progress(100, "Excel处理完成！")
            
            # 如果有尺寸列，询问用户是否整理尺寸并另存桌面
            if self.size_column in df.columns:
                result = self.show_size_processing_dialog(len(sku_list))
                if result == "yes":
                    # 显示表格导出弹窗，传递SKU列表
                    custom_name = self.show_excel_export_dialog(sku_list)
                    if custom_name is not None:  # 用户点击了确认
                        self.process_and_export_excel(df, excel_path, custom_name=custom_name)
                elif result == "process_and_open":
                    # 显示表格导出弹窗，传递SKU列表
                    custom_name = self.show_excel_export_dialog(sku_list)
                    if custom_name is not None:  # 用户点击了确认
                        self.process_and_export_excel(df, excel_path, custom_name=custom_name, open_after_export=True)
            else:
                processing_info = ""
                if hasattr(self, 'temp_processing_info'):
                    if self.temp_processing_info['processing']:
                        processing_info += f"\n加工方式：{self.temp_processing_info['processing']}"
                    if self.temp_processing_info['material']:
                        processing_info += f"\n材质：{self.temp_processing_info['material']}"
                
                messagebox.showinfo("成功", f"已成功处理Excel文件！\n找到 {len(sku_list)} 个SKU\n已自动填入搜索框{processing_info}")
                
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("错误", f"继续处理Excel文件时出错：{str(e)}")

    def continue_excel_processing_multi(self, sku_list, df, excel_path):
        """专用于多表格处理的Excel处理函数，不修改搜索框内容"""
        try:
            # 显示进度
            self.show_progress()
            self.update_progress(90, "应用临时设置并处理...")
            
            # 注意：多表格处理时不修改搜索框内容，保持用户第一次拖入时填入的内容
            
            self.update_progress(100, "Excel处理完成！")
            
            # 如果有尺寸列，询问用户是否整理尺寸并另存桌面
            if self.size_column in df.columns:
                result = self.show_size_processing_dialog(len(sku_list))
                if result == "yes":
                    # 显示表格导出弹窗
                    custom_name = self.show_excel_export_dialog(sku_list)
                    if custom_name is not None:  # 用户点击了确认
                        self.process_and_export_excel(df, excel_path, custom_name=custom_name)
                elif result == "process_and_open":
                    # 显示表格导出弹窗
                    custom_name = self.show_excel_export_dialog(sku_list)
                    if custom_name is not None:  # 用户点击了确认
                        self.process_and_export_excel(df, excel_path, custom_name=custom_name, open_after_export=True)
            else:
                processing_info = ""
                if hasattr(self, 'temp_processing_info'):
                    if self.temp_processing_info['processing']:
                        processing_info += f"\n加工方式：{self.temp_processing_info['processing']}"
                    if self.temp_processing_info['material']:
                        processing_info += f"\n材质：{self.temp_processing_info['material']}"
                
                # 多表格处理时的静默提示，不显示弹窗
                print(f"多表格处理：已成功处理Excel文件！找到 {len(sku_list)} 个SKU{processing_info}")
                
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("错误", f"多表格处理Excel文件时出错：{str(e)}")

    def parse_g_cell_to_sku_list(self, g_cell_value, skuid_count):
        """
        解析G列内容为SKU列表，支持序列性SKU、多行SKU和描述性字符串
        
        Args:
            g_cell_value: G列原始值
            skuid_count: SKUID数量（B列SKU数量）
            
        Returns:
            list: 解析后的SKU列表
        """
        if not g_cell_value:
            return [''] * skuid_count
        
        g_str = str(g_cell_value).strip()
        
        # 首先检查是否包含换行符（多行SKU）
        if '\n' in g_str:
            # 分割多行SKU
            lines = [line.strip() for line in g_str.split('\n') if line.strip()]
            if len(lines) >= skuid_count:
                # 如果行数足够，直接返回前N行
                return lines[:skuid_count]
            else:
                # 如果行数不够，用最后一行补齐
                result = lines[:]
                while len(result) < skuid_count:
                    result.append(lines[-1] if lines else '')
                return result
        
        # 检查是否为序列性SKU（包含数字后缀）
        # 改进的正则匹配：寻找最后一段连续数字作为序号
        # 匹配模式：任意字符 + 最后的连续数字（1-4位，避免匹配过长的数字串）
        sequential_match = re.match(r'^(.+?)(\d{1,4})$', g_str)
        if sequential_match and len(sequential_match.group(2)) <= 4:
            base_sku = sequential_match.group(1)
            start_num = int(sequential_match.group(2))
            
            # 生成序列性SKU列表
            sku_list = []
            num_digits = len(sequential_match.group(2))  # 保持原有的数字位数
            for i in range(skuid_count):
                new_num = start_num + i
                # 保持原有的数字位数格式（如01, 02, 03）
                formatted_num = str(new_num).zfill(num_digits)
                sku_list.append(f"{base_sku}{formatted_num}")
            
            return sku_list
        else:
            # 非序列性，所有行使用相同的描述性字符串
            return [g_str] * skuid_count

    def split_multi_sku_row_advanced(self, worksheet, row_idx, skuid_list, size_list, sku_blocks, size_column_index, processing_column_index, material_column_index, quantity_column_index, order_no_column_index, sku_column_index):
        """拆分包含多个SKUID的行（高级版本）"""
        print(f"进入split_multi_sku_row_advanced：行{row_idx}，SKU数量={len(skuid_list) if skuid_list else 0}")
        try:
            # 参数验证和边界情况处理
            if not skuid_list or len(skuid_list) <= 1:
                print(f"警告：第{row_idx}行SKU列表为空或只有一个SKU，跳过拆分。SKU列表: {skuid_list}")
                return
            
            if len(size_list) != len(skuid_list):
                print(f"警告：第{row_idx}行尺寸列表长度({len(size_list)})与SKU列表长度({len(skuid_list)})不匹配")
                # 自动调整尺寸列表长度
                if len(size_list) < len(skuid_list):
                    # 如果尺寸不够，用最后一个尺寸填充
                    last_size = size_list[-1] if size_list else ''
                    size_list.extend([last_size] * (len(skuid_list) - len(size_list)))
                    print(f"已自动填充尺寸列表: {size_list}")
                else:
                    # 如果尺寸太多，截断到SKU数量
                    size_list = size_list[:len(skuid_list)]
                    print(f"已截断尺寸列表: {size_list}")
            
            print(f"开始拆分第{row_idx}行，SKU数量: {len(skuid_list)}, 尺寸数量: {len(size_list)}")
            # 获取原行的所有数据
            original_data = {}
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                original_data[col_idx] = {
                    'value': cell.value,
                    'font': cell.font,
                    'alignment': cell.alignment,
                    'fill': cell.fill,
                    'border': cell.border
                }
            
            # 需要插入的新行数量（SKU数量-1）
            new_rows_count = len(skuid_list) - 1

            # ==== I列GroupID 标记处理 ====
            # 使用I列（第9列）进行GroupID标记
            group_col_idx = 9  # I列

            if not hasattr(self, '_next_group_id'):
                self._next_group_id = 1
            group_id = self._next_group_id
            self._next_group_id += 1

            # 在原始行的I列写入 GroupID
            worksheet.cell(row=row_idx, column=group_col_idx, value=group_id)
            print(f"原行{row_idx}在I列标记GroupID: {group_id}")
            # ==== I列GroupID 标记处理 ====
            
            # 先处理图片位置调整：
            # 1. 收集位于当前拆分行 (row_idx) 的所有图片，稍后重新分配；
            # 2. 其余位于当前行之后(包括当前行)的图片整体下移 new_rows_count 行，以给新增行腾出空间。
            orig_row_images = []  # 存放当前行（被拆分行）中的所有图片
            if hasattr(worksheet, '_images') and worksheet._images and new_rows_count > 0:
                for image in worksheet._images:
                    if not (hasattr(image, 'anchor') and hasattr(image.anchor, '_from')):
                        continue  # 跳过无法识别锚点的图片

                    img_from = image.anchor._from  # openpyxl.utils.cell.CellMarker
                    img_row0 = img_from.row  # 0-based 行索引

                    # 步骤1: 如果图片位于被拆分的原始行，则暂存以便后续重新分配
                    if img_row0 == row_idx - 1:
                        orig_row_images.append(image)
                        continue  # 先不移动

                    # 步骤2: 其余图片如果在原始行之后，则统一下移 new_rows_count 行
                    if img_row0 >= row_idx:
                        img_from.row += new_rows_count
                        if hasattr(image.anchor, 'to') and image.anchor.to:
                            image.anchor.to.row += new_rows_count

            # 在当前行后插入新行
            for i in range(new_rows_count):
                worksheet.insert_rows(row_idx + 1 + i)

            # 步骤3: 将 orig_row_images 依次分配到拆分后的行的 H 列
            if orig_row_images:
                # 按原始列索引排序，确保分配顺序符合 H, I, J...
                orig_row_images.sort(key=lambda img: getattr(img.anchor._from, 'col', 7))

                for idx_img, img in enumerate(orig_row_images):
                    # 目标行 (0-based) = 原始行起始 + idx_img
                    if idx_img <= new_rows_count:
                        target_row0 = (row_idx - 1) + idx_img  # openpyxl内部0基
                    else:
                        # 如果图片数量超过拆分行数，全部定位到最后一行
                        target_row0 = (row_idx - 1) + new_rows_count

                    target_col0 = 7  # H 列 0-based

                    # 计算跨度，保持原始大小
                    row_span = 0
                    col_span = 0
                    if hasattr(img.anchor, 'to') and img.anchor.to:
                        row_span = img.anchor.to.row - img.anchor._from.row
                        col_span = img.anchor.to.col - img.anchor._from.col

                    # 设置新的锚点
                    img.anchor._from.row = target_row0
                    img.anchor._from.col = target_col0
                    if hasattr(img.anchor, 'to') and img.anchor.to:
                        img.anchor.to.row = target_row0 + row_span
                        img.anchor.to.col = target_col0 + col_span
            
            # --- 预处理数量列平均分配逻辑 ---
            equal_qty_distribution = False
            total_rows_to_fill = len(skuid_list)
            if quantity_column_index and quantity_column_index in original_data:
                try:
                    original_qty_val = original_data[quantity_column_index]['value']
                    if original_qty_val is not None and str(original_qty_val).strip() != '':
                        qty_numeric = int(float(original_qty_val))
                        if qty_numeric == total_rows_to_fill:
                            # 当原数量与拆分后行数一致，说明每行数量应为1
                            equal_qty_distribution = True
                except Exception as e:
                    print(f"数量列预处理解析失败: {e}")

            # 填充所有相关行的数据
            for i, skuid in enumerate(skuid_list):
                current_row = row_idx + i
                
                # 如果是新增行（i > 0），先复制原始行的格式
                if i > 0:
                    # 复制原始行的所有格式到新增行
                    for col in range(1, worksheet.max_column + 1):
                        source_cell = worksheet.cell(row=row_idx, column=col)
                        target_cell = worksheet.cell(row=current_row, column=col)
                        
                        # 复制格式
                        if source_cell.font:
                            target_cell.font = copy(source_cell.font)
                        if source_cell.alignment:
                            target_cell.alignment = copy(source_cell.alignment)
                        if source_cell.border:
                            target_cell.border = copy(source_cell.border)
                        if source_cell.fill:
                            target_cell.fill = copy(source_cell.fill)
                        if source_cell.number_format:
                            target_cell.number_format = source_cell.number_format
                        
                        # 复制值（除了特殊列）
                        if col == 1:  # 序号列清空并标记为新增行
                            target_cell.value = "SKIP_SEQUENCE"  # 标记新增行跳过序号
                        elif col == order_no_column_index:  # D列（订单号）- 新增行清空，原始行保持
                            target_cell.value = None
                        elif col not in [sku_column_index, size_column_index, quantity_column_index]:  # 排除SKU、尺寸、数量列
                            target_cell.value = source_cell.value
                
                # 行高设置已移至边框设置之后统一处理
                
                # 处理原始行（i==0）- 确保D列订单号不被清空
                if i == 0:
                    # 写入GroupID标记到原始行
                    worksheet.cell(row=current_row, column=group_col_idx, value=group_id)
                    # 确保原始行的D列订单号保持不变（不需要额外操作，因为原始行不在上面的复制逻辑中）
                
                # 处理序号列（A列）- 新增行不显示序号但保留SKIP_SEQUENCE标记
                if i > 0:
                    sequence_cell = worksheet.cell(row=current_row, column=1)
                    sequence_cell.value = "SKIP_SEQUENCE"  # 保持标记，用于后续序号填充时跳过
                    # 写入GroupID标记
                    worksheet.cell(row=current_row, column=group_col_idx, value=group_id)
                
                # 填充SKU（G列）- 产品代码SKU（统一行为）
                if sku_column_index:
                    # 从原始G列解析SKU列表
                    original_g_value = original_data.get(sku_column_index, {}).get('value')
                    parsed_sku_list = self.parse_g_cell_to_sku_list(original_g_value, len(skuid_list))
                    
                    # 验证G列处理逻辑：确保G列SKU处理与B列提取的信息保持一致
                    if i == 0:  # 只在第一次处理时打印验证信息
                        print(f"🔍 [G列验证] 原始G列值: {original_g_value}")
                        print(f"🔍 [G列验证] B列SKUID数量: {len(skuid_list)}")
                        print(f"🔍 [G列验证] 解析后G列SKU数量: {len(parsed_sku_list)}")
                        print(f"🔍 [G列验证] 解析后G列SKU列表: {parsed_sku_list}")
                        
                        if len(parsed_sku_list) != len(skuid_list):
                            print(f"⚠️ [G列验证] 警告：G列SKU数量({len(parsed_sku_list)})与B列SKUID数量({len(skuid_list)})不匹配")
                        else:
                            print(f"✅ [G列验证] G列SKU数量与B列SKUID数量匹配")
                    
                    sku_value_to_write = parsed_sku_list[i] if i < len(parsed_sku_list) else ''

                    sku_cell = worksheet.cell(row=current_row, column=sku_column_index)
                    sku_cell.value = sku_value_to_write
                    # 保持原有格式，只需要确保居中对齐
                    if not sku_cell.font:
                        sku_cell.font = Font(name='宋体', size=18)
                    if not sku_cell.alignment:
                        sku_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 若M≠N，则仅在第一原始拆分行（i==0）为G列背景填充红色，不改变文字
                    try:
                        m_count = len(parsed_sku_list)
                        n_count = len(skuid_list)
                        if i == 0 and m_count != n_count:
                            sku_cell.fill = PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                    except Exception:
                        pass
                    
                    # 填充加工方式和材质信息（与单件订单保持一致的逻辑）
                    if sku_value_to_write:
                        processing, material = self.get_processing_info_by_sku(sku_value_to_write)
                        
                        # 填充加工方式
                        if processing and processing_column_index:
                            processing_cell = worksheet.cell(row=current_row, column=processing_column_index)
                            processing_cell.value = processing
                            # 设置加工方式字体为宋体、24号
                            processing_cell.font = Font(name='宋体', size=24)
                            print(f"填入第{current_row}行加工方式: {processing}")
                        
                        # 填充材质
                        if material and material_column_index:
                            material_cell = worksheet.cell(row=current_row, column=material_column_index)
                            material_cell.value = material
                            # 设置材质字体为宋体、24号、红色
                            material_cell.font = Font(name='宋体', size=24, color='FF0000')
                            print(f"填入第{current_row}行材质: {material}")
                
                # 填充尺寸（B列）- 直接保留原始Variants内容，不进行尺寸映射
                if size_column_index and i < len(size_list):
                    size_cell = worksheet.cell(row=current_row, column=size_column_index)
                    
                    # 🔧 修复多行拆分尺寸映射问题：直接使用size_list中的原始Variants内容
                    # size_list现在已经包含了原始的Variants内容，无需重复提取
                    original_size_for_mapping = size_list[i] if i < len(size_list) and size_list[i] else None
                    
                    # 如果没有尺寸，跳过
                    if not original_size_for_mapping:
                        print(f"⚠️ 第{current_row}行没有找到有效尺寸，跳过")
                        continue
                    
                    print(f"🔍 使用size_list中的原始Variants格式: '{original_size_for_mapping}'")
                    
                    # 🔧 关键修复：多行SKU处理时直接保留原始Variants内容，不进行尺寸映射
                    size_cell.value = original_size_for_mapping
                    # 保持原有格式，只需要确保居中对齐
                    if not size_cell.font:
                        size_cell.font = Font(name='宋体', size=18)
                    if not size_cell.alignment:
                        size_cell.alignment = Alignment(horizontal='center', vertical='center')
                    print(f"✅ 填入第{current_row}行B列原始尺寸(无映射): {original_size_for_mapping}")
                    
                    # 🔧 多行SKU处理时跳过尺寸映射，直接保留原始Variants格式
                    print(f"🔧 多行SKU处理：跳过尺寸映射，保留原始Variants格式")
                    
                    # 二次尺寸处理：仅在用户手动选择了预设时触发（与单件订单保持一致）
                    try:
                        secondary_flag = getattr(self, '_secondary_processing_triggered_manually', False)
                        print(f"调试ADV：二次处理标志 = {secondary_flag}, 尺寸列索引 = {size_column_index}")
                        if size_column_index and secondary_flag and sku_value_to_write:
                            processing, material = self.get_processing_info_by_sku(sku_value_to_write)
                            original_size = str(size_cell.value) if size_cell.value else ""
                            print(f"调试ADV：行{current_row} 原始尺寸 = '{original_size}', 加工方式 = '{processing}'")
                            if size_cell.value:
                                # 判断产品类型，只对窗帘和床上三件套进行二次处理
                                current_preset = getattr(self, 'last_selected_preset', '')
                                should_process = self._should_apply_secondary_processing(processing, current_preset)
                                print(f"调试ADV：产品类型判断 - 是否需要二次处理: {should_process}")
                                
                                if should_process:
                                    new_size = self.apply_secondary_size_processing(str(size_cell.value), current_preset)
                                    print(f"调试ADV：二次处理结果 = '{new_size}'")
                                    if new_size:
                                        size_cell.value = new_size
                                        size_cell.font = Font(name='宋体', size=18)
                                        print(f"调试ADV：已更新尺寸为 '{new_size}'")
                                else:
                                    print(f"调试ADV：画或其他产品类型，保留原格式 '{original_size}'")
                    except Exception as e:
                        print(f"二次尺寸处理集成出错ADV: {e}")
                elif size_column_index:
                    # 如果没有对应的尺寸，填入空值
                    size_cell = worksheet.cell(row=current_row, column=size_column_index)
                    size_cell.value = ""
                    print(f"第{current_row}行B列填入空值")
                
                # === 处理F列数量（产品数量）===
                if quantity_column_index:
                    quantity_cell = worksheet.cell(row=current_row, column=quantity_column_index)
                    # 始终保持水平、垂直居中对齐
                    quantity_cell.alignment = Alignment(horizontal='center', vertical='center')

                    if equal_qty_distribution:
                        # 当需要平均分配数量时，每一行固定写入 1，字体保持黑色
                        quantity_cell.value = 1
                        # 若原单元格已有字体则拷贝其属性，仅重置颜色为黑色；否则创建默认字体
                        base_font = quantity_cell.font if quantity_cell.font else Font(name='宋体', size=36)
                        quantity_cell.font = Font(name=base_font.name or '宋体', size=base_font.sz or 36, bold=base_font.b, italic=base_font.i, vertAlign=base_font.vertAlign, underline=base_font.underline, strike=base_font.strike, color='000000')
                    else:
                        if i == 0:
                            # 第一行保留原数量
                            if original_data[quantity_column_index]['value'] not in (None, ''):
                                quantity_cell.value = original_data[quantity_column_index]['value']
                                # 如果原有字体为空，设置默认字体
                                if not quantity_cell.font or quantity_cell.font == Font():
                                    quantity_cell.font = Font(name='宋体', size=36)
                                try:
                                    if int(float(quantity_cell.value)) > 1:
                                        quantity_cell.font = Font(name=quantity_cell.font.name or '宋体', size=quantity_cell.font.sz or 36, color='FF0000')
                                except (ValueError, TypeError) as e:
                                    print(f"数量格式转换错误: {quantity_cell.value}, 错误: {e}")
                        else:
                            # 非平均分配模式下，其余行清空数量
                            quantity_cell.value = None

            # 注意：合并操作已移至主处理逻辑中统一执行
            
        except Exception as e:
            print(f"拆分多SKUID行失败：{str(e)}")



    def split_multi_sku_row(self, worksheet, row_idx, sku_list, size_column_index, processing_column_index, material_column_index, quantity_column_index, order_no_column_index, sku_column_index):
        """拆分包含多个SKU的行"""
        try:
            # 调试：检查拆分前原始行D列的值
            original_d_value = worksheet.cell(row=row_idx, column=order_no_column_index).value
            print(f"🔍 拆分前原始行{row_idx} D列值: {original_d_value!r} (类型: {type(original_d_value)})")
            
            # 获取原行的所有数据
            original_data = {}
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                original_data[col_idx] = {
                    'value': cell.value,
                    'font': cell.font,
                    'alignment': cell.alignment,
                    'fill': cell.fill,
                    'border': cell.border
                }
            
            # 获取B列尺寸数据并按SKU数量分割
            size_data = []
            if size_column_index and original_data[size_column_index]['value']:
                size_content = str(original_data[size_column_index]['value']).strip()
                print(f"B列原始尺寸内容: '{size_content}'")
                print(f"B列数据类型: {type(original_data[size_column_index]['value'])}")
                
                # 检查B列是否包含SKUID格式的尺寸数据
                if 'SKUID:' in size_content:
                    print("从B列提取每个SKUID对应的尺寸")
                    # 按SKUID分割B列尺寸数据
                    b_size_blocks = re.split(r'(?=SKUID:)', size_content)
                    b_size_blocks = [block.strip() for block in b_size_blocks if block.strip()]
                    
                    for i, block in enumerate(b_size_blocks):
                        # 提取该SKUID对应的尺寸 - 从Variants行中提取尺寸，支持双引号格式
                        # 先尝试匹配Variants行中的尺寸（不需要斜杠前缀）
                        variants_size_match = re.search(r'Variants:\s*([^/\n\r]*(?:\d+(?:\.\d+)?\s*(?:cm|inch|inches|pulgadas|\'\')\s*[/\\]?[\d.]*\s*(?:in|inch|inches|pulgadas)?\s*[*×x]\s*\d+(?:\.\d+)?\s*(?:cm|inch|inches|pulgadas|\'\')\s*[/\\]?[\d.]*\s*(?:in|inch|inches|pulgadas)?\s*(?:[*×x]\s*\d+)?)[^/\n\r]*)', block, re.IGNORECASE)
                        if variants_size_match:
                            size_raw = variants_size_match.group(1).strip()
                        else:
                            # 如果Variants行没有匹配到，再尝试原来的斜杠格式
                            slash_size_match = re.search(r'/\s*(\d+(?:\.\d+)?\s*(?:inch|inches|cm|\'\')\s*[x*×]\s*\d+(?:\.\d+)?\s*(?:inch|inches|cm|\'\')\s*(?:[x*×]\s*\d+)?)', block, re.IGNORECASE)
                            if slash_size_match:
                                size_raw = slash_size_match.group(1).strip()
                            else:
                                size_raw = None
                                print(f"SKUID {i+1} 未匹配到任何尺寸格式")
                        
                        if size_raw:
                            # 标准化尺寸格式 - 处理双引号
                            size_normalized = re.sub(r'\s+', '', size_raw)
                            size_normalized = re.sub(r'inches?', 'inch', size_normalized, flags=re.IGNORECASE)
                            size_normalized = re.sub(r'\'\'', 'inch', size_normalized)  # 将双引号转换为inch
                            size_normalized = re.sub(r'[x×]', '*', size_normalized)
                            size_data.append(size_normalized)
                            print(f"SKUID {i+1} 尺寸: {size_normalized}")
                        else:
                            size_data.append('')
                            print(f"SKUID {i+1} 未找到尺寸")
                else:
                    # 尝试按不同分隔符分割尺寸
                    size_parts = [s.strip() for s in re.split(r'[\s\n\t]+', size_content) if s.strip()]
                    
                    # 如果尺寸数量与SKU数量匹配，则一一对应
                    if len(size_parts) == len(sku_list):
                        size_data = size_parts
                        print(f"尺寸数量匹配SKU数量，一一对应: {size_data}")
                    else:
                        # 否则所有行使用相同尺寸
                        size_data = [size_content] * len(sku_list)
                        print(f"尺寸数量不匹配，所有行使用相同尺寸: {size_data}")
            else:
                print(f"B列为空或不存在，size_column_index: {size_column_index}")
                if size_column_index:
                    print(f"B列原始值: {original_data[size_column_index]['value']}")
                    print(f"B列原始值类型: {type(original_data[size_column_index]['value'])}")
                size_data = [''] * len(sku_list)
                print(f"B列无尺寸数据，填充空值: {size_data}")
            
            # 需要插入的新行数量（SKU数量-1）
            new_rows_count = len(sku_list) - 1
            
            # 在当前行后插入新行
            for i in range(new_rows_count):
                worksheet.insert_rows(row_idx + 1 + i)
            
            # 填充所有相关行的数据
            for i, sku in enumerate(sku_list):
                current_row = row_idx + i
                
                # 行高设置已移至边框设置之后统一处理
                
                # 复制原始行的格式到新增行（除了第一行）
                if i > 0:
                    for col in range(1, worksheet.max_column + 1):
                        source_cell = worksheet.cell(row=row_idx, column=col)
                        target_cell = worksheet.cell(row=current_row, column=col)
                        
                        # 复制格式
                        if source_cell.font:
                            target_cell.font = copy(source_cell.font)
                        if source_cell.alignment:
                            target_cell.alignment = copy(source_cell.alignment)
                        if source_cell.fill:
                            target_cell.fill = copy(source_cell.fill)
                        if source_cell.border:
                            target_cell.border = copy(source_cell.border)
                        if source_cell.number_format:
                            target_cell.number_format = source_cell.number_format
                        
                        # 清空序号列（A列）的值并标记为新增行
                        if col == 1:
                            target_cell.value = "SKIP_SEQUENCE"  # 标记新增行跳过序号
                        # 跳过D列（订单号），不复制到新增行
                        elif col == order_no_column_index:  # D列通常是4
                            target_cell.value = None
                        # 对于其他列，如果原始单元格有值且不是特殊列，则复制值
                        elif col not in [sku_column_index, size_column_index, quantity_column_index]:
                            if source_cell.value is not None:
                                target_cell.value = source_cell.value
                
                # 填充SKU（G列）- 产品代码SKU（统一行为）
                if sku_column_index:
                    # 从原始G列解析SKU列表
                    original_g_value = original_data.get(sku_column_index, {}).get('value')
                    parsed_sku_list = self.parse_g_cell_to_sku_list(original_g_value, len(sku_list))
                    sku_value_to_write = parsed_sku_list[i] if i < len(parsed_sku_list) else ''

                    sku_cell = worksheet.cell(row=current_row, column=sku_column_index)
                    sku_cell.value = sku_value_to_write
                    sku_cell.font = Font(name='宋体', size=18)
                    sku_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 若M≠N，则仅在第一原始拆分行（i==0）为G列背景填充红色，不改变文字
                    try:
                        m_count = len(parsed_sku_list)
                        n_count = len(sku_list)
                        if i == 0 and m_count != n_count:
                            sku_cell.fill = PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                    except Exception:
                        pass
                    
                    # 填充加工方式和材质信息（与单件订单保持一致的逻辑）
                    if sku_value_to_write:
                        processing, material = self.get_processing_info_by_sku(sku_value_to_write)
                        
                        # 填充加工方式
                        if processing and processing_column_index:
                            processing_cell = worksheet.cell(row=current_row, column=processing_column_index)
                            processing_cell.value = processing
                            # 设置加工方式字体为宋体、24号
                            processing_cell.font = Font(name='宋体', size=24)
                            print(f"填入第{current_row}行加工方式: {processing}")
                        
                        # 填充材质
                        if material and material_column_index:
                            material_cell = worksheet.cell(row=current_row, column=material_column_index)
                            material_cell.value = material
                            # 设置材质字体为宋体、24号、红色
                            material_cell.font = Font(name='宋体', size=24, color='FF0000')
                            print(f"填入第{current_row}行材质: {material}")
                
                # 填充尺寸（B列）
                if size_column_index and i < len(size_data):
                    size_cell = worksheet.cell(row=current_row, column=size_column_index)
                    raw_size = size_data[i]
                    normalized_size = self.normalize_size_to_standard(raw_size)
                    if normalized_size:
                        normalized_size = normalized_size.replace(' x ', '*').replace('×', '*')
                        size_cell.value = normalized_size
                        print(f"填入第{current_row}行B列尺寸(标准化): {normalized_size}")
                    else:
                        size_cell.value = raw_size
                        print(f"填入第{current_row}行B列尺寸(原始): {raw_size}")
                    size_cell.font = Font(name='宋体', size=18)
                    size_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 应用尺寸映射 - 使用与process_and_export_excel一致的完整匹配逻辑
                    print(f"调试信息2 - self.size_mapping: {bool(self.size_mapping)}")
                    print(f"调试信息2 - self.size_column: {self.size_column}")
                    print(f"调试信息2 - size_column_index: {size_column_index}")
                    if self.size_mapping and self.size_column and size_column_index:
                        # 使用统一尺寸映射函数进行处理，替代旧版单维逻辑
                        size_cell.value = self._map_size_with_rules(str(size_cell.value))
                        size_str = str(size_cell.value)
                        mapped = True  # 标记为已映射，跳过旧版尺寸逻辑
                        # 旧版匹配细节已被统一方法取代
                        
                        # 第二优先级：标准化后匹配
                        if not mapped:
                            filename_standard_size = self.normalize_size_to_standard(size_str)
                            if filename_standard_size:
                                # 尝试完整标准化匹配
                                for original_size, new_size in self.size_mapping.items():
                                    standard_size = self.normalize_size_to_standard(original_size)
                                    if standard_size and filename_standard_size == standard_size:
                                        size_cell.value = new_size
                                        mapped = True
                                        print(f"✅ 标准化完整匹配成功2: '{size_str}' (标准化为 '{filename_standard_size}') -> '{new_size}'")
                                        break
                                
                                # 尝试基础尺寸匹配（忽略数量后缀）
                                if not mapped:
                                    base_std = filename_standard_size.split('*')[0] if '*' in filename_standard_size else filename_standard_size
                                    for original_size, new_size in self.size_mapping.items():
                                        standard_size = self.normalize_size_to_standard(original_size)
                                        if standard_size:
                                            std_base = standard_size.split('*')[0] if '*' in standard_size else standard_size
                                            if base_std == std_base:
                                                size_cell.value = new_size
                                                mapped = True
                                                print(f"✅ 基础尺寸匹配成功2: '{size_str}' (基础尺寸 '{base_std}') -> '{new_size}'")
                                                break
                        
                        # 第三优先级：相似度匹配（数字必须精确，只允许符号差异）
                        if not mapped:
                            for original_size, new_size in self.size_mapping.items():
                                # 使用与单件订单相同的数字精确匹配逻辑
                                if self._check_numerical_exactness(size_str, original_size):
                                    size_cell.value = new_size
                                    mapped = True
                                    print(f"✅ 数字精确匹配成功2: '{size_str}' -> '{new_size}' (数字相同，符号不同)")
                                    break
                        
                        # 最后：如果没有找到映射，使用标准化结果或备用逻辑
                        if not mapped:
                            filename_standard_size = self.normalize_size_to_standard(size_str)
                            if filename_standard_size:
                                # 优先使用新的标准化结果，统一维度分隔符为 *
                                size_cell.value = filename_standard_size.replace(' x ', '*').replace('×', '*')
                                print(f"⚠️ 无映射匹配2，使用标准化结果: '{size_str}' -> '{size_cell.value}'")
                            else:
                                # 尝试从原始内容中提取variants并标准化
                                variants_content = self.extract_variants_content(size_str)
                                if variants_content:
                                    vc_std = self.normalize_size_to_standard(variants_content)
                                    if vc_std:
                                        size_cell.value = vc_std.replace(' x ', '*').replace('×', '*')
                                        print(f"⚠️ 无映射匹配2，使用variants标准化结果: '{size_str}' -> '{size_cell.value}'")
                                    else:
                                        size_cell.value = variants_content
                                        print(f"⚠️ 无映射匹配2，使用variants原始结果: '{size_str}' -> '{size_cell.value}'")
                                else:
                                    # 最后才使用旧的process_variants_content作为备选
                                    auto_converted = self.process_variants_content(size_str)
                                    if auto_converted and auto_converted != size_str:
                                        size_cell.value = auto_converted.replace(' x ', '*').replace('×', '*')
                                        print(f"⚠️ 无映射匹配2，使用旧逻辑结果: '{size_str}' -> '{size_cell.value}'")
                                    else:
                                        print(f"❌ 无法处理尺寸2: '{size_str}'，保持原样")
                    elif self.size_column and size_column_index:
                        print("自动转换功能已禁用，保持原始尺寸")
                    else:
                        print("未进入尺寸映射逻辑2")
                    
                    # 二次尺寸处理：仅在用户手动选择了预设时触发（与单件订单保持一致）
                    try:
                        secondary_flag = getattr(self, '_secondary_processing_triggered_manually', False)
                        print(f"调试：二次处理标志 = {secondary_flag}, 尺寸列索引 = {size_column_index}")
                        if size_column_index and secondary_flag and sku_value_to_write:
                            processing, material = self.get_processing_info_by_sku(sku_value_to_write)
                            original_size = str(size_cell.value) if size_cell.value else ""
                            print(f"调试：行{current_row} 原始尺寸 = '{original_size}', 加工方式 = '{processing}'")
                            if size_cell.value:
                                # 判断产品类型，只对窗帘和床上三件套进行二次处理
                                current_preset = getattr(self, 'last_selected_preset', '')
                                should_process = self._should_apply_secondary_processing(processing, current_preset)
                                print(f"调试：产品类型判断 - 是否需要二次处理: {should_process}")
                                
                                if should_process:
                                    new_size = self.apply_secondary_size_processing(str(size_cell.value), current_preset)
                                    print(f"调试：二次处理结果 = '{new_size}'")
                                    if new_size:
                                        size_cell.value = new_size
                                        size_cell.font = Font(name='宋体', size=18)
                                        print(f"调试：已更新尺寸为 '{new_size}'")
                                else:
                                    print(f"调试：画或其他产品类型，保留原格式 '{original_size}'")
                    except Exception as e:
                        print(f"二次尺寸处理集成出错: {e}")
                
                # 只在第一行保留F列数量，其他行清空
                if quantity_column_index:
                    quantity_cell = worksheet.cell(row=current_row, column=quantity_column_index)
                    if i == 0:
                        # 第一行保留原数量
                        if original_data[quantity_column_index]['value']:
                            quantity_cell.value = original_data[quantity_column_index]['value']
                            quantity_cell.font = Font(name='宋体', size=36)
                            quantity_cell.alignment = Alignment(horizontal='center', vertical='center')
                            try:
                                if int(float(quantity_cell.value)) > 1:
                                    quantity_cell.font = Font(name='宋体', size=36, color='FF0000')
                            except (ValueError, TypeError) as e:
                                print(f"数量格式转换错误: {quantity_cell.value}, 错误: {e}")
                    else:
                        # 其他行清空数量
                        quantity_cell.value = None
            
            # --- 添加I列GroupID标记和合并逻辑 ---
            # 使用I列（第9列）进行GroupID标记
            group_col_idx = 9  # I列
            
            # 初始化GroupID计数器
            if not hasattr(self, '_next_group_id'):
                self._next_group_id = 1
            
            # 生成GroupID并递增
            group_id = self._next_group_id
            self._next_group_id += 1
            
            # 在原行的I列写入GroupID
            worksheet.cell(row=row_idx, column=group_col_idx).value = group_id
            print(f"原行{row_idx}在I列标记GroupID: {group_id}")
            
            # 为新插入的行在I列标记相同的GroupID
            for i in range(len(sku_list) - 1):
                new_row = row_idx + i + 1
                worksheet.cell(row=new_row, column=group_col_idx).value = group_id
                print(f"新行{new_row}在I列标记GroupID: {group_id}")
            
            # 注意：合并操作已移至主处理逻辑中统一执行
            
        except Exception as e:
            print(f"拆分多SKU行失败：{str(e)}")

    def merge_cells_by_i_column(self, worksheet, order_no_column_index):
        """基于I列相同数字合并A列和D列单元格
        
        Args:
            worksheet: 工作表对象
            order_no_column_index: D列(订单号)索引
        """
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
            
            print("开始基于I列相同数字合并A、D列单元格")
            
            # I列索引
            group_col_idx = 9
            
            # 获取工作表的最大行数
            max_row = worksheet.max_row
            
            # 收集所有I列的值和对应的行号
            group_data = {}  # {group_id: [row1, row2, ...]}
            
            for row in range(2, max_row + 1):  # 从第2行开始，跳过标题行
                cell_value = worksheet.cell(row=row, column=group_col_idx).value
                if cell_value is not None and str(cell_value).strip() != '':
                    group_id = str(cell_value).strip()
                    if group_id not in group_data:
                        group_data[group_id] = []
                    group_data[group_id].append(row)
            
            print(f"找到的分组数据: {group_data}")
            
            # 对每个分组进行合并
            for group_id, rows in group_data.items():
                if len(rows) <= 1:
                    print(f"分组 {group_id} 只有 {len(rows)} 行，跳过合并")
                    continue
                
                # 按行号排序
                rows.sort()
                merge_start = rows[0]
                merge_end = rows[-1]
                
                print(f"分组 {group_id} 包含行: {rows}, 合并范围: {merge_start}-{merge_end}")
                
                # 调试：打印合并前每行D列的内容
                print(f"🔍 合并前D列内容检查:")
                for row_num in rows:
                    d_value = worksheet.cell(row=row_num, column=order_no_column_index).value
                    print(f"  行{row_num} D列值: {d_value!r} (类型: {type(d_value)})")
                
                # 构建合并范围
                col_a_range = f"A{merge_start}:A{merge_end}"
                col_d_range = f"{get_column_letter(order_no_column_index)}{merge_start}:{get_column_letter(order_no_column_index)}{merge_end}"
                
                # 解除可能的重叠合并
                for rng in list(worksheet.merged_cells.ranges):
                    # 检查A列重叠
                    if rng.coord.startswith("A") and not (rng.max_row < merge_start or rng.min_row > merge_end):
                        worksheet.unmerge_cells(str(rng))
                        print(f"解除A列重叠合并: {str(rng)}")
                    # 检查D列重叠
                    if rng.coord.startswith(get_column_letter(order_no_column_index)) and not (rng.max_row < merge_start or rng.min_row > merge_end):
                        worksheet.unmerge_cells(str(rng))
                        print(f"解除D列重叠合并: {str(rng)}")
                
                # 执行合并（Excel会自动保留第一个单元格的内容）
                worksheet.merge_cells(col_a_range)
                worksheet.merge_cells(col_d_range)
                
                # 设置对齐
                worksheet.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                worksheet.cell(row=merge_start, column=order_no_column_index).alignment = Alignment(horizontal='center', vertical='center')
                
                print(f"完成分组 {group_id} 的A列和D列合并: {col_a_range} & {col_d_range}")
            
        except Exception as e:
            print(f"基于I列合并失败: {e}")

    def get_processing_info_by_sku(self, sku):
        """根据SKU获取加工方式和材质信息（大小写不敏感，优先级：手动选择的临时映射 > 预设前缀 > 全局临时设置 > 空）"""
        try:
            print(f"调试：查询SKU '{sku}' 的加工信息")
            print(f"调试：temp_sku_mapping存在: {hasattr(self, 'temp_sku_mapping')}")
            if hasattr(self, 'temp_sku_mapping'):
                print(f"调试：temp_sku_mapping内容: {getattr(self, 'temp_sku_mapping', {})}")
                print(f"调试：temp_sku_mapping对象ID = {id(getattr(self, 'temp_sku_mapping', {}))}")
            
            # 最高优先级：特定SKU的临时映射（手动选择的预设）
            if hasattr(self, 'temp_sku_mapping') and getattr(self, 'temp_sku_mapping', {}):
                temp_mapping = self.temp_sku_mapping
                
                # 首先尝试直接匹配
                if sku in temp_mapping:
                    temp_info = temp_mapping[sku]
                    print(f"调试：SKU {sku} 直接匹配临时映射 - 加工方式: {temp_info.get('processing', '')}, 材质: {temp_info.get('material', '')}")
                    return temp_info.get('processing', ''), temp_info.get('material', '')
                
                # 如果直接匹配失败，尝试在组合键中查找
                for combined_key, temp_info in temp_mapping.items():
                    # 检查是否为包含换行符或空格的组合键
                    if '\n' in combined_key or ' ' in combined_key:
                        # 分解组合键，支持换行符和空格分隔
                        individual_skus = []
                        for line in combined_key.split('\n'):
                            line = line.strip()
                            if line:
                                if ' ' in line:
                                    individual_skus.extend([s.strip() for s in line.split() if s.strip()])
                                else:
                                    individual_skus.append(line)
                        
                        # 检查当前SKU是否在分解后的列表中
                        if sku in individual_skus:
                            print(f"调试：SKU {sku} 在组合键 '{combined_key}' 中找到匹配 - 加工方式: {temp_info.get('processing', '')}, 材质: {temp_info.get('material', '')}")
                            return temp_info.get('processing', ''), temp_info.get('material', '')
                
                print(f"调试：SKU {sku} 在temp_sku_mapping中未找到匹配")
            
            # 其次：预设配置（processing_config.json）- 前缀匹配，大小写不敏感
            config_file = os.path.join(get_app_directory(), 'processing_config.json')
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                sku_prefix = self.extract_sku_prefix(sku)
                for prefix, info in config.items():
                    if sku_prefix.upper() == str(prefix).upper():
                        print(f"调试：SKU {sku} 使用预设配置 - 加工方式: {info.get('processing', '')}, 材质: {info.get('material', '')}")
                        return info.get('processing', ''), info.get('material', '')
            
            # 最后：全局临时设置
            if hasattr(self, 'temp_processing_info') and getattr(self, 'temp_processing_info', None):
                print(f"调试：SKU {sku} 使用全局临时设置 - 加工方式: {self.temp_processing_info.get('processing', '')}, 材质: {self.temp_processing_info.get('material', '')}")
                return self.temp_processing_info.get('processing', ''), self.temp_processing_info.get('material', '')
            
            print(f"调试：SKU {sku} 未找到任何配置，返回空值")
            return '', ''
        except Exception as e:
            print(f"获取加工方式信息失败：{str(e)}")
            return '', ''

    def apply_secondary_size_processing(self, size_text, processing_text):
        """二次尺寸处理：仅在手动预设触发时调用，并且未命中预设映射的情况下。
        规则：
        - 窗帘（含“窗帘”或CL）：要求格式为 Wcm*Hcm*N（必须有数量N），输出：宽{W*N}高{H}
        - 床上三件套（含“三件套”或“床上三件套”或SJT）：要求格式为 Wcm*Hcm，W<260 -> 枕套50x75x2，否则枕套51x91x2；输出分两行：被套{W}x{H}\n枕套{...}
        - 画（含“画”或cft）：不处理，保持原样
        限制：size_text若已包含“被套/枕套/宽/高”则视为已按预设输出，跳过。
        """
        try:
            import re
            if not size_text:
                return None
            s = str(size_text).strip()
            if not s:
                return None
            # 获取当前选择的预设信息来判断产品类型
            current_preset = getattr(self, 'last_selected_preset', '')
            
            # 若已是完整的预设输出格式（包含被套、枕套、床笠），跳过处理
            if any(k in s for k in ['被套', '枕套', '床笠']):
                print(f"调试：'{s}' 已是完整预设格式，跳过二次处理")
                return None
            
            # 对于三件套和床笠预设，即使包含"宽高"也要继续处理成完整格式
            if '三件套' in current_preset:
                print(f"调试：三件套预设，即使包含宽高也继续处理：'{s}'")
            elif '床笠' in current_preset:
                print(f"调试：床笠预设，即使包含宽高也继续处理：'{s}'")
            # 对于其他预设，如果已包含"宽高"则跳过
            elif any(k in s for k in ['宽', '高']):
                print(f"调试：'{s}' 已是预设格式，跳过二次处理")
                return None
            # 归一化分隔符
            s_norm = re.sub(r'[×xX]', '*', s)
            
            print(f"调试：当前预设='{current_preset}', 尺寸='{s_norm}'")
            
            # 完全基于用户选择的预设来决定处理方式
            if '窗帘' in current_preset:
                print(f"调试：用户选择了窗帘预设，按窗帘格式处理")
                # 先尝试标准格式 Wcm*Hcm*N（数量可选）
                m = re.match(r'^\s*(\d+)\s*cm\s*\*\s*(\d+)\s*cm\s*(?:\*\s*(\d+))?\s*$', s_norm, re.IGNORECASE)
                if not m:
                    m = re.match(r'^\s*(\d+)\s*\*\s*(\d+)\s*(?:\*\s*(\d+))?\s*$', s_norm)
                if m:
                    w = int(m.group(1)); h = int(m.group(2)); 
                    qty = int(m.group(3)) if m.group(3) else 2  # 数量默认为2（窗帘宽度翻倍）
                    print(f"调试：解析到 宽={w}, 高={h}, 数量={qty}")
                    if qty <= 0:
                        return None
                    w_total = w * qty
                    result = f"宽{w_total}高{h}"
                    print(f"调试：窗帘处理结果='{result}'")
                    return result
                
                print("调试：窗帘格式不匹配")
                return None
                
            # 使用外部的产品类型判断方法
            elif '三件套' in current_preset:
                print(f"调试：用户选择了三件套预设，按三件套格式处理")
                m = re.match(r'^\s*(\d+)\s*cm\s*\*\s*(\d+)\s*cm\s*$', s_norm, re.IGNORECASE)
                if not m:
                    m = re.match(r'^\s*(\d+)\s*\*\s*(\d+)\s*$', s_norm)
                if m:
                    w = int(m.group(1)); h = int(m.group(2))
                    pillow = "50x75x2" if w < 260 else "51x91x2"
                    result = f"被套{w}x{h}\n枕套{pillow}"
                    print(f"调试：三件套处理结果='{result}'")
                    return result
                return None
            
            # 床笠处理逻辑
            elif '床笠' in current_preset:
                print(f"调试：用户选择了床笠预设，按床笠格式处理")
                m = re.match(r'^\s*(\d+)\s*cm\s*\*\s*(\d+)\s*cm\s*$', s_norm, re.IGNORECASE)
                if not m:
                    m = re.match(r'^\s*(\d+)\s*\*\s*(\d+)\s*$', s_norm)
                if m:
                    w = int(m.group(1)); h = int(m.group(2))
                    # 床笠算法：宽度和高度都+50
                    w_final = w + 50
                    h_final = h + 50
                    # 床笠的枕套规格固定为50x75x2
                    result = f"床笠{w_final}x{h_final}\n枕套50x75x2"
                    print(f"调试：床笠处理结果='{result}' (原始尺寸: {w}x{h}, +50后: {w_final}x{h_final})")
                    return result
                return None
            
            # 用户选择的是其他预设，不进行处理
            print(f"调试：用户选择了其他预设，不进行二次处理")
            return None
        except Exception as e:
            print(f"二次尺寸处理出错: {e}")
            return None
    
    def add_mapping_entry(self):
        """添加新的映射条目"""
        entry_frame = ttk.Frame(self.mapping_entries_frame)
        entry_frame.pack(fill=tk.X, pady=2)
        
        # 原始尺寸输入框
        original_entry = ttk.Entry(entry_frame, width=20)
        original_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        # 对应尺寸输入框
        target_entry = ttk.Entry(entry_frame, width=20)
        target_entry.pack(side=tk.LEFT, padx=(0, 20))
        
        # 删除按钮
        delete_btn = ttk.Button(entry_frame, text="🗑", width=3,
                               command=lambda: self.delete_mapping_entry(entry_frame))
        delete_btn.pack(side=tk.LEFT)
        
        # 保存条目信息
        entry_info = {
            'frame': entry_frame,
            'original': original_entry,
            'target': target_entry
        }
        self.current_mapping_entries.append(entry_info)
        
        return entry_info
    
    def delete_mapping_entry(self, entry_frame):
        """删除映射条目"""
        # 从列表中移除
        self.current_mapping_entries = [
            entry for entry in self.current_mapping_entries 
            if entry['frame'] != entry_frame
        ]
        # 销毁界面元素
        entry_frame.destroy()
    
    def load_existing_mappings(self):
        """加载现有的映射配置"""
        # 清空现有条目
        for entry in self.current_mapping_entries:
            entry['frame'].destroy()
        self.current_mapping_entries.clear()
        
        # 加载现有配置
        for original, target in self.size_mapping.items():
            entry_info = self.add_mapping_entry()
            entry_info['original'].insert(0, original)
            entry_info['target'].insert(0, target)
        
        # 如果没有配置，添加几个空条目
        if not self.size_mapping:
            for _ in range(3):
                self.add_mapping_entry()
    
    def save_settings_mapping(self, silent=False):
        """保存设置页面的映射配置"""
        new_mapping = {}
        
        for entry in self.current_mapping_entries:
            original = entry['original'].get().strip()
            target = entry['target'].get().strip()
            
            if original and target:
                new_mapping[original] = target
        
        self.size_mapping = new_mapping
        
        if not silent:
            # 直接保存配置
            self.save_config()
            messagebox.showinfo("保存成功", f"已保存 {len(new_mapping)} 条尺寸映射配置")
    
    def process_settings_excel_file(self, excel_path, parent_window):
        """处理设置页面的Excel文件"""
        try:
            # 读取Excel文件
            pd = _import_pandas()
            df = pd.read_excel(excel_path)
            
            # 查找包含尺寸数据的列
            size_columns = []
            for col in df.columns:
                col_data = df[col].astype(str).str.lower()
                if col_data.str.contains('inch|variants', na=False, case=False).any():
                    size_columns.append(col)
            
            if len(size_columns) == 0:
                messagebox.showwarning("警告", "未找到包含尺寸数据的列！\n请确保Excel文件包含尺寸信息。", parent=parent_window)
                return
            
            # 提取尺寸映射
            new_mappings = {}
            
            # 如果只有一列包含尺寸数据，从多行文本中提取原始尺寸和转换后的尺寸
            if len(size_columns) == 1:
                size_col = size_columns[0]
                
                for _, row in df.iterrows():
                    size_text = str(row[size_col]).strip()
                    
                    if size_text and size_text != 'nan':
                        # 从多行文本中提取原始尺寸
                        original_size = self.extract_original_size_from_multiline(size_text)
                        # 转换为标准格式
                        converted_size = self.extract_converted_size_from_multiline(size_text)
                        
                        if original_size and converted_size:
                            new_mappings[original_size] = converted_size
            
            # 如果有两列或更多列，按原来的逻辑处理
            elif len(size_columns) >= 2:
                original_col = size_columns[0]
                mapped_col = size_columns[1]
                
                for _, row in df.iterrows():
                    original = str(row[original_col]).strip()
                    mapped = str(row[mapped_col]).strip()
                    
                    if original and mapped and original != 'nan' and mapped != 'nan':
                        new_mappings[original] = mapped
            
            if new_mappings:
                # 确认导入
                result = messagebox.askyesno("确认导入", 
                                           f"从Excel文件中识别到 {len(new_mappings)} 条尺寸映射配置。\n\n是否导入这些配置？", 
                                           parent=parent_window)
                if result:
                    self.size_mapping.update(new_mappings)
                    messagebox.showinfo("成功", f"成功导入 {len(new_mappings)} 条尺寸映射配置！", parent=parent_window)
            else:
                messagebox.showwarning("警告", "未从Excel文件中找到有效的尺寸映射数据！", parent=parent_window)
                
        except Exception as e:
            messagebox.showerror("错误", f"处理Excel文件时发生错误：\n{str(e)}", parent=parent_window)
    
    def extract_original_size_from_multiline(self, text):
        """从多行文本中提取原始尺寸信息"""
        if not text or pd.isna(text):
            return None
        
        text = str(text)
        
        # 查找Variants行中的尺寸信息
        variants_patterns = [
            r'Variants:\s*([^\n\r]+)',  # 匹配 Variants: 后面的内容
            r'Variant:\s*([^\n\r]+)',   # 匹配 Variant: 后面的内容
        ]
        
        for pattern in variants_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                variants_text = match.group(1).strip()
                
                # 移除颜色信息 (如 "/ Multicolor", "/ Verde", "/ Red")
                variants_text = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff]+\s*$', '', variants_text)
                
                # 提取原始尺寸格式
                size_patterns = [
                    # 匹配 "21 inches x 46 inches x 2" 格式
                    r'(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+)',
                    # 匹配 "21inch*46inch*2" 格式
                    r'(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+)',
                    # 匹配 "21''*46''*2" 格式
                    r'(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+)',
                    # 匹配 "21 inches x 46 inches" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+(?:\.\d+)?)\s+inches?',
                    # 匹配 "21inch*46inch" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+(?:\.\d+)?)\s*inch(?:es)?',
                    # 匹配 "21''*46''" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:\'\'|")',
                ]
                
                for size_pattern in size_patterns:
                    size_match = re.search(size_pattern, variants_text, re.IGNORECASE)
                    if size_match:
                        # 返回原始格式的尺寸
                        return size_match.group(0)
        
        return None
    
    def extract_converted_size_from_multiline(self, text):
        """从多行文本中提取并转换尺寸信息为标准格式"""
        if not text or pd.isna(text):
            return None
        
        text = str(text)
        
        # 查找Variants行中的尺寸信息
        variants_patterns = [
            r'Variants:\s*([^\n\r]+)',  # 匹配 Variants: 后面的内容
            r'Variant:\s*([^\n\r]+)',   # 匹配 Variant: 后面的内容
        ]
        
        for pattern in variants_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                variants_text = match.group(1).strip()
                
                # 移除颜色信息 (如 "/ Multicolor", "/ Verde", "/ Red")
                variants_text = re.sub(r'\s*/\s*[A-Za-z\u4e00-\u9fff]+\s*$', '', variants_text)
                
                # 尺寸模式匹配
                size_patterns = [
                    # 匹配 "21 inches x 46 inches x 2" 格式
                    r'(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+)',
                    # 匹配 "21inch*46inch*2" 格式
                    r'(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+)',
                    # 匹配 "21''*46''*2" 格式
                    r'(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+)',
                    # 匹配 "21 inches x 46 inches" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s+inches?\s*[×x]\s*(\d+(?:\.\d+)?)\s+inches?',
                    # 匹配 "21inch*46inch" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s*inch(?:es)?\s*[*×x]\s*(\d+(?:\.\d+)?)\s*inch(?:es)?',
                    # 匹配 "21''*46''" 格式（无倍数）
                    r'(\d+(?:\.\d+)?)\s*(?:\'\'|")\s*[*×x]\s*(\d+(?:\.\d+)?)\s*(?:\'\'|")',
                ]
                
                for size_pattern in size_patterns:
                    size_match = re.search(size_pattern, variants_text, re.IGNORECASE)
                    if size_match:
                        width = float(size_match.group(1))
                        height = float(size_match.group(2))
                        
                        # 检查是否有倍数
                        multiplier_value = 1
                        if len(size_match.groups()) >= 3 and size_match.group(3):
                            multiplier_value = int(size_match.group(3))
                        
                        # 转换为厘米
                        width_cm = round(width * 2.54)
                        height_cm = round(height * 2.54)
                        
                        # 检查是否为窗帘产品（通过检查当前处理的文件夹名称或产品类型）
                        is_curtain = self._is_curtain_product()
                        
                        # 所有产品都返回CM格式，让二次处理来决定最终格式
                        multiplier = f"*{multiplier_value}" if multiplier_value > 1 else ""
                        return f"{width_cm}CM*{height_cm}CM{multiplier}"
        
        return None
    
    def _should_apply_secondary_processing(self, processing_text, preset_text=""):
        """
        判断是否需要进行二次处理
        完全基于用户选择的预设类型来决定处理方式，不依赖Excel中的加工方式内容
        
        Args:
            processing_text: 加工方式文本（仅用于兼容性，实际不使用）
            preset_text: 预设文本（主要判断依据）
            
        Returns:
            bool: True表示需要二次处理，False表示不需要
        """
        try:
            preset = str(preset_text).strip() if preset_text else ""
            
            # 完全基于用户选择的预设来决定处理方式
            if preset:
                # 如果用户选择了三件套预设，则按三件套处理
                if '三件套' in preset:
                    return True
                # 如果用户选择了窗帘预设，则按窗帘处理
                if '窗帘' in preset:
                    return True
                # 如果用户选择了床笠预设，则按床笠处理
                if '床笠' in preset:
                    return True
            
            # 如果用户选择的是画或其他预设，则不进行二次处理
            return False
        except Exception as e:
            print(f"判断是否需要二次处理时出错: {e}")
            return False

    def _is_curtain_product(self):
        """判断当前处理的是否为窗帘产品"""
        try:
            # 优先级1：检查最后选择的预设（最高优先级）
            if hasattr(self, 'config') and self.config:
                last_preset = self.config.get('last_selected_preset', '').lower()
                # 如果预设明确包含"画"，则不是窗帘产品
                if '画' in last_preset or 'painting' in last_preset:
                    return False
                # 如果预设明确包含"窗帘"，则是窗帘产品
                if '窗帘' in last_preset or 'curtain' in last_preset:
                    return True
            
            # 优先级2：检查最后的手动处理配置
            if hasattr(self, 'config') and self.config:
                last_manual = self.config.get('last_manual_processing', {})
                processing = last_manual.get('processing', '').lower()
                # 如果手动设置包含画相关关键词，则不是窗帘产品
                if '画' in processing or 'painting' in processing:
                    return False
                # 如果手动设置包含穿杆，则是窗帘产品
                if '穿杆' in processing:  # 穿杆是窗帘特有的处理方式
                    return True
            
            # 优先级3：检查当前文件夹名称（最低优先级）
            if hasattr(self, 'current_folder') and self.current_folder:
                folder_name = self.current_folder.lower()
                # 如果文件夹名称包含画相关关键词，则不是窗帘产品
                if '画' in folder_name or 'painting' in folder_name:
                    return False
                # 如果文件夹名称包含窗帘相关关键词，则是窗帘产品
                if '窗帘' in folder_name or 'curtain' in folder_name or 'cl' in folder_name:
                    return True
            
            return False
        except Exception as e:
            print(f"判断窗帘产品时出错: {e}")
            return False
    
    def export_size_mapping_excel(self):
        """导出尺寸映射配置为Excel文件"""
        try:
            import pandas as pd
            from tkinter import filedialog
            
            if not self.size_mapping:
                messagebox.showwarning("无数据", "当前没有尺寸映射配置可导出")
                return
            
            # 选择保存位置
            file_path = filedialog.asksaveasfilename(
                title="导出尺寸映射配置",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )
            
            if file_path:
                # 创建DataFrame
                pd = _import_pandas()
                data = {
                    '原始尺寸': list(self.size_mapping.keys()),
                    '对应尺寸': list(self.size_mapping.values())
                }
                df = pd.DataFrame(data)
                
                # 保存到Excel
                df.to_excel(file_path, index=False, sheet_name='尺寸映射配置')
                messagebox.showinfo("导出成功", f"配置已导出到：{file_path}")
                
        except Exception as e:
            messagebox.showerror("导出失败", f"导出配置时出错：{str(e)}")
    
    def import_size_mapping_excel(self):
        """从Excel文件导入尺寸映射配置"""
        try:
            # 移除原有的导入语句，使用延迟导入
            from tkinter import filedialog
            
            # 临时释放设置窗口的grab_set，确保文件对话框能正常显示
            settings_window = None
            if hasattr(self, 'settings_window') and self.settings_window:
                settings_window = self.settings_window
                settings_window.grab_release()
            
            # 选择文件
            file_path = filedialog.askopenfilename(
                title="导入尺寸映射配置",
                filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
            )
            
            # 恢复设置窗口的grab_set
            if settings_window:
                settings_window.grab_set()
            
            if file_path:
                # 读取Excel文件
                pd = _import_pandas()
                df = pd.read_excel(file_path)
                
                # 查找包含映射数据的列
                original_col = None
                target_col = None
                
                for col in df.columns:
                    col_str = str(col).lower()
                    if '原始' in col_str or 'original' in col_str:
                        original_col = col
                    elif '对应' in col_str or '目标' in col_str or 'target' in col_str:
                        target_col = col
                
                # 如果没有找到标准列名，使用前两列
                if original_col is None or target_col is None:
                    if len(df.columns) >= 2:
                        original_col = df.columns[0]
                        target_col = df.columns[1]
                    else:
                        messagebox.showerror("格式错误", "Excel文件至少需要包含两列数据")
                        return
                
                # 导入数据
                new_mappings = {}
                for _, row in df.iterrows():
                    original = str(row[original_col]).strip()
                    target = str(row[target_col]).strip()
                    if original and target and original != 'nan' and target != 'nan':
                        new_mappings[original] = target
                
                if new_mappings:
                    # 询问是否覆盖现有配置
                    if self.size_mapping:
                        result = messagebox.askyesnocancel(
                            "导入配置",
                            f"发现 {len(new_mappings)} 条新配置。\n\n"
                            "是：覆盖现有配置\n"
                            "否：合并到现有配置\n"
                            "取消：取消导入"
                        )
                        if result is None:  # 取消
                            return
                        elif result:  # 是，覆盖
                            self.size_mapping = new_mappings
                        else:  # 否，合并
                            self.size_mapping.update(new_mappings)
                    else:
                        self.size_mapping = new_mappings
                    
                    # 保存配置并重新加载界面
                    self.save_config()
                    if hasattr(self, 'current_mapping_entries'):
                        self.load_existing_mappings()
                    messagebox.showinfo("导入成功", f"成功导入 {len(new_mappings)} 条尺寸映射配置")
                else:
                    messagebox.showwarning("无有效数据", "Excel文件中没有找到有效的尺寸映射数据")
                    
        except Exception as e:
            messagebox.showerror("导入失败", f"导入配置时出错：{str(e)}")
    
    def clear_size_mapping(self):
        """清空尺寸映射配置"""
        if messagebox.askyesno("确认清空", "确定要清空所有尺寸映射配置吗？"):
            # 销毁所有条目框架
            for entry_info in self.current_mapping_entries:
                entry_info['frame'].destroy()
            # 清空条目列表
            self.current_mapping_entries.clear()
            # 清空映射配置
            self.size_mapping.clear()
            # 保存配置
            self.save_config()
            messagebox.showinfo("成功", "尺寸映射配置已清空！")
    
    def setup_autocomplete(self, entry_widget, history_list):
        """为输入框设置自动完成功能"""
        # 绑定事件
        entry_widget.bind('<KeyRelease>', on_key_release)
        entry_widget.bind('<FocusOut>', on_focus_out)
    
    def show_autocomplete_listbox(self, entry_widget, matches):
        """显示自动完成下拉列表"""
        # 如果已经存在列表框，先销毁
        if hasattr(entry_widget, 'autocomplete_listbox'):
            entry_widget.autocomplete_listbox.destroy()
        
        # 创建列表框
        listbox = tk.Listbox(entry_widget.master, height=min(5, len(matches)),
                            font=('Microsoft YaHei UI', 9),
                            relief='solid', bd=1)
        
        # 添加匹配项
        for match in matches:
            listbox.insert(tk.END, match)
        
        # 计算位置
        x = entry_widget.winfo_x()
        y = entry_widget.winfo_y() + entry_widget.winfo_height()
        
        # 放置列表框
        listbox.place(x=x, y=y, width=entry_widget.winfo_width())
        
        # 绑定选择事件
        listbox.bind('<Double-Button-1>', on_select)
        listbox.bind('<Return>', on_select)
        
        # 保存引用
        entry_widget.autocomplete_listbox = listbox
    
    def hide_autocomplete_listbox(self, entry_widget):
        """隐藏自动完成下拉列表"""
        if hasattr(entry_widget, 'autocomplete_listbox'):
            entry_widget.autocomplete_listbox.destroy()
            delattr(entry_widget, 'autocomplete_listbox')
                
    def set_source_folder(self, folder_path):
        """设置源文件夹"""
        self.source_folder = folder_path
        self.folder_var.set(os.path.basename(folder_path))
        self.save_config()
        
    def refresh_folder(self):
        """刷新文件夹选择"""
        folder_path = filedialog.askdirectory(title="选择图片文件夹")
        if folder_path:
            self.set_source_folder(folder_path)
            
    def on_text_focus_in(self, event):
        """搜索框获得焦点时清除占位符"""
        current_text = self.search_entry.get('1.0', tk.END).strip()
        if current_text == "输入图片编号，每行一个\n例如：\nCL6453-1\nCL6812-4":
            self.search_entry.delete('1.0', tk.END)
            self.search_entry.config(fg='#212529')
    
    def on_text_focus_out(self, event):
        """搜索框失去焦点时恢复占位符"""
        current_text = self.search_entry.get('1.0', tk.END).strip()
        if not current_text:
            placeholder_text = "输入图片编号，每行一个\n例如：\nCL6453-1\nCL6812-4"
            self.search_entry.insert('1.0', placeholder_text)
            self.search_entry.config(fg='#6C757D')
    
    def on_text_change(self, event):
        """文本变化时的处理"""
        pass
    
    def on_window_resize(self, event):
        """窗口大小变化时的处理"""
        # 只处理主窗口的大小变化事件
        if event.widget == self.root:
            self.update_grid_layout()
    
    def calculate_grid_columns(self):
        """根据窗口宽度计算最佳网格列数"""
        try:
            # 获取内容区域的实际宽度
            canvas_width = self.canvas.winfo_width()
            if canvas_width <= 1:
                # 如果还没有渲染完成，使用窗口宽度估算
                window_width = self.root.winfo_width()
                if window_width > 1:
                    # 估算内容区域宽度（减去侧边栏和边距）
                    canvas_width = max(600, window_width - 400)
                else:
                    return 6  # 使用默认值
            
            # 计算可以容纳的列数（考虑边距和滚动条）
            available_width = canvas_width - 40  # 减去边距
            columns = max(1, available_width // self.min_item_width)
            
            # 限制最大和最小列数
            columns = min(max(columns, 2), 10)  # 最少2列，最多10列
            
            return int(columns)
        except Exception as e:
            print(f"计算网格列数时出错: {e}")
            return 6  # 出错时返回默认值
    
    def update_grid_layout(self):
        """更新网格布局"""
        # 延迟执行，确保窗口大小变化完成
        self.root.after(100, self._update_grid_layout_delayed)
    
    def _update_grid_layout_delayed(self):
        """延迟执行的网格布局更新"""
        try:
            new_columns = self.calculate_grid_columns()
            
            # 如果列数发生变化，重新排列图片
            if new_columns != self.grid_columns:
                self.grid_columns = new_columns
                
                # 更新滚动区域的列权重配置
                self.configure_grid_columns()
                
                # 重新排列所有可见的图片
                self.rearrange_visible_images()
        except Exception as e:
            print(f"更新网格布局时出错: {e}")
    
    def configure_grid_columns(self):
        """配置网格列权重"""
        try:
            # 清除旧的列配置
            for col in range(20):  # 清除可能存在的旧列配置
                self.scrollable_frame.columnconfigure(col, weight=0)
            
            # 配置当前列数的权重
            for col in range(self.grid_columns):
                self.scrollable_frame.columnconfigure(col, weight=1)
        except Exception as e:
            print(f"配置网格列时出错: {e}")
    
    def show_progress(self):
        """显示进度条（已禁用主界面进度条）"""
        # 主界面进度条已移除，此函数保留以避免调用错误
        pass
    
    def hide_progress(self):
        """隐藏进度条（已禁用主界面进度条）"""
        # 主界面进度条已移除，此函数保留以避免调用错误
        pass
    
    def update_progress(self, value, text=""):
        """更新进度条（已禁用主界面进度条）"""
        # 主界面进度条已移除，此函数保留以避免调用错误
        pass
    
    def remove_suffix(self):
        """去除序号尾缀功能"""
        current_text = self.search_entry.get('1.0', tk.END).strip()
        if current_text and current_text != "输入图片编号，每行一个\n例如：\nCL6453-1\nCL6812-4":
            lines = current_text.split('\n')
            processed_lines = []
            for line in lines:
                line = line.strip()
                if line:
                    # 按空格分割，处理每个SKU（支持同一行多个SKU的情况）
                    skus = line.split()
                    processed_skus = []
                    for sku in skus:
                        sku = sku.strip()
                        if sku:
                            # 去除每个SKU的序号尾缀（如 -1, -2, -3 等）
                            processed_sku = re.sub(r'-\d+$', '', sku)
                            processed_skus.append(processed_sku)
                    if processed_skus:
                        processed_lines.append(' '.join(processed_skus))
            
            # 检查处理后的内容是否与当前内容相同，避免不必要的重新填入
            processed_text = '\n'.join(processed_lines)
            if processed_text != current_text:
                self.search_entry.delete('1.0', tk.END)
                self.search_entry.insert('1.0', processed_text)
                self.search_entry.config(fg='#212529')
    
    def start_search(self):
        """开始搜索"""
        current_text = self.search_entry.get('1.0', tk.END).strip()
        if current_text and current_text != "输入图片编号，每行一个\n例如：\nCL6453-1\nCL6812-4":
            search_terms = [line.strip() for line in current_text.split('\n') if line.strip()]
            # 显示进度条
            self.show_progress()
            # 使用线程处理搜索，避免界面卡顿
            search_thread = threading.Thread(target=self.search_images_threaded, args=(search_terms,))
            search_thread.daemon = True
            search_thread.start()
            
    def search_images_threaded(self, search_terms):
        """优化的线程化搜索方法（支持多进程和智能缓存）"""
        try:
            if not self.source_folder or not search_terms:
                self.hide_progress()
                return
            
            # 生成更精确的搜索缓存键（包含文件夹修改时间）
            folder_mtime = os.path.getmtime(self.source_folder)
            search_key = f"{self.source_folder}:{folder_mtime}:{':'.join(sorted(search_terms))}"
            
            # 检查搜索结果缓存
            if hasattr(self, 'search_cache') and search_key in self.search_cache:
                cache_data = self.search_cache[search_key]
                # 检查缓存是否过期（10分钟，延长缓存时间）
                if time.time() - cache_data['timestamp'] < 600:
                    self.update_progress(50, "使用缓存的搜索结果...")
                    found_images = cache_data['results']
                    if found_images:
                        self.root.after(0, lambda: self.display_search_results(found_images))
                    else:
                        self.root.after(0, self.show_no_results)
                    self.update_progress(100, f"搜索完成 (缓存)")
                    self.root.after(1500, self.hide_progress)
                    return
            
            # 清空之前的搜索结果
            self.root.after(0, self.clear_search_results)
            
            start_time = time.time()
            
            # 第一阶段：获取所有图片文件（使用缓存或多进程扫描）
            self.update_progress(5, "正在获取文件列表...")
            
            # 使用线程池异步获取文件列表（1.6版本的优化）
            def get_files_task():
                return self.get_all_image_files_optimized()
            
            # 提交到搜索线程池
            future = self.search_executor.submit(get_files_task)
            
            try:
                # 等待文件获取完成（30秒超时）
                all_files = future.result(timeout=30)
            except Exception as e:
                print(f"获取文件列表失败: {e}")
                self.update_progress(100, f"获取文件列表失败: {str(e)}")
                self.root.after(2000, self.hide_progress)
                return
            
            if not all_files:
                self.update_progress(100, "未找到图片文件")
                self.root.after(1000, self.hide_progress)
                self.root.after(0, self.show_no_results)
                return
            
            total_files = len(all_files)
            self.update_progress(30, f"找到{total_files}个图片文件，开始搜索匹配项...")
            
            # 第二阶段：并行搜索匹配项
            found_images = self.parallel_search_optimized(all_files, search_terms)
            
            # 缓存搜索结果（智能缓存管理）
            if not hasattr(self, 'search_cache'):
                self.search_cache = {}
            
            # 清理过期缓存（超过1小时的缓存）
            current_time = time.time()
            expired_keys = [k for k, v in self.search_cache.items() 
                          if current_time - v['timestamp'] > 3600]
            for key in expired_keys:
                del self.search_cache[key]
            
            # 限制缓存大小，最多保存30个搜索结果（增加缓存容量）
            if len(self.search_cache) >= 30:
                # 删除最旧的缓存条目
                oldest_key = min(self.search_cache.keys(), 
                               key=lambda k: self.search_cache[k]['timestamp'])
                del self.search_cache[oldest_key]
            
            self.search_cache[search_key] = {
                'results': found_images,
                'timestamp': time.time(),
                'search_terms': search_terms,  # 保存搜索词用于调试
                'file_count': len(found_images)  # 保存结果数量
            }
            
            # 第三阶段：显示结果
            self.update_progress(95, f"找到{len(found_images)}个匹配文件，正在加载...")
            
            if found_images:
                self.root.after(0, lambda: self.display_search_results(found_images))
            else:
                self.root.after(0, self.show_no_results)
            
            # 完成
            elapsed_time = time.time() - start_time
            self.update_progress(100, f"搜索完成 ({elapsed_time:.2f}秒)")
            self.root.after(1500, self.hide_progress)
            
        except Exception as e:
            self.update_progress(100, f"搜索出错: {str(e)}")
            self.root.after(2000, self.hide_progress)
    
    def clear_search_results(self):
        """清空搜索结果"""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        # 重置网格位置
        self.current_row = 0
        self.current_col = 0
        
        # 更新滑动条显示状态
        self.update_scrollbar_visibility()
        
    def display_search_results(self, results):
        """显示搜索结果"""
        print(f"开始显示搜索结果，共{len(results)}个文件")
        self.search_results = results
        self.selected_images.clear()
        
        # 清空之前的统计框内容
        for widget in self.stats_container.winfo_children():
            widget.destroy()
        
        # 显示专业结果统计（固定在顶部，不滚动）
        stats_frame = tk.Frame(self.stats_container, bg='#E7F3FF', 
                              relief='solid', bd=1, padx=20, pady=12)
        stats_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=0, pady=0)
        stats_frame.columnconfigure(0, weight=1)
        
        # 统计信息布局
        stats_content = tk.Frame(stats_frame, bg='#E7F3FF')
        stats_content.pack(fill=tk.X)
        
        # 主要统计
        main_stats = tk.Label(stats_content, text=f"🎯 {len(results)} 个图片文件", 
                             bg='#E7F3FF', fg='#0D6EFD',
                             font=('Microsoft YaHei UI', 12, 'bold'))
        main_stats.pack(side=tk.LEFT)
        
        # 右侧操作区域
        right_frame = tk.Frame(stats_content, bg='#E7F3FF')
        right_frame.pack(side=tk.RIGHT, padx=(20, 0))  # 添加左边距，减少右边距
        
        # 一键勾选功能
        select_frame = tk.Frame(right_frame, bg='#E7F3FF')
        select_frame.pack(side=tk.LEFT, padx=(0, 10))  # 减少右边距
        
        # 全选/取消全选按钮
        self.select_all_btn = tk.Button(select_frame, text="全选",
                                       command=self.toggle_select_all,
                                       bg='#198754', fg='white',
                                       font=('Microsoft YaHei UI', 9),
                                       relief='flat', padx=8, pady=2,
                                       cursor='hand2')
        self.select_all_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Label(select_frame, text="勾选包含:", 
                bg='#E7F3FF', fg='#6C757D',
                font=('Microsoft YaHei UI', 9)).pack(side=tk.LEFT)
        
        self.select_keyword_var = tk.StringVar()
        self.select_entry = tk.Entry(select_frame, textvariable=self.select_keyword_var,
                               width=12, font=('Microsoft YaHei UI', 9),
                               relief='solid', bd=1)
        self.select_entry.pack(side=tk.LEFT, padx=(5, 5))
        
        # 移除自动完成功能
        # self.setup_autocomplete(self.select_entry, self.select_keyword_history)
        
        select_btn = tk.Button(select_frame, text="勾选",
                              command=self.select_by_keyword,
                              bg='#0D6EFD', fg='white',
                              font=('Microsoft YaHei UI', 9),
                              relief='flat', padx=8, pady=2,
                              cursor='hand2')
        select_btn.pack(side=tk.LEFT)
        
        
        # 一键隐藏功能
        hide_frame = tk.Frame(right_frame, bg='#E7F3FF')
        hide_frame.pack(side=tk.LEFT, padx=(10, 8))  # 减少左右间距
        
        tk.Label(hide_frame, text="隐藏不包含:", 
                bg='#E7F3FF', fg='#6C757D',
                font=('Microsoft YaHei UI', 9)).pack(side=tk.LEFT)
        
        self.hide_keyword_var = tk.StringVar()
        self.hide_entry = tk.Entry(hide_frame, textvariable=self.hide_keyword_var,
                             width=12, font=('Microsoft YaHei UI', 9),
                             relief='solid', bd=1)
        self.hide_entry.pack(side=tk.LEFT, padx=(5, 5))
        
        # 移除自动完成功能
        # self.setup_autocomplete(self.hide_entry, self.hide_keyword_history)
        
        hide_btn = tk.Button(hide_frame, text="隐藏",
                            command=self.hide_by_keyword,
                            bg='#DC3545', fg='white',
                            font=('Microsoft YaHei UI', 9),
                            relief='flat', padx=8, pady=2,
                            cursor='hand2')
        hide_btn.pack(side=tk.LEFT)
        
        # 显示全部按钮
        show_all_btn = tk.Button(hide_frame, text="显示全部",
                                command=self.show_all_images,
                                bg='#28A745', fg='white',
                                font=('Microsoft YaHei UI', 9),
                                relief='flat', padx=8, pady=2,
                                cursor='hand2')
        show_all_btn.pack(side=tk.LEFT, padx=(5, 0))
        
        # 操作提示
        tip_stats = tk.Label(right_frame, text="勾选图片", 
                            bg='#E7F3FF', fg='#6C757D',
                            font=('Microsoft YaHei UI', 10))
        tip_stats.pack(side=tk.LEFT, padx=(8, 0))  # 添加左边距
        
        # 配置网格列权重（响应式）
        self.configure_grid_columns()
        
        # 清空滚动区域的内容
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        # 显示结果（网格布局）- 分批加载优化
        self.current_row = 0  # 从第0行开始，统计信息已移到固定区域
        self.current_col = 0
        
        # 重置全选按钮状态
        if hasattr(self, 'select_all_btn'):
            self.select_all_btn.config(text="全选")
        
        print(f"当前网格列数: {self.grid_columns}")
        
        # 分批加载优化：初始只加载前50个图片，其余延迟加载
        self.batch_size = 50  # 每批加载数量
        self.current_batch = 0
        self.total_results = results
        
        # 立即加载第一批
        self.load_next_batch()
        
        # 自动检测同组图片并自动隐藏功能
        self.auto_detect_and_hide_duplicates(results)
        
        # 延迟检测低像素图片，等待隐藏功能完成
        self.root.after(500, lambda: self.check_low_resolution_images(results))
        
        # 延迟更新滑动条显示状态，确保内容已完全加载
        self.root.after(100, self.update_scrollbar_visibility)
        #     
        #     # 第一步：基于文件名相似度进行分组
        #     image_groups = self._group_similar_images(results)
        #     
        #     # 第二步：分析每组内的文件特征，识别高清版本
        #     for group_name, file_list in image_groups.items():
        #         if len(file_list) == 1:
        #             # 单独文件，检查是否可能需要高清处理
        #             file_info = file_list[0]
        #             if self._should_upscale_single_image(file_info):
        #                 standalone_images.append(file_info['path'])
        #         else:
        #             # 多个相似文件，智能识别哪些需要高清处理
        #             candidates = self._analyze_group_for_upscale_candidates(file_list)
        #             standalone_images.extend(candidates)
        #     
        #     # 添加调试信息
        #     print(f"智能检测结果:")
        #     print(f"  - 分组数量: {len(image_groups)}")
        #     print(f"  - 检测到的单独图片: {len(standalone_images)}")
        #     for img in standalone_images:
        #         print(f"    * {os.path.basename(img)}")
        #     
        #     # 如果检测到需要高清处理的图片，询问用户是否一键勾选
        #     if len(standalone_images) >= 1:
        #         self.show_standalone_images_dialog(standalone_images)
        #         
        # except Exception as e:
        #     print(f"智能检测单独图片时出错: {e}")
    
    def check_low_resolution_images(self, results):
        """检测隐藏处理后像素低于4000x4000的图片并弹窗提示"""
        try:
            # 获取当前显示的图片（未被隐藏的）
            visible_images = []
            for i, image_path in enumerate(results):
                # 检查图片项是否被隐藏
                if hasattr(self, 'image_items') and i in self.image_items:
                    image_item = self.image_items[i]
                    if image_item.winfo_viewable():  # 检查是否可见
                        visible_images.append(image_path)
                else:
                    # 如果图片项还未创建，默认认为是可见的
                    visible_images.append(image_path)
            
            # 过滤出单张图片（不属于任何组的图片）
            standalone_images = self._filter_standalone_images(visible_images)
            
            # 检测低像素图片
            low_resolution_images = []
            for image_path in standalone_images:
                width, height = self._get_image_dimensions_safe(image_path)
                if width > 0 and height > 0:
                    # 检查是否低于4000x4000像素
                    if width < 4000 or height < 4000:
                        low_resolution_images.append({
                            'path': image_path,
                            'width': width,
                            'height': height,
                            'filename': os.path.basename(image_path)
                        })
            
            # 如果发现低像素图片，显示弹窗
            if low_resolution_images:
                self.show_low_resolution_dialog(low_resolution_images)
                
        except Exception as e:
            print(f"检测低像素图片时出错: {e}")
    
    def _filter_standalone_images(self, image_paths):
        """过滤出单张图片（不属于任何组的图片）"""
        try:
            # 获取所有文件名（不含路径）
            filenames = [os.path.basename(path) for path in image_paths]
            
            # 创建文件名到完整路径的映射
            filename_to_path = {os.path.basename(path): path for path in image_paths}
            
            # 分析文件名，找出基础名称
            base_name_groups = {}
            
            for filename in filenames:
                # 移除扩展名
                name_without_ext = os.path.splitext(filename)[0]
                
                # 尝试提取基础名称（移除括号内容和特殊后缀）
                base_name = self._extract_base_name(name_without_ext)
                
                if base_name not in base_name_groups:
                    base_name_groups[base_name] = []
                base_name_groups[base_name].append(filename)
            
            # 找出只有单张图片的组
            standalone_filenames = []
            for base_name, file_list in base_name_groups.items():
                if len(file_list) == 1:
                    standalone_filenames.extend(file_list)
            
            # 转换回完整路径
            standalone_paths = [filename_to_path[filename] for filename in standalone_filenames]
            
            print(f"检测到 {len(standalone_paths)} 张单独图片，共 {len(image_paths)} 张可见图片")
            
            return standalone_paths
            
        except Exception as e:
            print(f"过滤单张图片时出错: {e}")
            return image_paths  # 出错时返回所有图片
    
    def _extract_base_name(self, name_without_ext):
        """提取文件的基础名称，移除括号内容和特殊后缀"""
        import re
        
        # 移除括号内容，如 CL001（x6） -> CL001
        base_name = re.sub(r'[（(][^）)]*[）)]', '', name_without_ext)
        
        # 移除常见的后缀分隔符及其后的内容，如 CL001_x6 -> CL001, CL001-高清 -> CL001
        base_name = re.sub(r'[-_\s][^-_\s]*$', '', base_name)
        
        # 清理多余的空格
        base_name = base_name.strip()
        
        # 如果处理后为空，使用原名称
        if not base_name:
            base_name = name_without_ext
            
        return base_name
    
    def show_low_resolution_dialog(self, low_resolution_images):
        """显示低像素图片提示弹窗"""
        try:
            count = len(low_resolution_images)
            
            # 创建弹窗
            dialog = tk.Toplevel(self.root)
            dialog.title("低像素图片检测")
            dialog.geometry("500x200")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.attributes('-topmost', True)
            dialog.focus_force()
            
            # 立即隐藏窗口，避免在左上角显示
            dialog.withdraw()
            
            # 居中显示 - 先更新布局但窗口仍然隐藏
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
            y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")
            
            # 设置好位置后再显示窗口，避免移动效果
            dialog.deiconify()
            
            # 主框架
            main_frame = ttk.Frame(dialog, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # 标题
            title_label = ttk.Label(main_frame, 
                                   text="🔍 低像素图片检测", 
                                   font=('Microsoft YaHei UI', 14, 'bold'))
            title_label.pack(pady=(0, 15))
            
            # 提示信息
            info_text = f"发现 {count} 张图片像素低于打印标准，是否一键勾选准备进行高清处理"
            warning_text = "⚠️ 基础像素高的图片要使用小倍数放大模式，否则等待时间会很长"
            
            info_label = ttk.Label(main_frame, text=info_text, 
                                  font=('Microsoft YaHei UI', 11))
            info_label.pack(pady=(0, 5))
            
            warning_label = ttk.Label(main_frame, text=warning_text, 
                                     font=('Microsoft YaHei UI', 10),
                                     foreground='#FF6B35')
            warning_label.pack(pady=(0, 15))
            

            
            # 按钮框架
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X)
            
            # 居中按钮容器
            center_frame = ttk.Frame(button_frame)
            center_frame.pack(expand=True)
            
            # 是按钮
            yes_btn = ttk.Button(center_frame, text="是", 
                                command=lambda: self.select_low_resolution_images(low_resolution_images, dialog),
                                width=10)
            yes_btn.pack(side=tk.LEFT, padx=(0, 20))
            
            # 否按钮
            no_btn = ttk.Button(center_frame, text="否", 
                               command=dialog.destroy,
                               width=10)
            no_btn.pack(side=tk.LEFT)
            
        except Exception as e:
            print(f"显示低像素图片弹窗时出错: {e}")
    
    def select_low_resolution_images(self, low_resolution_images, dialog):
        """一键勾选低像素图片"""
        try:
            # 获取低像素图片的路径列表
            low_res_paths = [img_info['path'] for img_info in low_resolution_images]
            print(f"需要勾选的低像素图片数量: {len(low_res_paths)}")
            
            # 在搜索结果中找到对应的图片并勾选
            selected_count = 0
            
            # 遍历所有图片项框架，找到对应的复选框
            for widget in self.scrollable_frame.winfo_children():
                if isinstance(widget, tk.Frame) and hasattr(widget, 'image_index'):
                    image_index = widget.image_index
                    
                    # 检查这个索引对应的图片是否在低像素列表中
                    if image_index < len(self.search_results):
                        image_path = self.search_results[image_index]
                        if image_path in low_res_paths:
                            print(f"找到匹配的图片: {image_path}")
                            
                            # 使用存储的复选框变量直接设置状态
                            if hasattr(widget, 'checkbox_var'):
                                print(f"找到复选框变量，勾选索引: {image_index}")
                                widget.checkbox_var.set(True)
                                # 手动调用toggle_selection更新选择状态
                                self.toggle_selection(image_index, True)
                                selected_count += 1
                                print(f"已勾选图片，当前计数: {selected_count}")
                            else:
                                print(f"未找到复选框变量: {image_index}")
            
            print(f"总共勾选了 {selected_count} 张图片")
            
            # 更新全选按钮状态
            self.update_select_all_button_state()
            
            # 关闭弹窗
            dialog.destroy()
            
        except Exception as e:
            print(f"一键勾选低像素图片时出错: {e}")
            messagebox.showerror("错误", f"勾选图片时出错：{str(e)}")

    def auto_detect_and_hide_duplicates(self, results):
        """自动检测同组图片并自动隐藏重复项"""
        try:
            # 收集所有文件名
            filenames = [os.path.basename(image_path) for image_path in results]
            
            # 检测同组图片的尾缀关键词
            detected_keywords = self._detect_suffix_keywords(filenames)
            
            if detected_keywords:
                # 将检测到的关键词填入隐藏功能输入框
                keywords_str = ' '.join(detected_keywords)
                self.hide_keyword_var.set(keywords_str)
                
                # 自动触发隐藏功能
                self.hide_by_keyword()
                
                print(f"自动检测到同组图片关键词: {keywords_str}")
            else:
                print("未检测到同组图片关键词")
                
        except Exception as e:
            print(f"自动检测同组图片时出错: {e}")
    
    def _detect_suffix_keywords(self, filenames):
        """检测文件名中的尾缀关键词"""
        # 常见的尾缀关键词模式
        suffix_patterns = [
            r'[（(]([^）)]+)[）)]',  # 括号内的内容，如（x6）、(高清)
            r'[-_\s]([^-_\s.]+)(?=\.[^.]*$)',  # 文件扩展名前的最后一个词，如_x6、-高清
        ]
        
        # 收集所有可能的关键词
        potential_keywords = set()
        
        for filename in filenames:
            # 移除文件扩展名
            name_without_ext = os.path.splitext(filename)[0]
            
            for pattern in suffix_patterns:
                matches = re.findall(pattern, name_without_ext, re.IGNORECASE)
                for match in matches:
                    # 过滤掉纯数字和过短的关键词
                    if len(match) >= 2 and not match.isdigit():
                        potential_keywords.add(match.strip())
        
        # 验证关键词：检查是否存在同名但不含关键词的文件
        valid_keywords = []
        
        for keyword in potential_keywords:
            # 检查是否存在包含此关键词的文件和不包含此关键词的同名文件
            has_with_keyword = False
            has_without_keyword = False
            
            for filename in filenames:
                name_without_ext = os.path.splitext(filename)[0]
                
                if keyword.lower() in filename.lower():
                    has_with_keyword = True
                    
                    # 尝试找到对应的不含关键词的文件
                    # 移除各种格式的关键词
                    patterns_to_remove = [
                        f'（{keyword}）', f'({keyword})', f' {keyword}', 
                        f'_{keyword}', f'-{keyword}',
                        f'（{keyword.upper()}）', f'({keyword.upper()})', f' {keyword.upper()}', 
                        f'_{keyword.upper()}', f'-{keyword.upper()}',
                        f'（{keyword.lower()}）', f'({keyword.lower()})', f' {keyword.lower()}', 
                        f'_{keyword.lower()}', f'-{keyword.lower()}'
                    ]
                    
                    for pattern in patterns_to_remove:
                        if pattern in name_without_ext:
                            clean_name = name_without_ext.replace(pattern, '')
                            # 检查是否存在对应的不含关键词的文件
                            for other_filename in filenames:
                                other_name_without_ext = os.path.splitext(other_filename)[0]
                                if other_name_without_ext == clean_name:
                                    has_without_keyword = True
                                    break
                            if has_without_keyword:
                                break
                    
                    if has_without_keyword:
                        break
            
            # 如果同时存在包含和不包含关键词的文件，则认为是有效关键词
            if has_with_keyword and has_without_keyword:
                valid_keywords.append(keyword)
        
        return valid_keywords
    
    def _get_image_dimensions_safe(self, file_path):
        """安全获取图片尺寸"""
        try:
            with Image.open(file_path) as img:
                return img.size  # (width, height)
        except (OSError, IOError, Exception):
            return (0, 0)
        # # 这种情况下可以选择最小的文件进行高清处理
        # print(f"  组内未检测到高清版本，选择最小文件: {smallest_file['filename']}")
        # candidates.append(smallest_file['path'])
        # 
        # return candidates
    


    def show_no_results(self):
        """显示无结果状态"""
        # 创建专业空状态提示
        empty_frame = tk.Frame(self.scrollable_frame, bg='#FFFFFF')
        empty_frame.grid(row=0, column=0, columnspan=self.grid_columns, 
                        sticky=(tk.W, tk.E, tk.N, tk.S), pady=80)
        
        # 空状态图标
        icon_label = tk.Label(empty_frame, text="🔍", 
                             bg='#FFFFFF', fg='#DEE2E6',
                             font=('Microsoft YaHei UI', 64))
        icon_label.pack(pady=(0, 24))
        
        # 空状态标题
        empty_title = tk.Label(empty_frame, text="未找到匹配的图片", 
                              bg='#FFFFFF', fg='#495057',
                              font=('Microsoft YaHei UI', 16, 'bold'))
        empty_title.pack(pady=(0, 12))
        
        # 提示文字
        tip_label = tk.Label(empty_frame, text="请检查搜索关键词或图库路径设置", 
                            bg='#FFFFFF', fg='#6C757D',
                            font=('Microsoft YaHei UI', 11))
        tip_label.pack(pady=(0, 8))
        
        # 建议文字
        suggest_label = tk.Label(empty_frame, text="• 尝试使用不同的关键词\n• 确认源文件夹包含图片文件\n• 支持格式：JPG, PNG, BMP, GIF", 
                                bg='#FFFFFF', fg='#ADB5BD',
                                font=('Microsoft YaHei UI', 10),
                                justify=tk.LEFT)
        suggest_label.pack()
            
    def load_next_batch(self):
        """分批加载图片项，提升UI响应速度"""
        try:
            start_idx = self.current_batch * self.batch_size
            end_idx = min(start_idx + self.batch_size, len(self.total_results))
            
            if start_idx >= len(self.total_results):
                return  # 已加载完所有图片
            
            print(f"加载第{self.current_batch + 1}批图片: {start_idx}-{end_idx-1}")
            
            # 加载当前批次的图片
            for i in range(start_idx, end_idx):
                image_path = self.total_results[i]
                print(f"创建图片项 {i}: {os.path.basename(image_path)}")
                self.create_image_item(i, image_path)
            
            self.current_batch += 1
            
            # 如果还有更多图片，延迟加载下一批（避免UI卡顿）
            if end_idx < len(self.total_results):
                self.root.after(100, self.load_next_batch)  # 100ms后加载下一批
                
        except Exception as e:
            print(f"分批加载图片时出错: {e}")

    def create_image_item(self, index, image_path, pre_selected=False):
        """创建网格布局的图片项卡片（响应式）- 使用线程池优化"""
        try:
            # 根据网格列数动态计算图片项大小
            item_width = max(140, self.min_item_width)
            thumb_size = max(100, item_width - 40)  # 缩略图比容器小40px
            
            # 创建卡片容器（响应式网格布局）
            item_frame = tk.Frame(self.scrollable_frame, bg='#FFFFFF', 
                                 relief='solid', bd=1, padx=8, pady=8)
            item_frame.grid(row=self.current_row, column=self.current_col, 
                           sticky=(tk.W, tk.E, tk.N, tk.S), padx=4, pady=4)
            
            # 添加索引属性用于一键勾选功能
            item_frame.image_index = index
            
            # 更新网格位置
            self.current_col += 1
            if self.current_col >= self.grid_columns:
                self.current_col = 0
                self.current_row += 1
            
            # 缩略图容器（动态尺寸）
            thumb_container = tk.Frame(item_frame, bg='#F8F9FA', 
                                      relief='solid', bd=1, 
                                      width=thumb_size, height=thumb_size)
            thumb_container.pack(pady=(0, 8))
            thumb_container.pack_propagate(False)
            
            # 使用线程池异步加载缩略图（1.6版本的优化）
            def load_thumbnail():
                try:
                    # 检查文件是否存在
                    if not os.path.exists(image_path):
                        print(f"缩略图加载失败：文件不存在 - {image_path}")
                        icon_size = max(24, thumb_size // 3)
                        self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                        return
                    
                    # 检查文件大小
                    try:
                        file_size = os.path.getsize(image_path)
                        if file_size == 0:
                            print(f"缩略图加载失败：文件为空 - {image_path}")
                            icon_size = max(24, thumb_size // 3)
                            self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                            return
                    except OSError as size_error:
                        print(f"缩略图加载失败：无法获取文件大小 - {image_path}, 错误: {size_error}")
                        icon_size = max(24, thumb_size // 3)
                        self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                        return
                    
                    with Image.open(image_path) as img:
                        # 根据容器大小调整缩略图
                        display_size = thumb_size - 10
                        img.thumbnail((display_size, display_size), Image.Resampling.LANCZOS)
                        photo = ImageTk.PhotoImage(img)
                        
                        # 在主线程中更新UI
                        self.root.after(0, lambda: self._update_thumbnail(thumb_container, photo))
                        print(f"缩略图加载成功: {os.path.basename(image_path)}")
                        
                except FileNotFoundError:
                    print(f"缩略图加载失败：文件未找到 - {image_path}")
                    icon_size = max(24, thumb_size // 3)
                    self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                except PermissionError:
                    print(f"缩略图加载失败：文件权限不足 - {image_path}")
                    icon_size = max(24, thumb_size // 3)
                    self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                except Image.UnidentifiedImageError:
                    print(f"缩略图加载失败：无法识别的图片格式 - {image_path}")
                    icon_size = max(24, thumb_size // 3)
                    self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
                except Exception as e:
                    print(f"缩略图加载失败：未知错误 - {image_path}, 错误类型: {type(e).__name__}, 错误信息: {str(e)}")
                    # 显示默认图标
                    icon_size = max(24, thumb_size // 3)
                    self.root.after(0, lambda: self._show_default_icon(thumb_container, icon_size))
            
            # 提交到线程池
            # 提交任务前检查容器是否有效
            if thumb_container and thumb_container.winfo_exists():
                self.thumbnail_executor.submit(load_thumbnail)
            
            # 文件信息容器
            info_frame = tk.Frame(item_frame, bg='#FFFFFF')
            info_frame.pack(fill=tk.X)
            
            # 文件名（截断显示）
            filename = os.path.basename(image_path)
            
            # 应用尺寸映射替换
            display_name = self.apply_size_mapping(filename)
            
            if len(display_name) > 20:
                display_name = display_name[:17] + "..."
                
            name_label = tk.Label(info_frame, text=display_name, 
                                 bg='#FFFFFF', fg='#212529',
                                 font=('Microsoft YaHei UI', 10, 'bold'),
                                 anchor='center')
            name_label.pack(fill=tk.X, pady=(0, 5))
            
            # 使用线程池异步加载文件信息（1.6版本的优化）
            def load_file_info():
                try:
                    # 获取文件大小
                    file_size = os.path.getsize(image_path)
                    if file_size < 1024:
                        size_text = f"{file_size} B"
                    elif file_size < 1024 * 1024:
                        size_text = f"{file_size / 1024:.1f} KB"
                    else:
                        size_text = f"{file_size / (1024 * 1024):.1f} MB"
                    
                    # 获取图片像素信息
                    pixel_text = ""
                    try:
                        with Image.open(image_path) as img:
                            width, height = img.size
                            pixel_text = f"{width}×{height}"
                    except Exception:
                        pixel_text = "未知尺寸"
                    
                    # 在主线程中更新UI
                    self.root.after(0, lambda: self._update_file_info(info_frame, size_text, pixel_text))
                    
                except Exception as e:
                    print(f"加载文件信息失败: {image_path}, 错误: {e}")
                    # 显示默认信息
                    self.root.after(0, lambda: self._update_file_info(info_frame, "未知大小", "未知尺寸"))
            
            # 提交到线程池
            self.info_executor.submit(load_file_info)
            
            # 选择框
            var = tk.BooleanVar()
            # 如果是预选中状态，设置为True
            if pre_selected:
                var.set(True)
            checkbox = tk.Checkbutton(info_frame, variable=var, text="选择",
                                     bg='#FFFFFF', activebackground='#FFFFFF',
                                     font=('Microsoft YaHei UI', 10),
                                     fg='#0D6EFD', activeforeground='#0D6EFD',
                                     selectcolor='#FFFFFF',
                                     command=lambda: self.toggle_selection(index, var.get()))
            checkbox.pack()
            
            # 将复选框变量存储到item_frame中以便后续访问
            item_frame.checkbox_var = var
            item_frame.checkbox = checkbox
            
            # 添加悬停效果
            item_frame.bind('<Enter>', on_enter)
            item_frame.bind('<Leave>', on_leave)
            
        except Exception as e:
            print(f"创建图片项时出错: {e}")
    
    def _update_thumbnail(self, container, photo):
        """在主线程中更新缩略图"""
        try:
            # 检查container是否仍然存在且有效
            if not container or not container.winfo_exists():
                return
            
            thumb_label = tk.Label(container, image=photo, bg='#F8F9FA')
            thumb_label.image = photo  # 保持引用
            thumb_label.pack(expand=True)
        except tk.TclError as e:
            # 处理widget已被销毁的情况
            if "bad window path name" in str(e):
                return  # 静默处理，widget已被销毁
            print(f"更新缩略图失败: {e}")
        except Exception as e:
            print(f"更新缩略图失败: {e}")
    
    def _show_default_icon(self, container, icon_size):
        """显示默认图标"""
        try:
            # 检查container是否仍然存在且有效
            if not container or not container.winfo_exists():
                return
                
            default_label = tk.Label(container, text="🖼️", 
                                    bg='#F8F9FA', fg='#6C757D',
                                    font=('Microsoft YaHei UI', icon_size))
            default_label.pack(expand=True)
        except tk.TclError as e:
            # 处理widget已被销毁的情况
            if "bad window path name" in str(e):
                return  # 静默处理，widget已被销毁
            print(f"显示默认图标失败: {e}")
        except Exception as e:
            print(f"显示默认图标失败: {e}")
    
    def _update_file_info(self, info_frame, size_text, pixel_text):
        """在主线程中更新文件信息"""
        try:
            # 检查info_frame是否仍然存在且有效
            if not info_frame or not info_frame.winfo_exists():
                return
                
            # 文件大小标签
            size_label = tk.Label(info_frame, text=size_text, 
                                 bg='#FFFFFF', fg='#6C757D',
                                 font=('Microsoft YaHei UI', 9),
                                 anchor='center')
            size_label.pack(fill=tk.X, pady=(0, 3))
            
            # 像素信息标签
            if pixel_text:
                pixel_label = tk.Label(info_frame, text=pixel_text, 
                                      bg='#FFFFFF', fg='#6C757D',
                                      font=('Microsoft YaHei UI', 9),
                                      anchor='center')
                pixel_label.pack(fill=tk.X, pady=(0, 8))
        except tk.TclError as e:
            # 处理widget已被销毁的情况
            if "bad window path name" in str(e):
                return  # 静默处理，widget已被销毁
            print(f"更新文件信息失败: {e}")
        except Exception as e:
            print(f"更新文件信息失败: {e}")
    
    def __del__(self):
        """析构函数，清理线程池资源"""
        try:
            if hasattr(self, 'thumbnail_executor'):
                self.thumbnail_executor.shutdown(wait=False)
            if hasattr(self, 'info_executor'):
                self.info_executor.shutdown(wait=False)
            if hasattr(self, 'search_executor'):
                self.search_executor.shutdown(wait=False)
        except Exception:
            pass
    
    def on_smart_upscale_enable_change(self):
        """智能高清处理启用状态改变事件"""
        self.smart_upscale_config['enabled'] = self.smart_upscale_enabled_var.get()
        self.save_config()
    
    def on_upscale_rename_change(self, *args):
        """处理重命名规则变化事件"""
        try:
            new_rule = self.upscale_rename_var.get()
            self.upscale_config['rename_rule'] = new_rule
            self.save_config()
        except Exception as e:
            print(f"保存重命名规则时出错: {e}")
    
    def on_skip_qualified_change(self):
        """跳过已达标图片选项改变事件"""
        self.smart_upscale_config['skip_qualified'] = self.skip_qualified_var.get()
        self.save_config()
    
    def save_smart_upscale_config(self):
        """保存智能高清处理配置"""
        try:
            # 验证输入
            target_width = int(self.target_width_var.get())
            target_height = int(self.target_height_var.get())
            # 移除超时变量的获取 - 不再需要超时机制
            
            if target_width <= 0 or target_height <= 0:
                messagebox.showerror("错误", "目标尺寸必须大于0")
                return
                
            # 移除超时验证 - 不再需要超时机制
            
            # 更新智能高清处理配置
            self.smart_upscale_config['target_width'] = target_width
            self.smart_upscale_config['target_height'] = target_height
            self.smart_upscale_config['enabled'] = self.smart_upscale_enabled_var.get()
            self.smart_upscale_config['skip_qualified'] = self.skip_qualified_var.get()
            # 移除超时配置的保存 - 不再需要超时机制
            
            # 更新API配置
            self.bigjpg_api_key = self.api_key_var.get()
            
            # 更新高清处理配置
            self.upscale_config['style'] = self.default_style_var.get()
            self.upscale_config['noise'] = self.default_noise_var.get()
            # 移除超时配置的保存 - 不再需要超时机制
            
            # 保存配置
            self.save_config()
            messagebox.showinfo("成功", "智能高清处理配置已保存")
            
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")
        except Exception as e:
            messagebox.showerror("错误", f"保存配置时出错: {str(e)}")
    
    def load_file_cache(self):
        try:
            if os.path.exists(self.cache_file):
                # 检查缓存文件大小，避免加载过大的缓存文件
                cache_size = os.path.getsize(self.cache_file)
                if cache_size > 50 * 1024 * 1024:  # 50MB限制
                    print("缓存文件过大，跳过加载")
                    self.file_cache = {}
                    self.last_scan_time = 0
                    return
                
                with open(self.cache_file, 'rb') as f:
                    cache_data = pickle.load(f)
                    self.file_cache = cache_data.get('files', {})
                    self.last_scan_time = cache_data.get('scan_time', 0)
                    
                # 清理过期的缓存条目（超过7天）
                current_time = time.time()
                expired_keys = []
                for key, data in self.file_cache.items():
                    if isinstance(data, dict) and 'timestamp' in data:
                        if current_time - data['timestamp'] > 7 * 24 * 3600:  # 7天
                            expired_keys.append(key)
                
                # 删除过期条目
                for key in expired_keys:
                    del self.file_cache[key]
                    
        except Exception as e:
            print(f"加载缓存失败: {e}")
            self.file_cache = {}
            self.last_scan_time = 0
    
    def save_file_cache(self):
        """保存文件缓存"""
        try:
            cache_data = {
                'files': self.file_cache,
                'scan_time': self.last_scan_time
            }
            with open(self.cache_file, 'wb') as f:
                pickle.dump(cache_data, f)
        except Exception as e:
            print(f"保存缓存失败: {e}")
    
    def get_all_image_files_optimized(self):
        """优化的文件扫描方法（实现 Smart Incremental Scan 增量扫描）"""
        folder_path = self.source_folder
        
        # 确保缓存结构正确
        if not isinstance(self.file_cache, dict):
            self.file_cache = {}
            
        if folder_path not in self.file_cache:
            self.file_cache[folder_path] = {'dir_data': {}, 'files': []}
            
        # 如果不是字典格式（旧版本缓存），则重置
        if 'dir_data' not in self.file_cache[folder_path]:
            self.file_cache[folder_path] = {'dir_data': {}, 'files': []}

        cached_dirs = self.file_cache[folder_path]['dir_data']
        all_files = []
        cache_updated = False
        
        self.update_progress(10, "正在执行增量扫描...")
        
        try:
            processed_dirs = 0
            for root, dirs, files in os.walk(folder_path):
                # 过滤掉隐藏目录和系统目录
                dirs[:] = [d for d in dirs if not d.startswith('.') and d not in ['__pycache__', 'node_modules', 'Thumbs.db']]
                
                try:
                    # 获取当前目录的修改时间
                    root_stat = os.stat(root)
                    current_mtime = root_stat.st_mtime
                    
                    # 检查缓存中是否有该目录且 mtime 未变
                    if root in cached_dirs and cached_dirs[root]['mtime'] == current_mtime:
                        # 目录未变，直接从缓存获取文件列表
                        dir_files = cached_dirs[root]['files']
                    else:
                        # 目录已变或不存在，重新扫描该目录下的图片
                        dir_files = []
                        for file in files:
                            if not file.startswith('.'):
                                file_ext = os.path.splitext(file)[1].lower()
                                if file_ext in self.supported_formats:
                                    dir_files.append(os.path.join(root, file))
                        
                        # 更新目录缓存
                        cached_dirs[root] = {
                            'mtime': current_mtime,
                            'files': dir_files
                        }
                        cache_updated = True
                    
                    all_files.extend(dir_files)
                    processed_dirs += 1
                    
                    # 进度反馈
                    if processed_dirs % 20 == 0:
                        self.update_progress(min(23, 10 + (processed_dirs / 100) * 5), f"已扫描 {len(all_files)} 个图片...")
                        
                except (OSError, PermissionError):
                    continue

            # 如果缓存有更新，保存到磁盘
            if cache_updated:
                self.file_cache[folder_path]['files'] = all_files
                self.file_cache[folder_path]['timestamp'] = time.time()
                self.save_file_cache()
            
            self.update_progress(25, f"扫描完成，找到 {len(all_files)} 个图片文件")
            return all_files
            
        except Exception as e:
            print(f"增量扫描失败，回退到原始方法: {e}")
            import traceback
            traceback.print_exc()
            return self.get_all_image_files_fallback()
    
    def get_all_image_files_fallback(self):
        """回退的单线程文件扫描方法"""
        all_files = []
        
        for root, dirs, files in os.walk(self.source_folder):
            for file in files:
                file_ext = os.path.splitext(file)[1].lower()
                if file_ext in self.supported_formats:
                    file_path = os.path.join(root, file)
                    all_files.append(file_path)
        
        return all_files
    
    def parallel_search_optimized(self, all_files, search_terms):
        """优化的并行搜索方法"""
        if not all_files:
            return []
        
        # 将文件列表分块，每个线程处理一部分
        chunk_size = max(100, len(all_files) // (self.cpu_count * 2))  # 使用更多线程
        file_chunks = [all_files[i:i + chunk_size] for i in range(0, len(all_files), chunk_size)]
        
        self.update_progress(50, f"使用{len(file_chunks)}个线程并行搜索...")
        
        found_images = []
        
        try:
            # 使用线程池并行搜索（IO密集型任务用线程更合适）
            with ThreadPoolExecutor(max_workers=self.cpu_count * 2) as executor:
                # 为每个块创建搜索任务
                futures = [executor.submit(parallel_search_files, chunk, search_terms) 
                          for chunk in file_chunks]
                
                # 收集结果
                for i, future in enumerate(futures):
                    try:
                        result = future.result(timeout=30)  # 30秒超时
                        found_images.extend(result)
                        
                        # 更新进度
                        progress = 50 + (i + 1) / len(futures) * 40
                        self.root.after(0, lambda p=progress: 
                                       self.update_progress(p, f"搜索进度: {i+1}/{len(futures)}"))
                    except Exception as e:
                        print(f"搜索块时出错: {e}")
            
            self.update_progress(90, f"搜索完成，找到{len(found_images)}个匹配文件")
            return found_images
            
        except Exception as e:
            print(f"并行搜索失败，回退到单线程: {e}")
            # 回退到单线程搜索
            return parallel_search_files(all_files, search_terms)
        
    def select_by_keyword(self):
        """根据关键词一键勾选图片（仅勾选可见的图片）"""
        keyword = self.select_keyword_var.get().strip()
        if not keyword:
            messagebox.showwarning("提示", "请输入要匹配的关键词")
            return
            
        # 保存到历史记录
        self.add_to_history(self.select_keyword_history, keyword)
            
        selected_count = 0
        
        # 遍历所有搜索结果
        for i, image_path in enumerate(self.search_results):
            filename = os.path.basename(image_path)
            print(f"检查文件: {filename}, 关键词: {keyword}")
            # 检查文件名是否包含关键词（不区分大小写）
            if keyword.lower() in filename.lower():
                print(f"匹配到文件: {filename}")
                # 找到对应的复选框并勾选
                for widget in self.scrollable_frame.winfo_children():
                    if hasattr(widget, 'image_index') and widget.image_index == i:
                        # 检查widget是否可见（未被隐藏）
                        if widget.grid_info():  # 只有可见的widget才有grid_info
                            # 递归查找复选框
                            checkbox = self.find_checkbox_in_widget(widget)
                            if checkbox:
                                checkbox.select()  # 勾选复选框
                                self.selected_images.add(i)
                                selected_count += 1
                                print(f"成功勾选: {filename}")
                            else:
                                print(f"未找到复选框: {filename}")
                        else:
                            print(f"跳过隐藏的文件: {filename}")
                        break
        
        # 取消非必要的成功提示弹窗
    
    def toggle_select_all(self):
        """切换全选/取消全选状态"""
        # 检查当前是否有选中的可见图片
        visible_selected_count = 0
        visible_total_count = 0
        
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                # 检查widget是否可见（未被隐藏）
                if widget.grid_info():  # 只有可见的widget才有grid_info
                    visible_total_count += 1
                    if widget.image_index in self.selected_images:
                        visible_selected_count += 1
        
        # 如果所有可见图片都已选中，则取消全选；否则全选
        if visible_selected_count == visible_total_count and visible_total_count > 0:
            self.deselect_all_visible()
        else:
            self.select_all_visible()
    
    def select_all_visible(self):
        """全选所有可见的图片"""
        selected_count = 0
        
        # 遍历所有图片widget
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                # 检查widget是否可见（未被隐藏）
                if widget.grid_info():  # 只有可见的widget才有grid_info
                    # 递归查找复选框
                    checkbox = self.find_checkbox_in_widget(widget)
                    if checkbox:
                        # 检查是否已经勾选，避免重复勾选
                        if widget.image_index not in self.selected_images:
                            checkbox.select()  # 勾选复选框
                            self.selected_images.add(widget.image_index)
                            selected_count += 1
        
        # 更新按钮文本
        self.select_all_btn.config(text="取消全选")
        
        # 取消非必要的成功提示弹窗
    
    def deselect_all_visible(self):
        """取消选择所有可见的图片"""
        deselected_count = 0
        
        # 遍历所有图片widget
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                # 检查widget是否可见（未被隐藏）
                if widget.grid_info():  # 只有可见的widget才有grid_info
                    # 递归查找复选框
                    checkbox = self.find_checkbox_in_widget(widget)
                    if checkbox:
                        # 检查是否已经勾选
                        if widget.image_index in self.selected_images:
                            checkbox.deselect()  # 取消勾选复选框
                            self.selected_images.discard(widget.image_index)
                            deselected_count += 1
        
        # 更新按钮文本
        self.select_all_btn.config(text="全选")
        
        # 取消非必要的成功提示弹窗
    
    def hide_by_keyword(self):
        """智能隐藏：只隐藏存在相似名称但包含关键词版本的图片"""
        keyword_input = self.hide_keyword_var.get().strip()
        if not keyword_input:
            messagebox.showwarning("提示", "请输入要保留的关键词")
            return
            
        # 保存到历史记录
        self.add_to_history(self.hide_keyword_history, keyword_input)
            
        # 支持多个关键词，以空格分隔
        keywords = [kw.strip() for kw in keyword_input.split() if kw.strip()]
        if not keywords:
            messagebox.showwarning("提示", "请输入有效的关键词")
            return
            
        # 显示进度条
        self.show_progress()
        
        # 使用线程处理隐藏操作，避免界面卡顿
        hide_thread = threading.Thread(target=self.hide_by_keyword_threaded, args=(keywords,))
        hide_thread.daemon = True
        hide_thread.start()
    
    def hide_by_keyword_threaded(self, keywords):
        """线程化的智能隐藏处理"""
        try:
            # 收集所有图片文件名
            all_filenames = []
            widget_map = {}  # 文件名到widget的映射
            
            self.update_progress(10, "正在收集图片信息...")
            
            for widget in self.scrollable_frame.winfo_children():
                if hasattr(widget, 'image_index'):
                    image_path = self.search_results[widget.image_index]
                    filename = os.path.basename(image_path)
                    all_filenames.append(filename)
                    widget_map[filename] = widget
            
            total_files = len(all_filenames)
            if total_files == 0:
                self.update_progress(100, "没有图片需要处理")
                self.root.after(1000, self.hide_progress)
                return
            
            hidden_count = 0
            processed_count = 0
            
            keywords_str = ' '.join(keywords)
            self.update_progress(20, f"正在分析 {total_files} 个图片（关键词：{keywords_str}）...")
            
            # 对每个不包含任何关键词的文件，检查是否存在包含关键词的相似文件
            for i, filename in enumerate(all_filenames):
                # 检查文件名是否包含任何一个关键词
                contains_keyword = any(keyword.lower() in filename.lower() for keyword in keywords)
                
                if not contains_keyword:
                    # 获取不含扩展名的基础名称
                    base_name, ext = os.path.splitext(filename)
                    
                    # 检查是否存在包含关键词的相似文件（忽略扩展名差异）
                    has_similar_with_keyword = False
                    
                    for other_filename in all_filenames:
                         # 检查其他文件是否包含任何一个关键词
                         other_contains_keyword = any(keyword.lower() in other_filename.lower() for keyword in keywords)
                         
                         if other_contains_keyword:
                             other_base_name, other_ext = os.path.splitext(other_filename)
                             
                             # 检查基础名称是否匹配（移除任何关键词后）
                             # 支持多种格式：（x6）、(x6)、 x6、_x6、-x6
                             for keyword in keywords:
                                 patterns_to_remove = [
                                     f"（{keyword}）", f"({keyword})", f" {keyword}", 
                                     f"_{keyword}", f"-{keyword}",
                                     f"（{keyword.upper()}）", f"({keyword.upper()})", f" {keyword.upper()}", 
                                     f"_{keyword.upper()}", f"-{keyword.upper()}",
                                     f"（{keyword.lower()}）", f"({keyword.lower()})", f" {keyword.lower()}", 
                                     f"_{keyword.lower()}", f"-{keyword.lower()}"
                                 ]
                                 
                                 for pattern in patterns_to_remove:
                                     if pattern in other_base_name:
                                         clean_other_name = other_base_name.replace(pattern, "")
                                         if clean_other_name == base_name:
                                             has_similar_with_keyword = True
                                             break
                                 
                                 if has_similar_with_keyword:
                                     break
                             
                             if has_similar_with_keyword:
                                 break
                    
                    # 如果存在相似的包含关键词的文件，则隐藏当前文件
                    if has_similar_with_keyword:
                        widget = widget_map[filename]
                        self.root.after(0, widget.grid_remove)
                        hidden_count += 1
                    else:
                        # 确保显示（可能之前被隐藏了）
                        widget = widget_map[filename]
                        if not widget.grid_info():  # 如果当前被隐藏了，重新显示
                            self.root.after(0, widget.grid)
                else:
                    # 包含关键词的文件确保显示
                    widget = widget_map[filename]
                    if not widget.grid_info():  # 如果当前被隐藏了，重新显示
                        self.root.after(0, widget.grid)
                
                processed_count += 1
                
                # 更新进度 - 每处理10个文件更新一次，并让出CPU时间
                if processed_count % 10 == 0 or processed_count == total_files:
                    progress = 20 + (processed_count / total_files) * 70
                    self.root.after(0, lambda p=progress, pc=processed_count, tc=total_files: 
                                   self.update_progress(p, f"正在处理... ({pc}/{tc})"))
                    # 让出CPU时间，避免界面卡顿
                    import time
                    time.sleep(0.001)
            
            # 完成处理后重新排列可见图片
            self.root.after(0, self.rearrange_visible_images)
            
            self.update_progress(100, "处理完成")
            
            # 显示结果
            if hidden_count > 0:
                # 取消非必要的成功提示弹窗
                pass
            else:
                pass
            
            # 隐藏进度条
            self.root.after(1000, self.hide_progress)
            
        except Exception as e:
            self.update_progress(100, f"处理出错: {str(e)}")
            self.root.after(2000, self.hide_progress)
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理过程中出现错误：{str(e)}"))
        
    def add_to_history(self, history_list, keyword):
        """添加关键词到历史记录"""
        if keyword and keyword not in history_list:
            history_list.insert(0, keyword)
            # 限制历史记录数量
            if len(history_list) > self.max_history_count:
                history_list.pop()
            # 保存配置
            self.save_config()
    
    def rearrange_visible_images(self):
        """重新排列所有可见的图片，消除空隙"""
        visible_widgets = []
        
        # 收集所有可见的图片widget
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                grid_info = widget.grid_info()
                if grid_info:  # 如果有grid_info，说明是可见的
                    visible_widgets.append(widget)
        
        # 按原始索引排序，保持原有顺序
        visible_widgets.sort(key=lambda w: w.image_index)
        
        # 重新排列可见的图片
        row = 0
        col = 0
        for widget in visible_widgets:
            widget.grid(row=row, column=col, sticky=(tk.W, tk.E, tk.N, tk.S), padx=4, pady=4)
            col += 1
            if col >= self.grid_columns:
                col = 0
                row += 1
    
    def show_all_images(self):
        """显示所有隐藏的图片"""
        shown_count = 0
        
        # 遍历所有图片项，重新显示所有图片
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                # 检查是否被隐藏（通过检查grid_info是否为空字典）
                grid_info = widget.grid_info()
                if not grid_info:  # 如果grid_info为空，说明被隐藏了
                    widget.grid()  # 先显示出来，位置稍后重新排列
                    shown_count += 1
        
        # 重新排列所有可见图片，确保布局整齐
        self.rearrange_visible_images()
        
        # 不再清空隐藏关键词输入框，保留用户输入的内容
        # self.hide_keyword_var.set("")  # 注释掉这行，保留输入框内容
        
        # 取消非必要的成功提示弹窗
    
    def find_checkbox_in_widget(self, widget):
        """递归查找组件中的复选框"""
        if isinstance(widget, tk.Checkbutton):
            return widget
        
        for child in widget.winfo_children():
            result = self.find_checkbox_in_widget(child)
            if result:
                return result
        
        return None
    
    def toggle_selection(self, index, selected):
        """切换选择状态"""
        if selected:
            self.selected_images.add(index)
        else:
            self.selected_images.discard(index)
        
        # 动态更新全选按钮状态
        self.update_select_all_button_state()
    
    def update_select_all_button_state(self):
        """根据当前选择状态更新全选按钮文本"""
        if not hasattr(self, 'select_all_btn'):
            return
        
        # 统计可见图片的选择状态
        visible_selected_count = 0
        visible_total_count = 0
        
        for widget in self.scrollable_frame.winfo_children():
            if hasattr(widget, 'image_index'):
                # 检查widget是否可见（未被隐藏）
                if widget.grid_info():  # 只有可见的widget才有grid_info
                    visible_total_count += 1
                    if widget.image_index in self.selected_images:
                        visible_selected_count += 1
        
        # 根据选择状态更新按钮文本
        if visible_selected_count == visible_total_count and visible_total_count > 0:
            self.select_all_btn.config(text="取消全选")
        else:
            self.select_all_btn.config(text="全选")
        
        # 更新高清处理按钮状态
        if hasattr(self, 'upscale_btn'):
            if self.selected_images:
                self.upscale_btn.config(state="normal")
            else:
                self.upscale_btn.config(state="disabled")
            
    def organize_images(self):
        """整理图片到用户配置的导出路径"""
        if not self.selected_images:
            messagebox.showwarning("警告", "请先选择图片")
            return
            
        folder_name = self.folder_name_var.get().strip()
        if not folder_name:
            messagebox.showwarning("警告", "请输入文件夹名称")
            return
            
        try:
            # 记录用户原始输入（作为目标文件夹名）
            final_folder_name = folder_name
            
            # 验证并获取有效的导出路径
            export_dir = self.validate_and_reset_export_path('image')
            target_folder = os.path.join(export_dir, final_folder_name)
            
            # 处理重名：只有当物理文件夹已存在时才添加序号
            counter = 2
            original_target = target_folder
            while os.path.exists(target_folder):
                target_folder = f"{original_target}-{counter}"
                counter += 1
            
            # 更新最终使用的文件夹名（用于记录和下次预设）
            actual_folder_name = os.path.basename(target_folder)
                
            # 创建文件夹
            os.makedirs(target_folder)
            
            # 复制文件
            copied_count = 0
            for item in self.selected_images:
                # 处理数据类型混用问题：既可能是索引(int)也可能是路径(str)
                if isinstance(item, int):
                    # 如果是索引，从search_results中获取路径
                    if item < len(self.search_results):
                        source_path = self.search_results[item]
                    else:
                        continue  # 跳过无效索引
                else:
                    # 如果是路径字符串，直接使用
                    source_path = item
                
                # 确保源文件存在
                if not os.path.exists(source_path):
                    continue
                
                file_name = os.path.basename(source_path)
                target_path = os.path.join(target_folder, file_name)
                
                shutil.copy2(source_path, target_path)
                copied_count += 1
                
            # 直接打开文件夹
            self.open_folder(target_folder)
            
            # 记录到时间轴以便下次 get_smart_name 识别
            self.record_to_timeline(actual_folder_name, 'image_packages')
            
            # 自动学习业务缩写 (从文件夹名中抓取前缀)
            if self.naming_center.get('auto_save_new_items', True):
                import re
                abbrev_match = re.search(r'^([A-Z]{2,4})', actual_folder_name)
                if abbrev_match:
                    new_a = abbrev_match.group(1).upper()
                    if 'business_abbreviations' not in self.naming_center:
                        self.naming_center['business_abbreviations'] = []
                    if new_a and new_a not in self.naming_center['business_abbreviations']:
                        self.naming_center['business_abbreviations'].append(new_a)
                        print(f"✨ 自动学习新业务缩写: {new_a}")
            
            # 生成下一个推荐名称（自动递增序号）
            next_name = self.get_smart_name(actual_folder_name, 'image_packages')
            
            # 更新输入框，为下一次操作做准备
            self.folder_name_var.set(next_name)
            
            # 保存配置
            self.save_config()
                
        except Exception as e:
            messagebox.showerror("错误", f"文件复制失败：{str(e)}")
            
    def on_mousewheel(self, event):
        """优化的鼠标滚轮处理，支持横向和纵向滚动"""
        # 检测操作系统和滚轮方向
        if event.delta:
            # Windows系统
            delta = -1 * (event.delta / 120)
        else:
            # Linux系统
            if event.num == 4:
                delta = -1
            elif event.num == 5:
                delta = 1
            else:
                return
        
        # 优化滚动速度和平滑度
        scroll_speed = 3  # 每次滚动的行数
        
        # 获取事件源组件
        widget = event.widget
        
        # 查找对应的Canvas
        canvas = None
        if hasattr(self, 'canvas') and (widget == self.canvas or widget in self.canvas.winfo_children()):
            canvas = self.canvas
        else:
            # 检查是否是设置页面或其他对话框中的Canvas
            parent = widget
            while parent and canvas is None:
                if isinstance(parent, tk.Canvas):
                    canvas = parent
                    break
                try:
                    parent = parent.master
                except (AttributeError, tk.TclError):
                    break
        
        if canvas:
            # 检查是否按住Shift键进行横向滚动
            if event.state & 0x1:  # Shift键被按下
                canvas.xview_scroll(int(delta * scroll_speed), "units")
            else:
                # 默认纵向滚动
                canvas.yview_scroll(int(delta * scroll_speed), "units")
            
            # 防止滚动过度
            canvas.update_idletasks()
        
    def bind_mousewheel(self, widget):
        """绑定鼠标滚轮事件到指定组件"""
        # Windows和MacOS
        widget.bind("<MouseWheel>", self.on_mousewheel)
        # Linux
        widget.bind("<Button-4>", self.on_mousewheel)
        widget.bind("<Button-5>", self.on_mousewheel)
    
    def on_scrollable_frame_configure(self, event):
        """当滚动框架内容变化时的回调函数"""
        # 更新滚动区域
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        # 检查是否需要显示滚动条
        self.update_scrollbar_visibility()
    
    def on_canvas_configure(self, event):
        """当Canvas大小变化时的回调函数"""
        # 检查是否需要显示滚动条
        self.update_scrollbar_visibility()
    
    def update_scrollbar_visibility(self):
        """更新滚动条的显示状态"""
        try:
            # 获取Canvas的实际大小和滚动区域
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            
            # 获取滚动区域的大小
            scroll_region = self.canvas.cget("scrollregion")
            if scroll_region:
                # 解析滚动区域 "x1 y1 x2 y2"
                x1, y1, x2, y2 = map(float, scroll_region.split())
                content_width = x2 - x1
                content_height = y2 - y1
                
                # 检查是否需要垂直滚动条
                need_v_scrollbar = content_height > canvas_height
                if need_v_scrollbar != self.v_scrollbar_visible:
                    if need_v_scrollbar:
                        self.v_scrollbar.grid(row=2, column=1, sticky=(tk.N, tk.S))
                    else:
                        self.v_scrollbar.grid_remove()
                    self.v_scrollbar_visible = need_v_scrollbar
                
                # 检查是否需要水平滚动条
                need_h_scrollbar = content_width > canvas_width
                if need_h_scrollbar != self.h_scrollbar_visible:
                    if need_h_scrollbar:
                        self.h_scrollbar.grid(row=3, column=0, sticky=(tk.W, tk.E))
                    else:
                        self.h_scrollbar.grid_remove()
                    self.h_scrollbar_visible = need_h_scrollbar
                    
        except Exception as e:
            # 如果出现错误，保持滚动条隐藏状态
            pass
        
    def bind_mousewheel(self, widget):
        """绑定鼠标滚轮事件到指定组件"""
        # Windows和MacOS
        widget.bind("<MouseWheel>", self.on_mousewheel)
        # Linux
        widget.bind("<Button-4>", self.on_mousewheel)
        widget.bind("<Button-5>", self.on_mousewheel)
        
    def load_config(self, apply_geometry=True):
        """加载配置"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.source_folder = config.get('source_folder', '')
                    # 加载窗口尺寸
                    window_geometry = config.get('window_geometry', '1320x800')
                    if apply_geometry:
                        self.root.geometry(window_geometry)
                    # 加载设置窗口尺寸
                    self.settings_window_geometry = config.get('settings_window_geometry', '700x650')
                    # 加载上次使用的文件夹名称
                    self.last_folder_name = config.get('last_folder_name', '')
                    # 加载Excel相关配置
                    if 'size_mapping' in config:
                        self.size_mapping.update(config['size_mapping'])
                    self.sku_column = config.get('sku_column', 'SKU')
                    self.size_column = config.get('size_column', '尺寸')
                    # 加载记忆功能历史记录
                    self.select_keyword_history = config.get('select_keyword_history', [])
                    self.hide_keyword_history = config.get('hide_keyword_history', [])
                    # 加载导出路径配置
                    self.excel_export_path = config.get('excel_export_path', os.path.join(os.path.expanduser("~"), "Desktop"))
                    self.image_export_path = config.get('image_export_path', os.path.join(os.path.expanduser("~"), "Desktop"))
                    
                    # 如果配置文件中的路径为空，使用默认桌面路径
                    if not self.excel_export_path:
                        self.excel_export_path = os.path.join(os.path.expanduser("~"), "Desktop")
                    if not self.image_export_path:
                        self.image_export_path = os.path.join(os.path.expanduser("~"), "Desktop")
                        
                    self.upscale_export_path = config.get('upscale_export_path', '')
                    # 加载BigJPG API密钥
                    self.bigjpg_api_key = config.get('bigjpg_api_key', '')
                    # 加载ImgBB API密钥 (计划书 1.8 核心改进)
                    self.imgbb_api_key = config.get('imgbb_api_key', '5d170edec64cef15aefa2540d93724cc')
                    # 加载高清处理配置
                    self.upscale_config.update(config.get('upscale_config', {}))
                    # 加载智能高清处理配置
                    self.smart_upscale_config.update(config.get('smart_upscale_config', {}))
                    # 加载上次手动处理弹窗的输入内容
                    self.last_manual_processing = config.get('last_manual_processing', {'processing': '', 'material': ''})
                    # 加载上次选择的预设
                    self.last_selected_preset = config.get('last_selected_preset', '')
                    # 加载上次表格导出名称
                    self.last_excel_export_name = config.get('last_excel_export_name', '')
                    # 加载自动搜索开关配置
                    self.auto_search_enabled = config.get('auto_search_enabled', False)
                    # 加载名称序号记录数据（新的数据结构：{基础名称: 最后序号}）
                    self.name_sequence_records = config.get('name_sequence_records', {})
                    # 每次启动时清理过期的名称序号记录，只保留今天的记录
                    try:
                        removed = self._cleanup_old_name_sequence_records()
                        if removed:
                            print(f"已清理 {removed} 条过期 name_sequence_records")
                    except Exception as e:
                        print(f"清理旧记录时出错: {e}")
                        
                    # 加载模块化命名配置
                    self.naming_center = config.get('naming_center', {
                        "business_abbreviations": ["CHX", "HX"],
                        "logistics_providers": ["Y2尊祐", "Y1尚为"],
                        "last_logistics_provider": "Y2尊祐",
                        "custom_suffixes": ["艺术家", "画家", "设计师"],
                        "last_custom_suffix": "艺术家",
                        "naming_template": "{prefix}-{date}-{provider}-{product}-{suffix}",
                        "waybill_template": "{abbreviation}-{date}-{tickets}票-Y2面单-{merchant}-{product}",
                        "auto_save_new_items": True
                    })
                    
                    # 确保 business_abbreviations 存在 (向下兼容)
                    if 'business_abbreviations' not in self.naming_center:
                        if 'business_abbreviation' in self.naming_center:
                            self.naming_center['business_abbreviations'] = [self.naming_center['business_abbreviation']]
                        else:
                            self.naming_center['business_abbreviations'] = ["CHX", "HX"]
        except Exception:
            pass
            
    def save_config(self):
        """保存配置 - 使用原子性写入"""
        try:
            # 获取当前窗口尺寸
            window_geometry = self.root.geometry()
            # 获取当前文件夹名称
            current_folder_name = self.folder_name_var.get().strip()
            
            config = {
                'source_folder': self.source_folder,
                'window_geometry': window_geometry,
                'settings_window_geometry': getattr(self, 'settings_window_geometry', '700x650'),
                'last_folder_name': current_folder_name if current_folder_name else self.last_folder_name,
                'size_mapping': self.size_mapping,
                'sku_column': self.sku_column,
                'size_column': self.size_column,
                'select_keyword_history': self.select_keyword_history,
                'hide_keyword_history': self.hide_keyword_history,
                'excel_export_path': self.excel_export_path,
                'image_export_path': self.image_export_path,
                'upscale_export_path': self.upscale_export_path,
                'bigjpg_api_key': getattr(self, 'bigjpg_api_key', ''),
                'imgbb_api_key': getattr(self, 'imgbb_api_key', '5d170edec64cef15aefa2540d93724cc'),
                'upscale_config': getattr(self, 'upscale_config', {
                    'style': 'art',
                    'noise': '1', 
                    'x2': '1'
                }),
                'smart_upscale_config': getattr(self, 'smart_upscale_config', {
                    'target_width': 8000,
                    'target_height': 8000,
                    'skip_qualified': True,
                    'enabled': True
                }),
                'last_manual_processing': getattr(self, 'last_manual_processing', {'processing': '', 'material': ''}),
                'last_selected_preset': getattr(self, 'last_selected_preset', ''),
                'last_excel_export_name': getattr(self, 'last_excel_export_name', ''),
                'auto_search_enabled': getattr(self, 'auto_search_enabled', False),
                'name_sequence_records': getattr(self, 'name_sequence_records', {}),
                'naming_center': getattr(self, 'naming_center', {
                    "business_abbreviation": "CHX",
                    "logistics_providers": ["Y2尊祐", "Y1尚为"],
                    "last_logistics_provider": "Y2尊祐",
                    "custom_suffixes": ["艺术家", "画家", "设计师"],
                    "last_custom_suffix": "艺术家",
                    "naming_template": "{prefix}-{date}-{provider}-{product}-{suffix}",
                    "waybill_template": "{abbreviation}-{date}-{tickets}票-Y2面单-{merchant}-{product}",
                    "auto_save_new_items": True
                })
            }
            
            # 使用临时文件实现原子性写入
            temp_file = self.config_file + '.tmp'
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            # 原子性重命名
            if os.path.exists(self.config_file):
                os.replace(temp_file, self.config_file)
            else:
                os.rename(temp_file, self.config_file)
                
        except Exception as e:
            # 清理临时文件
            temp_file = self.config_file + '.tmp'
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except (OSError, IOError):
                    pass
            print(f"保存配置失败: {e}")
    

    # --------------------------------------------------
    # 新增功能：每日启动清理旧的 name_sequence_records 记录
    # --------------------------------------------------
    def _cleanup_old_name_sequence_records(self):
        """删除非今日日期的 name_sequence_records 条目，返回删除数量"""
        import datetime
        today_str = datetime.date.today().strftime("%m-%d")
        if not hasattr(self, 'name_sequence_records'):
            return 0
        keys_to_delete = [k for k in list(self.name_sequence_records.keys()) if f'-{today_str}-' not in k]
        for k in keys_to_delete:
            self.name_sequence_records.pop(k, None)
        return len(keys_to_delete)

    def _remove_sequence_number(self, name):
        """
        移除名称中的序号，保留基础名称和特殊后缀
        """
        import re
        
        # 检查是否有特殊后缀 (从配置中动态获取)
        special_suffixes = []
        if hasattr(self, 'naming_center') and 'custom_suffixes' in self.naming_center:
            # 为每个自定义后缀加上横杠
            special_suffixes = [f"-{s}" for s in self.naming_center['custom_suffixes']]
        
        # 默认回退列表
        if not special_suffixes:
            special_suffixes = ['-艺术家', '-画家', '-设计师']
            
        suffix = ''
        name_part = name
        
        # 优先匹配较长的后缀，防止部分匹配
        for special_suffix in sorted(special_suffixes, key=len, reverse=True):
            if name.endswith(special_suffix):
                suffix = special_suffix
                name_part = name[:-len(special_suffix)]
                break
        
        # 移除序号：匹配末尾的 -数字 模式
        if suffix:
            # 有特殊后缀的情况：从name_part中移除序号
            cleaned_name_part = re.sub(r'-\d+$', '', name_part)
            return cleaned_name_part + suffix
        else:
            # 没有特殊后缀的情况：直接从name中移除序号
            return re.sub(r'-\d+$', '', name)
    
    def get_smart_name(self, name, export_type):
        """
        智能名称处理函数 - 基于基础名称的最后序号生成新序号
        
        Args:
            name: 原始名称
            export_type: 导出类型，'excel_exports' 或 'image_packages'
        
        Returns:
            处理后的智能名称
        """
        import re
        
        # 确保name_sequence_records存在
        if not hasattr(self, 'name_sequence_records'):
            self.name_sequence_records = {}
        
        # 提取基础名称（移除可能存在的序号）
        base_name = self._remove_sequence_number(name)
        
        # 构建记录键（基础名称 + 导出类型）
        record_key = f"{base_name}#{export_type}"
        
        # 获取该基础名称的最后序号
        last_sequence = self.name_sequence_records.get(record_key, 0)
        
        # 生成下一个序号
        next_sequence = last_sequence + 1
        
        # 如果是第一次（序号为1），直接返回基础名称
        if next_sequence == 1:
            return base_name
        
        # 否则返回带序号的名称
        return f"{base_name}-{next_sequence}"
    
    def record_to_timeline(self, name, export_type):
        """
        记录基础名称和对应的序号
        
        Args:
            name: 导出/打包的名称
            export_type: 导出类型，'excel_exports' 或 'image_packages'
        """
        import re
        
        # 确保name_sequence_records存在
        if not hasattr(self, 'name_sequence_records'):
            self.name_sequence_records = {}
        
        # 提取基础名称和序号
        base_name = self._remove_sequence_number(name)
        
        # 提取序号（如果存在）
        sequence_match = re.search(r'-(\d+)$', name)
        if sequence_match:
            sequence_number = int(sequence_match.group(1))
        else:
            sequence_number = 1  # 如果没有序号，认为是第一次
        
        # 构建记录键（基础名称 + 导出类型）
        record_key = f"{base_name}#{export_type}"
        
        # 更新该基础名称的最后序号
        self.name_sequence_records[record_key] = sequence_number
        
        # 保存配置
        self.save_config()
    
    def on_closing(self):
        """窗口关闭时保存配置并清空隐藏关键词输入框"""
        # 在程序关闭时清空隐藏关键词输入框
        if hasattr(self, 'hide_keyword_var'):
            self.hide_keyword_var.set("")
        
        # 清理线程池资源
        try:
            if hasattr(self, 'thumbnail_executor') and self.thumbnail_executor:
                self.thumbnail_executor.shutdown(wait=False)
            if hasattr(self, 'info_executor') and self.info_executor:
                self.info_executor.shutdown(wait=False)
            if hasattr(self, 'search_executor') and self.search_executor:
                self.search_executor.shutdown(wait=False)
        except Exception as e:
            print(f"清理线程池时出错: {e}")
        
        self.save_config()
        self.root.destroy()
    
    def open_folder(self, folder_path):
        """跨平台打开文件夹"""
        import platform
        import subprocess
        
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(folder_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", folder_path])
            elif system == "Linux":
                subprocess.run(["xdg-open", folder_path])
            else:
                # 如果无法识别系统，尝试使用默认方法
                os.startfile(folder_path)
        except Exception as e:
            messagebox.showerror("错误", f"无法打开文件夹: {str(e)}")
    
    def start_upscale_process(self):
        """开始高清处理流程"""
        # 获取选中的图片索引
        selected_indices = list(self.selected_images)
        
        if not selected_indices:
            messagebox.showwarning("提示", "请先选择要进行高清处理的图片")
            return
        
        # 检查是否启用智能模式
        if self.smart_upscale_config.get('enabled', True):
            try:
                # 分析处理计划
                plan_data = self.analyze_upscale_batch(selected_indices)
                
                # 如果有需要处理的图片，显示智能计划对话框
                if plan_data['statistics']['to_process'] > 0:
                    self.show_smart_upscale_plan_dialog(plan_data)
                else:
                    # 所有图片都已达标
                    messagebox.showinfo("提示", "所有选中的图片都已达到目标尺寸，无需处理")
                    
            except Exception as e:
                print(f"智能分析失败: {e}")
                # 智能分析失败，回退到传统模式
                messagebox.showwarning("提示", "智能分析失败，将使用传统配置模式")
                selected_images = [self.search_results[i] for i in selected_indices if i < len(self.search_results)]
                self.show_upscale_config_dialog(selected_images)
        else:
            # 传统模式
            selected_images = [self.search_results[i] for i in selected_indices if i < len(self.search_results)]
            self.show_upscale_config_dialog(selected_images)
    
    def show_smart_upscale_plan_dialog(self, plan_data):
        """显示智能高清处理计划对话框"""
        try:
            from smart_upscale_plan_dialog import show_smart_upscale_plan_dialog
            show_smart_upscale_plan_dialog(
                parent=self.root,
                plan_data=plan_data,
                start_callback=lambda modified_data: self.start_smart_upscale_processing(modified_data)
            )
        except Exception as e:
            print(f"显示智能计划对话框失败: {e}")
            messagebox.showerror("错误", f"显示处理计划时出错: {str(e)}")
    
    def start_smart_upscale_processing(self, plan_data):
        """开始智能高清处理"""
        try:
            processing_list = plan_data.get('processing_list', [])
            
            if not processing_list:
                messagebox.showinfo("提示", "没有需要处理的图片")
                return
            
            # 检查API密钥
            if not hasattr(self, 'bigjpg_api_key') or not self.bigjpg_api_key:
                messagebox.showerror("错误", "请先在设置页面配置BigJPG API密钥")
                return
            
            # 保存处理列表供后续使用
            self.pending_processing_list = processing_list
            
            # 直接开始处理，不显示配置窗口
            self.start_smart_upscale_processing_direct(processing_list)
            
        except Exception as e:
            print(f"启动智能处理失败: {e}")
            messagebox.showerror("错误", f"启动智能处理时出错: {str(e)}")
    
    def start_smart_upscale_processing_direct(self, processing_list):
        """直接开始智能高清处理（版本1.6内置进度窗口）"""
        # 检查API密钥
        if not self.bigjpg_api_key:
            messagebox.showerror("错误", "BigJPG API密钥未设置")
            return
        
        if not self.imgbb_api_key:
            messagebox.showerror("错误", "ImgBB API密钥未设置")
            return
        
        # 保存处理列表
        self.processing_list = processing_list
        
        # 创建进度窗口
        total_count = len(processing_list)
        self.create_upscale_progress_window(total_count)
        
        # 在后台线程中处理
        def process_images():
            try:
                upscaler = BigJPGUpscaler(self.bigjpg_api_key)  # 只传递BigJPG API密钥，使用默认base_url
                self.upscale_tasks = {}
                self.completed_tasks = 0
                self.total_tasks = total_count
                
                # 第一阶段：上传所有图片
                self.update_upscale_progress(0, f"开始上传 {self.total_tasks} 张图片...", 
                                           stage="上传图片", upload_count=0, process_count=0, 
                                           download_count=0, total_count=self.total_tasks)
                
                for i, image_info in enumerate(processing_list):
                    try:
                        # 智能模式：image_info 是包含路径和倍数的字典
                        image_path = image_info['path']
                        upscale_factor = image_info['factor']
                        
                        # 根据倍数设置x2参数 (BigJPG API: 1=2x, 2=4x, 3=8x, 4=16x)
                        factor_to_x2 = {2: '1', 4: '2', 8: '3', 16: '4'}
                        x2_value = factor_to_x2.get(upscale_factor, '2')  # 默认4倍
                        current_params = {
                            'style': self.smart_upscale_config.get('default_type', 'art'),
                            'noise': self.upscale_config.get('noise', '-1'),
                            'x2': x2_value
                        }
                        
                        # 更新上传进度
                        self.update_upscale_progress(
                            (i / self.total_tasks) * 30,  # 上传阶段占30%
                            f"正在上传: {os.path.basename(image_path)} ({i+1}/{self.total_tasks})",
                            stage="上传图片", upload_count=i+1, process_count=0, 
                            download_count=0, total_count=self.total_tasks
                        )
                        
                        # 上传图片
                        print(f"发送给BigJPG API的参数: style={current_params['style']}, noise={current_params['noise']}, x2={current_params['x2']}")
                        result = upscaler.upload_image(
                            image_path, 
                            style=current_params['style'],
                            noise=current_params['noise'], 
                            x2=current_params['x2']
                        )
                        
                        # 检查是否有错误
                        if 'error' in result:
                            error_msg = result['error']
                            print(f"[高清处理错误] 上传失败: {error_msg}")
                            
                            # 在进度条中显示错误信息
                            self.update_upscale_progress(
                                (i / self.total_tasks) * 30,
                                f"上传失败: {os.path.basename(image_path)}",
                                stage="上传失败", 
                                upload_count=i, process_count=0, 
                                download_count=0, total_count=self.total_tasks,
                                error_message=error_msg
                            )
                            
                            # 显示错误弹窗
                            self.root.after(0, lambda msg=error_msg: 
                                          messagebox.showerror("上传失败", f"图片上传失败: {msg}"))
                            return  # 停止处理
                        else:
                            # 检查BigJPG API的响应格式
                            print(f"处理图片 {os.path.basename(image_path)} 的API响应: {result}")
                            
                            # 尝试多种可能的任务ID字段名，优先使用tid
                            task_id = None
                            for field in ['tid', 'task_id', 'id', 'taskId', 'task', 'uuid']:
                                if field in result:
                                    task_id = result[field]
                                    break
                            
                            if task_id:
                                self.upscale_tasks[task_id] = {
                                    'original_path': image_path,
                                    'status': 'uploaded',
                                    'progress': 0
                                }
                                print(f"成功获取任务ID: {task_id}")
                            else:
                                # 如果没有找到任务ID，检查是否有状态字段
                                if 'status' in result:
                                    status = result['status']
                                    if status == 'param_error':
                                        error_msg = "参数错误，请检查API密钥和图片URL"
                                    else:
                                        error_msg = f"API返回状态: {status}"
                                else:
                                    error_msg = f"无法获取任务ID，完整响应: {result}"
                                
                                self.root.after(0, lambda msg=error_msg: 
                                              messagebox.showerror("上传失败", msg))
                    
                    except Exception as e:
                        self.root.after(0, lambda err=str(e): 
                                      messagebox.showerror("上传错误", f"上传过程中出错: {err}"))
                
                if not self.upscale_tasks:
                    self.root.after(0, lambda: messagebox.showerror("错误", "没有成功上传的图片"))
                    self.close_upscale_progress()
                    return
                
                # 第二阶段：监控处理进度
                self.update_upscale_progress(30, "图片上传完成，开始处理...", 
                                           stage="云端处理", upload_count=self.total_tasks, 
                                           process_count=0, download_count=0, total_count=self.total_tasks)
                self.monitor_upscale_progress(upscaler)
                
            except Exception as e:
                self.root.after(0, lambda err=str(e): 
                              messagebox.showerror("处理错误", f"处理过程中出错: {err}"))
                self.close_upscale_progress()
        
        # 启动后台线程
        threading.Thread(target=process_images, daemon=True).start()
    
    def show_upscale_config_dialog(self, selected_images):
        """显示高清处理配置对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("放大配置")
        dialog.geometry("500x450")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.attributes('-topmost', True)
        dialog.focus_force()
        dialog.configure(bg='#f0f0f0')
        
        # 立即隐藏窗口，避免在左上角显示
        dialog.withdraw()
        
        # 居中显示 - 先更新布局但窗口仍然隐藏
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # 设置好位置后再显示窗口，避免移动效果
        dialog.deiconify()
        
        # 主框架
        main_frame = tk.Frame(dialog, bg='#f0f0f0', padx=30, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 图片类型选择
        type_frame = tk.Frame(main_frame, bg='#f0f0f0')
        type_frame.pack(fill=tk.X, pady=(0, 20))
        
        type_label = tk.Label(type_frame, text="图片类型", bg='#f0f0f0', 
                             font=('Microsoft YaHei UI', 12, 'bold'))
        type_label.pack(anchor=tk.W, pady=(0, 10))
        
        style_var = tk.StringVar(value=self.upscale_config.get('style', 'art'))
        type_radio_frame = tk.Frame(type_frame, bg='#f0f0f0')
        type_radio_frame.pack(anchor=tk.W)
        
        art_radio = tk.Radiobutton(type_radio_frame, text="卡通 / 插画", 
                                  variable=style_var, value="art", bg='#f0f0f0',
                                  font=('Microsoft YaHei UI', 10))
        art_radio.pack(side=tk.LEFT, padx=(0, 30))
        
        photo_radio = tk.Radiobutton(type_radio_frame, text="照片", 
                                    variable=style_var, value="photo", bg='#f0f0f0',
                                    font=('Microsoft YaHei UI', 10))
        photo_radio.pack(side=tk.LEFT)
        
        # 放大倍数选择
        scale_frame = tk.Frame(main_frame, bg='#f0f0f0')
        scale_frame.pack(fill=tk.X, pady=(0, 20))
        
        scale_label = tk.Label(scale_frame, text="放大倍数", bg='#f0f0f0',
                              font=('Microsoft YaHei UI', 12, 'bold'))
        scale_label.pack(anchor=tk.W, pady=(0, 10))
        
        x2_var = tk.StringVar(value=self.upscale_config.get('x2', '1'))
        scale_radio_frame = tk.Frame(scale_frame, bg='#f0f0f0')
        scale_radio_frame.pack(anchor=tk.W)
        
        for i, (value, text) in enumerate([("1", "2x"), ("2", "4x"), ("3", "8x"), ("4", "16x")]):
            radio = tk.Radiobutton(scale_radio_frame, text=text, 
                                  variable=x2_var, value=value, bg='#f0f0f0',
                                  font=('Microsoft YaHei UI', 10))
            radio.pack(side=tk.LEFT, padx=(0, 30) if i < 3 else (0, 0))
        
        # 降噪程度选择
        noise_frame = tk.Frame(main_frame, bg='#f0f0f0')
        noise_frame.pack(fill=tk.X, pady=(0, 20))
        
        noise_label = tk.Label(noise_frame, text="降噪程度", bg='#f0f0f0',
                              font=('Microsoft YaHei UI', 12, 'bold'))
        noise_label.pack(anchor=tk.W, pady=(0, 10))
        
        noise_var = tk.StringVar(value=self.upscale_config.get('noise', '1'))
        noise_radio_frame = tk.Frame(noise_frame, bg='#f0f0f0')
        noise_radio_frame.pack(anchor=tk.W)
        
        noise_options = [("-1", "无"), ("0", "低"), ("1", "中"), ("2", "高"), ("3", "最高")]
        for i, (value, text) in enumerate(noise_options):
            radio = tk.Radiobutton(noise_radio_frame, text=text, 
                                  variable=noise_var, value=value, bg='#f0f0f0',
                                  font=('Microsoft YaHei UI', 10))
            radio.pack(side=tk.LEFT, padx=(0, 20) if i < 4 else (0, 0))
        

        
        # 选中图片数量提示
        count_label = tk.Label(main_frame, 
                              text=f"已选择 {len(selected_images)} 张图片进行高清处理",
                              bg='#f0f0f0', font=('Microsoft YaHei UI', 10))
        count_label.pack(pady=(10, 20))
        
        # 按钮区域
        button_frame = tk.Frame(main_frame, bg='#f0f0f0')
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        def on_cancel():
            dialog.destroy()
        
        # 确定按钮 - 开始处理按钮
        confirm_btn = tk.Button(button_frame, 
                               text="🚀 开始处理",
                               command=on_start,
                               bg='#28a745',
                               fg='white', 
                               font=('Microsoft YaHei UI', 11, 'bold'),
                               padx=30, pady=12, 
                               relief=tk.FLAT, 
                               cursor='hand2')
        confirm_btn.pack(side=tk.RIGHT, padx=(15, 0))
        
        # 取消按钮
        cancel_btn = tk.Button(button_frame, text="取消", command=on_cancel,
                              bg='#6c757d', fg='white', font=('Microsoft YaHei UI', 10),
                              padx=20, pady=10, relief=tk.FLAT, cursor='hand2')
        cancel_btn.pack(side=tk.RIGHT, padx=(0, 10))
    
    def start_async_upscale(self, selected_images, params=None, processing_plan=None):
        """开始异步高清处理
        
        Args:
            selected_images: 选中的图片列表
            params: 处理参数（传统模式）
            processing_plan: 智能处理计划（智能模式）
        """
        # 初始化计数器和错误标记
        self._current_download_count = 0
        self._download_error_shown = False
        
        # 创建BigJPG处理器
        if not self.bigjpg_api_key:
            messagebox.showerror("错误", "API密钥未设置")
            return
        
        upscaler = BigJPGUpscaler(self.bigjpg_api_key)  # 只传递BigJPG API密钥
        
        # 确定处理模式和参数
        if processing_plan:
            # 智能处理模式
            images_to_process = processing_plan['to_process']
            smart_config = self.smart_upscale_config
            
            # 从智能配置生成处理参数
            params = {
                'style': smart_config.get('default_type', 'art'),
                'noise': self.upscale_config.get('noise', '-1'),
                'x2': '1'  # 智能模式下，倍数由计划决定
            }
            
            # 创建进度窗口
            self.create_upscale_progress_window(len(images_to_process))
            
            # 显示智能处理信息
            self.update_upscale_progress(0, f"智能处理模式：{len(images_to_process)} 张图片需要处理")
        else:
            # 传统处理模式
            images_to_process = selected_images
            if not params:
                messagebox.showerror("错误", "处理参数未设置")
                return
            
            # 创建进度窗口
            self.create_upscale_progress_window(len(images_to_process))
        
        # 在后台线程中处理
        def process_images():
            try:
                self.upscale_tasks = {}
                self.completed_tasks = 0
                self.total_tasks = len(images_to_process)
                
                # 第一阶段：上传所有图片
                self.update_upscale_progress(0, f"开始上传 {self.total_tasks} 张图片...", 
                                           stage="上传图片", upload_count=0, process_count=0, 
                                           download_count=0, total_count=self.total_tasks)
                
                for i, image_info in enumerate(images_to_process):
                    try:
                        # 获取图片路径和处理参数
                        if processing_plan:
                            # 智能模式：image_info 是包含路径和倍数的字典
                            image_path = image_info['path']
                            upscale_factor = image_info['factor']
                            
                            # 根据倍数设置x2参数 (BigJPG API: 1=2x, 2=4x, 3=8x, 4=16x)
                            factor_to_x2 = {2: '1', 4: '2', 8: '3', 16: '4'}
                            x2_value = factor_to_x2.get(upscale_factor, '2')  # 默认4倍
                            current_params = {
                                'style': params['style'],
                                'noise': params['noise'],
                                'x2': x2_value
                            }
                        else:
                            # 传统模式：image_info 就是图片路径
                            image_path = image_info
                            current_params = params
                        
                        # 更新上传进度
                        self.update_upscale_progress(
                            (i / self.total_tasks) * 30,  # 上传阶段占30%
                            f"正在上传: {os.path.basename(image_path)} ({i+1}/{self.total_tasks})",
                            stage="上传图片", upload_count=i+1, process_count=0, 
                            download_count=0, total_count=self.total_tasks
                        )
                        
                        # 上传图片
                        print(f"发送给BigJPG API的参数: style={current_params['style']}, noise={current_params['noise']}, x2={current_params['x2']}")
                        result = upscaler.upload_image(
                            image_path, 
                            style=current_params['style'],
                            noise=current_params['noise'], 
                            x2=current_params['x2']
                        )
                        
                        # 检查是否有错误
                        if 'error' in result:
                            error_msg = result['error']
                            print(f"[高清处理错误] 上传失败: {error_msg}")
                            
                            # 在进度条中显示错误信息
                            self.update_upscale_progress(
                                (i / self.total_tasks) * 30,
                                f"上传失败: {os.path.basename(image_path)}",
                                stage="上传失败", 
                                upload_count=i, process_count=0, 
                                download_count=0, total_count=self.total_tasks,
                                error_message=error_msg
                            )
                            
                            # 显示错误弹窗
                            self.root.after(0, lambda msg=error_msg: 
                                          messagebox.showerror("上传失败", f"图片上传失败: {msg}"))
                            return  # 停止处理
                        else:
                            # 检查BigJPG API的响应格式
                            print(f"处理图片 {os.path.basename(image_path)} 的API响应: {result}")
                            
                            # 尝试多种可能的任务ID字段名，优先使用tid
                            task_id = None
                            for field in ['tid', 'task_id', 'id', 'taskId', 'task', 'uuid']:
                                if field in result:
                                    task_id = result[field]
                                    break
                            
                            if task_id:
                                self.upscale_tasks[task_id] = {
                                    'original_path': image_path,
                                    'status': 'uploaded',
                                    'progress': 0
                                }
                                print(f"成功获取任务ID: {task_id}")
                            else:
                                # 如果没有找到任务ID，检查是否有状态字段
                                if 'status' in result:
                                    status = result['status']
                                    if status == 'param_error':
                                        error_msg = "参数错误，请检查API密钥和图片URL"
                                    else:
                                        error_msg = f"API返回状态: {status}"
                                else:
                                    error_msg = f"无法获取任务ID，完整响应: {result}"
                                
                                self.root.after(0, lambda msg=error_msg: 
                                              messagebox.showerror("上传失败", msg))
                    
                    except Exception as e:
                        self.root.after(0, lambda err=str(e): 
                                      messagebox.showerror("上传错误", f"上传过程中出错: {err}"))
                
                if not self.upscale_tasks:
                    self.root.after(0, lambda: messagebox.showerror("错误", "没有成功上传的图片"))
                    self.close_upscale_progress()
                    return
                
                # 第二阶段：监控处理进度
                self.update_upscale_progress(30, "图片上传完成，开始处理...", 
                                           stage="云端处理", upload_count=self.total_tasks, 
                                           process_count=0, download_count=0, total_count=self.total_tasks)
                self.monitor_upscale_progress(upscaler)
                
            except Exception as e:
                self.root.after(0, lambda err=str(e): 
                              messagebox.showerror("处理错误", f"处理过程中出错: {err}"))
                self.close_upscale_progress()
        
        # 启动后台线程
        threading.Thread(target=process_images, daemon=True).start()
    
    def monitor_upscale_progress(self, upscaler):
        """监控高清处理进度"""
        
        def check_progress():
            """检查处理进度的内部函数"""
            try:
                if not hasattr(self, 'upscale_tasks') or not self.upscale_tasks:
                    self.close_upscale_progress()
                    return
                
                # 初始化已下载任务集合
                if not hasattr(self, '_downloaded_tasks'):
                    self._downloaded_tasks = set()
                
                # 动态获取仍需监控的任务（排除已下载的）
                pending_task_ids = [task_id for task_id in self.upscale_tasks.keys() 
                                  if task_id not in self._downloaded_tasks]
                
                # 如果没有待监控的任务，直接退出
                if not pending_task_ids:
                    print("所有任务已完成并下载，退出监控")
                    total_tasks = getattr(self, 'total_tasks', len(self.upscale_tasks))
                    self.update_upscale_progress(
                        100, 
                        "所有图片处理完成！",
                        stage="完成",
                        upload_count=total_tasks,
                        process_count=total_tasks,
                        download_count=len(self._downloaded_tasks),
                        total_count=total_tasks
                    )
                    self.root.after(3000, self.close_upscale_progress)
                    return
                
                # 只查询仍在处理中的任务状态
                print(f"监控中的任务: {pending_task_ids} (共{len(pending_task_ids)}个)")
                status_result = upscaler.check_status(pending_task_ids)
                print(f"任务状态查询结果: {status_result}")
                
                completed_count = 0
                processing_count = 0
                failed_count = 0
                
                # 处理状态查询结果
                if isinstance(status_result, dict) and not status_result.get('error'):
                    # 确保已下载任务集合存在
                    if not hasattr(self, '_downloaded_tasks'):
                        self._downloaded_tasks = set()
                    
                    # 确保已完成任务集合存在（包括成功和失败的任务）
                    if not hasattr(self, '_completed_tasks'):
                        self._completed_tasks = set()
                    
                    # 检查每个任务的状态
                    for task_id, task_info in status_result.items():
                        if task_id in self.upscale_tasks and task_id not in self._completed_tasks:
                            status = task_info.get('status', 'unknown')
                            progress = task_info.get('progress', 0)
                            
                            if status == 'success' or status == 'finished':
                                # 任务完成，立即下载
                                download_url = task_info.get('download_url') or task_info.get('url')
                                if download_url:
                                    print(f"任务完成，开始下载: {task_id}, URL: {download_url}")
                                    try:
                                        task_data = self.upscale_tasks.get(task_id)
                                        if isinstance(task_data, dict):
                                            original_path = task_data.get('original_path')
                                        elif isinstance(task_data, str):
                                            original_path = task_data
                                        else:
                                            print(f"警告: 任务 {task_id} 的数据格式异常")
                                            self._completed_tasks.add(task_id)  # 标记为已完成（失败）
                                            failed_count += 1
                                            continue
                                        
                                        if original_path and os.path.exists(original_path):
                                            # 立即下载并添加到已下载集合
                                            self.download_upscaled_image(download_url, original_path, 
                                                                       len(self._downloaded_tasks) + 1, 
                                                                       len(self.upscale_tasks), completed_count)
                                            self._downloaded_tasks.add(task_id)
                                            self._completed_tasks.add(task_id)  # 标记为已完成（成功）
                                            completed_count += 1
                                    except Exception as e:
                                        print(f"下载任务 {task_id} 时出错: {e}")
                                        self._completed_tasks.add(task_id)  # 标记为已完成（失败）
                                        failed_count += 1
                                else:
                                    print(f"警告：任务 {task_id} 完成但缺少下载URL")
                                    self._completed_tasks.add(task_id)  # 标记为已完成（失败）
                                    failed_count += 1
                                    
                            elif status == 'processing' or status == 'waiting':
                                # 任务正在处理中
                                processing_count += 1
                                if isinstance(self.upscale_tasks[task_id], dict):
                                    self.upscale_tasks[task_id]['status'] = status
                                    self.upscale_tasks[task_id]['progress'] = progress
                                    
                            elif status == 'failed' or status == 'error':
                                # 任务失败，直接标记为已完成
                                print(f"任务 {task_id} 处理失败")
                                self._completed_tasks.add(task_id)  # 标记为已完成（失败）
                                failed_count += 1
                
                # 计算总体完成情况
                total_tasks = getattr(self, 'total_tasks', len(self.upscale_tasks))
                total_completed = len(self._downloaded_tasks)  # 成功下载的任务数
                total_finished = len(self._completed_tasks)    # 已完成的任务数（包括成功和失败）
                remaining_tasks = total_tasks - total_finished  # 使用已完成任务数计算剩余任务
                
                if remaining_tasks == 0:
                    # 所有任务完成
                    total_failed = total_finished - total_completed  # 失败任务数 = 已完成任务数 - 成功任务数
                    if total_failed > 0:
                        self.update_upscale_progress(100, f"处理完成！成功: {total_completed}, 失败: {total_failed}", 
                                                   stage="完成", upload_count=total_tasks, 
                                                   process_count=total_tasks, download_count=total_completed, 
                                                   total_count=total_tasks)
                        self.root.after(5000, self.close_upscale_progress)  # 5秒后关闭，确保图片加载完成
                    else:
                        self.update_upscale_progress(100, f"全部处理完成！共处理 {total_completed} 张图片", 
                                                   stage="完成", upload_count=total_tasks, 
                                                   process_count=total_tasks, download_count=total_completed, 
                                                   total_count=total_tasks)
                        self.root.after(4000, self.close_upscale_progress)  # 4秒后关闭，确保图片加载完成
                else:
                    # 继续监控
                    progress_percent = 30 + (total_finished / total_tasks) * 70  # 30%用于上传，70%用于处理和下载
                    
                    # 构建详细状态文本
                    status_parts = []
                    if completed_count > 0:
                        status_parts.append(f"{completed_count} 新完成")
                    if processing_count > 0:
                        status_parts.append(f"{processing_count} 处理中")
                    if failed_count > 0:
                        status_parts.append(f"{failed_count} 失败")
                    
                    status_text = f"处理中... ({'/'.join(status_parts)}, 剩余{remaining_tasks}个)"
                    
                    self.update_upscale_progress(progress_percent, 
                                               status_text, 
                                               stage="云端处理", upload_count=total_tasks, 
                                               process_count=total_completed, download_count=total_completed, 
                                               total_count=total_tasks)
                    
                    # 继续检查
                    self.root.after(2000, check_progress)  # 2秒后再次检查
                    
            except Exception as e:
                print(f"监控进度时出错: {e}")
                self.root.after(0, lambda: messagebox.showerror("监控错误", f"监控处理进度时出错: {e}"))
                self.close_upscale_progress()
        
        # 开始监控
        self.root.after(1000, check_progress)  # 1秒后开始检查
    
    def download_upscaled_image_sync(self, download_url, original_path, task_id):
        """同步下载放大后的图片"""
        try:
            import requests
            import os
            
            # 获取原始文件信息
            original_dir = os.path.dirname(original_path)
            original_name, original_ext = os.path.splitext(os.path.basename(original_path))
            
            # 生成高清图片文件名
            rename_rule = self.upscale_config.get('rename_rule', '（高清）')
            upscaled_filename = f"{original_name}{rename_rule}{original_ext}"
            upscaled_path = os.path.join(original_dir, upscaled_filename)
            
            # 确保目录存在
            os.makedirs(original_dir, exist_ok=True)
            
            # 下载图片
            print(f"开始下载高清图片: {download_url}")
            response = requests.get(download_url, stream=True, timeout=30)
            response.raise_for_status()
            
            # 保存图片
            with open(upscaled_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # 验证文件完整性并添加到搜索结果
            if os.path.exists(upscaled_path) and os.path.getsize(upscaled_path) > 0:
                print(f"高清图片下载完成: {upscaled_path}")
                
                # 使用验证方法确保文件完整性后再添加到搜索结果
                self._verify_and_add_upscaled_image(upscaled_path)
                
                return True
            else:
                print(f"下载的文件无效: {upscaled_path}")
                return False
                
        except Exception as e:
            print(f"下载高清图片时出错: {e}")
            return False
    
    def download_upscaled_image(self, download_url, original_path, current_index=None, total_count=None, download_total=None):
        """下载放大后的图片（在后台线程中执行）"""
        def download_in_background():
            """后台下载函数"""
            try:
                import requests
                import os
                
                # 获取原始文件信息
                original_dir = os.path.dirname(original_path)
                original_name, original_ext = os.path.splitext(os.path.basename(original_path))
                
                # 生成高清图片文件名
                rename_rule = self.upscale_config.get('rename_rule', '（高清）')
                upscaled_filename = f"{original_name}{rename_rule}{original_ext}"
                upscaled_path = os.path.join(original_dir, upscaled_filename)
                
                # 确保目录存在
                os.makedirs(original_dir, exist_ok=True)
                
                # 下载图片
                print(f"开始下载高清图片: {download_url}")
                response = requests.get(download_url, stream=True, timeout=30)
                response.raise_for_status()
                
                # 保存图片
                with open(upscaled_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                # 验证文件完整性并添加到搜索结果
                if os.path.exists(upscaled_path) and os.path.getsize(upscaled_path) > 0:
                    print(f"高清图片下载完成: {upscaled_path}")
                    
                    # 更新下载计数器
                    if not hasattr(self, '_current_download_count'):
                        self._current_download_count = 0
                    self._current_download_count += 1
                    
                    # 如果有下载总数，更新进度显示
                    if download_total and download_total > 0:
                        self.root.after(0, lambda:
                            self.update_upscale_progress(
                                90,  # 保持在90%
                                f"下载完成... ({self._current_download_count}/{download_total})",
                                stage="下载图片",
                                upload_count=total_count if 'total_count' in locals() else download_total,
                                process_count=total_count if 'total_count' in locals() else download_total,
                                download_count=self._current_download_count,
                                total_count=total_count if 'total_count' in locals() else download_total
                            ))
                    
                    # 使用验证方法确保文件完整性后再添加到搜索结果
                    self._verify_and_add_upscaled_image(upscaled_path)
                else:
                    print(f"下载的文件无效: {upscaled_path}")
                    # 使用全局标记防止重复弹窗
                    if not hasattr(self, '_download_error_shown'):
                        self._download_error_shown = True
                        self.root.after(0, lambda: 
                                      messagebox.showerror("下载失败", f"下载图片失败，请检查网络连接或重试"))
                    
            except Exception as e:
                print(f"下载高清图片时出错: {e}")
                # 使用全局标记防止重复弹窗
                if not hasattr(self, '_download_error_shown'):
                    self._download_error_shown = True
                    self.root.after(0, lambda: messagebox.showerror("下载错误", f"下载高清图片时出错: {e}"))
        
        # 在后台线程中执行下载操作，避免阻塞主线程
        import threading
        download_thread = threading.Thread(target=download_in_background, daemon=True)
        download_thread.start()
    
    def _verify_and_add_upscaled_image(self, save_path):
        """验证文件完整性后添加到搜索结果（在后台线程中执行）"""
        def verify_file():
            """验证文件并添加到搜索结果"""
            try:
                import os
                import time
                
                # 等待文件系统完成写入（减少延迟）
                time.sleep(0.1)
                
                # 检查文件是否存在
                if not os.path.exists(save_path):
                    print(f"文件不存在，等待写入完成: {save_path}")
                    time.sleep(0.3)
                    if not os.path.exists(save_path):
                        print(f"文件仍不存在: {save_path}")
                        return
                
                # 检查文件大小是否合理（至少1KB）
                file_size = os.path.getsize(save_path)
                if file_size < 1024:
                    print(f"文件大小异常，等待下载完成: {save_path}, 当前大小: {file_size}")
                    time.sleep(0.3)
                    file_size = os.path.getsize(save_path)
                    if file_size < 1024:
                        print(f"文件大小仍然异常: {save_path}, 大小: {file_size}")
                        return
                
                # 尝试打开图片验证完整性
                try:
                    from PIL import Image
                    with Image.open(save_path) as img:
                        # 验证图片可以正常读取
                        img.verify()
                    print(f"文件完整性验证通过: {save_path}")
                except Exception as img_error:
                    print(f"图片文件损坏，等待重新下载: {save_path}, 错误: {img_error}")
                    time.sleep(0.3)
                    try:
                        with Image.open(save_path) as img:
                            img.verify()
                        print(f"重新验证成功: {save_path}")
                    except Exception:
                        print(f"图片文件仍然损坏: {save_path}")
                        return
                
                # 文件验证通过，在主线程中安全地添加到搜索结果
                if hasattr(self, 'root'):
                    self.root.after(0, lambda: self.add_upscaled_to_results(save_path))
                    
            except Exception as e:
                print(f"文件验证过程出错: {save_path}, 错误: {e}")
        
        # 在后台线程中进行文件验证，避免阻塞主线程
        import threading
        threading.Thread(target=verify_file, daemon=True).start()
    
    def update_checkbox_state(self, index, selected):
        """更新指定索引的复选框状态"""
        try:
            # 遍历所有图片项，找到对应索引的复选框
            for widget in self.scrollable_frame.winfo_children():
                if hasattr(widget, 'image_index') and widget.image_index == index:
                    # 找到对应的复选框
                    checkbox = self.find_checkbox_in_widget(widget)
                    if checkbox:
                        # 获取复选框的变量并设置状态
                        var = checkbox.cget('variable')
                        if var:
                            checkbox.tk.globalsetvar(var, selected)
                        break
        except Exception as e:
            print(f"更新复选框状态时出错: {e}")

    def add_upscaled_to_results(self, upscaled_path):
        """将高清图片添加到搜索结果中并自动勾选"""
        try:
            # 检查文件是否存在
            if not os.path.exists(upscaled_path):
                return
            
            # 检查是否已经在搜索结果中
            if upscaled_path in self.search_results:
                # 如果已存在，直接勾选
                index = self.search_results.index(upscaled_path)
                self.selected_images.add(upscaled_path)  # 使用图片路径而不是索引
                # 找到对应的复选框并设置状态
                self.update_checkbox_state(index, True)
                self.update_select_all_button_state()
                return
            
            # 添加到搜索结果列表
            self.search_results.append(upscaled_path)
            new_index = len(self.search_results) - 1
            
            # 自动勾选新添加的图片
            self.selected_images.add(upscaled_path)  # 使用图片路径而不是索引
            
            # 查找并取消对应原图的勾选状态
            self.unselect_original_image(upscaled_path)
            
            # 创建新的图片项并显示（预选中状态）
            self.create_image_item(new_index, upscaled_path, pre_selected=True)
            
            # 更新统计信息
            self.update_stats()
            
            # 更新全选按钮状态
            self.update_select_all_button_state()
            
            print(f"高清图片已添加到搜索结果并自动勾选: {upscaled_path}")
            
            # 高清处理完成后自动触发隐藏功能
            self.auto_hide_after_upscale(upscaled_path)
            
        except Exception as e:
            print(f"添加高清图片到搜索结果时出错: {e}")
    
    def auto_hide_after_upscale(self, upscaled_path):
        """高清处理完成后自动隐藏功能"""
        try:
            # 获取当前的重命名规则
            rename_rule = self.upscale_config.get('rename_rule', '（高清）')
            
            # 从重命名规则中提取关键词，去除括号
            keyword_to_add = rename_rule.strip().replace('（', '').replace('）', '').replace('(', '').replace(')', '')
            
            if keyword_to_add:
                print(f"高清处理完成，检查是否需要添加关键词: {keyword_to_add}")
                
                # 获取当前隐藏输入框的内容
                current_content = ""
                if hasattr(self, 'hide_entry'):
                    current_content = self.hide_entry.get().strip()
                
                # 检查当前内容是否已包含该关键词
                keywords_list = [kw.strip() for kw in current_content.split() if kw.strip()]
                
                if keyword_to_add not in keywords_list:
                    # 如果不包含，则添加到现有内容后面
                    if current_content:
                        new_content = f"{current_content} {keyword_to_add}"
                    else:
                        new_content = keyword_to_add
                    
                    # 更新输入框内容
                    if hasattr(self, 'hide_entry'):
                        self.hide_entry.delete(0, tk.END)
                        self.hide_entry.insert(0, new_content)
                    
                    print(f"已添加关键词 '{keyword_to_add}' 到隐藏条件")
                else:
                    print(f"关键词 '{keyword_to_add}' 已存在，无需添加")
                
                # 无论是否添加新关键词，都重新触发隐藏功能
                final_keywords = [kw.strip() for kw in self.hide_entry.get().split() if kw.strip()] if hasattr(self, 'hide_entry') else []
                if final_keywords:
                    self.hide_by_keyword_threaded(final_keywords)
                    print(f"已重新启动隐藏功能，使用关键词: {' '.join(final_keywords)}")
            else:
                print("重命名规则为空，跳过自动隐藏")
                
        except Exception as e:
            print(f"自动隐藏处理时出错: {e}")

    def unselect_original_image(self, upscaled_path):
        """取消高清图片对应原图的勾选状态"""
        try:
            # 获取高清图片的文件名
            upscaled_filename = os.path.basename(upscaled_path)
            upscaled_name, upscaled_ext = os.path.splitext(upscaled_filename)
            
            # 获取当前的重命名规则
            rename_rule = self.upscale_config.get('rename_rule', '（高清）')
            
            # 如果高清图片文件名包含重命名规则，去除它来获取原图文件名
            if rename_rule in upscaled_name:
                original_name = upscaled_name.replace(rename_rule, '')
                original_filename = original_name + upscaled_ext
                
                # 在搜索结果中查找对应的原图
                for i, image_path in enumerate(self.search_results):
                    if os.path.basename(image_path) == original_filename:
                        # 找到原图，取消其勾选状态
                        # 修复：使用索引而不是路径来移除选择，因为selected_images中存储的是索引
                        if i in self.selected_images:
                            self.selected_images.discard(i)  # 使用索引移除
                            # 更新复选框状态
                            self.update_checkbox_state(i, False)
                            print(f"已取消原图勾选: {original_filename} (索引: {i})")
                        # 同时检查是否有路径形式的选择（兼容性处理）
                        elif image_path in self.selected_images:
                            self.selected_images.discard(image_path)  # 移除路径形式
                            # 更新复选框状态
                            self.update_checkbox_state(i, False)
                            print(f"已取消原图勾选(路径形式): {original_filename}")
                        break
                        
        except Exception as e:
            print(f"取消原图勾选时出错: {e}")
    
    def calculate_optimal_upscale_factor(self, original_width, original_height, target_width=None, target_height=None):
        """
        计算最佳放大倍数
        
        Args:
            original_width: 原始宽度
            original_height: 原始高度
            target_width: 目标宽度（可选）
            target_height: 目标高度（可选）
            
        Returns:
            int: 最佳放大倍数 (2, 4, 8, 16)
        """
        try:
            # 使用配置的目标尺寸或传入的参数
            if target_width is None:
                target_width = self.smart_upscale_config.get('target_width', 8000)
            if target_height is None:
                target_height = self.smart_upscale_config.get('target_height', 8000)
            
            # 如果图片已经达到或超过目标尺寸，返回1（不需要放大）
            if original_width >= target_width and original_height >= target_height:
                return 1
            
            # 计算需要的放大倍数
            width_factor = target_width / original_width
            height_factor = target_height / original_height
            
            # 取较大的倍数确保两个维度都能达到目标
            required_factor = max(width_factor, height_factor)
            
            # 选择最接近且不小于required_factor的标准倍数
            standard_factors = [2, 4, 8, 16]
            for factor in standard_factors:
                if factor >= required_factor:
                    return factor
            
            # 如果需要的倍数超过16，返回16
            return 16
            
        except Exception as e:
            print(f"计算最佳放大倍数时出错: {e}")
            return 4  # 默认返回4倍
    
    def extract_sku_prefix(self, sku):
        """
        从SKU中提取前缀
        
        Args:
            sku: SKU字符串
            
        Returns:
            str: 匹配的前缀，如果没有匹配则返回空字符串
        """
        try:
            sku_str = str(sku).strip()
            if not sku_str:
                return ""
            
            # 加载处理配置以获取所有可能的前缀
            app_dir = get_app_directory()
            config_file = os.path.join(app_dir, 'processing_config.json')
            
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    processing_config = json.load(f)
                
                # 按前缀长度降序排列，优先匹配较长的前缀
                prefixes = sorted(processing_config.keys(), key=len, reverse=True)
                
                for prefix in prefixes:
                    # 使用正则表达式进行“全词前缀匹配”
                    # 匹配规则：SKU 必须以 prefix 开头，且 prefix 后面不能紧跟字母（防止 F 误匹配 FL）
                    # 但允许后面跟数字（如 FL02057）或非字母数字字符（如 FL-001）
                    pattern = rf"^{re.escape(prefix)}(?![A-Z])"
                    if re.match(pattern, sku_str, re.IGNORECASE):
                        return prefix
            
            return ""
            
        except Exception as e:
            print(f"提取SKU前缀时出错: {e}")
            return ""

    def analyze_sku_prefixes(self, data):
        """
        分析SKU前缀并加载处理配置
        
        Args:
            data: DataFrame或SKU列表
            
        Returns:
            tuple: (analysis_result, processing_config)
        """
        try:
            # 加载处理配置
            app_dir = get_app_directory()
            config_file = os.path.join(app_dir, 'processing_config.json')
            processing_config = {}
            
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    processing_config = json.load(f)
            
            # 分析SKU前缀
            analysis_result = {}
            
            # 如果是DataFrame，提取SKU列
            if hasattr(data, 'columns'):
                if hasattr(self, 'sku_column') and self.sku_column in data.columns:
                    sku_list = data[self.sku_column].dropna().astype(str).tolist()
                else:
                    sku_list = []
            else:
                # 如果是列表，直接使用
                sku_list = data if isinstance(data, list) else []
            
            # 统计每个前缀的出现次数
            prefix_counts = {}
            # 按前缀长度降序排列，优先匹配较长的前缀 (防止短前缀如'F'误匹配'FL')
            sorted_prefixes = sorted(processing_config.keys(), key=len, reverse=True)
            
            for sku in sku_list:
                sku_str = str(sku).strip()
                if sku_str:
                    # 查找匹配的前缀
                    for prefix in sorted_prefixes:
                        # 使用正则表达式进行“全词前缀匹配”
                        # 匹配规则：SKU 必须以 prefix 开头，且 prefix 后面不能紧跟字母（防止 F 误匹配 FL）
                        # 但允许后面跟数字（如 FL02057）或非字母数字字符（如 FL-001）
                        pattern = rf"^{re.escape(prefix)}(?![A-Z])"
                        if re.match(pattern, sku_str, re.IGNORECASE):
                            prefix_counts[prefix] = prefix_counts.get(prefix, 0) + 1
                            break
            
            analysis_result = {
                'prefix_counts': prefix_counts,
                'total_skus': len(sku_list),
                'matched_skus': sum(prefix_counts.values())
            }
            
            return analysis_result, processing_config
            
        except Exception as e:
            print(f"分析SKU前缀时出错: {e}")
            return {}, {}

    def analyze_upscale_batch(self, selected_images):
        """
        分析批量高清处理计划
        
        Args:
            selected_images: 选中的图片索引列表
            
        Returns:
            dict: 包含统计信息和处理计划的字典
        """
        try:
            statistics = {
                'total_images': len(selected_images),
                'to_process': 0,
                'qualified': 0
            }
            
            processing_list = []
            
            target_width = self.smart_upscale_config.get('target_width', 8000)
            target_height = self.smart_upscale_config.get('target_height', 8000)
            skip_qualified = self.smart_upscale_config.get('skip_qualified', True)
            
            for idx in selected_images:
                if idx >= len(self.search_results):
                    continue
                    
                image_path = self.search_results[idx]
                filename = os.path.basename(image_path)
                
                try:
                    # 获取图片尺寸
                    with Image.open(image_path) as img:
                        original_width, original_height = img.size
                    
                    # 计算最佳放大倍数
                    scale_factor = self.calculate_optimal_upscale_factor(
                        original_width, original_height, target_width, target_height
                    )
                    
                    # 计算放大后的尺寸
                    target_w = original_width * scale_factor
                    target_h = original_height * scale_factor
                    
                    # 判断是否需要处理
                    if scale_factor == 1 and skip_qualified:
                        statistics['qualified'] += 1
                        continue
                    else:
                        statistics['to_process'] += 1
                        
                        processing_list.append({
                            'filename': filename,
                            'path': image_path,  # 修改为path字段，与处理代码保持一致
                            'original_width': original_width,
                            'original_height': original_height,
                            'target_width': target_w,
                            'target_height': target_h,
                            'factor': scale_factor  # 修改为factor字段，与处理代码保持一致
                        })
                        
                except Exception as e:
                    print(f"分析图片 {filename} 时出错: {e}")
                    continue
            
            return {
                'statistics': statistics,
                'processing_list': processing_list
            }
            
        except Exception as e:
            print(f"分析批量处理计划时出错: {e}")
            return {
                'statistics': {'total_images': 0, 'to_process': 0, 'qualified': 0},
                'processing_list': []
            }
    
    def update_stats(self):
        """更新统计信息"""
        try:
            total_count = len(self.search_results)
            selected_count = len(self.selected_images)
            
            # 更新统计标签
            if hasattr(self, 'stats_label'):
                stats_text = f"共找到 {total_count} 个文件，已选择 {selected_count} 个"
                self.stats_label.config(text=stats_text)
        except Exception as e:
            print(f"更新统计信息时出错: {e}")
    
    def show_auto_dismiss_message(self, message, duration=3000):
        """显示自动消失的提示信息"""
        try:
            # 创建一个临时的提示标签
            if hasattr(self, 'temp_message_label'):
                self.temp_message_label.destroy()
            
            # 在状态栏位置显示提示信息
            self.temp_message_label = tk.Label(
                self.root, 
                text=message,
                bg='#28a745',  # 绿色背景
                fg='white',
                font=('Microsoft YaHei UI', 10, 'bold'),
                pady=8
            )
            self.temp_message_label.pack(side=tk.BOTTOM, fill=tk.X)
            
            # 设置自动消失
            self.root.after(duration, self._hide_auto_dismiss_message)
            
        except Exception as e:
            print(f"显示自动消失提示时出错: {e}")
    
    def update_folder_name_date(self, folder_name):
        """
        检测文件夹名称中的日期并更新为今天的日期
        支持格式：MM-DD 或 MM-dd
        """
        import datetime
        
        # 获取今天的日期
        today = datetime.date.today()
        today_month = today.strftime("%m")
        today_day = today.strftime("%d")
        
        # 匹配日期模式 MM-DD 或 MM-dd
        date_pattern = r'(\d{2})-(\d{2})'
        
        def replace_date(match):
            # 返回今天的日期格式
            return f"{today_month}-{today_day}"
        
        # 替换匹配到的日期
        updated_name = re.sub(date_pattern, replace_date, folder_name)
        
        return updated_name

    def check_and_update_folder_date(self):
        """
        智能检查并更新文件夹名称中的日期
        只在日期发生变化时才更新，避免重复操作
        """
        import datetime
        
        # 获取今天的日期
        today = datetime.date.today()
        
        # 检查是否需要更新（日期是否发生变化）
        if self.last_date_check == today:
            return  # 今天已经检查过了，无需重复更新
        
        # 获取当前文件夹名称
        current_folder_name = self.folder_name_var.get().strip()
        
        # 如果文件夹名称为空，不进行处理
        if not current_folder_name:
            self.last_date_check = today
            return
        
        # 检查文件夹名称中是否包含日期模式
        date_pattern = r'(\d{2})-(\d{2})'
        if re.search(date_pattern, current_folder_name):
            # 更新文件夹名称中的日期
            updated_name = self.update_folder_name_date(current_folder_name)
            
            # 如果日期确实发生了变化，更新输入框
            if updated_name != current_folder_name:
                # 应用智能名称处理 - 检查今日重复并自动添加序号
                updated_name = self.get_smart_name(updated_name, 'image_packages')
                self.folder_name_var.set(updated_name)
        
        # 记录本次检查的日期
        self.last_date_check = today

    def create_upscale_progress_window(self, total_images):
        """创建高清处理进度窗口"""
        self.upscale_progress_window = tk.Toplevel(self.root)
        self.upscale_progress_window.title("高清处理进度")
        self.upscale_progress_window.geometry("580x320")  # 增加窗口尺寸以容纳更多信息
        self.upscale_progress_window.resizable(False, False)
        self.upscale_progress_window.transient(self.root)
        
        # 初始化期间忽略Configure事件，避免错误位置被保存
        self._upscale_positioning = True
        
        # 绑定主窗口的点击事件，用于取消进度条置顶
        self.root.bind('<Button-1>', self.on_main_window_click, add=True)
        self.root.bind('<FocusIn>', self.on_main_window_focus, add=True)
        
        # 绑定窗口移动事件，用于保存位置（初始化期间会被忽略）
        self.upscale_progress_window.bind('<Configure>', self.on_upscale_progress_window_move)
        
        # 主框架
        main_frame = ttk.Frame(self.upscale_progress_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="🚀 高清处理进行中", 
                               font=('Microsoft YaHei UI', 14, 'bold'))
        title_label.pack(pady=(0, 15))
        
        # 处理信息框架
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        # 基本信息
        self.upscale_info_label = ttk.Label(info_frame, 
                                           text=f"正在处理 {total_images} 张图片，请耐心等待...",
                                           font=('Microsoft YaHei UI', 10))
        self.upscale_info_label.pack()
        
        # 当前阶段信息
        self.upscale_stage_label = ttk.Label(info_frame, 
                                            text="当前阶段：准备上传",
                                            font=('Microsoft YaHei UI', 9),
                                            foreground="#0066CC")
        self.upscale_stage_label.pack(pady=(5, 0))
        
        # 进度条框架
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 进度条容器，用于居中显示
        progress_container = ttk.Frame(progress_frame)
        progress_container.pack(anchor=tk.CENTER)
        
        # 进度条
        self.upscale_progress_var = tk.DoubleVar()
        self.upscale_progress_bar = ttk.Progressbar(progress_container, 
                                                   variable=self.upscale_progress_var,
                                                   maximum=100,
                                                   length=480,  # 增加进度条长度
                                                   style="TProgressbar")
        self.upscale_progress_bar.pack()
        
        # 进度百分比容器，用于居中显示
        percent_container = ttk.Frame(progress_frame)
        percent_container.pack(anchor=tk.CENTER, pady=(5, 0))
        
        # 进度百分比
        self.upscale_progress_percent = ttk.Label(percent_container, 
                                                 text="0%",
                                                 font=('Microsoft YaHei UI', 9, 'bold'),
                                                 foreground="#0066CC")
        self.upscale_progress_percent.pack()
        
        # 详细状态信息
        self.upscale_progress_text = ttk.Label(main_frame, 
                                              text="准备开始...",
                                              font=('Microsoft YaHei UI', 9),
                                              foreground="#666666")
        self.upscale_progress_text.pack(pady=(0, 10))
        
        # 任务详情框架
        details_frame = ttk.LabelFrame(main_frame, text="任务详情", padding="10")
        details_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # 任务统计信息
        self.upscale_stats_label = ttk.Label(details_frame, 
                                            text=f"上传：0/{total_images} | 处理：0/{total_images} | 下载：0/{total_images}",
                                            font=('Microsoft YaHei UI', 9))
        self.upscale_stats_label.pack()
        
        # 预计剩余时间
        self.upscale_time_label = ttk.Label(details_frame, 
                                           text="预计剩余时间：计算中...",
                                           font=('Microsoft YaHei UI', 9),
                                           foreground="#666666")
        self.upscale_time_label.pack(pady=(5, 0))
        
        # 初始化时间记录
        self.upscale_start_time = time.time()
        
        # 相对于主程序窗口居中显示 - 在创建完所有组件后再设置位置
        self.upscale_progress_window.update_idletasks()
        
        # 尝试加载保存的窗口位置
        saved_position = self.load_upscale_progress_window_position()
        
        if saved_position:
            # 使用保存的位置
            x, y = saved_position
            progress_width = 580
            progress_height = 320
            
            # 确保窗口不会超出屏幕边界
            screen_width = self.upscale_progress_window.winfo_screenwidth()
            screen_height = self.upscale_progress_window.winfo_screenheight()
            
            if x < 0:
                x = 0
            elif x + progress_width > screen_width:
                x = screen_width - progress_width
                
            if y < 0:
                y = 0
            elif y + progress_height > screen_height:
                y = screen_height - progress_height
            
            # 设置窗口位置
            self.upscale_progress_window.geometry(f"{progress_width}x{progress_height}+{x}+{y}")
        else:
            # 使用默认的居中位置
            # 强制更新窗口以获取准确的尺寸信息
            self.upscale_progress_window.update_idletasks()
            
            # 获取屏幕尺寸
            screen_width = self.upscale_progress_window.winfo_screenwidth()
            screen_height = self.upscale_progress_window.winfo_screenheight()
            
            # 使用固定的窗口尺寸
            progress_width = 580
            progress_height = 320
            
            # 计算屏幕居中位置
            x = (screen_width - progress_width) // 2
            y = (screen_height - progress_height) // 2
            
            # 确保窗口不会超出屏幕边界
            if x < 0:
                x = 0
            elif x + progress_width > screen_width:
                x = screen_width - progress_width
                
            if y < 0:
                y = 0
            elif y + progress_height > screen_height:
                y = screen_height - progress_height
            
            # 设置窗口位置
            self.upscale_progress_window.geometry(f"{progress_width}x{progress_height}+{x}+{y}")
            
            print(f"设置进度窗口位置: {x}, {y} (屏幕尺寸: {screen_width}x{screen_height})")
        
        # 强制更新窗口位置
        self.upscale_progress_window.update()
        
        # 移除进度条窗口的强制置顶，避免遮挡用户使用其他程序
        # self.upscale_progress_window.attributes('-topmost', True)
        # self.progress_topmost_active = True
        
        # 初始化完成后，短暂延迟再允许处理移动事件
        try:
            self.root.after(300, lambda: setattr(self, '_upscale_positioning', False))
        except Exception as e:
            print(f"取消初始化定位标志时出错: {e}")
        
        # 添加关闭事件处理
        self.upscale_progress_window.protocol("WM_DELETE_WINDOW", self.on_upscale_progress_close)
    
    def create_actual_progress_window(self):
        """创建实际的进度窗口"""
        self.actual_progress_window = tk.Toplevel(self.root)
        self.actual_progress_window.title("🚀 高清处理进度")
        self.actual_progress_window.geometry("520x220")
        self.actual_progress_window.resizable(False, False)
        self.actual_progress_window.transient(self.root)
        
        # 居中显示
        self.actual_progress_window.update_idletasks()
        x = (self.actual_progress_window.winfo_screenwidth() // 2) - (self.actual_progress_window.winfo_width() // 2)
        y = (self.actual_progress_window.winfo_screenheight() // 2) - (self.actual_progress_window.winfo_height() // 2)
        self.actual_progress_window.geometry(f"+{x}+{y}")
        
        # 主框架
        main_frame = ttk.Frame(self.actual_progress_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(main_frame, text="🚀 高清处理进行中", 
                               font=('Microsoft YaHei UI', 14, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 处理信息
        total_images = len(getattr(self, 'pending_processing_list', []))
        info_label = ttk.Label(main_frame, 
                              text=f"正在处理 {total_images} 张图片，请耐心等待...",
                              font=('Microsoft YaHei UI', 10))
        info_label.pack(pady=(0, 15))
        
        # 进度条
        self.upscale_progress_var = tk.DoubleVar()
        self.upscale_progress_bar = ttk.Progressbar(main_frame, 
                                                   variable=self.upscale_progress_var,
                                                   maximum=100,
                                                   length=420,
                                                   style="TProgressbar")
        self.upscale_progress_bar.pack(pady=(0, 10))
        
        # 进度文本
        self.upscale_progress_text = ttk.Label(main_frame, 
                                              text="准备开始...",
                                              font=('Microsoft YaHei UI', 9),
                                              foreground="#666666")
        self.upscale_progress_text.pack()
        
        # 添加关闭确认对话框
        self.actual_progress_window.protocol("WM_DELETE_WINDOW", self.on_upscale_progress_close)
    
    def update_upscale_progress(self, progress, text, stage="", upload_count=0, process_count=0, download_count=0, total_count=0, error_message=None):
        """更新高清处理进度"""
        try:
            if hasattr(self, 'upscale_progress_window') and self.upscale_progress_window and self.upscale_progress_window.winfo_exists():
                # 更新进度条
                if hasattr(self, 'upscale_progress_var'):
                    self.upscale_progress_var.set(progress)
                
                # 更新进度百分比
                if hasattr(self, 'upscale_progress_percent'):
                    self.upscale_progress_percent.config(text=f"{progress:.1f}%")
                
                # 更新详细状态文本
                if hasattr(self, 'upscale_progress_text'):
                    if error_message:
                        # 如果有错误信息，显示红色错误文本
                        self.upscale_progress_text.config(text=f"❌ 错误：{error_message}", foreground="#CC0000")
                        # 更新阶段为错误状态
                        if hasattr(self, 'upscale_stage_label'):
                            self.upscale_stage_label.config(text="当前阶段：处理失败", foreground="#CC0000")
                    else:
                        # 正常状态，显示灰色文本
                        self.upscale_progress_text.config(text=text, foreground="#666666")
                        # 更新当前阶段
                        if stage and hasattr(self, 'upscale_stage_label'):
                            self.upscale_stage_label.config(text=f"当前阶段：{stage}", foreground="#0066CC")
                
                # 更新任务统计
                if total_count > 0 and hasattr(self, 'upscale_stats_label'):
                    stats_text = f"上传：{upload_count}/{total_count} | 处理：{process_count}/{total_count} | 下载：{download_count}/{total_count}"
                    self.upscale_stats_label.config(text=stats_text)
                
                # 计算并更新预计剩余时间
                if hasattr(self, 'upscale_start_time') and hasattr(self, 'upscale_time_label') and progress > 0:
                    elapsed_time = time.time() - self.upscale_start_time
                    if progress < 100 and not error_message:
                        estimated_total_time = elapsed_time * (100 / progress)
                        remaining_time = estimated_total_time - elapsed_time
                        
                        if remaining_time > 60:
                            time_text = f"预计剩余时间：约 {int(remaining_time // 60)} 分 {int(remaining_time % 60)} 秒"
                        else:
                            time_text = f"预计剩余时间：约 {int(remaining_time)} 秒"
                    elif error_message:
                        time_text = "处理已停止"
                    else:
                        time_text = "处理完成！"
                    
                    self.upscale_time_label.config(text=time_text)

                # 立即刷新进度窗口，确保细粒度下载信息及时显示
                try:
                    self.upscale_progress_window.update_idletasks()
                except Exception:
                    pass
                
                self.upscale_progress_window.update()
        except Exception as e:
            print(f"更新进度时出错: {str(e)}")
            # 如果更新进度出错，不影响主流程继续执行

    def close_upscale_progress(self):
        """关闭高清处理进度窗口"""
        try:
            if hasattr(self, 'upscale_progress_window') and self.upscale_progress_window and self.upscale_progress_window.winfo_exists():
                self.upscale_progress_window.destroy()
                self.upscale_progress_window = None
            # 清理事件绑定和状态标记
            if hasattr(self, 'progress_topmost_active'):
                self.progress_topmost_active = False
        except Exception as e:
            print(f"关闭进度窗口时出错: {str(e)}")
            # 强制设置为None，防止后续访问出错
            if hasattr(self, 'upscale_progress_window'):
                self.upscale_progress_window = None
            if hasattr(self, 'progress_topmost_active'):
                self.progress_topmost_active = False
    
    def auto_select_all_after_upscale(self):
        """批量高清处理完成后自动全选所有可见图片"""
        try:
            # 调用现有的全选功能
            self.select_all_visible()
            print("批量高清处理完成，已自动全选所有可见图片")
        except Exception as e:
            print(f"自动全选图片时出错: {str(e)}")
    
    def on_upscale_progress_close(self):
        """处理进度窗口关闭事件"""
        # 直接关闭，不中断处理
        self.close_upscale_progress()
    
    def on_main_window_click(self, event):
        """主窗口点击事件处理"""
        # 移除进度条置顶相关逻辑，因为进度条已不再强制置顶
        pass
    
    def on_main_window_focus(self, event):
        """主窗口获得焦点事件处理"""
        # 移除进度条置顶相关逻辑，因为进度条已不再强制置顶
        pass
    
    def on_upscale_progress_window_move(self, event):
        """处理高清处理进度窗口移动事件，保存窗口位置"""
        try:
            # 初始化定位阶段忽略一切Configure事件
            if getattr(self, '_upscale_positioning', False):
                # 调试输出
                try:
                    gx, gy = self.upscale_progress_window.winfo_x(), self.upscale_progress_window.winfo_y()
                    print(f"[忽略初始化Configure] 当前几何: x={gx}, y={gy}, 事件: width={event.width}, height={event.height}")
                except Exception:
                    pass
                return
            
            # 只在窗口移动时保存位置，忽略其他Configure事件
            if (event.widget == self.upscale_progress_window and 
                hasattr(self, 'upscale_progress_window') and 
                self.upscale_progress_window and 
                self.upscale_progress_window.winfo_exists()):
                
                # 延迟保存，避免频繁保存
                if hasattr(self, '_save_position_after_id'):
                    self.root.after_cancel(self._save_position_after_id)
                
                # 调试输出
                gx, gy = self.upscale_progress_window.winfo_x(), self.upscale_progress_window.winfo_y()
                print(f"[移动事件] 计划保存位置: x={gx}, y={gy}")
                
                self._save_position_after_id = self.root.after(500, self.save_upscale_progress_window_position)
        except Exception as e:
            print(f"处理进度窗口移动事件时出错: {str(e)}")
    
    def save_upscale_progress_window_position(self):
        """保存高清处理进度窗口位置"""
        try:
            if (hasattr(self, 'upscale_progress_window') and 
                self.upscale_progress_window and 
                self.upscale_progress_window.winfo_exists()):
                
                x = self.upscale_progress_window.winfo_x()
                y = self.upscale_progress_window.winfo_y()
                
                # 获取当前配置
                config = self.config_manager.load_config('main', {})
                
                # 保存窗口位置
                config['upscale_progress_window_position'] = {'x': x, 'y': y}
                
                # 保存配置
                self.config_manager.save_config('main', config, silent=True)
                
        except Exception as e:
            print(f"保存进度窗口位置时出错: {str(e)}")
    
    def load_upscale_progress_window_position(self):
        """加载高清处理进度窗口位置"""
        try:
            config = self.config_manager.load_config('main', {})
            position_data = config.get('upscale_progress_window_position')
            
            if position_data and 'x' in position_data and 'y' in position_data:
                return (position_data['x'], position_data['y'])
                
        except Exception as e:
            print(f"加载进度窗口位置时出错: {str(e)}")
        
        return None
    
    def process_upscale_with_progress(self):
        """使用进度条处理高清放大"""
        # 在后台线程中处理
        import threading
        thread = threading.Thread(target=process_smart_upscale, daemon=True)
        thread.start()
    
    def on_auto_search_toggle(self):
        """处理自动搜索开关变化事件"""
        try:
            self.auto_search_enabled = self.auto_search_var.get()
            self.save_config()
        except Exception as e:
            print(f"保存自动搜索配置时出错: {e}")

    def check_for_updates(self):
        """检查软件更新"""
        try:
            from update_module import check_for_updates
            check_for_updates(self.root)
        except Exception as e:
            messagebox.showerror("检查更新失败", f"无法检查更新: {e}")

    def get_name_matching_config(self):
        """获取名称匹配配置 - 从processing_config中提取product_name"""
        try:
            # 优先从当前界面获取配置
            config = {}
            if hasattr(self, 'current_processing_entries'):
                for entry_info in self.current_processing_entries:
                    prefix = entry_info['prefix_entry'].get().strip()
                    product_name = entry_info['product_name_entry'].get().strip()
                    if prefix and product_name:
                        config[prefix] = product_name
            
            # 如果界面没有配置，从processing_config.json文件加载
            if not config:
                app_dir = get_app_directory()
                config_file = os.path.join(app_dir, 'processing_config.json')
                if os.path.exists(config_file):
                    with open(config_file, 'r', encoding='utf-8') as f:
                        processing_config = json.load(f)
                        # 从processing_config中提取product_name
                        for prefix, info in processing_config.items():
                            if isinstance(info, dict) and 'product_name' in info:
                                product_name = info['product_name'].strip()
                                if product_name:
                                    config[prefix] = product_name
            
            print(f"🔍 获取到的名称匹配配置: {config}")
            return config
        except Exception as e:
            print(f"读取名称匹配配置失败：{str(e)}")
            return {}

    def apply_name_matching(self, folder_name, analysis_result, name_matching_config):
        """根据SKU前缀分析结果应用名称匹配 - 基于位置的灵活替换逻辑"""
        print(f"🔍 apply_name_matching 被调用:")
        print(f"  folder_name: {folder_name}")
        print(f"  analysis_result: {analysis_result}")
        print(f"  name_matching_config: {name_matching_config}")
        try:
            if not analysis_result or not name_matching_config:
                print("  ❌ 未配置名称匹配规则或分析结果为空")
                return folder_name, "未配置名称匹配规则"
            
            # 处理analysis_result可能是tuple的情况
            if isinstance(analysis_result, tuple) and len(analysis_result) == 2:
                # 如果是tuple，取第一个元素作为真正的分析结果
                actual_analysis_result = analysis_result[0]
                print(f"  📝 检测到tuple格式，提取分析结果: {actual_analysis_result}")
            else:
                actual_analysis_result = analysis_result
            
            # 获取所有前缀统计信息
            prefix_counts = actual_analysis_result.get('prefix_counts', {})
            print(f"  📊 前缀统计: {prefix_counts}")
            
            # 查找表格中存在的预设前缀
            found_prefixes = []
            for prefix in name_matching_config.keys():
                if prefix in prefix_counts and prefix_counts[prefix] > 0:
                    found_prefixes.append(prefix)
            
            print(f"  🎯 表格中发现的预设前缀: {found_prefixes}")
            
            if not found_prefixes:
                print("  ❌ 表格中未发现任何预设前缀")
                return folder_name, "表格中未发现任何预设前缀"
            
            # 获取对应的产品类型
            target_products = []
            for prefix in found_prefixes:
                product = name_matching_config.get(prefix)
                if product:
                    target_products.append(product)
            
            if not target_products:
                print("  ❌ 未找到对应的产品类型")
                return folder_name, "未找到对应的产品类型"
            
            # 去除重复的产品类型，只保留唯一值
            unique_products = list(set(target_products))
            print(f"  📋 去重前产品类型: {target_products}")
            print(f"  ✨ 去重后产品类型: {unique_products}")
            
            # 组合多个产品类型（用+连接）
            combined_product = "+".join(unique_products)
            print(f"  🎯 目标产品类型: {combined_product}")
            
            # 使用灵活的位置匹配进行产品类型替换
            import re
            
            print(f"  🔍 开始基于位置的灵活产品类型替换...")
            
            # 分析名称结构，寻找产品类型的位置
            modified_name, reason = self._flexible_name_replacement(folder_name, combined_product, found_prefixes)
            
            return modified_name, reason
                
        except Exception as e:
            print(f"应用名称匹配时出错：{str(e)}")
            return folder_name, f"名称匹配处理出错：{str(e)}"
    
    def _flexible_name_replacement(self, folder_name, target_product, found_prefixes):
        """灵活的名称替换逻辑，基于位置而非固定文本"""
        import re
        
        # 按"-"分割名称，分析结构
        parts = folder_name.split('-')
        print(f"  📋 名称分割结果: {parts}")
        
        if len(parts) < 3:
            # 如果分割后少于3部分，直接在末尾添加
            modified_name = f"{folder_name}-{target_product}"
            reason = f"根据表格中的预设前缀 {found_prefixes}，在文件夹名称末尾添加产品类型 '{target_product}'"
            print(f"  ✅ 添加产品类型: {modified_name}")
            return modified_name, reason
        
        # 尝试多种灵活匹配策略
        
        # 策略1: 查找倒数第二个位置（通常是产品类型的位置）
        if len(parts) >= 2:
            # 检查倒数第二个部分是否可能是产品类型
            potential_product_index = len(parts) - 2
            potential_product = parts[potential_product_index]
            
            # 如果倒数第二个部分看起来像产品类型（不是日期格式，不是纯数字，不是特殊后缀）
            special_suffixes = ["艺术家", "画家", "设计师"]  # 不应被替换的特殊后缀
            if (not re.match(r'^\d{2}-\d{2}$', potential_product) and 
                not potential_product.isdigit() and 
                potential_product not in special_suffixes):
                print(f"  🎯 在倒数第二个位置找到疑似产品类型: '{potential_product}'")
                
                if potential_product != target_product:
                    # 替换倒数第二个位置的内容
                    new_parts = parts.copy()
                    new_parts[potential_product_index] = target_product
                    modified_name = '-'.join(new_parts)
                    reason = f"根据表格中的预设前缀 {found_prefixes}，将位置 {potential_product_index + 1} 的产品类型从 '{potential_product}' 替换为 '{target_product}'"
                    print(f"  ✅ 替换产品类型: {modified_name}")
                    return modified_name, reason
                else:
                    reason = f"位置 {potential_product_index + 1} 的产品类型已经是目标类型 '{target_product}'（预设前缀：{found_prefixes}）"
                    print(f"  ℹ️ 产品类型已经是目标类型")
                    return folder_name, reason
        
        # 策略2: 查找包含已知产品类型关键词的位置
        known_products = ["床上三件套", "窗帘", "画", "SJT", "CL", "cft"]
        special_suffixes = ["艺术家", "画家", "设计师"]  # 不应被替换的特殊后缀
        for i, part in enumerate(parts):
            # 跳过特殊后缀
            if part in special_suffixes:
                print(f"  🚫 跳过特殊后缀: '{part}'")
                continue
                
            if any(product in part for product in known_products):
                print(f"  🎯 在位置 {i + 1} 找到包含已知产品类型的部分: '{part}'")
                
                if part != target_product:
                    new_parts = parts.copy()
                    new_parts[i] = target_product
                    modified_name = '-'.join(new_parts)
                    reason = f"根据表格中的预设前缀 {found_prefixes}，将位置 {i + 1} 的产品类型从 '{part}' 替换为 '{target_product}'"
                    print(f"  ✅ 替换产品类型: {modified_name}")
                    return modified_name, reason
                else:
                    reason = f"位置 {i + 1} 的产品类型已经是目标类型 '{target_product}'（预设前缀：{found_prefixes}）"
                    print(f"  ℹ️ 产品类型已经是目标类型")
                    return folder_name, reason
        
        # 策略3: 智能位置推断 - 查找最可能是产品类型的位置
        # 通常产品类型不会是：日期、纯数字、单个字母、过短的字符串、特殊后缀
        special_suffixes = ["艺术家", "画家", "设计师"]  # 不应被替换的特殊后缀
        for i, part in enumerate(parts):
            # 跳过明显不是产品类型的部分
            if (re.match(r'^\d{2}-\d{2}$', part) or  # 日期格式
                part.isdigit() or  # 纯数字
                len(part) <= 2 or  # 过短
                re.match(r'^[A-Z]{1,3}\d*$', part) or  # 像代码的格式
                part in special_suffixes):  # 特殊后缀
                continue
            
            # 如果这个位置看起来合理，且不是第一个位置（通常是前缀）
            if i > 0:
                print(f"  🎯 推断位置 {i + 1} 可能是产品类型: '{part}'")
                
                if part != target_product:
                    new_parts = parts.copy()
                    new_parts[i] = target_product
                    modified_name = '-'.join(new_parts)
                    reason = f"根据表格中的预设前缀 {found_prefixes}，智能推断并将位置 {i + 1} 的内容从 '{part}' 替换为产品类型 '{target_product}'"
                    print(f"  ✅ 智能替换产品类型: {modified_name}")
                    return modified_name, reason
        
        # 策略4: 如果所有策略都失败，在末尾添加
        modified_name = f"{folder_name}-{target_product}"
        reason = f"根据表格中的预设前缀 {found_prefixes}，无法确定产品类型位置，在文件夹名称末尾添加产品类型 '{target_product}'"
        print(f"  ✅ 末尾添加产品类型: {modified_name}")
        return modified_name, reason

    def unified_size_extraction(self, text):
        """统一的尺寸提取方法，避免多重逻辑产生不一致结果"""
        try:
            if not text:
                return []
            
            text = str(text).strip()
            print(f"🔧 [统一提取] 开始处理文本: {text}")
            
            # 检查是否包含多个SKUID
            skuid_matches = re.findall(r'SKUID:', text)
            
            if len(skuid_matches) > 1:
                # 多SKUID情况：按SKUID分割并分别处理
                print(f"📦 检测到多SKUID ({len(skuid_matches)}个)，分别处理")
                sku_blocks = re.split(r'(?=SKUID:)', text)
                sku_blocks = [block.strip() for block in sku_blocks if block.strip()]
                
                size_list = []
                for i, block in enumerate(sku_blocks):
                    print(f"🔍 处理SKUID块 {i+1}: '{block[:50]}...'")
                    extracted_size = self._extract_single_size(block)
                    size_list.append(extracted_size if extracted_size else '')
                    print(f"✅ SKUID {i+1} 提取结果: {extracted_size}")
                
                return size_list
            else:
                # 单SKUID情况
                print(f"📦 检测到单SKUID，直接处理")
                extracted_size = self._extract_single_size(text)
                return [extracted_size] if extracted_size else []
                
        except Exception as e:
            print(f"❌ [统一提取] 处理时出错: {e}")
            return []
    
    def _extract_single_size(self, text):
        """从单个文本块中提取尺寸信息"""
        try:
            # 方法1: 优先使用标准化方法
            standardized = self.normalize_size_to_standard(text)
            if standardized and standardized != text:
                # 标准化成功且有实际转换
                standardized = standardized.replace(' x ', '*').replace('×', '*')
                print(f"🎯 [方法1] 标准化成功: {standardized}")
                return standardized
            
            # 方法2: 斜杠格式解析
            slash_size_match = re.search(r'/\s*(\d+(?:\.\d+)?\s*(?:inch|inches|cm|\'\')\s*[x*×\s]+\s*\d+(?:\.\d+)?\s*(?:inch|inches|cm|\'\')\s*(?:[x*×\s]+\s*\d+)?)', text, re.IGNORECASE)
            if slash_size_match:
                size_raw = slash_size_match.group(1).strip()
                size_normalized = re.sub(r'\s+', '', size_raw)
                size_normalized = re.sub(r'inches?', 'inch', size_normalized, flags=re.IGNORECASE)
                size_normalized = re.sub(r'\'\'', 'inch', size_normalized)
                size_normalized = re.sub(r'[x×]', '*', size_normalized)
                print(f"🎯 [方法2] 斜杠格式成功: {size_normalized}")
                return size_normalized
            
            # 方法3: extract_variants_content方法
            variants_result = self.extract_variants_content(text)
            if variants_result:
                first_line = variants_result.split('\n')[0] if '\n' in variants_result else variants_result
                print(f"🎯 [方法3] variants提取成功: {first_line}")
                return first_line
            
            # 方法4: 如果标准化返回了原文本，说明无法识别，返回空
            print(f"❌ [统一提取] 所有方法都失败")
            return None
            
        except Exception as e:
            print(f"❌ [单个提取] 处理时出错: {e}")
            return None

def main():
    """主函数"""
    # 修复Windows下multiprocessing程序自动多开的问题
    mp.freeze_support()
    
    # 设置工作目录为应用程序所在目录，确保相对路径文件能正确找到
    app_dir = get_app_directory()
    os.chdir(app_dir)
    print(f"[DEBUG] 工作目录已设置为: {app_dir}")
    
    root = tkdnd.TkinterDnD.Tk()
    app = ImageOrganizerApp(root)
    
    # 设置窗口图标（可选）
    try:
        root.iconbitmap(default='logo.ico')
    except FileNotFoundError:
        print("图标文件 logo.ico 未找到")
    except Exception as e:
        print(f"设置主窗口图标时出错: {e}")
        
    root.mainloop()

if __name__ == "__main__":
    main()