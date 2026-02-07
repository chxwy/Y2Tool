#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import os

def check_merge_result():
    # 查找输出文件
    output_file = None
    for file in os.listdir('.'):
        if file.startswith('急采CHX-10-22-Y2尊祐-画-艺术家') and file.endswith('.xlsx'):
            output_file = file
            break
    
    if not output_file:
        print("未找到输出文件")
        return
    
    print(f"检查输出文件: {output_file}")
    
    try:
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active
        
        print("\n=== 检查D列合并结果 ===")
        
        # 检查所有行的D列值
        for row in range(1, worksheet.max_row + 1):
            d_cell = worksheet.cell(row=row, column=4)  # D列
            i_cell = worksheet.cell(row=row, column=9)  # I列 (GroupID)
            
            print(f"行{row}: D列='{d_cell.value}', I列='{i_cell.value}'")
            
            # 检查是否有合并的单元格
            if hasattr(d_cell, 'coordinate'):
                for merged_range in worksheet.merged_cells.ranges:
                    if d_cell.coordinate in merged_range:
                        print(f"  -> D列在合并范围: {merged_range}")
                        break
        
        print(f"\n=== 合并范围信息 ===")
        for merged_range in worksheet.merged_cells.ranges:
            print(f"合并范围: {merged_range}")
            
        workbook.close()
        
    except Exception as e:
        print(f"检查文件时出错: {e}")

if __name__ == "__main__":
    check_merge_result()