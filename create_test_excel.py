#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

print("开始创建测试Excel文件...")

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "测试数据"

# 设置表头
headers = ["序号", "尺寸", "加工方式", "订单号", "材质", "数量", "SKU"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

print("表头设置完成")

# 添加测试数据 - 包含多SKU的行
test_data = [
    [1, "90CM*30CM", "正常加工", "TEST001", "磨毛布", 2, "SKU001\nSKU002"],  # 多SKU行
    [2, "80CM*25CM", "正常加工", "TEST002", "纯棉", 1, "SKU003"],  # 单SKU行
]

for row_idx, row_data in enumerate(test_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

print("测试数据添加完成")

# 保存文件
filename = "test_multi_sku.xlsx"
wb.save(filename)
print(f"✅ 测试文件已创建: {filename}")
print("文件包含:")
print("- 第2行: 多SKU数据 (SKU001\\nSKU002)")
print("- 第3行: 单SKU数据 (SKU003)")